"""
Evaluate the chatbot against the 10,000-row training dataset,
store responses, measure accuracy, and retrain the classifiers.
"""
import os
import sys
import json
import time
import re
from collections import Counter, defaultdict

# Add parent dir to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
os.chdir(os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

XLSX_PATH = "/Users/abdulrhmanbakr/Downloads/brightsmile_training_10000.xlsx"
OUTPUT_PATH = "/Users/abdulrhmanbakr/Downloads/brightsmile_training_10000_evaluated.xlsx"

# ── Step 1: Load classifiers directly (no Flask server needed) ──
print("Loading classifiers...")
import intent_classifier
import sklearn_classifier

# Map training file intents → our classifier intent names
# The training file uses: booking, availability, doctor_info, treatment_question, etc.
# Our intent_classifier uses: booking, availability_question, question, lead_capture, cancel, greeting, farewell
# Our sklearn_classifier uses: book_appointment, check_availability, services_info, pricing_info, etc.

# Both classifiers now use the same 19 intent names — direct match
ALL_19_INTENTS = [
    "booking", "availability", "doctor_info", "treatment_question", "emergency",
    "greeting", "farewell", "cancellation", "pricing_insurance", "lead_capture",
    "waitlist", "promotions", "loyalty", "clinic_info", "pre_visit_form",
    "recall", "symptom_question", "human_handoff", "complaint",
]

# Classic classifier: exact match expected
INTENT_MAP_CLASSIC = {intent: [intent] for intent in ALL_19_INTENTS}

# Sklearn classifier: exact match expected (retrained on same 19 intents)
INTENT_MAP_SKLEARN = {intent: [intent] for intent in ALL_19_INTENTS}

# Also check action-based accuracy using keyword/pattern matching on the response
ACTION_PATTERNS = {
    "start_booking_flow": [r"book", r"appointment", r"full name", r"specialty", r"what type"],
    "show_availability": [r"available", r"slot", r"time", r"schedule", r"open"],
    "show_doctor_comparison": [r"doctor", r"dr\.", r"specialist", r"comparison"],
    "provide_education": [r"root canal|veneer|whitening|implant|brace|filling|crown|extraction|cleaning|procedure|treatment"],
    "activate_emergency_flow": [r"emergency|urgent|immediate|rinse|pain relief|call|911"],
    "send_greeting": [r"hello|welcome|hi|how can i help|assist"],
    "send_farewell": [r"thank|bye|goodbye|see you|take care|wonderful day"],
    "cancel_or_reschedule": [r"cancel|reschedule|sorry to hear"],
    "provide_pricing_info": [r"cost|price|sar|fee|payment|insurance|\$|\d+"],
    "capture_lead": [r"name|number|phone|contact|follow up|reach"],
    "manage_waitlist": [r"waitlist|waiting list|notify|spot opens"],
    "handle_promotion": [r"discount|offer|promo|code|special|off|%"],
    "show_loyalty_info": [r"point|loyalty|reward|balance|redeem"],
    "provide_clinic_info": [r"hour|open|close|location|address|sunday|monday|saturday|friday"],
    "handle_form_query": [r"form|fill|pre-visit|before your|prepare"],
    "handle_recall": [r"remind|recall|checkup|6 month|come back|follow.up"],
    "provide_symptom_advice": [r"pain|cavity|crack|root|gum|swollen|bleeding|sensitive|hurt"],
    "initiate_handoff": [r"connect|transfer|human|team|staff|hold|moment"],
    "handle_complaint": [r"sorry|apologize|feedback|improve|concern|experience"],
}


def evaluate_action(expected_action, response):
    """Check if the chatbot response matches the expected action via keyword patterns."""
    if not expected_action or not response:
        return False
    response_lower = response.lower()
    patterns = ACTION_PATTERNS.get(expected_action, [])
    if not patterns:
        return True  # Unknown action, skip
    matches = sum(1 for p in patterns if re.search(p, response_lower))
    return matches >= 1  # At least 1 keyword pattern matches


print("=" * 70)
print("CHATBOT EVALUATION — 10,000 Training Questions")
print("=" * 70)

# ── Step 2: Load the Excel data ──
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active
total_rows = ws.max_row - 1  # Exclude header
print(f"Loaded {total_rows} questions from {XLSX_PATH}")

# Add output columns if not present
output_cols = {
    12: "Classic Intent",
    13: "Classic Conf",
    14: "Sklearn Intent",
    15: "Sklearn Conf",
    16: "Classic Match",
    17: "Sklearn Match",
    18: "Action Match",
}
for col, header in output_cols.items():
    ws.cell(1, col, header)
    ws.cell(1, col).font = Font(bold=True)

# ── Step 3: Run classification on every row ──
classic_correct = 0
sklearn_correct = 0
action_correct = 0
total = 0
errors_by_intent = defaultdict(list)
intent_stats = defaultdict(lambda: {"total": 0, "classic_ok": 0, "sklearn_ok": 0, "action_ok": 0})

print("\nRunning classification...")
start_time = time.time()

for row_idx in range(2, ws.max_row + 1):
    user_input = ws.cell(row_idx, 2).value
    expected_intent = ws.cell(row_idx, 3).value
    expected_action = ws.cell(row_idx, 9).value

    if not user_input or not expected_intent:
        continue

    total += 1
    user_input_str = str(user_input).strip()

    # Classify with both classifiers
    classic_intent, classic_conf = intent_classifier.classify(user_input_str)
    sk_intent, sk_conf = sklearn_classifier.classify(user_input_str)

    # Check correctness
    classic_ok = classic_intent in INTENT_MAP_CLASSIC.get(expected_intent, [])
    sklearn_ok = (sk_intent in INTENT_MAP_SKLEARN.get(expected_intent, [])) if sk_intent else False

    # Check if the classified intents would lead to the correct action
    ACTION_TO_INTENTS = {
        "start_booking_flow": ["booking"],
        "show_availability": ["availability"],
        "send_greeting": ["greeting"],
        "send_farewell": ["farewell"],
        "cancel_or_reschedule": ["cancellation"],
        "capture_lead": ["lead_capture"],
        "activate_emergency_flow": ["emergency"],
        "provide_education": ["treatment_question"],
        "provide_symptom_advice": ["symptom_question", "emergency"],
        "provide_pricing_info": ["pricing_insurance"],
        "provide_clinic_info": ["clinic_info"],
        "show_doctor_comparison": ["doctor_info"],
        "handle_promotion": ["promotions"],
        "show_loyalty_info": ["loyalty"],
        "handle_form_query": ["pre_visit_form"],
        "handle_recall": ["recall"],
        "initiate_handoff": ["human_handoff"],
        "handle_complaint": ["complaint"],
        "manage_waitlist": ["waitlist"],
    }
    action_ok = False
    valid_intents = ACTION_TO_INTENTS.get(expected_action, [])
    if valid_intents:
        action_ok = classic_intent in valid_intents or (sk_intent in valid_intents if sk_intent else False)

    if classic_ok:
        classic_correct += 1
    if sklearn_ok:
        sklearn_correct += 1
    if action_ok:
        action_correct += 1

    # Track per-intent stats
    intent_stats[expected_intent]["total"] += 1
    if classic_ok:
        intent_stats[expected_intent]["classic_ok"] += 1
    if sklearn_ok:
        intent_stats[expected_intent]["sklearn_ok"] += 1
    if action_ok:
        intent_stats[expected_intent]["action_ok"] += 1

    # Store misclassifications for analysis
    if not classic_ok and not sklearn_ok:
        errors_by_intent[expected_intent].append({
            "input": user_input_str[:80],
            "classic": classic_intent,
            "sklearn": sk_intent,
        })

    # Write results to Excel
    ws.cell(row_idx, 12, classic_intent)
    ws.cell(row_idx, 13, round(classic_conf, 3))
    ws.cell(row_idx, 14, sk_intent or "N/A")
    ws.cell(row_idx, 15, round(sk_conf, 3) if sk_conf else 0)
    ws.cell(row_idx, 16, "✓" if classic_ok else "✗")
    ws.cell(row_idx, 17, "✓" if sklearn_ok else "✗")
    ws.cell(row_idx, 18, "✓" if action_ok else "✗")

    # Color code
    green = PatternFill(start_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", fill_type="solid")
    ws.cell(row_idx, 16).fill = green if classic_ok else red
    ws.cell(row_idx, 17).fill = green if sklearn_ok else red
    ws.cell(row_idx, 18).fill = green if action_ok else red

    if total % 1000 == 0:
        elapsed = time.time() - start_time
        print(f"  Processed {total}/{total_rows} ({elapsed:.1f}s)")

elapsed = time.time() - start_time

# ── Step 4: Save results ──
wb.save(OUTPUT_PATH)
print(f"\nResults saved to: {OUTPUT_PATH}")

# ── Step 5: Print accuracy report ──
print("\n" + "=" * 70)
print("ACCURACY REPORT")
print("=" * 70)
print(f"Total questions evaluated: {total}")
print(f"Time: {elapsed:.1f}s")
print()
print(f"Classic Classifier Accuracy:  {classic_correct}/{total} = {100*classic_correct/total:.1f}%")
print(f"Sklearn Classifier Accuracy:  {sklearn_correct}/{total} = {100*sklearn_correct/total:.1f}%")
print(f"Action Routing Accuracy:      {action_correct}/{total} = {100*action_correct/total:.1f}%")

print("\n── Per-Intent Breakdown ──")
print(f"{'Intent':<22} {'Total':>6} {'Classic':>10} {'Sklearn':>10} {'Action':>10}")
print("-" * 62)
for intent in sorted(intent_stats.keys()):
    s = intent_stats[intent]
    t = s["total"]
    c_pct = 100 * s["classic_ok"] / t if t else 0
    sk_pct = 100 * s["sklearn_ok"] / t if t else 0
    a_pct = 100 * s["action_ok"] / t if t else 0
    print(f"{intent:<22} {t:>6} {c_pct:>9.1f}% {sk_pct:>9.1f}% {a_pct:>9.1f}%")

print("\n── Worst Misclassifications (samples) ──")
for intent in sorted(errors_by_intent.keys(), key=lambda k: -len(errors_by_intent[k])):
    errs = errors_by_intent[intent]
    print(f"\n{intent} ({len(errs)} both-miss):")
    for e in errs[:3]:
        print(f"  \"{e['input']}\" → classic={e['classic']}, sklearn={e['sklearn']}")

# ── Step 6: Save accuracy summary to JSON for retraining ──
summary = {
    "total": total,
    "classic_accuracy": round(100 * classic_correct / total, 2),
    "sklearn_accuracy": round(100 * sklearn_correct / total, 2),
    "action_accuracy": round(100 * action_correct / total, 2),
    "per_intent": {k: dict(v) for k, v in intent_stats.items()},
    "error_samples": {k: v[:5] for k, v in errors_by_intent.items()},
}
with open("evaluation_results.json", "w") as f:
    json.dump(summary, f, indent=2)
print(f"\nDetailed results saved to evaluation_results.json")

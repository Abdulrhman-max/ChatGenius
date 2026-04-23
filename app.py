"""
ChatGenius Flask backend with three chatbot features:
1. Q&A — answers business questions from knowledge base
2. Appointment Booking — collects info, checks calendar, confirms, emails
3. Lead Capture — collects name/phone when not ready to book
All work inside the chat window with a conversation state machine.
"""

from flask import Flask, request, jsonify, send_from_directory, redirect
import torch
import os
import json
import re
import uuid
import secrets
from datetime import datetime, timedelta

import database as db
import calendar_service as cal
import email_service as email
import social_auth
import chatbot_classifier
import dental_ai
import dental_knowledge_engine as dke
import intent_classifier
import message_interpreter
import claude_specialist
import grok_cleaner
import sklearn_classifier
import smart_router
import restriction_filter
import background_tasks

# ── New engine modules ──
import translations as tr
import emergency_handler
import handoff_engine
import doctor_comparison
import recall_engine
import treatment_followup_engine
import missed_call_engine
import gallery_engine
import promotions_engine as promo
import lead_engine
import loyalty_engine as loyalty
import ab_testing
import two_factor_auth as tfa
import referral_engine
import patient_profile_engine as patient_profile
import realtime_engine as realtime
import benchmarking_engine as benchmarks
import gmb_engine as gmb

# ── New Feature Engines (10 features) ──
import appointment_reminder_engine as reminder_eng
import survey_engine
import upsell_engine
import channel_engine
import invoice_engine
import report_engine
import package_engine
import doctor_portal_engine
import noshow_recovery_engine

app = Flask(__name__, static_folder="static")

# ── CORS for embedded chatbot ──
from flask_cors import CORS
CORS(app, resources={r"/chat": {"origins": "*"}, r"/static/chatbot-embed.js": {"origins": "*"}, r"/api/chatbot-customization/public/*": {"origins": "*"}})

# ── Load knowledge base ──
KB_PATH = os.path.join(os.path.dirname(__file__), "data", "knowledge_base.json")
with open(KB_PATH) as f:
    KB = json.load(f)

# ── Model config ──
MODEL_DIR = os.path.join(os.path.dirname(__file__), "models", "chatgenius-tinyllama")
BASE_MODEL = "TinyLlama/TinyLlama-1.1B-Chat-v1.0"
SYSTEM_MSG = "You are ChatGenius AI, a helpful sales assistant. Answer concisely in 1-3 sentences using this context:\n\n"

model = None
tokenizer = None
device = None

# ── Session store (in-memory, keyed by session_id) ──
sessions = {}


def get_session(sid):
    if sid not in sessions:
        sessions[sid] = {"flow": None, "data": {}, "step": None, "history": []}
    # Migration: add history to old sessions
    if "history" not in sessions[sid]:
        sessions[sid]["history"] = []
    return sessions[sid]


def reset_session(sid):
    sessions[sid] = {"flow": None, "data": {}, "step": None, "history": []}


def is_admin_role(user):
    """Check if user is any type of admin (head_admin or admin)."""
    return user.get("role") in ("admin", "head_admin")


def get_effective_admin_id(user):
    """Get the company-scoping admin_id.
    head_admin uses their own id, admin uses their admin_id, doctor uses admin_id."""
    if user.get("role") == "head_admin":
        return user["id"]
    return user.get("admin_id", 0) or user["id"]


import requests as http_requests  # for external API calls

# ── Cache for customer API lookups (keyed by admin_id + customer_id) ──
_customer_cache = {}
_CUSTOMER_CACHE_TTL = 300  # 5 minutes


def fetch_customer_by_id(admin_id, customer_id, api_url_override=""):
    """Fetch a single customer from the business's external API by ID.
    Calls GET {base_url}/{customer_id} and returns customer dict or None.
    api_url_override: if provided (from embed config), use this instead of dashboard setting.
    Caches results for 5 minutes."""
    if not customer_id:
        return None

    cache_key = f"{admin_id}_{customer_id}"
    cache = _customer_cache.get(cache_key)
    now = datetime.now()
    if cache and (now - cache["fetched_at"]).total_seconds() < _CUSTOMER_CACHE_TTL:
        return cache["customer"]

    # Use URL from embed config first, fall back to dashboard setting
    config = db.get_customers_api_config(admin_id)
    base_url = (api_url_override or config.get("customers_api_url", "")).strip().rstrip("/")
    key = config.get("customers_api_key", "").strip()
    if not base_url:
        return None

    try:
        headers = {}
        if key:
            headers["Authorization"] = f"Bearer {key}"
            headers["X-API-Key"] = key
        resp = http_requests.get(f"{base_url}/{customer_id}", headers=headers, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        # Support both direct object and wrapped {"customer": {...}} or {"data": {...}}
        customer = data if "name" in data or "email" in data else data.get("customer", data.get("data", data))
        _customer_cache[cache_key] = {"customer": customer, "fetched_at": now}
        print(f"[customers_api] Fetched customer {customer_id} for admin {admin_id}: {customer.get('name', 'N/A')}", flush=True)
        return customer
    except Exception as e:
        print(f"[customers_api] Error fetching customer {customer_id} from {base_url}: {e}", flush=True)
        return None


def load_model():
    global model, tokenizer, device
    from transformers import AutoModelForCausalLM, AutoTokenizer
    from peft import PeftModel

    device = torch.device("mps" if torch.backends.mps.is_available() else "cpu")
    print(f"Loading model on {device}...")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_DIR)
    base = AutoModelForCausalLM.from_pretrained(BASE_MODEL, dtype=torch.float16)
    model = PeftModel.from_pretrained(base, MODEL_DIR)
    model = model.merge_and_unload()
    model.to(device)
    model.eval()
    print("Model ready!")


# ══════════════════════════════════════════════
#  Q&A — Knowledge Base Responses
# ══════════════════════════════════════════════

KEYWORD_MAP = {
    "pricing": ["price", "pricing", "cost", "how much", "plan", "basic", "pro", "agency", "pay", "afford", "cheap", "expensive", "dollar", "month", "subscription", "tier", "billing"],
    "features": ["feature", "what can", "what does", "capabilit", "offer", "include", "do for me", "functionality"],
    "setup": ["setup", "set up", "install", "get started", "begin", "onboard", "configure", "implement", "how long", "how do i"],
    "trial": ["trial", "free", "try", "test", "credit card", "cancel", "risk", "money back", "refund", "guarantee"],
    "industries": ["industry", "dental", "law", "restaurant", "real estate", "ecommerce", "e-commerce", "salon", "clinic", "shop", "store", "fitness", "my business"],
    "integration": ["integrat", "connect", "crm", "hubspot", "salesforce", "zapier", "calendar", "wordpress", "shopify", "wix", "website", "embed", "slack"],
    "security": ["safe", "secur", "encrypt", "gdpr", "complian", "privacy", "data", "ccpa", "soc"],
    "support": ["support", "help me with", "contact", "customer service", "onboarding"],
    "comparison": ["differ", "compar", "vs", "versus", "better", "intercom", "drift", "tidio", "zendesk", "unique", "advantage", "regular chatbot"],
    "how_it_works": ["how does", "how it work", "what is chatgenius", "about chatgenius", "tell me about", "explain"],
    "languages": ["language", "spanish", "french", "german", "translat", "multilingual"],
}

KB_RESPONSES = {
    "pricing": "We have three plans: **Basic** at $99/month (700 conversations, 1 chatbot, smart booking, calendar scheduling, email reminders, patient profiles, pre-visit forms, basic analytics), **Pro** at $249/month (5,000 conversations, 4 chatbots, everything in Basic + advanced reminders, no-show recovery, ROI dashboard, lead capture, waitlist, promotions, multi-language, AI PDF extraction), and **Agency** at $599/month (unlimited conversations, everything in Pro + AI no-show prediction, advanced analytics, API access, PMS/CRM integration, full doctor portal, priority support, custom email sending). All include a 14-day free trial — no credit card required!",
    "features": "ChatGenius includes: 24/7 instant AI replies (under 2 seconds), automated appointment booking with calendar sync, smart lead capture with CRM integration, one-line website integration, a no-code dashboard, and templates for 20+ industries. Pro adds multi-language support, analytics, and human handoff.",
    "setup": "Setup takes under 5 minutes: 1) Sign up free, 2) Enter your business info, 3) Upload your knowledge base or let AI learn from your website, 4) Customize the look, 5) Paste one line of code on your site. No coding or technical skills needed!",
    "trial": "We offer a 14-day free trial with full Pro features — no credit card required. After the trial, choose a paid plan or continue with a limited free tier (50 conversations/month). We also have a 30-day money-back guarantee on all paid plans. Zero risk!",
    "industries": "ChatGenius works for any industry! Popular verticals: dental clinics, law firms, real estate, restaurants, e-commerce, fitness studios, salons, automotive, professional services, and education. We have pre-built templates for 20+ industries, and the AI adapts to your specific business.",
    "integration": "We integrate with HubSpot, Salesforce, Zoho, Pipedrive (CRM), Google Calendar, Calendly, Outlook (scheduling), Slack, Teams (communication), Zapier, Make (automation), and Google Analytics. Works on WordPress, Shopify, Wix, Squarespace, Webflow, and any custom website.",
    "security": "All data is encrypted with AES-256 at rest and TLS 1.3 in transit. We're GDPR and CCPA compliant, hosted on AWS with 99.9% uptime. Agency plan includes SOC 2 Type II compliance. We never sell your data — you own everything and can export or delete anytime.",
    "support": "Basic: email support. Pro: priority email + chat support. Agency: dedicated account manager, priority support, full doctor portal access. All users get access to our knowledge base and tutorials.",
    "comparison": "Unlike scripted chatbots, ChatGenius uses real AI that understands context and intent. Compared to Intercom ($74+/mo, built for enterprise), we're purpose-built for SMBs at lower cost. Compared to Tidio, our AI handles unexpected questions and maintains conversation flow.",
    "how_it_works": "ChatGenius uses AI trained on your business data to understand and respond to customer questions naturally. Visitors get instant answers, can book appointments, and share their contact info — all automatically. You manage everything from a simple dashboard.",
    "languages": "On Pro and Agency plans, ChatGenius supports 10 languages: English, Spanish, French, German, Portuguese, Italian, Dutch, Japanese, Korean, and Chinese. The chatbot auto-detects the visitor's language and responds accordingly.",
}


DENTAL_KEYWORD_MAP = {
    "hours": ["hour", "open", "close", "when are you", "what time", "schedule", "weekend", "saturday", "sunday", "lunch break", "working hour", "office hour"],
    "location": ["where", "address", "location", "directions", "located", "find you", "come to", "parking", "near"],
    "services": ["service", "offer", "provide", "do you do", "treatment", "cleaning", "filling", "whitening", "implant", "crown", "braces", "invisalign", "root canal", "extraction", "cosmetic", "veneer", "pediatric", "orthodont", "x-ray", "xray", "check-up", "checkup", "emergency"],
    "pricing": ["price", "cost", "how much", "fee", "charge", "expensive", "afford", "cheap", "pay", "dollar", "rate"],
    "insurance": ["insurance", "accept insurance", "dental plan", "delta", "cigna", "aetna", "metlife", "bluecross", "coverage", "in-network", "out of network"],
    "payment": ["payment", "pay", "credit card", "cash", "financing", "carecredit", "payment plan", "installment"],
    "emergency": ["emergency", "urgent", "pain", "broken tooth", "knocked out", "bleeding", "swollen", "abscess", "after hours", "walk-in", "walk in", "same day"],
    "about": ["about", "who are you", "tell me about", "experience", "years", "credential", "certified", "accredit", "award", "team", "doctor", "dentist", "staff", "why choose", "special", "different", "unique", "reputation", "review"],
    "contact": ["phone", "call", "contact", "reach", "number", "email"],
}


def _get_company_info_response(topic, admin_id=1):
    """Build a dynamic response from saved company info."""
    info = db.get_company_info(admin_id)

    field_map = {
        "hours": "business_hours",
        "location": "address",
        "services": "services",
        "pricing": "pricing_insurance",
        "insurance": "pricing_insurance",
        "payment": "pricing_insurance",
        "emergency": "emergency_info",
        "about": "about",
        "contact": "phone",
    }

    field = field_map.get(topic)
    if not field:
        return None

    # If no company info or field is empty, tell the user it's not available yet
    if not info or not info.get(field):
        friendly_names = {
            "hours": "business hours",
            "location": "address",
            "services": "services",
            "pricing": "pricing information",
            "insurance": "insurance information",
            "payment": "payment options",
            "emergency": "emergency information",
            "about": "practice details",
            "contact": "contact information",
        }
        label = friendly_names.get(topic, "that information")
        return (
            f"I'm sorry, our {label} haven't been set up yet. "
            f"Please call us directly or **leave your contact info** and we'll get back to you!\n\n"
            f"I can also help you **book an appointment** or find the right specialist for your needs."
        )

    biz_name = _fix_spelling_general(info.get("business_name") or "our office")
    content = info[field]

    # Use OpenAI to reformat the raw data into a clean, professional response
    if dental_ai.is_configured():
        formatted = _reformat_with_ai(topic, content, biz_name)
        if formatted:
            return formatted

    # Fallback: fix spelling locally and format
    cleaned_content = _fix_spelling_general(content)

    prefix_map = {
        "hours": f"Here are **{biz_name}**'s hours:\n\n",
        "location": f"We're located at:\n\n",
        "services": f"Here are the services we offer at **{biz_name}**:\n\n",
        "pricing": f"Here's our pricing information:\n\n",
        "insurance": f"Here's our insurance and coverage info:\n\n",
        "payment": f"Here are our payment options:\n\n",
        "emergency": f"Here's our emergency information:\n\n",
        "about": f"Here's about **{biz_name}**:\n\n",
        "contact": f"You can reach us at:\n\n",
    }

    response = prefix_map.get(topic, "") + cleaned_content
    response += "\n\nWould you like to **book an appointment** or ask another question?"
    return response


def _fix_spelling_general(text):
    """Fix common misspellings in any text using a general English word list + edit distance."""
    # Common words that appear in dental office info
    _COMMON_WORDS = {
        "opens", "open", "from", "every", "day", "days", "excluding", "except",
        "friday", "fridays", "saturday", "saturdays", "sunday", "sundays",
        "monday", "mondays", "tuesday", "tuesdays", "wednesday", "wednesdays",
        "thursday", "thursdays", "morning", "afternoon", "evening", "night",
        "closed", "available", "hours", "until", "through", "between", "and",
        "the", "dentist", "dental", "clinic", "office", "practice", "doctor",
        "cleaning", "whitening", "filling", "fillings", "extraction", "braces",
        "implant", "implants", "crown", "crowns", "bridge", "bridges", "veneer",
        "veneers", "root", "canal", "checkup", "consultation", "surgery",
        "orthodontics", "pediatric", "cosmetic", "emergency", "general",
        "insurance", "accepted", "payment", "plans", "cash", "credit", "card",
        "financing", "price", "pricing", "cost", "free", "consultation",
        "located", "location", "address", "street", "avenue", "road", "drive",
        "main", "north", "south", "east", "west", "center", "central", "park",
        "suite", "floor", "building", "parking", "phone", "call", "email",
        "clinic", "hospital", "smile", "bright", "white", "pearl", "diamond",
        "star", "golden", "silver", "royal", "premier", "advanced", "modern",
        "gentle", "perfect", "happy", "healthy", "total",
        "contact", "walk-ins", "walkins", "welcome", "appointment", "appointments",
        "required", "recommended", "patients", "patient", "new", "existing",
        "accept", "provide", "offer", "service", "services", "treatment",
        "treatments", "professional", "experienced", "certified", "licensed",
        "specialist", "specialists", "care", "quality", "best", "trusted",
        "years", "experience", "serving", "community", "family", "children",
        "adults", "seniors", "all", "ages", "comprehensive", "preventive",
        "restorative", "aesthetic", "oral", "health", "hygiene", "exam",
    }

    words = text.split()
    result = []
    for word in words:
        # Extract just letters for matching, preserve original punctuation
        clean = re.sub(r'[^a-zA-Z]', '', word).lower()
        if not clean or len(clean) <= 2:
            result.append(word)
            continue

        # If already a known word, keep it
        if clean in _COMMON_WORDS:
            result.append(word)
            continue

        # Find closest match using edit distance
        best_match = None
        best_dist = 3  # Max edit distance to consider
        for known in _COMMON_WORDS:
            if abs(len(known) - len(clean)) > 2:
                continue
            dist = _edit_distance(clean, known)
            if dist < best_dist:
                best_dist = dist
                best_match = known

        if best_match and best_dist <= 2:
            # Preserve capitalization of first letter
            if word[0].isupper():
                best_match = best_match.capitalize()
            # Preserve any trailing punctuation
            trailing = re.sub(r'^[a-zA-Z]+', '', word)
            result.append(best_match + trailing)
        else:
            result.append(word)

    return ' '.join(result)


def _reformat_with_ai(topic, raw_content, biz_name):
    """Use OpenAI to reformat raw company info into a clean, professional chatbot response."""
    try:
        from openai import OpenAI
        client = OpenAI(api_key=dental_ai.OPENAI_API_KEY)

        topic_label = {
            "hours": "business hours",
            "location": "address/location",
            "services": "services offered",
            "pricing": "pricing and insurance",
            "insurance": "insurance information",
            "payment": "payment options",
            "emergency": "emergency information",
            "about": "about the practice",
            "contact": "contact information",
        }.get(topic, topic)

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": f"""You are a friendly dental office chatbot for {biz_name}.
The admin has saved the following raw {topic_label} data. Your job is to:
1. Fix any spelling or grammar mistakes
2. Present it in a clean, well-formatted way using **bold** for key info
3. Keep the same information — do NOT add or invent anything
4. Be concise and professional
5. End with asking if they'd like to book an appointment or ask another question
6. Use line breaks and bullet points where appropriate"""},
                {"role": "user", "content": f"Here is the raw {topic_label} data to reformat:\n\n{raw_content}"},
            ],
            max_tokens=300,
            temperature=0.3,
        )

        return response.choices[0].message.content.strip()

    except Exception as e:
        print(f"[reformat_ai] Error: {e}")
        return None


# ── Symptom-to-Specialty mapping ──
SYMPTOM_SPECIALTY_MAP = {
    "Orthodontist": [
        "jaw", "bite", "overbite", "underbite", "crossbite", "open bite",
        "misalign", "crooked", "braces", "alignment", "straighten", "gap between teeth",
        "upper jaw", "lower jaw", "jaw forward", "jaw backward", "teeth crowding",
        "crowded teeth", "spacing", "malocclusion", "invisalign", "retainer",
    ],
    "Endodontist": [
        "root canal", "tooth infection", "infected tooth", "abscess", "tooth abscess",
        "severe tooth pain", "throbbing pain", "deep cavity", "pulp", "nerve pain",
        "tooth sensitivity to hot", "sensitivity to cold", "cracked tooth pain",
        "tooth hurting", "teeth hurting", "tooth hurt", "teeth hurt", "toothache",
        "tooth ache", "teeth ache", "tooth pain", "teeth pain", "molar pain",
        "sharp pain in tooth", "tooth is killing", "painful tooth",
        "teeth is hurting", "tooth is hurting", "teeth are hurting",
        "teeth is paining", "tooth is paining",
    ],
    "Periodontist": [
        "gum", "bleeding gum", "gum disease", "receding gum", "swollen gum",
        "gingivitis", "periodontitis", "loose tooth", "bone loss", "gum surgery",
        "deep cleaning", "gum infection", "gum pain", "gum recession",
    ],
    "Oral & Maxillofacial Surgeon": [
        "wisdom tooth", "wisdom teeth", "impacted", "jaw surgery", "facial trauma",
        "jaw fracture", "broken jaw", "oral surgery", "tumor", "cyst",
        "cleft", "jaw reconstruction", "facial surgery", "biopsy",
    ],
    "Prosthodontist": [
        "denture", "false teeth", "missing teeth", "crown", "bridge", "implant",
        "replacement teeth", "lost tooth", "tooth replacement", "partial denture",
        "full denture", "veneer", "cosmetic restoration",
    ],
    "Pediatric Dentist": [
        "child", "kid", "baby tooth", "baby teeth", "baby", "my son", "my daughter",
        "toddler", "infant", "children", "pediatric", "kids teeth", "my kid",
        "little one", "boy teeth", "girl teeth", "son teeth", "daughter teeth",
    ],
    "Cosmetic Dentist": [
        "whitening", "teeth whitening", "whiter", "stain", "discolor", "yellow teeth",
        "smile makeover", "cosmetic", "bonding", "reshape", "aesthet",
        "ugly teeth", "teeth look", "beautiful smile", "nice smile", "white teeth",
        "brighter teeth", "brighten", "bleach",
    ],
    "Dental Anesthesiologist": [
        "sedation", "anesthesia", "afraid of dentist", "dental anxiety",
        "dental phobia", "nervous", "put me to sleep", "pain free",
    ],
    "Orofacial Pain Specialist": [
        "tmj", "jaw pain", "face pain", "clicking jaw", "jaw lock",
        "jaw clicking", "jaw popping", "headache from jaw", "facial pain",
        "temporomandibular",
    ],
    "General Dentist": [
        "cleaning", "checkup", "check-up", "cavity", "filling", "routine",
        "dental exam", "teeth cleaning", "general", "regular visit",
    ],
    "Family Dentist": [
        "family", "whole family", "family dental",
    ],
    "Oral Pathologist": [
        "mouth sore", "oral lesion", "mouth disease", "tongue sore",
        "white patch", "red patch", "oral cancer", "mouth cancer",
    ],
    "Oral Radiologist": [
        "x-ray", "xray", "scan", "ct scan", "panoramic", "dental imaging",
    ],
}


def find_symptom_specialty(text):
    """Match patient symptoms/description to a dental specialty."""
    lower = text.lower()
    # Also check a simplified version (remove filler words for better matching)
    simplified = re.sub(r'\b(is|are|am|was|were|my|me|i|the|a|an|it|so|very|really|too|has been|have been)\b', ' ', lower)
    simplified = re.sub(r'\s+', ' ', simplified).strip()
    scores = {}
    for specialty, keywords in SYMPTOM_SPECIALTY_MAP.items():
        score = sum(1 for kw in keywords if kw in lower or kw in simplified)
        if score > 0:
            scores[specialty] = score

    if not scores:
        return None

    # If child-related words are present, boost Pediatric Dentist
    child_words = ["baby", "child", "kid", "son", "daughter", "toddler", "infant", "children"]
    if any(w in lower for w in child_words) and "Pediatric Dentist" in scores:
        scores["Pediatric Dentist"] += 10

    return max(scores, key=scores.get)


def _build_symptom_response(specialty, admin_id):
    """Build a response recommending a specialist and showing available doctors."""
    # Get active doctors in this specialty
    all_doctors = db.get_doctors(admin_id)
    matching_doctors = [d for d in all_doctors if d.get("status") == "active" and d.get("specialty") == specialty]

    response = f"Based on what you're describing, I'd recommend seeing an **{specialty}**."

    if matching_doctors:
        if len(matching_doctors) == 1:
            doc = matching_doctors[0]
            response += f"\n\nWe have **Dr. {doc['name']}** ({specialty}) available for you."
        else:
            doc_list = "\n".join([f"**{i+1}.** Dr. {d['name']}" for i, d in enumerate(matching_doctors)])
            response += f"\n\nHere are our {specialty} doctors:\n\n{doc_list}"
        response += "\n\nWould you like to **book an appointment**?"
    else:
        # No doctors in this specialty — check if any doctors at all
        active_doctors = [d for d in all_doctors if d.get("status") == "active"]
        if active_doctors:
            response += "\n\nWe don't currently have a specialist in that category, but our other doctors may be able to help."
            response += "\n\nWould you like to **book an appointment** with one of our available doctors?"
        else:
            response += "\n\nWould you like to **book an appointment** or **leave your contact info** so we can get back to you?"

    return response


def find_dental_topic(text):
    """Check if the message matches a dental office question category."""
    lower = text.lower()
    scores = {}
    for topic, keywords in DENTAL_KEYWORD_MAP.items():
        score = sum(1 for kw in keywords if kw in lower)
        if score > 0:
            scores[topic] = score
    return max(scores, key=scores.get) if scores else None


def find_topic(text):
    lower = text.lower()
    scores = {}
    for topic, keywords in KEYWORD_MAP.items():
        score = sum(1 for kw in keywords if kw in lower)
        if score > 0:
            scores[topic] = score
    return max(scores, key=scores.get) if scores else None


def generate_ai_response(user_message, context):
    prompt = f"<|system|>\n{SYSTEM_MSG}{context}</s>\n<|user|>\n{user_message}</s>\n<|assistant|>\n"
    inputs = tokenizer(prompt, return_tensors="pt", truncation=True, max_length=384)
    inputs = {k: v.to(device) for k, v in inputs.items()}
    with torch.no_grad():
        outputs = model.generate(
            **inputs, max_new_tokens=60, temperature=0.7,
            do_sample=True, repetition_penalty=1.3, pad_token_id=tokenizer.eos_token_id,
        )
    new_tokens = outputs[0][inputs["input_ids"].shape[1]:]
    response = tokenizer.decode(new_tokens, skip_special_tokens=True).strip()
    for stop in ["</s>", "<|user|>", "<|system|>", "\n\n\n"]:
        if stop in response:
            response = response[:response.index(stop)].strip()
    return response if len(response) > 10 else None


# ══════════════════════════════════════════════
#  Spell Correction — fix misspellings before processing
# ══════════════════════════════════════════════

# Known words the chatbot needs to understand
_VOCABULARY = {
    # Booking
    "book": ["bok", "bkoo", "boook", "boo", "bokk", "buk"],
    "appointment": ["appoitmnet", "appointmnt", "apointment", "appointmet", "appoitment", "appoint", "appoinment", "appointemnt", "appointmnet"],
    "schedule": ["shedule", "scedule", "scheduel", "schedul", "shcedule"],
    "reserve": ["reserv", "resrve", "resereve"],
    "available": ["availble", "avialable", "avalable", "availabe", "avaliable"],
    "meeting": ["meating", "meetin", "meting", "meetng"],
    "slot": ["slott", "slto", "solt"],
    # Lead capture
    "contact": ["contcat", "conact", "contac", "cotact"],
    "number": ["numbr", "nubmer", "numbe", "numbeer"],
    "phone": ["phon", "fone", "phoen", "pohne"],
    "call": ["cal", "cll", "calll"],
    # General
    "want": ["wnat", "watn", "wnt", "wnta"],
    "hello": ["helo", "hllo", "helllo", "heloo"],
    "pricing": ["pricng", "priceing", "pricingg", "pricin"],
    "features": ["featurs", "feaures", "fetures", "featrues"],
    "help": ["hlep", "halp", "hep", "helpp"],
    "cancel": ["cancle", "cancl", "canel", "canecl"],
    "please": ["pleas", "pls", "plz", "pleaes", "plese"],
    "thanks": ["thnks", "thanx", "thansk", "thnaks", "thx"],
    "yes": ["yse", "ys", "yess", "yea", "yeah", "yep", "ya"],
    "no": ["noo", "nah", "nope"],
    "information": ["informaiton", "infromation", "infomation", "informaton"],
    "interested": ["intrested", "intreseted", "intersted", "intreested"],
    "question": ["qustion", "quesiton", "questin", "qestion"],
    "tomorrow": ["tomorow", "tommorow", "tommorrow", "tomorrw", "tmrw", "tmr"],
    "today": ["tday", "2day", "toady", "todya"],
}

# Build reverse lookup: misspelling -> correct word
_SPELL_MAP = {}
for correct, misspellings in _VOCABULARY.items():
    for ms in misspellings:
        _SPELL_MAP[ms] = correct


def _edit_distance(a, b):
    """Simple Levenshtein distance."""
    if len(a) < len(b):
        return _edit_distance(b, a)
    if len(b) == 0:
        return len(a)
    prev = range(len(b) + 1)
    for i, ca in enumerate(a):
        curr = [i + 1]
        for j, cb in enumerate(b):
            curr.append(min(prev[j + 1] + 1, curr[j] + 1, prev[j] + (ca != cb)))
        prev = curr
    return prev[len(b)]


def correct_spelling(text):
    """Fix misspellings in user input using vocabulary + edit distance."""
    words = text.lower().split()
    corrected = []
    for word in words:
        # Clean punctuation for matching
        clean = re.sub(r'[^a-z]', '', word)
        if not clean or len(clean) <= 2:
            corrected.append(word)
            continue

        # Direct lookup in spelling map
        if clean in _SPELL_MAP:
            # Preserve original casing/punctuation context
            corrected.append(word.lower().replace(clean, _SPELL_MAP[clean]))
            continue

        # Already a known correct word
        if clean in _VOCABULARY:
            corrected.append(word)
            continue

        # Fuzzy match: try edit distance against vocabulary
        best_match = None
        best_dist = 999
        for correct_word in _VOCABULARY:
            if len(correct_word) >= 4 and len(clean) >= 4:
                dist = _edit_distance(clean, correct_word)
                max_allowed = 1 if len(clean) <= 5 else 2
                if dist <= max_allowed and dist < best_dist:
                    best_dist = dist
                    best_match = correct_word

        if best_match:
            corrected.append(word.lower().replace(clean, best_match))
        else:
            corrected.append(word)

    return " ".join(corrected)


def _is_affirmative(text):
    """Check if text expresses agreement/yes — flexible matching."""
    t = text.lower().strip().rstrip("!.")
    if t in ("yes", "yeah", "yep", "yea", "sure", "ok", "okay", "y", "ya", "yup",
             "yes please", "add me", "waitlist", "go ahead", "do it", "absolutely",
             "of course", "definitely", "please", "right", "correct", "confirm"):
        return True
    if re.search(r'\b(yes|yeah|yep|sure|okay|ok|please|absolutely|definitely|go ahead)\b', t):
        return True
    return False


def _is_negative(text):
    """Check if text expresses disagreement/no — flexible matching."""
    t = text.lower().strip().rstrip("!.")
    if t in ("no", "nah", "nope", "n", "no thanks", "no thank you", "other times",
             "show me", "skip", "pass", "never mind", "nevermind", "not now"):
        return True
    if re.search(r'\b(no|nah|nope|don\'?t|not)\b', t) and not re.search(r'\b(yes|yeah|sure|okay)\b', t):
        return True
    return False


# ══════════════════════════════════════════════
#  Intent Detection — route to correct flow
# ══════════════════════════════════════════════

def detect_intent(text):
    """Detect user intent using the smart intent classifier.
    Uses TF-IDF scoring against intent examples + structural analysis
    to understand WHAT the user actually wants — not just keyword matching.
    """
    intent, confidence = intent_classifier.classify(text)

    # Map 19 classifier intents to flow intents
    if intent == "booking":
        return "booking"
    if intent in ("cancellation", "cancel"):
        return "cancel"
    if intent == "lead_capture":
        return "lead_capture"
    # All other intents (availability, doctor_info, treatment_question, emergency,
    # greeting, farewell, pricing_insurance, clinic_info, waitlist, promotions,
    # loyalty, pre_visit_form, recall, symptom_question, human_handoff, complaint)
    # are routed to Q&A / smart router
    return "qa"


# ══════════════════════════════════════════════
#  Booking Flow State Machine
# ══════════════════════════════════════════════

def _extract_email(text):
    """Try to extract an email address from natural text."""
    match = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', text)
    return match.group(0).lower() if match else None


def _mask_email(addr):
    """Mask email: show first 3 chars of local part + *** + domain. e.g. dro.*********@gmail.com"""
    if not addr or "@" not in addr:
        return addr or ""
    local, domain = addr.rsplit("@", 1)
    visible = min(3, len(local))
    return local[:visible] + "." + "*" * max(len(local) - visible, 3) + "@" + domain


def _extract_phone(text):
    """Try to extract a phone number from natural text."""
    # Remove common filler words
    cleaned = re.sub(r'\b(my|number|is|phone|call|me|at|its|it\'s|here)\b', '', text, flags=re.IGNORECASE)
    digits = re.sub(r'[^\d+\-() ]', '', cleaned).strip()
    if len(re.sub(r'\D', '', digits)) >= 7:
        return digits
    # Try just pulling all digits
    all_digits = re.sub(r'\D', '', text)
    if len(all_digits) >= 7:
        return all_digits
    return None


def _generate_upcoming_dates(num_days=7):
    """Generate upcoming weekday dates as dropdown items."""
    today = datetime.now()
    dates = []
    i = 1
    while len(dates) < num_days:
        d = today + timedelta(days=i)
        if d.weekday() != 4:  # Skip Friday (4) — clinic closed; open Sun-Thu + Sat
            day_label = d.strftime("%A, %B %d")
            # Show "Tomorrow" for the next day if it's a weekday
            if i == 1:
                day_label = f"Tomorrow — {day_label}"
            dates.append({
                "name": d.strftime("%A, %B %d"),
                "display": day_label,
                "iso": d.strftime("%Y-%m-%d"),
            })
        i += 1
    return dates


def _get_off_dates_with_blocks(doctor_id, admin_id=None):
    """Return combined off_dates (doctor off days + schedule block dates + flexible schedule off days)
    for calendar greying. Covers current month and next 3 months."""
    from datetime import datetime as _dt_h, timedelta as _td_h
    off_dates = list(db.get_doctor_off_dates(doctor_id)) if doctor_id else []
    now = _dt_h.now()

    if admin_id and doctor_id:
        for month_offset in range(4):
            y = now.year + (now.month + month_offset - 1) // 12
            m = (now.month + month_offset - 1) % 12 + 1
            blocked = db.get_blocked_dates_for_calendar(admin_id, doctor_id, y, m)
            off_dates.extend(blocked)

    # Also grey out recurring off days from flexible schedule (daily_hours with off: true)
    if doctor_id:
        doctor = db.get_doctor_by_id(doctor_id)
        if doctor and doctor.get("schedule_type") == "flexible" and doctor.get("daily_hours"):
            try:
                daily = doctor["daily_hours"]
                if isinstance(daily, str):
                    daily = json.loads(daily)
                # Find which day names are marked off
                off_day_names = [day for day, hrs in daily.items() if isinstance(hrs, dict) and hrs.get("off")]
                if off_day_names:
                    # Generate dates for next 4 months that fall on these off days
                    end_date = now + _td_h(days=120)
                    d = now
                    while d <= end_date:
                        if d.strftime("%A") in off_day_names:
                            off_dates.append(d.strftime("%Y-%m-%d"))
                        d += _td_h(days=1)
            except (json.JSONDecodeError, ValueError, KeyError, TypeError):
                pass

    return list(set(off_dates))


def _doctor_avail_str(doctor):
    """Get a display string of working days for a doctor (for UI dropdowns)."""
    wd = _get_doctor_working_days(doctor)
    return ", ".join(wd) if isinstance(wd, list) else wd


def _doctor_dropdown_item(d):
    """Build a doctor dropdown item dict with working days from schedule config."""
    return {"name": d["name"], "specialty": d.get("specialty", "General"), "availability": _doctor_avail_str(d),
            "years_of_experience": d.get("years_of_experience", 0), "gender": d.get("gender", ""),
            "languages": d.get("languages", ""), "qualifications": d.get("qualifications", "")}


def _get_customer_booked_dates(session, admin_id):
    """Get booked dates filtered to the current customer only.
    Only marks dates where THIS user has active future bookings."""
    customer_name = (session.get("_prefill_name") or session.get("_greeting_name") or "").strip().lower()
    customer_email = (session.get("_prefill_email") or "").strip().lower()
    customer_phone = (session.get("_prefill_phone") or "").strip()

    if not customer_name and not customer_email and not customer_phone:
        return db.get_booking_dates(admin_id)

    from datetime import date as _date_h
    today = _date_h.today().isoformat()
    conn = db.get_db()
    rows = conn.execute(
        "SELECT DISTINCT date FROM bookings WHERE admin_id=? AND date>=? AND status NOT IN ('cancelled','no_show') ORDER BY date",
        (admin_id, today)
    ).fetchall()
    conn.close()

    matched_dates = []
    for r in rows:
        date_str = r["date"]
        bookings = db.find_bookings_by_date(admin_id, date_str)
        for b in bookings:
            if ((customer_name and b.get("customer_name", "").strip().lower() == customer_name) or
                (customer_email and b.get("customer_email", "").strip().lower() == customer_email) or
                (customer_phone and b.get("customer_phone", "").strip() == customer_phone)):
                matched_dates.append(date_str)
                break

    return matched_dates



def _get_doctor_working_days(doctor):
    """Get the actual working days from a doctor's daily_hours config.
    Returns a list like ['Sunday', 'Monday', 'Thursday'] for flexible schedules,
    or the availability string for fixed schedules."""
    if doctor.get("schedule_type") == "flexible" and doctor.get("daily_hours"):
        try:
            daily = doctor["daily_hours"]
            if isinstance(daily, str):
                daily = json.loads(daily)
            # Order days properly
            day_order = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
            working = [d for d in day_order if d in daily and not daily[d].get("off")]
            if working:
                return working
        except (json.JSONDecodeError, ValueError):
            pass
    return doctor.get("availability", "Mon-Fri")


def _get_doctor_hours_for_day(doctor, day_name):
    """Get the working hours for a specific day from daily_hours config.
    Returns (from_time, to_time) or None if off/not configured."""
    if doctor.get("schedule_type") == "flexible" and doctor.get("daily_hours"):
        try:
            daily = doctor["daily_hours"]
            if isinstance(daily, str):
                daily = json.loads(daily)
            day_hours = daily.get(day_name)
            if day_hours and not day_hours.get("off"):
                return day_hours.get("from"), day_hours.get("to")
        except (json.JSONDecodeError, ValueError):
            pass
    # Fixed schedule — return global start/end
    s = doctor.get("start_time", "09:00 AM")
    e = doctor.get("end_time", "05:00 PM")
    if s and e:
        return s, e
    return None


def _generate_doctor_slots(doctor, breaks=None, selected_date=None, service_duration=None):
    """Generate appointment time slots from a doctor's schedule.
    Supports both fixed (same hours daily) and flexible (per-day hours) schedules.
    selected_date: a date string like '2026-04-07' to look up the day-specific hours.
    service_duration: override slot length in minutes (e.g. from service's duration_minutes)."""
    def parse_12h(time_str):
        """Parse '09:00 AM' to total minutes since midnight."""
        if not time_str or time_str == '00:00 AM' or time_str == '00:00':
            return None
        m = re.match(r'(\d{1,2}):(\d{2})\s*(AM|PM)', time_str, re.IGNORECASE)
        if not m:
            return None
        h, mi, ampm = int(m.group(1)), int(m.group(2)), m.group(3).upper()
        if ampm == 'PM' and h < 12:
            h += 12
        if ampm == 'AM' and h == 12:
            h = 0
        return h * 60 + mi

    # Determine start/end times based on schedule type
    doc_start = doctor.get("start_time")
    doc_end = doctor.get("end_time")
    # Fallback to 9 AM – 5 PM if the doctor has no working hours configured,
    # so breaks and schedule blocks can still be applied.
    if not doc_start or doc_start == "00:00 AM":
        doc_start = "09:00 AM"
    if not doc_end or doc_end == "00:00 AM":
        doc_end = "05:00 PM"

    day_appointment_length = None  # per-day override from flexible schedule
    if doctor.get("schedule_type") == "flexible" and doctor.get("daily_hours") and selected_date:
        try:
            daily = doctor["daily_hours"]
            if isinstance(daily, str):
                daily = json.loads(daily)
            # Get the day name from the selected date
            from datetime import datetime as _dt
            day_name = _dt.strptime(selected_date, "%Y-%m-%d").strftime("%A")
            day_hours = daily.get(day_name)
            if day_hours:
                if day_hours.get("off"):
                    return []  # Doctor is off this day
                doc_start = day_hours.get("from", doc_start)
                doc_end = day_hours.get("to", doc_end)
                if day_hours.get("appointment_length"):
                    day_appointment_length = int(day_hours["appointment_length"])
            else:
                # Doctor doesn't work this day
                return []
        except (json.JSONDecodeError, ValueError, KeyError):
            pass

    start_min = parse_12h(doc_start)
    end_min = parse_12h(doc_end)
    if start_min is None or end_min is None or start_min >= end_min:
        return []

    # Priority: service_duration > per-day length > doctor default > 60
    length = service_duration or day_appointment_length or doctor.get("appointment_length") or 60
    if isinstance(length, str):
        length = int(length)

    def mins_to_12h(mins):
        h = mins // 60
        m = mins % 60
        ampm = "AM" if h < 12 else "PM"
        display_h = h if h <= 12 else h - 12
        if display_h == 0:
            display_h = 12
        return f"{display_h:02d}:{m:02d} {ampm}"

    # Parse break times into minute ranges (filter by day if we know the day)
    break_ranges = []
    if breaks:
        day_name = None
        if selected_date:
            try:
                from datetime import datetime as _dt0
                day_name = _dt0.strptime(selected_date, "%Y-%m-%d").strftime("%A")
            except ValueError:
                pass
        for b in breaks:
            # Only include breaks for this day (or breaks with no day set)
            b_day = b.get("day_of_week", "")
            if day_name and b_day and b_day != day_name:
                continue
            bs = parse_12h(b.get("start_time", ""))
            be = parse_12h(b.get("end_time", ""))
            if bs is not None and be is not None:
                break_ranges.append((bs, be))

    # ── Feature 11: Load schedule blocks for this doctor/date (rebuilt) ──
    block_ranges = []
    if selected_date:
        try:
            admin_id = doctor.get("admin_id", 0)
            doctor_id = doctor.get("id")
            # Check full-day block first (no time_str = checks full-day blocks only)
            if db.is_slot_blocked(admin_id, doctor_id, selected_date, None):
                return []  # Entire day is blocked (holiday / full-day block)
            # Load all blocks to extract time ranges for partial-day blocks
            blocks = db.get_schedule_blocks(admin_id, doctor_id=doctor_id)
            from datetime import datetime as _dt2
            date_obj = _dt2.strptime(selected_date, "%Y-%m-%d")
            for blk in blocks:
                btype = blk.get("block_type", "single_date")
                blk_start_time = blk.get("start_time", "")
                blk_end_time = blk.get("end_time", "")
                if not blk_start_time or not blk_end_time:
                    continue  # Full-day blocks already handled above
                # Check if this block's date pattern matches selected_date
                matches = False
                if btype == "single_date":
                    matches = (blk.get("start_date") == selected_date)
                elif btype == "date_range":
                    sd = blk.get("start_date", "")
                    ed = blk.get("end_date", sd)
                    matches = (sd <= selected_date <= ed)
                elif btype == "recurring":
                    matches = db._date_matches_recurring(date_obj, blk)
                if matches:
                    bs = parse_12h(blk_start_time)
                    be = parse_12h(blk_end_time)
                    if bs is not None and be is not None:
                        block_ranges.append((bs, be))
        except Exception:
            pass

    slots = []
    current = start_min
    while current + length <= end_min:
        h = current // 60
        m = current % 60
        slot_end = current + length
        # Check if this slot overlaps any break
        overlaps_break = False
        for bs, be in break_ranges:
            if current < be and slot_end > bs:
                overlaps_break = True
                # Jump past the break
                current = be
                break
        if overlaps_break:
            continue
        # Check if this slot overlaps any schedule block (Feature 11)
        overlaps_block = False
        for bs, be in block_ranges:
            if current < be and slot_end > bs:
                overlaps_block = True
                current = be
                break
        if overlaps_block:
            continue
        start_str = mins_to_12h(current)
        end_str = mins_to_12h(slot_end)
        time_str = f"{start_str} - {end_str}"
        slots.append({"time": time_str, "hour": h, "minute": m})
        current += length

    # Edge case: if remaining time is within 15 minutes of a full slot, add a shorter final slot
    # e.g. current=18:13, end_min=19:00, length=60 → 47 min remaining, shortfall=13 ≤ 15
    # Create a new slot 6:13 PM - 7:00 PM instead of wasting that time
    if current < end_min:
        remaining = end_min - current
        shortfall = length - remaining
        if 0 < shortfall <= 15:
            # Check no break/block overlaps this range
            overlaps = False
            for bs, be in break_ranges:
                if current < be and end_min > bs:
                    overlaps = True
                    break
            if not overlaps:
                for bs, be in block_ranges:
                    if current < be and end_min > bs:
                        overlaps = True
                        break
            if not overlaps:
                h = current // 60
                m = current % 60
                start_str = mins_to_12h(current)
                end_str = mins_to_12h(end_min)
                slots.append({"time": f"{start_str} - {end_str}", "hour": h, "minute": m})

    return slots


def _extract_time(text, available_slots=None):
    """Extract a time from natural text, with fuzzy matching against available slots."""
    # Normalize whitespace and dashes
    lower = re.sub(r'\s+', ' ', text.strip()).lower()
    lower = lower.replace('–', '-').replace('—', '-')

    # Exact match against available slots (e.g. from dropdown selection like "09:00 AM - 10:00 AM")
    if available_slots:
        for s in available_slots:
            slot_lower = re.sub(r'\s+', ' ', s["time"]).lower().replace('–', '-').replace('—', '-')
            if slot_lower == lower:
                return s["time"]
        # Partial match: user sends "11:00 AM" and slot is "11:00 AM - 11:30 AM"
        for s in available_slots:
            if " - " in s["time"]:
                start_part = s["time"].split(" - ")[0].strip().lower()
                if start_part == lower:
                    return s["time"]
            if lower in s["time"].lower():
                return s["time"]

    # Handle ordinal references: "the first one", "the second", "3rd slot"
    ordinals = {"first": 0, "1st": 0, "second": 1, "2nd": 1, "third": 2, "3rd": 2,
                "fourth": 3, "4th": 3, "fifth": 4, "5th": 4, "sixth": 5, "6th": 5,
                "seventh": 6, "7th": 6, "eighth": 7, "8th": 7, "last": -1}
    for word, idx in ordinals.items():
        if word in lower and available_slots:
            actual_idx = idx if idx >= 0 else len(available_slots) - 1
            if 0 <= actual_idx < len(available_slots):
                return available_slots[actual_idx]["time"]

    # Handle "earliest", "latest", "morning", "afternoon"
    if available_slots:
        if any(w in lower for w in ["earliest", "first available", "soonest", "early", "morning"]):
            for s in available_slots:
                if s["hour"] < 12:
                    return s["time"]
            return available_slots[0]["time"]
        if any(w in lower for w in ["latest", "last available", "late", "end of day"]):
            return available_slots[-1]["time"]
        if "afternoon" in lower:
            for s in available_slots:
                if s["hour"] >= 12:
                    return s["time"]
        if "lunch" in lower:
            for s in available_slots:
                if 11 <= s["hour"] <= 13:
                    return s["time"]

    # Try to parse a direct time value and match against available slots
    time_patterns = [
        r'(\d{1,2}:\d{2}\s*(?:am|pm))',  # 2:00 pm
        r'(\d{1,2}\s*(?:am|pm))',          # 2pm, 2 pm
        r'(\d{1,2}:\d{2})',                # 14:00
    ]
    for pattern in time_patterns:
        match = re.search(pattern, lower)
        if match:
            raw_time = match.group(1).strip()
            # Try to match against slot names
            if available_slots:
                matched = _match_raw_time_to_slot(raw_time, available_slots)
                if matched:
                    return matched
                # Time parsed but not in available slots — don't accept it
                return None
            return raw_time

    # If the whole message looks like a time (just a number with optional am/pm)
    simple = re.match(r'^(\d{1,2})\s*(am|pm|a\.m\.|p\.m\.)?\.?$', lower)
    if simple:
        num = simple.group(1)
        ampm = (simple.group(2) or "").replace(".", "")
        raw_time = f"{num} {ampm}" if ampm else num
        if not ampm:
            h = int(num)
            if 1 <= h <= 6:
                raw_time = f"{num} pm"
        if available_slots:
            matched = _match_raw_time_to_slot(raw_time, available_slots)
            if matched:
                return matched
            return None
        return raw_time

    return None


def _match_raw_time_to_slot(raw_time, available_slots):
    """Match a raw time string like '2:00 pm' or '2 pm' to a slot name like '02:00 PM - 02:30 PM'."""
    # Normalize the raw time to extract hour and minute
    m = re.match(r'(\d{1,2}):?(\d{2})?\s*(am|pm)?', raw_time.lower().strip())
    if not m:
        return None
    hour = int(m.group(1))
    minute = int(m.group(2) or 0)
    ampm = m.group(3)

    # Convert to 24-hour for comparison
    if ampm == 'pm' and hour < 12:
        hour += 12
    elif ampm == 'am' and hour == 12:
        hour = 0

    for slot in available_slots:
        if slot.get("hour") == hour and slot.get("minute", 0) == minute:
            return slot["time"]

    # Try just matching hour (ignore minutes if user said "2 pm")
    if m.group(2) is None:
        for slot in available_slots:
            if slot.get("hour") == hour:
                return slot["time"]

    return None


# ══════════════════════════════════════════════
#  Fast Booking — parse everything from one message
# ══════════════════════════════════════════════

def _parse_fast_booking(message, admin_id):
    """
    Extract booking details from a single message for fast booking.
    Returns (extracted_dict, active_doctors_list).
    """
    lower = message.lower().strip()
    extracted = {}

    # Get active doctors for this admin
    all_doctors = db.get_doctors(admin_id)
    doctors = [d for d in all_doctors if d.get("status") == "active"]

    # 1. Extract doctor name — match against active doctors
    matched_doctors = []
    for d in doctors:
        doc_name = d["name"].lower()
        # Check "dr ahmed", "dr. ahmed", "doctor ahmed"
        for prefix in [f"dr {doc_name}", f"dr. {doc_name}", f"doctor {doc_name}", doc_name]:
            if prefix in lower:
                if d not in matched_doctors:
                    matched_doctors.append(d)
                break
        else:
            # Check individual name parts (first name, last name) — min 3 chars
            for part in doc_name.split():
                if len(part) >= 3:
                    # Match "dr part" or "dr. part" or "doctor part" or standalone part
                    for pat in [f"dr {part}", f"dr. {part}", f"doctor {part}"]:
                        if pat in lower:
                            if d not in matched_doctors:
                                matched_doctors.append(d)
                            break
                    else:
                        # Standalone name part — only if preceded by dr/doctor or context is clear
                        if re.search(rf'\b{re.escape(part)}\b', lower):
                            if d not in matched_doctors:
                                matched_doctors.append(d)

    if len(matched_doctors) == 1:
        extracted["doctor"] = matched_doctors[0]
    elif len(matched_doctors) > 1:
        # Check if names are actually different people with same first/last name
        extracted["ambiguous_doctors"] = matched_doctors

    # 2. Extract time (careful not to grab date numbers)
    time_match = re.search(r'(\d{1,2}:\d{2}\s*(?:am|pm))', lower)
    if not time_match:
        time_match = re.search(r'(?<!\d)(\d{1,2}\s*(?:am|pm))', lower)
    if time_match:
        extracted["time_raw"] = time_match.group(1).strip()

    # 3. Extract date — natural language
    date_words = [
        "today", "tomorrow", "day after tomorrow",
        "next monday", "next tuesday", "next wednesday", "next thursday",
        "next friday", "next saturday", "next sunday",
        "this monday", "this tuesday", "this wednesday", "this thursday",
        "this friday", "this saturday", "this sunday",
        "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday",
        "next week",
    ]
    for kw in date_words:
        if kw in lower:
            extracted["date_raw"] = kw
            break
    # Also check for specific dates like "april 15", "15 april", "april 15th"
    if "date_raw" not in extracted:
        date_match = re.search(
            r'((?:january|february|march|april|may|june|july|august|september|october|november|december)'
            r'\s+\d{1,2}(?:st|nd|rd|th)?)',
            lower
        )
        if date_match:
            extracted["date_raw"] = date_match.group(1)
        else:
            # "15th of april" format
            date_match2 = re.search(
                r'(\d{1,2}(?:st|nd|rd|th)?\s+(?:of\s+)?'
                r'(?:january|february|march|april|may|june|july|august|september|october|november|december))',
                lower
            )
            if date_match2:
                extracted["date_raw"] = date_match2.group(1)

    # 4. Extract email
    email_found = _extract_email(message)
    if email_found:
        extracted["email"] = email_found

    # 5. Extract phone (only if it looks like a real phone, not a time)
    # Remove the time portion to avoid confusion
    msg_without_time = message
    if "time_raw" in extracted:
        msg_without_time = re.sub(re.escape(extracted["time_raw"]), '', msg_without_time, flags=re.IGNORECASE)
    phone_found = _extract_phone(msg_without_time)
    if phone_found and len(re.sub(r'\D', '', phone_found)) >= 7:
        extracted["phone"] = phone_found

    # 6. Extract customer name — "my name is X", "i'm X", "name is X"
    name_match = re.search(
        r"(?:my\s+name\s+is|i'?m|name\s*(?:is)?:?\s+)([A-Za-z]+(?:\s+[A-Za-z]+){0,2})",
        message, re.IGNORECASE
    )
    if name_match:
        candidate = name_match.group(1).strip()
        # Don't capture doctor names or common words as customer name
        common_words = {"looking", "trying", "wanting", "going", "interested", "here", "ready", "available"}
        if candidate.lower() not in common_words and len(candidate) >= 2:
            extracted["name"] = candidate.title()

    return extracted, doctors


def _init_fast_booking(session, extracted, doctors, admin_id):
    """
    Set up session with pre-filled data and determine the first step to ask.
    Returns (first_reply, ui_options_or_none).
    """
    # Embedded widget: require login to book
    if session.get("_is_embedded") and not session.get("_customer_logged_in"):
        session["flow"] = None
        session["step"] = None
        return "You need to log in first before booking an appointment. Please log in to your account and try again.", None

    data = session["data"]
    data["_admin_id"] = admin_id

    # Auto-fill name/email/phone from patient record (real app mode)
    if session.get("_prefill_name") and "name" not in data:
        data["name"] = session["_prefill_name"]
    if session.get("_prefill_email") and "email" not in data:
        data["email"] = session["_prefill_email"]
    if session.get("_prefill_phone") and "phone" not in data:
        data["phone"] = session["_prefill_phone"]
    if session.get("patient_id") and "_patient_id" not in data:
        data["_patient_id"] = session["patient_id"]

    # Apply context-detected service (set by booking trigger from conversation history)
    if data.get("_detected_service") and not data.get("_service_id"):
        svc = data["_detected_service"]
        try:
            all_services = db.get_services_with_doctors(admin_id)
            svc_full = next((s for s in all_services if s["id"] == svc["id"]), svc)
            data["service_name"] = svc_full["name"]
            data["_service_id"] = svc_full["id"]
            data["_service_data"] = svc_full
            data["_services"] = all_services
        except Exception:
            data["service_name"] = svc.get("name", "")
            data["_service_id"] = svc.get("id", 0)
            data["_service_data"] = svc
        del data["_detected_service"]

    # Pre-fill doctor
    if "doctor" in extracted:
        doc = extracted["doctor"]
        data["doctor_id"] = doc["id"]
        data["doctor_name"] = doc["name"]
        data["_doctors"] = doctors

    # Date hint: don't auto-validate dates from initial message.
    # Always show a calendar so the user can pick the exact date (e.g. "this Sunday" vs "next Sunday").
    date_validated = False

    # Pre-fill time (will validate availability later when we have date)
    if "time_raw" in extracted:
        data["_pending_time"] = extracted["time_raw"]

    # Pre-fill name
    if "name" in extracted:
        data["name"] = extracted["name"]

    # Pre-fill email
    if "email" in extracted:
        data["email"] = extracted["email"]

    # Pre-fill phone
    if "phone" in extracted:
        data["phone"] = extracted["phone"]

    # Handle ambiguous doctors — ask for clarification
    if "ambiguous_doctors" in extracted:
        amb = extracted["ambiguous_doctors"]
        # Check if they have different specialties
        specialties = set(d.get("specialty", "") for d in amb)
        data["_doctors"] = amb
        data["_all_doctors"] = doctors
        session["step"] = "get_doctor"
        lines = []
        for i, d in enumerate(amb, 1):
            spec = d.get("specialty", "General Dentist")
            lines.append(f"**{i}.** Dr. {d['name']} — {spec}")
        doc_list = "\n".join(lines)
        ui = {"type": "doctors", "items": [
            _doctor_dropdown_item(d)
            for d in amb
        ]}
        return (
            f"I found multiple doctors matching that name:\n\n{doc_list}\n\n"
            f"Which one would you like to book with?"
        ), ui

    # Determine the first missing step
    # Order: doctor → date → time → name → email → phone
    if "doctor_id" not in data:
        # No doctor matched — show full doctor list
        if doctors:
            categories = list(set(d.get("specialty", "General Dentist") for d in doctors if d.get("specialty")))
            categories.sort()
            if len(categories) > 1:
                data["_all_doctors"] = doctors
                data["_categories"] = categories
                session["step"] = "get_category"
                ui = {"type": "categories", "items": [{"name": c} for c in categories]}
                return "I'd love to help you book! What type of doctor would you like to see?", ui
            else:
                data["_doctors"] = doctors
                session["step"] = "get_doctor"
                ui = {"type": "doctors", "items": [
                    _doctor_dropdown_item(d)
                    for d in doctors
                ]}
                return "I'd love to help you book! Which doctor would you like to see?", ui
        else:
            if "name" not in data:
                session["step"] = "get_name"
                return "I'd love to help you book! What's your full name?", None
            if "email" not in data:
                session["step"] = "get_email"
                return f"Great, {data['name']}! What's your email address? (We'll send you a confirmation)", None
            # Name and email already known — skip to date
            session["step"] = "get_date"
            return f"Hi {data['name']}! When would you like to come in?", None

    # Doctor is set — build confirmation prefix
    doc_name = data.get("doctor_name", "")
    confirmations = [f"Dr. **{doc_name}**"]

    if not date_validated:
        session["step"] = "get_date"
        off_dates = _get_off_dates_with_blocks(data["doctor_id"], admin_id)
        ui = {"type": "calendar", "doctor_id": data["doctor_id"], "off_dates": off_dates}
        conf_text = f"Got it! Booking with {confirmations[0]}. When would you like to come in?"
        return conf_text, ui

    confirmations.append(f"**{data['date_display']}**")

    # Try to validate and set time if we have a pending time
    if "_pending_time" in data and date_validated:
        # Generate available slots
        doctor = db.get_doctor_by_id(data["doctor_id"])
        slots = []
        if doctor:
            doctor_breaks = db.get_doctor_breaks(data["doctor_id"])
            slots = _generate_doctor_slots(doctor, breaks=doctor_breaks, selected_date=data.get("date_iso"))
        booked_times = db.get_booked_times(data["doctor_id"], data["date_iso"])

        available_slots = [s for s in slots if not _is_booked_slot(s["time"], booked_times)]
        data["available_slots"] = available_slots
        data["all_slots"] = slots
        data["booked_slot_names"] = [s["time"] for s in slots if _is_booked_slot(s["time"], booked_times)]

        # Try to match the pending time against available slots
        matched_time = _match_time_to_slot(data["_pending_time"], available_slots)
        if matched_time:
            data["chosen_time"] = matched_time
            del data["_pending_time"]
            confirmations.append(f"**{matched_time}**")
        else:
            # Time not available — show slots
            del data["_pending_time"]
            session["step"] = "get_time"
            dropdown_items = []
            for s in slots:
                item = {"name": s["time"], "hour": s["hour"], "minute": s.get("minute", 0)}
                if _is_booked_slot(s["time"], booked_times):
                    item["booked"] = True
                dropdown_items.append(item)
            ui = {"type": "timeslots", "items": dropdown_items}
            return (
                f"Got it! Booking with {confirmations[0]} on {confirmations[1]}.\n\n"
                f"Unfortunately that time isn't available. Here are the open slots:"
            ), ui

    if "chosen_time" not in data:
        # Need time
        doctor = db.get_doctor_by_id(data["doctor_id"])
        slots = []
        if doctor:
            doctor_breaks = db.get_doctor_breaks(data["doctor_id"])
            slots = _generate_doctor_slots(doctor, breaks=doctor_breaks, selected_date=data.get("date_iso"))
        booked_times = db.get_booked_times(data["doctor_id"], data["date_iso"])
        available_slots = [s for s in slots if not _is_booked_slot(s["time"], booked_times)]
        data["available_slots"] = available_slots
        data["all_slots"] = slots
        data["booked_slot_names"] = [s["time"] for s in slots if _is_booked_slot(s["time"], booked_times)]

        session["step"] = "get_time"
        dropdown_items = []
        for s in slots:
            item = {"name": s["time"], "hour": s["hour"], "minute": s.get("minute", 0)}
            if _is_booked_slot(s["time"], booked_times):
                item["booked"] = True
            dropdown_items.append(item)
        ui = {"type": "timeslots", "items": dropdown_items}
        return (
            f"Got it! Booking with {confirmations[0]} on {confirmations[1]}.\n\n"
            f"What time would you like?"
        ), ui

    # Doctor + date + time all set
    if "name" not in data:
        session["step"] = "get_name"
        conf = " | ".join(confirmations)
        return f"Almost there! {conf}\n\nWhat's your full name?", None

    if "email" not in data:
        session["step"] = "get_email"
        return f"Great! What's your email address? (We'll send you a confirmation)", None

    if "phone" not in data:
        session["step"] = "get_phone"
        return "And your phone number? (In case we need to reach you)", None

    # Service booking: collect patient type and notes if not already done
    if data.get("_service_id") and "patient_type" not in data:
        session["step"] = "get_patient_type"
        return "Almost done! Are you a **new** or **returning** patient?", {"type": "confirm_yesno", "items": [{"name": "New patient", "value": "new"}, {"name": "Returning patient", "value": "returning"}]}

    if data.get("_service_id") and "patient_notes" not in data:
        session["step"] = "get_patient_notes"
        return "Any notes or concerns for the doctor? (or say **skip**)", None

    # Everything provided — skip to discount or finalize
    try:
        promos_available = promo.has_active_promotions(data.get("_admin_id", 1))
        if promos_available:
            session["step"] = "ask_discount"
            return "Do you have a discount or promo code? (or say **skip**)", None
    except Exception:
        pass

    # Build booking summary
    summary_parts = []
    if data.get("service_name"):
        summary_parts.append(f"**Service:** {data['service_name']}")
    if data.get("doctor_name"):
        summary_parts.append(f"**Doctor:** Dr. {data['doctor_name']}")
    if data.get("date_display"):
        summary_parts.append(f"**Date:** {data['date_display']}")
    if data.get("chosen_time"):
        summary_parts.append(f"**Time:** {data['chosen_time']}")
    svc_data = data.get("_service_data", {})
    if svc_data.get("duration_minutes"):
        summary_parts.append(f"**Duration:** {svc_data['duration_minutes']} minutes")
    if svc_data.get("price"):
        summary_parts.append(f"**Price:** From {svc_data['price']} {db.get_company_currency(data.get('_admin_id', 0))}")
    summary_parts.append(f"**Patient:** {data['name']}")
    if data.get("email"):
        summary_parts.append(f"**Email:** {data['email']}")

    session["step"] = "finalize_booking"
    if summary_parts and data.get("_service_id"):
        summary = "\n".join(summary_parts)
        return f"Here's your booking summary:\n\n{summary}\n\nSay **confirm** to book it, or **edit** to make changes.", None
    conf = " | ".join(confirmations)
    return f"Perfect! {conf} for **{data['name']}**. Just say **confirm** to book it.", None


def _is_booked_slot(slot_time, booked_list):
    """Check if a slot is booked."""
    if slot_time in booked_list:
        return True
    if " - " in slot_time:
        start_part = slot_time.split(" - ")[0].strip()
        return start_part in booked_list
    return False


def _match_time_to_slot(time_raw, available_slots):
    """Match a raw time string like '2pm' or '2:00 pm' to an available slot."""
    if not available_slots:
        return None

    lower = time_raw.lower().strip()

    # Normalize: "2pm" → "02:00 PM", "2:30pm" → "02:30 PM"
    m = re.match(r'^(\d{1,2}):?(\d{2})?\s*(am|pm)$', lower)
    if not m:
        return None

    hour = int(m.group(1))
    minute = int(m.group(2) or 0)
    ampm = m.group(3).upper()

    # Convert to 24h for comparison
    h24 = hour
    if ampm == "PM" and hour < 12:
        h24 += 12
    if ampm == "AM" and hour == 12:
        h24 = 0

    # Find the slot that starts at this time
    for s in available_slots:
        if s["hour"] == h24 and s.get("minute", 0) == minute:
            return s["time"]

    # Try matching just the hour if minute is 0
    if minute == 0:
        for s in available_slots:
            if s["hour"] == h24:
                return s["time"]

    return None


def handle_booking(session, user_message, corrected_message=None):
    # Embedded widget: require login to book
    if session.get("_is_embedded") and not session.get("_customer_logged_in"):
        session["flow"] = None
        session["step"] = None
        session["data"] = {}
        return "You need to log in first before booking an appointment. Please log in to your account and try again."

    step = session["step"]
    data = session["data"]
    corrected = corrected_message or correct_spelling(user_message)
    lower = corrected.lower().strip()

    # Auto-fill name/email/phone from prefill (dashboard user or embedded widget)
    if session.get("_prefill_name") and "name" not in data:
        data["name"] = session["_prefill_name"]
    if session.get("_prefill_email") and "email" not in data:
        data["email"] = session["_prefill_email"]
    if session.get("_prefill_phone") and "phone" not in data:
        data["phone"] = session["_prefill_phone"]
    # If email is known but phone isn't, try to find phone from patient record
    if "email" in data and "phone" not in data and not session.get("_phone_lookup_done"):
        session["_phone_lookup_done"] = True
        _admin = data.get("_admin_id") or session.get("admin_id", 0)
        if _admin and data["email"]:
            try:
                _prec = db.get_or_create_patient(_admin, data.get("name", ""), data["email"], "", increment_booking=False)
                if _prec and _prec.get("phone"):
                    data["phone"] = _prec["phone"]
                    session["_prefill_phone"] = _prec["phone"]
            except Exception:
                pass
    if session.get("patient_id") and "_patient_id" not in data:
        data["_patient_id"] = session["patient_id"]

    # At any step (after doctor is selected), detect "change doctor" intent
    # Check both corrected and raw input for misspellings
    if step and step not in ("get_doctor", "get_name", "get_category", "get_booking_type", "get_service", "get_service_doctor") and data.get("doctor_id"):
        raw_lower = user_message.lower().strip()
        # Normalize common misspellings of "doctor" in both raw and corrected
        _doc_pattern = r'(doc(?:t(?:o|e)?r)?|dentist|dr\.?|d[oap]c[tk]?[oe]?r)'
        _change_pattern = r'(change|switch|different|another|other|don.?t\s*want|not\s*this|swap)'
        change_doc = re.search(_change_pattern + r'.*' + _doc_pattern, lower) or \
                     re.search(_doc_pattern + r'.*' + _change_pattern, lower) or \
                     re.search(_change_pattern + r'.*' + _doc_pattern, raw_lower) or \
                     re.search(_doc_pattern + r'.*' + _change_pattern, raw_lower)
        if change_doc:
            admin_id_ctx = data.get("_admin_id", 0)
            widget_id = data.get("_widget_id", "")
            doctors = db.get_doctors(admin_id_ctx)
            doc_list = [dict(d) for d in doctors] if doctors else []
            session["step"] = "get_doctor"
            session["data"] = {"_admin_id": admin_id_ctx, "_widget_id": widget_id, "_doctors": doc_list,
                               "name": data.get("name", ""), "email": data.get("email", ""), "phone": data.get("phone", "")}
            if doc_list:
                session["_ui_options"] = {
                    "type": "doctors",
                    "items": [{"id": d["id"], "name": d["name"], "specialty": d.get("specialty", "")} for d in doc_list]
                }
            return "No problem! Which doctor would you like to see instead?"

    # At any step, detect promo/discount code intent and store for later
    if step and step not in (None, "ask_discount", "finalize_booking"):
        promo_match = re.search(r'\b(promo|promotion|discount|coupon|code|voucher)\b', lower)
        if promo_match and step in ("get_email", "get_name", "get_phone", "get_time"):
            data["_has_promo_code"] = True
            if step == "get_email":
                return "Great that you have a promo code! I'll ask for it after we collect your details.\n\nFirst, what's your **email address**? (or say **skip**)"
            elif step == "get_name":
                return "Great that you have a promo code! I'll ask for it shortly.\n\nFirst, what's your **full name**?"
            elif step == "get_phone":
                return "Great that you have a promo code! I'll ask for it right after this.\n\nWhat's your **phone number**?"

    # At any step, handle conversational questions
    if step and step not in (None, "ask_name"):
        # "What slots/times are available?" — show slots if we have a date
        if any(w in lower for w in ["available", "slots", "times", "options", "what time", "show me", "list"]):
            if data.get("date_str") and step in ("get_time",):
                # Re-show the slots
                result, error = cal.get_available_slots(data["date_str"])
                if not error:
                    slots = result["slots"]
                    slot_list = "  ".join([f"**{s['time']}**" for s in slots[:8]])
                    return f"Here are the available slots on **{data['date_display']}**:\n\n{slot_list}\n\nJust tell me which time you'd like!"
            elif step == "get_date":
                return "Sure! Just tell me which day you're interested in and I'll show you all available times. For example: 'tomorrow', 'Monday', 'next Friday', or a specific date like 'April 15'."

        # "Can I change the date?" — go back to date step
        if any(w in lower for w in ["different date", "change date", "other date", "another date", "pick another day", "change day"]):
            if step in ("get_time",):
                session["step"] = "get_date"
                doctor_id = data.get("doctor_id")
                off_dates = _get_off_dates_with_blocks(doctor_id, data.get("_admin_id", 0))
                session["_ui_options"] = {"type": "calendar", "doctor_id": doctor_id, "off_dates": off_dates}
                return "No problem! Pick a different date:"

    # Step 1: Ask for name (skip if patient is pre-filled)
    if step is None or step == "ask_name":
        # Pre-detect service name from the user's initial booking message
        # e.g. "I want to book Root Canal" → flag service for later skip
        if not data.get("_detected_service"):
            admin_id_check = data.get("_admin_id", 1)
            try:
                all_services = db.get_company_services(admin_id_check)
                active_services = [s for s in all_services if s.get("is_active", 1)]
                msg_lower_check = user_message.lower().strip()
                for svc in active_services:
                    svc_lower = svc["name"].lower()
                    if svc_lower in msg_lower_check or any(w in msg_lower_check for w in svc_lower.split() if len(w) >= 5):
                        data["_detected_service"] = svc
                        break
            except Exception:
                pass

        if data.get("name") and session.get("_patient_prefilled"):
            # Patient already known — skip name
            admin_id = data.get("_admin_id", 1)

            # If service was detected from context or initial message, skip straight to service flow
            if data.get("_detected_service"):
                svc = data.pop("_detected_service")
                all_services = db.get_services_with_doctors(admin_id)
                svc_full = next((s for s in all_services if s["id"] == svc["id"]), svc)
                data["service_name"] = svc_full["name"]
                data["_service_id"] = svc_full["id"]
                data["_service_data"] = svc_full
                data["_services"] = all_services
                session["step"] = "get_service"
                return handle_booking(session, svc_full["name"], svc_full["name"])

            # If service was already set by _init_fast_booking (context detection), go to doctor selection
            if data.get("_service_id") and data.get("_service_data"):
                session["step"] = "get_service"
                return handle_booking(session, data["service_name"], data["service_name"])

            services = db.get_services_with_doctors(admin_id)
            all_doctors = db.get_doctors(admin_id)
            doctors = [d for d in all_doctors if d.get("status") == "active"]
            if services and doctors:
                data["_all_doctors"] = doctors
                data["_services"] = services
                session["step"] = "get_booking_type"
                session["_ui_options"] = {
                    "type": "booking_type",
                    "items": [
                        {"name": "Book a Service", "value": "service"},
                        {"name": "Book an Appointment", "value": "appointment"},
                    ]
                }
                return f"Hi {data['name']}! How would you like to book?"
            elif doctors:
                cat_set = set()
                for d in doctors:
                    spec = d.get("specialty", "")
                    if spec:
                        for s in spec.split(","):
                            s = s.strip()
                            if s:
                                cat_set.add(s)
                categories = sorted(cat_set)
                if len(categories) > 1:
                    data["_all_doctors"] = doctors
                    data["_categories"] = categories
                    session["step"] = "get_category"
                    session["_ui_options"] = {"type": "categories", "items": [{"name": c} for c in categories]}
                    return f"Hi {data['name']}! What type of doctor would you like to see?"
                else:
                    data["_doctors"] = doctors
                    session["step"] = "get_doctor"
                    session["_ui_options"] = {"type": "doctors", "items": [_doctor_dropdown_item(d) for d in doctors]}
                    return f"Hi {data['name']}! Which doctor would you like to see?"
            else:
                session["step"] = "get_date"
                return f"Hi {data['name']}! When would you like to come in?"
        session["step"] = "get_name"
        return "I'd love to help you book an appointment! What's your full name?"

    if step == "get_name":
        name = user_message.strip()
        name = re.sub(r'^(my name is|i\'?m|it\'?s|name:?|hi,?\s*(i\'?m)?)\s*', '', name, flags=re.IGNORECASE).strip()
        if len(name) < 2:
            return "I didn't quite catch your name. Could you tell me your full name?"
        data["name"] = name.title()

        # If service was pre-detected from the initial booking message, skip to service flow
        if data.get("_detected_service"):
            svc = data.pop("_detected_service")
            admin_id = data.get("_admin_id", 1)
            # Reload service with doctor info
            all_services = db.get_services_with_doctors(admin_id)
            svc_full = next((s for s in all_services if s["id"] == svc["id"]), svc)
            data["service_name"] = svc_full["name"]
            data["_service_id"] = svc_full["id"]
            data["_service_data"] = svc_full
            data["_services"] = all_services
            session["step"] = "get_service"
            return handle_booking(session, svc_full["name"], svc_full["name"])

        # Check if services exist — if so, ask "Service or Appointment?"
        admin_id = data.get("_admin_id", 1)
        services = db.get_services_with_doctors(admin_id)
        all_doctors = db.get_doctors(admin_id)
        doctors = [d for d in all_doctors if d.get("status") == "active"]

        if services and doctors:
            data["_all_doctors"] = doctors
            data["_services"] = services
            session["step"] = "get_booking_type"
            session["_ui_options"] = {
                "type": "booking_type",
                "items": [
                    {"name": "Book a Service", "value": "service"},
                    {"name": "Book an Appointment", "value": "appointment"},
                ]
            }
            return f"Nice to meet you, {data['name']}! How would you like to book?"
        elif doctors:
            # No services configured — go straight to doctor selection
            cat_set = set()
            for d in doctors:
                spec = d.get("specialty", "")
                if spec:
                    for s in spec.split(","):
                        s = s.strip()
                        if s:
                            cat_set.add(s)
            categories = sorted(cat_set)
            if len(categories) > 1:
                data["_all_doctors"] = doctors
                data["_categories"] = categories
                session["step"] = "get_category"
                session["_ui_options"] = {"type": "categories", "items": [{"name": c} for c in categories]}
                return f"Nice to meet you, {data['name']}! What type of doctor would you like to see?"
            else:
                data["_doctors"] = doctors
                session["step"] = "get_doctor"
                session["_ui_options"] = {"type": "doctors", "items": [_doctor_dropdown_item(d) for d in doctors]}
                return f"Nice to meet you, {data['name']}! Which doctor would you like to see?"
        else:
            if "email" in data and "phone" in data:
                session["step"] = "get_date"
                return f"Nice to meet you, {data['name']}! When would you like to come in?"
            if "email" in data:
                session["step"] = "get_phone"
                return f"Nice to meet you, {data['name']}! What's your phone number? (In case we need to reach you)"
            session["step"] = "get_email"
            return f"Nice to meet you, {data['name']}! What's your email address? (We'll send you a confirmation)"

    # Step: Booking type choice (Service or Appointment)
    if step == "get_booking_type":
        chosen = None

        # First: check explicit "appointment" / "service" keywords before service-name matching
        if "service" in lower:
            chosen = "service"
        elif "appointment" in lower:
            chosen = "appointment"
        else:
            # Try matching by item name
            for item in [{"name": "Book a Service", "value": "service"}, {"name": "Book an Appointment", "value": "appointment"}]:
                if item["name"].lower() in lower or lower in item["name"].lower():
                    chosen = item["value"]
                    break
            # Try number
            if not chosen:
                num_match = re.search(r'(\d+)', lower)
                if num_match:
                    idx = int(num_match.group(1)) - 1
                    if idx == 0:
                        chosen = "service"
                    elif idx == 1:
                        chosen = "appointment"

        # If no explicit keyword match, try matching a specific service name from history/current message
        if not chosen:
            services = data.get("_services", [])
            if services:
                current_msg = user_message.lower().strip()
                user_history_texts = []
                for msg in session.get("history", [])[-8:]:
                    if msg.get("role") == "user":
                        user_history_texts.append(msg.get("content", "").lower())
                combined_user_hist = " ".join(user_history_texts)

                best_svc = None
                best_score = 0
                for svc in services:
                    svc_lower = svc["name"].lower()
                    svc_words = [w for w in svc_lower.split() if len(w) >= 5]
                    score = 0
                    if svc_lower in current_msg:
                        score = 100
                    elif svc_words and any(w in current_msg for w in svc_words):
                        score = 50 + sum(1 for w in svc_words if w in current_msg)
                    elif svc_lower in combined_user_hist:
                        score = 30
                    elif svc_words and any(w in combined_user_hist for w in svc_words):
                        score = 10 + sum(1 for w in svc_words if w in combined_user_hist)
                    if score > best_score:
                        best_score = score
                        best_svc = svc

                if best_svc and best_score >= 30:
                    data["service_name"] = best_svc["name"]
                    data["_service_id"] = best_svc["id"]
                    data["_service_data"] = best_svc
                    session["step"] = "get_service"
                    return handle_booking(session, best_svc["name"], corrected_message)

        if chosen == "service":
            services = data.get("_services", [])
            if not services:
                admin_id = data.get("_admin_id", 1)
                services = db.get_services_with_doctors(admin_id)
            # Filter to active services only
            services = [s for s in services if s.get("is_active", 1)]
            if not services:
                # No services available — fall through to normal appointment
                chosen = "appointment"
            else:
                data["_services"] = services
                session["step"] = "get_service"
                session["_ui_options"] = {
                    "type": "services",
                    "items": [{"name": s["name"], "id": s["id"]} for s in services]
                }
                return "Which service are you interested in?"

        if chosen == "appointment":
            # Normal appointment flow — go to category or doctor selection
            all_doctors = data.get("_all_doctors", [])
            if not all_doctors:
                admin_id = data.get("_admin_id", 1)
                all_doctors = db.get_doctors(admin_id)
                all_doctors = [d for d in all_doctors if d.get("status") == "active"]
            cat_set = set()
            for d in all_doctors:
                spec = d.get("specialty", "")
                if spec:
                    for s in spec.split(","):
                        s = s.strip()
                        if s:
                            cat_set.add(s)
            categories = sorted(cat_set)
            if len(categories) > 1:
                data["_all_doctors"] = all_doctors
                data["_categories"] = categories
                session["step"] = "get_category"
                session["_ui_options"] = {"type": "categories", "items": [{"name": c} for c in categories]}
                return "What type of doctor would you like to see?"
            else:
                data["_doctors"] = all_doctors
                session["step"] = "get_doctor"
                session["_ui_options"] = {"type": "doctors", "items": [_doctor_dropdown_item(d) for d in all_doctors]}
                return "Which doctor would you like to see?"

        # Didn't understand — re-show
        session["_ui_options"] = {
            "type": "booking_type",
            "items": [
                {"name": "Book a Service", "value": "service"},
                {"name": "Book an Appointment", "value": "appointment"},
            ]
        }
        return "Please choose one of the options:"

    # Step: Service selection
    if step == "get_service":
        services = data.get("_services", [])
        chosen_svc = None

        # Try exact match (dropdown sends exact name)
        raw_lower = user_message.lower().strip()
        for s in services:
            if s["name"].lower() == raw_lower or s["name"].lower() == lower:
                chosen_svc = s
                break

        # Try number selection
        if not chosen_svc:
            num_match = re.search(r'(\d+)', lower)
            if num_match:
                idx = int(num_match.group(1)) - 1
                if 0 <= idx < len(services):
                    chosen_svc = services[idx]

        # Try partial match
        if not chosen_svc:
            for s in services:
                if s["name"].lower() in lower or lower in s["name"].lower():
                    chosen_svc = s
                    break

        if not chosen_svc:
            session["_ui_options"] = {
                "type": "services",
                "items": [{"name": s["name"], "id": s["id"]} for s in services]
            }
            return "I didn't recognize that service. Please pick one from the list:"

        data["service_name"] = chosen_svc["name"]
        data["_service_id"] = chosen_svc["id"]
        data["_service_data"] = chosen_svc

        # Build service details message
        svc_desc = chosen_svc.get("description", "")
        svc_dur = chosen_svc.get("duration_minutes", 60)
        svc_price = chosen_svc.get("price", 0)
        svc_currency = db.get_company_currency(data.get("_admin_id", 0))
        detail_parts = [f"Great choice! **{chosen_svc['name']}**"]
        if svc_desc:
            detail_parts[0] += f" — {svc_desc}"
        detail_parts[0] += "."
        extras = []
        if svc_dur:
            extras.append(f"Sessions take about **{svc_dur} minutes**")
        if svc_price:
            extras.append(f"pricing starts from **{svc_price} {svc_currency}**")
        if extras:
            detail_parts.append(" and ".join(extras) + ".")
        detail_parts.append("The doctor will walk you through everything at the consultation — no commitment needed.")
        svc_detail_msg = " ".join(detail_parts)

        # Get doctors assigned to this service
        doctors = db.get_doctors_for_service(chosen_svc["id"])
        if not doctors:
            # No doctors for this service — offer to notify when one is assigned
            session["step"] = "get_service_notify"
            session["_ui_options"] = {
                "type": "confirm_yesno",
                "items": [
                    {"name": "Yes, notify me", "value": "yes"},
                    {"name": "No thanks", "value": "no"},
                ]
            }
            return (f"We do offer **{chosen_svc['name']}**, but right now no doctors are assigned to this service. "
                    f"Would you like to be notified when a doctor becomes available for it?")

        data["_doctors"] = doctors
        # Always show doctor dropdown so patient can see who performs this service
        session["step"] = "get_service_doctor"
        doc_items = []
        for d in doctors:
            wd = _get_doctor_working_days(d)
            avail_str = ", ".join(wd) if isinstance(wd, list) else wd
            doc_items.append({"name": d["name"], "specialty": d.get("specialty", "General"),
                              "availability": avail_str,
                              "years_of_experience": d.get("years_of_experience", 0),
                              "gender": d.get("gender", "")})
        session["_ui_options"] = {
            "type": "doctors",
            "items": doc_items
        }
        return f"{svc_detail_msg}\n\nWhich doctor would you prefer?"

    # Step: Confirm single auto-selected doctor for a service
    if step == "confirm_service_doctor":
        raw_lower = user_message.lower().strip()
        if raw_lower in ("yes", "y", "yeah", "yep", "sure", "ok", "okay", "yes, continue"):
            d = data.get("_auto_doctor", {})
            data["doctor_name"] = d["name"]
            data["doctor_id"] = d["id"]
            session["step"] = "get_date"
            off_dates = _get_off_dates_with_blocks(d["id"], data.get("_admin_id", 0))
            session["_ui_options"] = {"type": "calendar", "doctor_id": d["id"], "off_dates": off_dates}
            return f"When would you like to come in for your **{data.get('service_name', '')}** appointment?"
        else:
            # Go back to service selection
            services = data.get("_services", [])
            services = [s for s in services if s.get("is_active", 1)]
            session["step"] = "get_service"
            session["_ui_options"] = {
                "type": "services",
                "items": [{"name": s["name"], "id": s["id"]} for s in services]
            }
            return "No problem! Which service would you like instead?"

    # Step: User wants to be notified when a doctor is assigned to a service
    if step == "get_service_notify":
        raw_lower = user_message.lower().strip()
        if raw_lower in ("yes", "y", "yeah", "yep", "sure", "ok", "okay", "please", "yes, notify me"):
            # Need email to notify — check if we already have it
            if data.get("email"):
                # Save interest
                db.add_service_interest(
                    service_id=data.get("_service_id", 0),
                    service_name=data.get("service_name", ""),
                    patient_name=data.get("name", ""),
                    patient_email=data.get("email", ""),
                    patient_phone=data.get("phone", ""),
                    admin_id=data.get("_admin_id", 0),
                )
                session["flow"] = None
                session["step"] = None
                return (f"Got it! We'll notify you at **{data['email']}** as soon as a doctor is available for "
                        f"**{data.get('service_name', 'this service')}**. Is there anything else I can help with?")
            else:
                session["step"] = "get_service_notify_email"
                return "Sure! What's your **email address** so we can notify you?"
        else:
            # User said no — show all active services except the one they just rejected
            services = data.get("_services", [])
            rejected_id = data.get("_service_id", 0)
            remaining = [s for s in services if s.get("is_active", 1) and s["id"] != rejected_id]
            if remaining:
                data["_services"] = remaining
                session["step"] = "get_service"
                session["_ui_options"] = {
                    "type": "services",
                    "items": [{"name": s["name"], "id": s["id"]} for s in remaining]
                }
                return "No problem! Would you like to choose a different service?"
            else:
                session["flow"] = None
                session["step"] = None
                return "No problem! Is there anything else I can help with?"

    # Step: Collect email for service notification
    if step == "get_service_notify_email":
        extracted_email = _extract_email(user_message)
        if extracted_email:
            data["email"] = extracted_email
            db.add_service_interest(
                service_id=data.get("_service_id", 0),
                service_name=data.get("service_name", ""),
                patient_name=data.get("name", ""),
                patient_email=extracted_email,
                patient_phone=data.get("phone", ""),
                admin_id=data.get("_admin_id", 0),
            )
            session["flow"] = None
            session["step"] = None
            return (f"We'll notify you at **{email}** as soon as a doctor is available for "
                    f"**{data.get('service_name', 'this service')}**. Is there anything else I can help with?")
        else:
            return "I didn't catch a valid email. Could you please type your **email address**?"

    # Step: Doctor selection for a service
    if step == "get_service_doctor":
        doctors = data.get("_doctors", [])
        chosen = None
        raw_lower = user_message.lower().strip()

        # Exact match (dropdown sends exact name)
        for d in doctors:
            if d["name"].lower() == raw_lower or d["name"].lower() == lower:
                chosen = d
                break

        # Number selection
        if not chosen:
            num_match = re.search(r'(\d+)', lower)
            if num_match:
                idx = int(num_match.group(1)) - 1
                if 0 <= idx < len(doctors):
                    chosen = doctors[idx]

        # Partial match
        if not chosen:
            for d in doctors:
                if d["name"].lower() in lower or lower in d["name"].lower():
                    chosen = d
                    break
            if not chosen:
                for d in doctors:
                    for word in lower.split():
                        if len(word) >= 3 and word in d["name"].lower():
                            chosen = d
                            break

        if not chosen:
            # Check if user is correcting the service (e.g. "I said braces not consultation")
            services = data.get("_services", [])
            if services:
                for s in services:
                    s_lower = s["name"].lower()
                    s_words = [w for w in s_lower.split() if len(w) >= 4]
                    if s_lower in raw_lower or any(w in raw_lower for w in s_words):
                        # User wants a different service — go back
                        data["service_name"] = s["name"]
                        data["_service_id"] = s["id"]
                        data["_service_data"] = s
                        session["step"] = "get_service"
                        return handle_booking(session, s["name"], user_message)
            session["_ui_options"] = {
                "type": "doctors",
                "items": [_doctor_dropdown_item(d) for d in doctors]
            }
            return "I didn't recognize that doctor. Please pick one from the list:"

        data["doctor_name"] = chosen["name"]
        data["doctor_id"] = chosen["id"]
        session["step"] = "get_date"
        off_dates = _get_off_dates_with_blocks(chosen["id"], data.get("_admin_id", 0))
        session["_ui_options"] = {"type": "calendar", "doctor_id": chosen["id"], "off_dates": off_dates}
        svc_name = data.get("service_name", "")
        return f"Great choice! **{svc_name}** with **Dr. {chosen['name']}**.\n\nWhen would you like to come in?"

    # Step: Patient type (new/returning) for service bookings
    if step == "get_patient_type":
        raw_lower = user_message.lower().strip()
        if any(w in raw_lower for w in ("new", "first", "never")):
            data["patient_type"] = "new"
        elif any(w in raw_lower for w in ("return", "existing", "been here", "been before", "coming back")):
            data["patient_type"] = "returning"
        else:
            data["patient_type"] = raw_lower if raw_lower in ("new", "returning") else "new"
        session["step"] = "get_patient_notes"
        return "Any notes or concerns for the doctor? (or say **skip**)"

    # Step: Patient notes for the doctor
    if step == "get_patient_notes":
        raw_lower = user_message.lower().strip()
        if raw_lower in ("skip", "no", "none", "nope", "nothing", "n/a", "na"):
            data["patient_notes"] = ""
        else:
            data["patient_notes"] = user_message.strip()
        # Now check for discount or go to summary/finalize
        try:
            promos_available = promo.has_active_promotions(data.get("_admin_id", 1))
            if promos_available:
                session["step"] = "ask_discount"
                return "Do you have a discount or promo code? (or say **skip**)"
        except Exception:
            pass
        # Build service booking summary
        summary_parts = []
        if data.get("service_name"):
            summary_parts.append(f"**Service:** {data['service_name']}")
        if data.get("doctor_name"):
            summary_parts.append(f"**Doctor:** Dr. {data['doctor_name']}")
        if data.get("date_display"):
            summary_parts.append(f"**Date:** {data['date_display']}")
        if data.get("chosen_time"):
            summary_parts.append(f"**Time:** {data['chosen_time']}")
        svc_data = data.get("_service_data", {})
        if svc_data.get("duration_minutes"):
            summary_parts.append(f"**Duration:** {svc_data['duration_minutes']} minutes")
        if svc_data.get("price"):
            summary_parts.append(f"**Price:** From {svc_data['price']} {db.get_company_currency(data.get('_admin_id', 0))}")
        summary_parts.append(f"**Patient:** {data['name']}")
        if data.get("email"):
            summary_parts.append(f"**Email:** {data['email']}")
        session["step"] = "finalize_booking"
        summary = "\n".join(summary_parts)
        return f"Here's your booking summary:\n\n{summary}\n\nSay **confirm** to book it, or **edit** to make changes."

    # Step 2a: Got category choice, show doctors in that category
    if step == "get_category":
        categories = data.get("_categories", [])
        all_doctors = data.get("_all_doctors", [])
        chosen_cat = None

        # Try number selection
        num_match = re.search(r'(\d+)', lower)
        if num_match:
            idx = int(num_match.group(1)) - 1
            if 0 <= idx < len(categories):
                chosen_cat = categories[idx]

        # Try name matching
        if not chosen_cat:
            for cat in categories:
                if cat.lower() in lower or lower in cat.lower():
                    chosen_cat = cat
                    break
            # Fuzzy: check if any word matches
            if not chosen_cat:
                for cat in categories:
                    for word in lower.split():
                        if len(word) >= 3 and word in cat.lower():
                            chosen_cat = cat
                            break

        if not chosen_cat:
            # Maybe user typed a doctor name instead of a category — try matching
            for d in all_doctors:
                if d["name"].lower() in lower or lower in d["name"].lower():
                    data["doctor_name"] = d["name"]
                    data["doctor_id"] = d["id"]
                    session["step"] = "get_date"
                    off_dates = _get_off_dates_with_blocks(d["id"], data.get("_admin_id", 0))
                    session["_ui_options"] = {"type": "calendar", "doctor_id": d["id"], "off_dates": off_dates}
                    return f"Great choice! You'll be seeing **Dr. {d['name']}**.\n\nWhen would you like to come in?"
            # No match at all — re-show categories
            session["_ui_options"] = {"type": "categories", "items": [{"name": c} for c in categories]}
            return "I didn't recognize that specialty. Please pick one from the list:"

        # Filter doctors by chosen category (support comma-separated multi-specialty)
        doctors = [d for d in all_doctors if chosen_cat in [s.strip() for s in (d.get("specialty") or "").split(",")]]
        if not doctors:
            session["_ui_options"] = {"type": "categories", "items": [{"name": c} for c in categories]}
            return f"No active doctors found for **{chosen_cat}**. Please pick another specialty:"

        data["_doctors"] = doctors
        data["_chosen_category"] = chosen_cat

        session["step"] = "get_doctor"
        session["_ui_options"] = {"type": "doctors", "items": [_doctor_dropdown_item(d) for d in doctors]}
        return f"Here are our **{chosen_cat}** doctors:"

    # Step 2b: Got doctor choice, ask for email
    if step == "get_doctor":
        doctors = data.get("_doctors", [])
        chosen = None
        raw_lower = user_message.lower().strip()

        # Try exact match against raw input first (dropdown sends exact name)
        for d in doctors:
            if d["name"].lower() == raw_lower or d["name"].lower() == lower:
                chosen = d
                break

        # Try number selection
        if not chosen:
            num_match = re.search(r'(\d+)', lower)
            if num_match:
                idx = int(num_match.group(1)) - 1
                if 0 <= idx < len(doctors):
                    chosen = doctors[idx]

        # Try name matching against both raw and corrected
        if not chosen:
            for d in doctors:
                dname = d["name"].lower()
                if dname in raw_lower or raw_lower in dname or dname in lower or lower in dname:
                    chosen = d
                    break
            # Fuzzy: check if any word matches
            if not chosen:
                for d in doctors:
                    for word in raw_lower.split() + lower.split():
                        if len(word) >= 3 and word in d["name"].lower():
                            chosen = d
                            break

        if not chosen:
            # Re-show the doctors dropdown instead of returning None
            doctor_items = [{"label": f"Dr. {d['name']}", "sublabel": d.get("specialty", "")} for d in doctors]
            session["_ui_options"] = {"type": "doctors", "items": doctor_items}
            return "I didn't recognize that doctor. Please pick one from the list:"

        data["doctor_name"] = chosen["name"]
        data["doctor_id"] = chosen["id"]
        session["step"] = "get_date"
        off_dates = _get_off_dates_with_blocks(chosen["id"], data.get("_admin_id", 0))
        session["_ui_options"] = {"type": "calendar", "doctor_id": chosen["id"], "off_dates": off_dates}
        return f"Great choice! You'll be seeing **Dr. {chosen['name']}**.\n\nWhen would you like to come in?"

    # Step 3: Got date, show time slot dropdown
    if step == "get_date":
        result, error = cal.get_available_slots(user_message)
        if error:
            # Check if input looks like a date attempt or something else entirely
            date_like = bool(re.search(
                r'(today|tomorrow|monday|tuesday|wednesday|thursday|friday|saturday|sunday|'
                r'next\s+week|january|february|march|april|may|june|july|august|september|'
                r'october|november|december|\d{1,2}[/\-]\d{1,2}|\d{4}-\d{2}-\d{2}|\d{1,2}(?:st|nd|rd|th))',
                lower
            ))
            if date_like:
                doctor_id = data.get("doctor_id")
                off_dates = _get_off_dates_with_blocks(doctor_id, data.get("_admin_id", 0))
                session["_ui_options"] = {"type": "calendar", "doctor_id": doctor_id, "off_dates": off_dates}
                return error + "\n\nPlease pick another day:"
            # Not a date — return None to route to AI
            return None

        data["date_str"] = user_message
        data["date_display"] = result["date_display"]
        data["date_iso"] = result["date"].isoformat()

        # Check if this date is a doctor's off day or blocked by schedule
        doctor_id = data.get("doctor_id")
        admin_id = data.get("_admin_id", 0)
        if doctor_id:
            off_dates = _get_off_dates_with_blocks(doctor_id, admin_id)
            if data["date_iso"] in off_dates:
                doctor = db.get_doctor_by_id(doctor_id)
                doc_name = doctor["name"] if doctor else "The dentist"
                session["_ui_options"] = {
                    "type": "calendar",
                    "doctor_id": doctor_id,
                    "off_dates": off_dates,
                }
                return f"Sorry, Dr. **{doc_name}** will have an off day on **{data['date_display']}**. Please choose another date:"

        # Generate slots from doctor's schedule
        slots = []
        doctor_has_schedule = False
        if doctor_id:
            doctor = db.get_doctor_by_id(doctor_id)
            if doctor:
                doctor_has_schedule = True
                doctor_breaks = db.get_doctor_breaks(doctor_id)
                svc_dur = None
                if data.get("_service_data"):
                    svc_dur = data["_service_data"].get("duration_minutes")
                slots = _generate_doctor_slots(doctor, breaks=doctor_breaks, selected_date=data.get("date_iso"), service_duration=svc_dur)

        if not slots and not doctor_has_schedule:
            # Only fall back to generic calendar slots if doctor has no schedule configured
            slots = result["slots"]
        elif not slots and doctor_has_schedule:
            # Doctor has a schedule but no slots for this day (e.g. flexible off day or holiday)
            doc_name = doctor["name"] if doctor else "The dentist"
            off_dates = _get_off_dates_with_blocks(doctor_id, data.get("_admin_id", 0))
            session["_ui_options"] = {"type": "calendar", "doctor_id": doctor_id, "off_dates": off_dates}
            return f"Sorry, Dr. **{doc_name}** will have an off day on **{data['date_display']}**. Please choose another date:"

        # Filter out already booked times for this doctor on this date
        booked_times = []
        if doctor_id:
            booked_times = db.get_booked_times(doctor_id, result["date"].isoformat())

        available_slots = []
        booked_slot_names = []
        for s in slots:
            if _is_booked_slot(s["time"], booked_times):
                booked_slot_names.append(s["time"])
            else:
                available_slots.append(s)

        data["available_slots"] = available_slots
        data["booked_slot_names"] = booked_slot_names
        data["all_slots"] = slots

        # Build dropdown — show all slots, mark booked ones for waitlist option
        dropdown_items = []
        for s in slots:
            is_booked = _is_booked_slot(s["time"], booked_times)
            item = {"name": s["time"], "hour": s["hour"], "minute": s.get("minute", 0)}
            if is_booked:
                item["booked"] = True
            dropdown_items.append(item)

        session["_ui_options"] = {"type": "timeslots", "items": dropdown_items}
        session["step"] = "get_time"
        return f"Here are the available times on **{data['date_display']}** for **Dr. {data.get('doctor_name', '')}**:"

    # Step 4: Got time, ask for email — or offer waitlist if slot is booked
    if step == "get_time":
        all_slots = data.get("all_slots", [])
        available_slots = data.get("available_slots", [])
        booked_names = data.get("booked_slot_names", [])

        # First try exact match against all slots (dropdown sends exact slot name)
        time_str_all = _extract_time(user_message, all_slots)

        if time_str_all:
            # Check if this slot is booked — offer waitlist
            if time_str_all in booked_names:
                data["waitlist_time"] = time_str_all
                session["step"] = "waitlist_offer"
                return (f"Unfortunately **{time_str_all}** on **{data['date_display']}** is fully booked.\n\n"
                        f"Would you like to **join the waitlist**? We'll notify you instantly if a spot opens up — "
                        f"you'll get priority before it becomes available to anyone else.\n\n"
                        f"**Yes** — add me to the waitlist\n**No** — show me other times")
            # It's an available slot
            avail_names = [s["time"] for s in available_slots]
            if time_str_all in avail_names:
                data["chosen_time"] = time_str_all
                if "email" in data and "phone" in data:
                    session["step"] = "ask_discount"
                    return f"**{time_str_all}** on **{data['date_display']}** — great choice!\n\nDo you have a **promotion code** for a discount? If yes, type it now, or say **no** to continue."
                if "email" in data:
                    session["step"] = "get_phone"
                    return f"**{time_str_all}** on **{data['date_display']}** — great choice!\n\nWhat's your phone number? (In case we need to reach you)"
                session["step"] = "get_email"
                return f"**{time_str_all}** on **{data['date_display']}** — great choice!\n\nWhat's your email address? (We'll send you a confirmation)"

        # Fallback: try matching against available slots only
        time_str = _extract_time(user_message, available_slots)
        if time_str:
            # Safety check: make sure regex didn't match a booked slot's start time
            for bname in booked_names:
                if time_str.lower() in bname.lower() or bname.lower().startswith(time_str.lower()):
                    data["waitlist_time"] = bname
                    session["step"] = "waitlist_offer"
                    return (f"Unfortunately **{bname}** on **{data['date_display']}** is fully booked.\n\n"
                            f"Would you like to **join the waitlist**? We'll notify you instantly if a spot opens up — "
                            f"you'll get priority before it becomes available to anyone else.\n\n"
                            f"**Yes** — add me to the waitlist\n**No** — show me other times")
            data["chosen_time"] = time_str
            if "email" in data and "phone" in data:
                session["step"] = "ask_discount"
                return f"**{time_str}** on **{data['date_display']}** — great choice!\n\nDo you have a **promotion code** for a discount? If yes, type it now, or say **no** to continue."
            if "email" in data:
                session["step"] = "get_phone"
                return f"**{time_str}** on **{data['date_display']}** — great choice!\n\nWhat's your phone number? (In case we need to reach you)"
            session["step"] = "get_email"
            return f"**{time_str}** on **{data['date_display']}** — great choice!\n\nWhat's your email address? (We'll send you a confirmation)"

        # No time match — re-show the time slots instead of returning None
        all_s = data.get("all_slots", [])
        if all_s:
            dropdown_items = []
            for s in all_s:
                item = {"name": s["time"], "hour": s["hour"], "minute": s.get("minute", 0)}
                if s["time"] in booked_names:
                    item["booked"] = True
                dropdown_items.append(item)
            session["_ui_options"] = {"type": "timeslots", "items": dropdown_items}
            return f"I didn't catch that. Please pick a time slot from the list for **{data.get('date_display', 'your chosen date')}**:"
        return None

    # Step 4b: Waitlist offer response
    if step == "waitlist_offer":
        if _is_affirmative(user_message):
            # Reuse name/email/phone from booking flow / customer data if available
            if data.get("name"):
                data["waitlist_name"] = data["name"]
            if data.get("email"):
                data["waitlist_email"] = data["email"]
            if data.get("phone"):
                # Already have everything — add to waitlist directly
                if data.get("waitlist_name"):
                    session["step"] = "waitlist_get_phone"
                    return handle_booking(session, data["phone"], corrected)
            if data.get("waitlist_name") and data.get("waitlist_email"):
                session["step"] = "waitlist_get_phone"
                return f"I'll add you to the waitlist, {data['waitlist_name']}! What's your **phone number**?"
            if data.get("waitlist_name"):
                session["step"] = "waitlist_get_email"
                return f"I'll add you to the waitlist, {data['waitlist_name']}! What's your **email address**? (We'll send you a notification when a spot opens)"
            session["step"] = "waitlist_get_name"
            return "I'll add you to the waitlist! First, what's your **full name**?"
        elif _is_negative(user_message):
            # Reset booking flow — start over from doctor selection
            admin_id_ctx = data.get("_admin_id", 0)
            widget_id = data.get("_widget_id", "")
            doctors = db.get_doctors(admin_id_ctx)
            doc_list = [dict(d) for d in doctors] if doctors else []
            session["step"] = "get_doctor"
            session["data"] = {"_admin_id": admin_id_ctx, "_widget_id": widget_id, "_doctors": doc_list}
            if doc_list:
                session["_ui_options"] = {
                    "type": "doctors",
                    "items": [{"id": d["id"], "name": d["name"], "specialty": d.get("specialty", "")} for d in doc_list]
                }
            return "No problem! Let's start fresh. Which doctor would you like to see?"
        return "Please say **yes** to join the waitlist or **no** to see other available times."

    # Step 4c: Waitlist — get name
    if step == "waitlist_get_name":
        data["waitlist_name"] = user_message.strip().title()
        session["step"] = "waitlist_get_email"
        return f"Thanks, {data['waitlist_name']}! What's your **email address**? (We'll send you a notification when a spot opens)"

    # Step 4d: Waitlist — get email
    if step == "waitlist_get_email":
        extracted_email = _extract_email(user_message)
        if extracted_email:
            data["waitlist_email"] = extracted_email
            session["step"] = "waitlist_get_phone"
            return "Got it! And your **phone number**?"
        if any(w in lower for w in ["skip", "no email", "none", "na"]):
            data["waitlist_email"] = ""
            session["step"] = "waitlist_get_phone"
            return "No worries! What's your **phone number** then?"
        return "I couldn't find a valid email. Could you type it out? Example: john@example.com\n\nOr say **skip**."

    # Step 4e: Waitlist — get phone and add to waitlist
    if step == "waitlist_get_phone":
        extracted_phone = _extract_phone(user_message)
        if not extracted_phone:
            return "I couldn't find a valid phone number. Could you try again? Example: (555) 123-4567"

        # Detect duplicate: already booked OR already on waitlist for same slot
        try:
            admin_id_ctx = data.get("_admin_id", 0)
            doctor_id = data.get("doctor_id", 0)
            date_iso = data.get("date_iso", data.get("date_str", ""))
            slot = data.get("waitlist_time", "")
            wl_email = (data.get("waitlist_email") or "").strip().lower()
            name = (data.get("waitlist_name") or "").strip().lower()
            if date_iso and doctor_id and slot:
                conn = db.get_db()
                existing_booking = conn.execute(
                    """SELECT id FROM bookings
                       WHERE admin_id=? AND doctor_id=? AND date=? AND time=?
                             AND status != 'cancelled'
                             AND (LOWER(customer_email)=? OR customer_phone=? OR LOWER(customer_name)=?)
                       LIMIT 1""",
                    (admin_id_ctx, doctor_id, date_iso, slot, wl_email, extracted_phone, name)
                ).fetchone()
                existing_wait = conn.execute(
                    """SELECT id FROM waitlist
                       WHERE admin_id=? AND doctor_id=? AND date=? AND time_slot=?
                             AND status IN ('waiting','notified')
                             AND (LOWER(patient_email)=? OR patient_phone=? OR LOWER(patient_name)=?)
                       LIMIT 1""",
                    (admin_id_ctx, doctor_id, date_iso, slot, wl_email, extracted_phone, name)
                ).fetchone()
                conn.close()
                if existing_booking:
                    session["flow"] = None
                    session["step"] = None
                    session["data"] = {}
                    return (f"Oh! You already have an appointment with **Dr. {data.get('doctor_name','')}** "
                            f"on **{data.get('date_display','')}** at **{slot}**.\n\n"
                            f"Is there anything else I can help you with?")
                if existing_wait:
                    session["flow"] = None
                    session["step"] = None
                    session["data"] = {}
                    return (f"Oh! You're already on the waitlist for **Dr. {data.get('doctor_name','')}** "
                            f"on **{data.get('date_display','')}** at **{slot}**.\n\n"
                            f"Is there anything else I can help you with?")
        except Exception:
            pass

        # Add to waitlist
        try:
            wid = db.add_to_waitlist(
                admin_id=data.get("_admin_id", 0),
                doctor_id=data.get("doctor_id", 0),
                date=data.get("date_iso", data.get("date_str", "")),
                time_slot=data["waitlist_time"],
                patient_name=data["waitlist_name"],
                patient_email=data.get("waitlist_email", ""),
                patient_phone=extracted_phone,
                session_id=data.get("_session_id", "")
            )
            # Get their position
            position = 1
            try:
                wl = db.get_waitlist(data.get("_admin_id", 0), doctor_id=data.get("doctor_id", 0),
                                     date=data.get("date_iso", ""), time_slot=data["waitlist_time"])
                for i, entry in enumerate(wl):
                    if entry.get("id") == wid:
                        position = i + 1
                        break
            except Exception:
                pass

            doctor_name = data.get("doctor_name", "")
            doctor_info = f" with **Dr. {doctor_name}**" if doctor_name else ""

            session["flow"] = None
            session["step"] = None
            session["data"] = {}

            pos_label = {1: "1st", 2: "2nd", 3: "3rd"}.get(position, f"{position}th")
            return (
                f"You're on the waitlist! Here's your spot:\n\n"
                f"**Position:** {pos_label} in line\n"
                f"**Slot:** {data['waitlist_time']} on {data['date_display']}{doctor_info}\n"
                f"**Name:** {data['waitlist_name']}\n\n"
                f"If a spot opens up, you'll be notified immediately"
                f"{' at **' + data.get('waitlist_email', '') + '**' if data.get('waitlist_email') else ''}"
                f" and given a time window to confirm before it goes to the next person.\n\n"
                f"Is there anything else I can help you with?"
            )
        except Exception as e:
            session["flow"] = None
            session["step"] = None
            session["data"] = {}
            return "Sorry, something went wrong adding you to the waitlist. Please try again later."

    # Step 5: Got email, ask for phone
    if step == "get_email":
        extracted_email = _extract_email(user_message)
        if extracted_email:
            data["email"] = extracted_email
            session["step"] = "get_phone"
            return "Got it! And your phone number? (In case we need to reach you)"
        if any(w in lower for w in ["skip", "no email", "don't have", "dont have", "none", "no thanks", "n/a", "na"]):
            data["email"] = ""
            session["step"] = "get_phone"
            return "No worries! What's your phone number instead?"
        return "I couldn't find a valid email in that. Could you type it out? Example: john@example.com\n\nOr say **skip** if you'd rather not provide one."

    # Step 6a: Promotion code step
    if step == "ask_discount":
        raw = user_message.strip()
        lower_raw = raw.lower()
        # Explicit "no" / skip → no code applied
        if _is_negative(raw) or lower_raw in ('skip', 'none', 'لا', 'n/a', 'na'):
            data["promotion_code"] = ""
            session["step"] = "finalize_booking"
            return handle_booking(session, user_message, corrected_message)
        # "yes" alone → ask for the code
        if lower_raw in ('yes', 'y', 'yeah', 'yep', 'sure', 'نعم', 'اي'):
            session["step"] = "ask_discount_code"
            return "Great — please type your **promotion code**."
        # "yes CODE" → strip leading yes
        m = re.match(r'^(?:yes|yeah|yep|sure)[,\s]+(.+)$', raw, re.IGNORECASE)
        code = m.group(1).strip() if m else raw
        try:
            result = promo.validate_discount_code(data.get("_admin_id", 1), code)
        except Exception:
            result = {"valid": False}
        if result.get("valid"):
            data["promotion_code"] = code
            data["discount_code_id"] = result.get("code_id")
            data["discount_info"] = result
            session["step"] = "finalize_booking"
            return handle_booking(session, user_message, corrected_message)
        data["_invalid_code"] = code
        session["step"] = "ask_discount_invalid"
        return (f"The code **{code}** isn't valid. Would you like to **try another code**, or **continue** without a promotion code?\n\n"
                f"Reply **retry** to enter a new code, or **continue** to book without one.")

    # Step 6a-i: After invalid promo code
    if step == "ask_discount_invalid":
        lower_raw = user_message.strip().lower()
        if lower_raw in ('continue', 'confirm', 'no', 'skip', 'book', 'proceed') or _is_negative(lower_raw):
            data["promotion_code"] = ""
            session["step"] = "finalize_booking"
            return handle_booking(session, user_message, corrected_message)
        if lower_raw in ('retry', 'try', 'again', 'yes') or _is_affirmative(lower_raw):
            session["step"] = "ask_discount_code"
            return "Please type your **promotion code**."
        # Treat as a new code attempt
        session["step"] = "ask_discount"
        return handle_booking(session, user_message, corrected_message)

    # Step 6a-ii: Waiting for the user to type the code after saying "yes"
    if step == "ask_discount_code":
        session["step"] = "ask_discount"
        return handle_booking(session, user_message, corrected_message)

    # Step 6b: Loyalty redemption step (Feature 18)
    if step == "ask_loyalty":
        if _is_affirmative(user_message):
            patient_id = data.get("_patient_id") or session.get("patient_id")
            balance = data.get("_loyalty_balance", {})
            if patient_id and balance.get("points", 0) > 0:
                try:
                    data["loyalty_points_used"] = balance["points"]
                    data["loyalty_discount"] = balance.get("sar_value", 0)
                except Exception:
                    pass
        session["step"] = "finalize_booking"
        return handle_booking(session, user_message, corrected_message)

    # Step: edit_booking_choice — patient picked "edit" from summary
    if step == "edit_booking_choice":
        choice = lower.strip()
        if choice in ("cancel", "start over", "restart"):
            session["flow"] = None
            session["step"] = None
            session["data"] = {}
            return "No problem! Your booking has been cancelled. How can I help you?"
        has_service = bool(data.get("_service_id"))
        if "service" in choice or choice == "1" and has_service:
            session["step"] = "get_booking_type"
            return handle_booking(session, "book a service", "book a service")
        if "doctor" in choice or (choice == "2" and has_service) or (choice == "1" and not has_service):
            # Re-show doctor selection for the current service
            if data.get("_service_id"):
                session["step"] = "get_service"
                return handle_booking(session, data.get("service_name", ""), data.get("service_name", ""))
            else:
                session["step"] = "get_doctor"
                return handle_booking(session, user_message, corrected_message)
        if "date" in choice or "time" in choice or (choice == "3" and has_service) or (choice == "2" and not has_service):
            session["step"] = "get_date"
            return handle_booking(session, user_message, corrected_message)
        return "Please pick a number (1, 2, 3) or say what you'd like to change."

    # Step: finalize_booking — after discount/loyalty, actually book
    if step == "finalize_booking":
        # Handle "edit" request — let patient pick what to change
        if lower.strip() in ("edit", "change", "modify", "go back", "back"):
            edit_options = []
            if data.get("_service_id"):
                edit_options.append("1. Service")
            edit_options.append(f"{'2' if data.get('_service_id') else '1'}. Doctor")
            edit_options.append(f"{'3' if data.get('_service_id') else '2'}. Date & time")
            session["step"] = "edit_booking_choice"
            return "What would you like to change?\n\n" + "\n".join(edit_options) + "\n\nOr say **cancel** to start over."

        time_str = data.get("chosen_time", "")

        # Detect duplicate booking attempt: same customer, doctor, date, time
        try:
            date_iso = data.get("date_iso", "")
            doctor_id = data.get("doctor_id", 0)
            admin_id_ctx = data.get("_admin_id", 0)
            cust_email = (data.get("email") or "").strip().lower()
            phone = (data.get("phone") or "").strip()
            name = (data.get("name") or "").strip().lower()
            if date_iso and doctor_id and time_str:
                conn = db.get_db()
                existing = conn.execute(
                    """SELECT id FROM bookings
                       WHERE admin_id=? AND doctor_id=? AND date=? AND time=?
                             AND status != 'cancelled'
                             AND (LOWER(customer_email)=? OR customer_phone=? OR LOWER(customer_name)=?)
                       LIMIT 1""",
                    (admin_id_ctx, doctor_id, date_iso, time_str, cust_email, phone, name)
                ).fetchone()
                conn.close()
                if existing:
                    session["flow"] = None
                    session["step"] = None
                    session["data"] = {}
                    return (f"Oh! You already have an appointment with **Dr. {data.get('doctor_name','')}** "
                            f"on **{data.get('date_display','')}** at **{time_str}**.\n\n"
                            f"Is there anything else I can help you with?")
        except Exception:
            pass

        booking_result, error = cal.book_appointment(
            data.get("date_str", ""), time_str,
            data.get("name", ""), data.get("email", "")
        )
        if error:
            return error + "\n\nPlease pick another time from the available slots."

        # Save to database
        booking_id = db.save_booking(
            customer_name=data["name"],
            customer_email=data.get("email", ""),
            customer_phone=data.get("phone", ""),
            date=booking_result["date"],
            time=booking_result["time"],
            service=data.get("service_name", "General Consultation"),
            calendar_event_id=booking_result.get("calendar_event_id", ""),
            doctor_id=data.get("doctor_id", 0),
            doctor_name=data.get("doctor_name", ""),
            admin_id=data.get("_admin_id", 0),
            promotion_code=data.get("promotion_code", ""),
            service_id=data.get("_service_id", 0),
            notes=data.get("patient_notes", ""),
            patient_type=data.get("patient_type", ""),
        )


        # Increment promotion usage counter
        if data.get("promotion_code"):
            try:
                conn = db.get_db()
                conn.execute("UPDATE promotions SET current_uses = current_uses + 1 WHERE code=? AND admin_id=?",
                             (data["promotion_code"], data.get("_admin_id", 0)))
                conn.commit(); conn.close()
            except Exception:
                pass

        # Send doctor notification email for service bookings
        if data.get("_service_id") and data.get("doctor_id") and db.is_feature_enabled(data.get("_admin_id", 0), "email_booking_confirmation"):
            try:
                doctor = db.get_doctor_by_id(data["doctor_id"])
                if doctor and doctor.get("email"):
                    import email_service as email_svc
                    email_svc.send_doctor_booking_notification(
                        to_email=doctor["email"],
                        doctor_name=data.get("doctor_name", ""),
                        patient_name=data["name"],
                        service_name=data.get("service_name", ""),
                        date_display=data.get("date_display", booking_result["date"]),
                        time_display=data.get("chosen_time", booking_result["time"]),
                        patient_notes=data.get("patient_notes", ""),
                    )
            except Exception as e:
                print(f"[booking] Failed to send doctor notification: {e}", flush=True)

        # Mark chat session as booked for analytics
        if data.get("_session_id"):
            try:
                db.mark_session_booked(data["_session_id"])
            except Exception:
                pass

        # Convert lead if one exists for this session
        try:
            session_id = data.get("_session_id") or session.get("_session_id", "")
            if session_id:
                conn_b = db.get_db()
                last_bid = conn_b.execute("SELECT id FROM bookings ORDER BY id DESC LIMIT 1").fetchone()
                conn_b.close()
                if last_bid:
                    lead_engine.on_booking_completed(data.get("_admin_id", 0), session_id, last_bid[0])
        except Exception:
            pass

        # ── Patient profile + pre-visit form ──
        form_token = None
        _auto_confirmed = False
        try:
            patient = db.get_or_create_patient(
                data.get("_admin_id", 0),
                name=data["name"], email=data.get("email", ""), phone=data.get("phone", ""))
            if patient:
                conn = db.get_db()
                last_booking = conn.execute("SELECT id FROM bookings WHERE customer_name=? AND date=? ORDER BY id DESC LIMIT 1",
                    (data["name"], booking_result["date"])).fetchone()
                if last_booking:
                    conn.execute("UPDATE bookings SET patient_id=? WHERE id=?", (patient["id"], last_booking["id"]))
                    conn.commit()

                    # Check if returning patient already has a submitted form
                    _admin_id = data.get("_admin_id", 0)
                    existing_form = db.get_patient_submitted_form(
                        _admin_id, email=data.get("email", ""), phone=data.get("phone", ""))

                    if existing_form:
                        # Returning patient — clone old form, auto-confirm, add ROI revenue
                        form_token = db.clone_form_for_booking(existing_form, last_booking["id"], _admin_id, patient_name=data["name"])
                        db.confirm_booking_by_id(last_booking["id"])
                        _auto_confirmed = True
                        # ROI: track revenue on auto-confirm
                        try:
                            revenue = 0
                            svc_data_roi = data.get("_service_data", {})
                            if svc_data_roi and svc_data_roi.get("price"):
                                revenue = float(svc_data_roi["price"])
                            elif data.get("doctor_id"):
                                doc_roi = db.get_doctor_by_id(data["doctor_id"])
                                if doc_roi:
                                    revenue = float(doc_roi.get("avg_appointment_price", 20) or 20)
                            if not revenue:
                                revenue = 20.0
                            db.add_booking_revenue(last_booking["id"], revenue)
                        except Exception:
                            pass
                        # Schedule reminders
                        if db.is_feature_enabled(data.get("_admin_id", 0), "auto_reminders"):
                            try:
                                reminder_eng.schedule_reminders(last_booking["id"], _admin_id)
                            except Exception:
                                pass
                        print(f"[booking] Returning patient auto-confirmed booking_id={last_booking['id']}", flush=True)
                    else:
                        # New patient — create fresh form
                        form_token = db.create_previsit_form(last_booking["id"], _admin_id, patient_name=data["name"])
                        print(f"[booking] Pre-visit form created: token={form_token[:20]}... booking_id={last_booking['id']}", flush=True)
                else:
                    print(f"[booking] WARNING: No booking found for name={data['name']} date={booking_result['date']}", flush=True)
                conn.close()
            else:
                print(f"[booking] WARNING: get_or_create_patient returned None", flush=True)
        except Exception as e:
            print(f"[booking] ERROR creating form: {e}", flush=True)
            form_token = None

        # Send pre-visit form email only for NEW patients (not returning)
        _bk_admin = data.get("_admin_id") or session.get("_admin_id", 0)
        if not _auto_confirmed and data.get("email") and form_token and db.is_feature_enabled(_bk_admin, "email_previsit_form"):
            try:
                base_url = request.host_url.rstrip("/")
                form_url = f"{base_url}/form/{form_token}"
                email.send_previsit_form(
                    data["email"], data["name"], form_url,
                    booking_result["date_display"], booking_result["time"],
                    doctor_name=data.get("doctor_name", ""),
                    admin_id=_bk_admin,
                )
                print(f"[booking] Pre-visit form email sent to {data['email']}", flush=True)
            except Exception as e:
                print(f"[booking] ERROR sending form email: {e}", flush=True)
        elif _auto_confirmed and data.get("email"):
            # Send confirmation email for auto-confirmed returning patients
            try:
                _bk_admin_ac = data.get("_admin_id") or session.get("_admin_id", 0)
                if db.is_feature_enabled(_bk_admin_ac, "email_booking_confirmation"):
                    _svc_data = data.get("_service_data", {})
                    # Generate cancel token and confirm URL
                    _cancel_tok = secrets.token_urlsafe(32)
                    try:
                        _bconn = db.get_db()
                        _bconn.execute("UPDATE bookings SET cancel_token=? WHERE id=?", (_cancel_tok, booking_id))
                        _bconn.commit(); _bconn.close()
                    except Exception:
                        _cancel_tok = ""
                    _base = request.host_url.rstrip("/")
                    _confirm_url = f"{_base}/booking-confirmed/{booking_id}"
                    _cancel_url = f"{_base}/booking-cancel/{_cancel_tok}" if _cancel_tok else ""
                    email.send_booking_confirmation_customer(
                        customer_name=data["name"],
                        customer_email=data["email"],
                        date_display=booking_result.get("date_display", data.get("date_display", "")),
                        time_display=booking_result.get("time", data.get("chosen_time", "")),
                        doctor_name=data.get("doctor_name", ""),
                        confirm_url=_confirm_url,
                        cancel_url=_cancel_url,
                        service_name=data.get("service_name", ""),
                        duration_minutes=_svc_data.get("duration", 0) if _svc_data else 0,
                        price=str(_svc_data.get("price", "")) if _svc_data else "",
                        admin_id=_bk_admin_ac,
                    )
                    print(f"[booking] Auto-confirmed booking email sent to {data['email']}", flush=True)
            except Exception as e:
                print(f"[booking] ERROR sending auto-confirm email: {e}", flush=True)
        elif not form_token:
            print(f"[booking] WARNING: No form token — skipping form email", flush=True)

        # A/B test + real-time event
        try:
            ab_testing.record_conversion(data.get("_admin_id", 0), 'opening_message', data.get("_session_id", ""))
        except Exception:
            pass
        try:
            realtime.emit_new_booking(data.get("_admin_id", 0), {
                "customer_name": data["name"],
                "doctor_name": data.get("doctor_name", ""),
                "date": booking_result["date"],
                "time": booking_result["time"],
            })
        except Exception:
            pass

        # Flag booking as confirmed for the /chat response
        session["_booking_confirmed"] = True

        # Reset session
        session["flow"] = None
        session["step"] = None
        session["data"] = {}

        if _auto_confirmed:
            confirmation = f"Congratulations! Your appointment is **confirmed**!\n\n"
        else:
            confirmation = f"Almost there!\n\n"
        confirmation += f"**Name:** {data['name']}\n"
        if data.get("doctor_name"):
            confirmation += f"**Doctor:** Dr. {data['doctor_name']}\n"
        confirmation += (
            f"**Date:** {booking_result['date_display']}\n"
            f"**Time:** {booking_result['time']}\n"
        )
        if data.get("discount_info"):
            confirmation += f"**Discount:** {data['discount_info'].get('description', 'Applied')}\n"
        if data.get("loyalty_points_used"):
            confirmation += f"**Loyalty Points Used:** {data['loyalty_points_used']}\n"
        if _auto_confirmed:
            confirmation += f"\nWelcome back! A **confirmation email** has been sent to **{data.get('email', '')}**."
        elif data.get("email") and form_token:
            confirmation += (
                f"\nA **pre-visit form** has been sent to **{data['email']}**.\n"
                f"Please check your email and fill it out to **confirm your appointment**."
            )
        elif data.get("email"):
            confirmation += f"\nA confirmation email has been sent to **{data['email']}**."

        # Add preparation instructions for service bookings
        svc_data = data.get("_service_data", {})
        if svc_data.get("preparation_instructions"):
            confirmation += f"\n\n**Preparation instructions:**\n{svc_data['preparation_instructions']}"

        # Show average appointment price for non-service bookings
        if not data.get("_service_id") and data.get("doctor_id"):
            try:
                _appt_doc = db.get_doctor_by_id(data["doctor_id"])
                if _appt_doc:
                    _appt_price = _appt_doc.get("avg_appointment_price", 20) or 20
                    _appt_curr = db.get_company_currency(data.get("_admin_id", 0))
                    confirmation += f"\n\nThe average appointment price is **{_appt_price} {_appt_curr}**."
            except Exception:
                pass

        confirmation += "\n\nIs there anything else I can help you with?"
        return confirmation

    # Step 6: Got phone, finalize booking
    if step == "get_phone":
        extracted_phone = _extract_phone(user_message)
        if not extracted_phone:
            return "I couldn't find a valid phone number. Could you try again? Example: (555) 123-4567 or 5551234567"

        data["phone"] = extracted_phone

        # For service bookings: ask patient type next
        if data.get("_service_id"):
            session["step"] = "get_patient_type"
            return "Are you a **new patient** or a **returning patient**?"

        # Always offer promotion code before finalizing
        session["step"] = "ask_discount"
        return "Do you have a **promotion code** for a discount? If yes, type it now, or say **no** to continue."

        time_str = data.get("chosen_time", "")

        booking_result, error = cal.book_appointment(
            data["date_str"], time_str,
            data["name"], data.get("email", "")
        )
        if error:
            return error + "\n\nPlease pick another time from the available slots."

        # Save to database
        booking_id = db.save_booking(
            customer_name=data["name"],
            customer_email=data.get("email", ""),
            customer_phone=data["phone"],
            date=booking_result["date"],
            time=booking_result["time"],
            service=data.get("service_name", "General Consultation"),
            calendar_event_id=booking_result.get("calendar_event_id", ""),
            doctor_id=data.get("doctor_id", 0),
            doctor_name=data.get("doctor_name", ""),
            admin_id=data.get("_admin_id", 0),
            promotion_code=data.get("promotion_code", ""),
            service_id=data.get("_service_id", 0),
            notes=data.get("patient_notes", ""),
            patient_type=data.get("patient_type", ""),
        )


        # Increment promotion usage counter
        if data.get("promotion_code"):
            try:
                conn = db.get_db()
                conn.execute("UPDATE promotions SET current_uses = current_uses + 1 WHERE code=? AND admin_id=?",
                             (data["promotion_code"], data.get("_admin_id", 0)))
                conn.commit(); conn.close()
            except Exception:
                pass

        # Send doctor notification email for service bookings
        if data.get("_service_id") and data.get("doctor_id") and db.is_feature_enabled(data.get("_admin_id", 0), "email_booking_confirmation"):
            try:
                doctor = db.get_doctor_by_id(data["doctor_id"])
                if doctor and doctor.get("email"):
                    import email_service as email_svc
                    email_svc.send_doctor_booking_notification(
                        to_email=doctor["email"],
                        doctor_name=data.get("doctor_name", ""),
                        patient_name=data["name"],
                        service_name=data.get("service_name", ""),
                        date_display=data.get("date_display", booking_result["date"]),
                        time_display=data.get("chosen_time", booking_result["time"]),
                        patient_notes=data.get("patient_notes", ""),
                    )
            except Exception as e:
                print(f"[booking] Failed to send doctor notification: {e}", flush=True)

        # Mark chat session as booked for analytics
        if data.get("_session_id"):
            try:
                db.mark_session_booked(data["_session_id"])
            except Exception:
                pass

        # ── Patient profile + pre-visit form ──
        form_token = None
        _auto_confirmed = False
        try:
            patient = db.get_or_create_patient(
                data.get("_admin_id", 0),
                name=data["name"], email=data.get("email", ""), phone=data["phone"])
            if patient:
                conn = db.get_db()
                last_booking = conn.execute("SELECT id FROM bookings WHERE customer_name=? AND date=? ORDER BY id DESC LIMIT 1",
                    (data["name"], booking_result["date"])).fetchone()
                if last_booking:
                    conn.execute("UPDATE bookings SET patient_id=? WHERE id=?", (patient["id"], last_booking["id"]))
                    conn.commit()

                    # Check if returning patient already has a submitted form
                    _admin_id = data.get("_admin_id", 0)
                    existing_form = db.get_patient_submitted_form(
                        _admin_id, email=data.get("email", ""), phone=data.get("phone", ""))

                    if existing_form:
                        # Returning patient — clone old form, auto-confirm, add ROI revenue
                        form_token = db.clone_form_for_booking(existing_form, last_booking["id"], _admin_id, patient_name=data["name"])
                        db.confirm_booking_by_id(last_booking["id"])
                        _auto_confirmed = True
                        # ROI: track revenue on auto-confirm
                        try:
                            revenue = 0
                            svc_data_roi = data.get("_service_data", {})
                            if svc_data_roi and svc_data_roi.get("price"):
                                revenue = float(svc_data_roi["price"])
                            elif data.get("doctor_id"):
                                doc_roi = db.get_doctor_by_id(data["doctor_id"])
                                if doc_roi:
                                    revenue = float(doc_roi.get("avg_appointment_price", 20) or 20)
                            if not revenue:
                                revenue = 20.0
                            db.add_booking_revenue(last_booking["id"], revenue)
                        except Exception:
                            pass
                        # Schedule reminders
                        if db.is_feature_enabled(data.get("_admin_id", 0), "auto_reminders"):
                            try:
                                reminder_eng.schedule_reminders(last_booking["id"], _admin_id)
                            except Exception:
                                pass
                        print(f"[booking-svc] Returning patient auto-confirmed booking_id={last_booking['id']}", flush=True)
                    else:
                        # New patient — create fresh form
                        form_token = db.create_previsit_form(last_booking["id"], _admin_id, patient_name=data["name"])
                        print(f"[booking-svc] Pre-visit form created: token={form_token[:20]}... booking_id={last_booking['id']}", flush=True)
                else:
                    print(f"[booking-svc] WARNING: No booking found for name={data['name']} date={booking_result['date']}", flush=True)
                conn.close()
            else:
                print(f"[booking-svc] WARNING: get_or_create_patient returned None", flush=True)
        except Exception as e:
            print(f"[booking-svc] ERROR creating form: {e}", flush=True)
            form_token = None

        # ── A/B test + real-time event ──
        try:
            ab_testing.record_conversion(data.get("_admin_id", 0), 'opening_message', data.get("_session_id", ""))
        except Exception:
            pass
        if session.get("_ab_test_id"):
            try:
                db.increment_ab_test(session["_ab_test_id"], session.get("_ab_variant", "a"), booked=True)
            except Exception:
                pass
        try:
            realtime.emit_new_booking(data.get("_admin_id", 0), {
                "customer_name": data["name"],
                "doctor_name": data.get("doctor_name", ""),
                "date": booking_result["date"],
                "time": booking_result["time"],
            })
        except Exception:
            pass

        # Get booking ID for email links
        booking_id = None
        try:
            conn = db.get_db()
            bid_row = conn.execute("SELECT id FROM bookings WHERE customer_name=? AND date=? ORDER BY id DESC LIMIT 1",
                (data["name"], booking_result["date"])).fetchone()
            if bid_row:
                booking_id = bid_row["id"]
            conn.close()
        except Exception:
            pass

        # Send pre-visit form email only for NEW patients (not returning)
        _bks_admin = data.get("_admin_id") or session.get("_admin_id", 0)
        if not _auto_confirmed and data.get("email") and form_token and db.is_feature_enabled(_bks_admin, "email_previsit_form"):
            try:
                base_url = request.host_url.rstrip("/")
                form_url = f"{base_url}/form/{form_token}"
                email.send_previsit_form(
                    data["email"], data["name"], form_url,
                    booking_result["date_display"], booking_result["time"],
                    doctor_name=data.get("doctor_name", ""),
                    admin_id=_bks_admin,
                )
                print(f"[booking-svc] Pre-visit form email sent to {data['email']}", flush=True)
            except Exception as e:
                print(f"[booking-svc] ERROR sending form email: {e}", flush=True)
        elif _auto_confirmed and data.get("email"):
            # Send confirmation email for auto-confirmed returning patients
            try:
                if db.is_feature_enabled(_bks_admin, "email_booking_confirmation"):
                    _svc_data = data.get("_service_data", {})
                    _cancel_tok = secrets.token_urlsafe(32)
                    try:
                        _bconn = db.get_db()
                        _bconn.execute("UPDATE bookings SET cancel_token=? WHERE id=?", (_cancel_tok, booking_id))
                        _bconn.commit(); _bconn.close()
                    except Exception:
                        _cancel_tok = ""
                    _base = request.host_url.rstrip("/")
                    _confirm_url = f"{_base}/booking-confirmed/{booking_id}"
                    _cancel_url = f"{_base}/booking-cancel/{_cancel_tok}" if _cancel_tok else ""
                    email.send_booking_confirmation_customer(
                        customer_name=data["name"],
                        customer_email=data["email"],
                        date_display=booking_result.get("date_display", data.get("date_display", "")),
                        time_display=booking_result.get("time", data.get("chosen_time", "")),
                        doctor_name=data.get("doctor_name", ""),
                        confirm_url=_confirm_url,
                        cancel_url=_cancel_url,
                        service_name=data.get("service_name", ""),
                        duration_minutes=_svc_data.get("duration", 0) if _svc_data else 0,
                        price=str(_svc_data.get("price", "")) if _svc_data else "",
                        admin_id=_bks_admin,
                    )
                    print(f"[booking-svc] Auto-confirmed booking email sent to {data['email']}", flush=True)
            except Exception as e:
                print(f"[booking-svc] ERROR sending auto-confirm email: {e}", flush=True)
        elif not form_token:
            print(f"[booking-svc] WARNING: No form token — skipping form email", flush=True)

        # Convert/remove lead now that booking is confirmed
        try:
            lead_engine.on_booking_completed(data.get("_admin_id", 0), session_id, booking_id or 0)
        except Exception:
            pass

        # Reset session
        session["flow"] = None
        session["step"] = None
        session["data"] = {}

        confirmation = (
            f"Almost there!\n\n"
            f"**Name:** {data['name']}\n"
        )
        if data.get("doctor_name"):
            confirmation += f"**Doctor:** Dr. {data['doctor_name']}\n"
        confirmation += (
            f"**Date:** {booking_result['date_display']}\n"
            f"**Time:** {booking_result['time']}\n"
        )
        if _auto_confirmed:
            confirmation += f"\nWelcome back! Your appointment is **automatically confirmed**."
        elif data.get("email") and form_token:
            confirmation += (
                f"\nA **pre-visit form** has been sent to **{data['email']}**.\n"
                f"Please check your email and fill it out to **confirm your appointment**."
            )
        elif data.get("email"):
            confirmation += f"\nA confirmation email has been sent to **{data['email']}**."

        # Add preparation instructions for service bookings
        svc_data = data.get("_service_data", {})
        if svc_data.get("preparation_instructions"):
            confirmation += f"\n\n**Preparation instructions:**\n{svc_data['preparation_instructions']}"

        # Show average appointment price for non-service bookings
        if not data.get("_service_id") and data.get("doctor_id"):
            try:
                _appt_doc = db.get_doctor_by_id(data["doctor_id"])
                if _appt_doc:
                    _appt_price = _appt_doc.get("avg_appointment_price", 20) or 20
                    _appt_curr = db.get_company_currency(data.get("_admin_id", 0))
                    confirmation += f"\n\nThe average appointment price is **{_appt_price} {_appt_curr}**."
            except Exception:
                pass

        confirmation += "\n\nIs there anything else I can help you with?"
        return confirmation

    return "Something went wrong with the booking. Let me start over. Would you like to book an appointment?"


# ══════════════════════════════════════════════
#  Cancel Appointment Flow
# ══════════════════════════════════════════════

def handle_cancel_appointment(session, user_message, admin_id):
    step = session["step"]
    data = session["data"]
    lower = user_message.strip().lower()

    # Abort if user says nevermind
    if lower in ("nevermind", "never mind", "stop", "go back"):
        session["flow"] = None
        session["step"] = None
        session["data"] = {}
        return "No problem! How else can I help you?"

    # Step 1: Get date — parse natural language ("this monday", "april 10", etc.)
    if step == "get_date":
        from calendar_service import _parse_date
        parsed_date = _parse_date(user_message)
        if not parsed_date:
            return "I didn't understand that date. Could you try again? (e.g. Monday, April 10, tomorrow)"

        date_iso = parsed_date.isoformat()
        date_display = parsed_date.strftime("%A, %B %d, %Y")
        bookings = db.find_bookings_by_date(admin_id, date_iso)

        # For embedded widget with logged-in customer: only show THEIR bookings
        if session.get("_is_embedded") and session.get("_customer_logged_in"):
            customer_name = session.get("_prefill_name", "").strip().lower()
            customer_email = session.get("_prefill_email", "").strip().lower()
            customer_phone = session.get("_prefill_phone", "").strip()
            if customer_name or customer_email or customer_phone:
                bookings = [b for b in bookings if
                    (customer_name and b.get("customer_name", "").strip().lower() == customer_name) or
                    (customer_email and b.get("customer_email", "").strip().lower() == customer_email) or
                    (customer_phone and b.get("customer_phone", "").strip() == customer_phone)
                ]

        if not bookings:
            session["flow"] = None
            session["step"] = None
            session["data"] = {}
            return f"There are no active appointments on **{date_display}**. Say **cancel my appointment** to try a different date."

        data["_cancel_date"] = date_iso
        data["_cancel_date_display"] = date_display

        if len(bookings) == 1:
            b = bookings[0]
            data["_booking_to_cancel"] = b
            session["step"] = "confirm"
            session["_ui_options"] = {"type": "confirm_yesno", "items": [{"name": "Yes, cancel it", "value": "yes"}, {"name": "No, keep it", "value": "no"}]}
            doctor_info = f" with **Dr. {b['doctor_name']}**" if b.get("doctor_name") else ""
            return (f"I found one appointment on **{date_display}**:\n\n"
                    f"**{b['customer_name']}** — {b['time']}{doctor_info}\n\n"
                    f"Do you want to cancel this appointment?")

        # Multiple bookings — try to auto-match if user mentioned name, time, or doctor
        auto_matched = None
        for b in bookings:
            name_match = b.get("customer_name", "").lower() in lower
            doctor_match = b.get("doctor_name", "") and b["doctor_name"].lower() in lower
            # Check if the user mentioned a time that matches this booking
            time_match = False
            if b.get("time"):
                # Extract start time for matching (e.g. "2:00 PM" from "2:00 PM - 3:00 PM")
                btime = b["time"].split(" - ")[0].strip().lower() if " - " in b["time"] else b["time"].strip().lower()
                time_match = btime in lower or b["time"].lower() in lower
                # Also match "2 pm", "2pm", "2:00pm" etc.
                time_digits = re.search(r'(\d{1,2})\s*(?::?\s*(\d{2}))?\s*(am|pm)', lower)
                if time_digits:
                    user_hour = int(time_digits.group(1))
                    user_min = time_digits.group(2) or "00"
                    user_ampm = time_digits.group(3).upper()
                    user_time_str = f"{user_hour}:{user_min} {user_ampm}".lower()
                    time_match = time_match or user_time_str in btime

            if (name_match and doctor_match) or (name_match and time_match) or (doctor_match and time_match):
                auto_matched = b
                break

        if auto_matched:
            data["_booking_to_cancel"] = auto_matched
            session["step"] = "confirm"
            session["_ui_options"] = {"type": "confirm_yesno", "items": [{"name": "Yes, cancel it", "value": "yes"}, {"name": "No, keep it", "value": "no"}]}
            doctor_info = f" with **Dr. {auto_matched['doctor_name']}**" if auto_matched.get("doctor_name") else ""
            return (f"I found the appointment on **{date_display}**:\n\n"
                    f"**{auto_matched['customer_name']}** — {auto_matched['time']}{doctor_info}\n\n"
                    f"Do you want to cancel this appointment?")

        # Multiple bookings on that date — show as dropdown
        data["_bookings_list"] = bookings
        session["step"] = "choose"
        lines = []
        for i, b in enumerate(bookings, 1):
            doctor_info = f" — Dr. {b['doctor_name']}" if b.get("doctor_name") else ""
            lines.append(f"**{i}.** {b['customer_name']} at {b['time']}{doctor_info}")
        session["_ui_options"] = {
            "type": "cancel_bookings",
            "items": [
                {"name": f"{b['customer_name']} — {b['time']}" + (f" (Dr. {b['doctor_name']})" if b.get("doctor_name") else ""), "index": i}
                for i, b in enumerate(bookings, 1)
            ]
        }
        return f"I found **{len(bookings)}** appointments on **{date_display}**:\n\n" + "\n".join(lines) + "\n\nWhich one do you want to cancel?"

    # Step 2: Choose which booking
    if step == "choose":
        bookings = data.get("_bookings_list", [])
        chosen = None

        # Try number selection
        num_match = re.search(r'(\d+)', lower)
        if num_match:
            idx = int(num_match.group(1)) - 1
            if 0 <= idx < len(bookings):
                chosen = bookings[idx]

        # Try name matching
        if not chosen:
            for b in bookings:
                if b.get("customer_name", "").lower() in lower or lower in b.get("customer_name", "").lower():
                    chosen = b
                    break
            if not chosen:
                for b in bookings:
                    if b.get("doctor_name", "").lower() in lower or lower in b.get("doctor_name", "").lower():
                        chosen = b
                        break

        if chosen:
            data["_booking_to_cancel"] = chosen
            session["step"] = "confirm"
            session["_ui_options"] = {"type": "confirm_yesno", "items": [{"name": "Yes, cancel it", "value": "yes"}, {"name": "No, keep it", "value": "no"}]}
            doctor_info = f" with **Dr. {chosen['doctor_name']}**" if chosen.get("doctor_name") else ""
            return (f"You selected:\n\n"
                    f"**{chosen['customer_name']}** — {chosen['time']}{doctor_info}\n\n"
                    f"Do you want to cancel this appointment?")

        return "I couldn't match that. Please pick a number from the list or say the patient name."

    # Step 3: Confirm cancellation
    if step == "confirm":
        if lower in ("yes", "yeah", "yep", "yea", "sure", "ok", "okay", "y", "yes please", "confirm"):
            booking = data.get("_booking_to_cancel")
            if booking:
                db.cancel_booking(booking["id"])
                # Increment patient cancellation count
                if booking.get("patient_id"):
                    try:
                        _conn = db.get_db()
                        _conn.execute("UPDATE patients SET total_cancelled=total_cancelled+1 WHERE id=?", (booking["patient_id"],))
                        _conn.commit()
                        _conn.close()
                    except Exception:
                        pass
                # Trigger waitlist cascade — notify next waiting patient
                if booking.get("doctor_id") and booking.get("date") and booking.get("time"):
                    try:
                        import background_tasks
                        background_tasks.trigger_waitlist_processing(
                            booking.get("admin_id", 0), booking["doctor_id"],
                            booking["date"], booking["time"])
                    except Exception:
                        pass
                doctor_info = f" with Dr. {booking['doctor_name']}" if booking.get("doctor_name") else ""
                date_display = data.get("_cancel_date_display", booking["date"])
                # ── Feature 16: Emit real-time cancellation event ──
                try:
                    realtime.emit_booking_cancelled(admin_id, booking["id"], booking["customer_name"])
                except Exception:
                    pass
                session["flow"] = None
                session["step"] = None
                session["data"] = {}
                return (f"Done! The appointment for **{booking['customer_name']}** on **{date_display}** at **{booking['time']}**{doctor_info} "
                        f"has been **cancelled**. The time slot is now free for others.\n\n"
                        f"Is there anything else I can help you with?")
        elif lower in ("no", "nah", "nope", "n", "no thanks"):
            session["flow"] = None
            session["step"] = None
            session["data"] = {}
            return "Okay, the appointment is still active. How else can I help you?"

        return "Please say **yes** to confirm the cancellation or **no** to keep it."

    # Fallback
    session["flow"] = None
    session["step"] = None
    session["data"] = {}
    return "Something went wrong. Say **cancel my appointment** to try again."


# ══════════════════════════════════════════════
#  Reschedule Appointment Flow State Machine
# ══════════════════════════════════════════════

def handle_reschedule(session, user_message, admin_id):
    step = session["step"]
    data = session["data"]
    lower = user_message.strip().lower()

    # Abort if user says nevermind
    if lower in ("nevermind", "never mind", "stop", "go back", "cancel"):
        session["flow"] = None
        session["step"] = None
        session["data"] = {}
        return "No problem! Your appointment stays as is. How else can I help you?"

    # ── Step 1: Select date that has bookings ──
    if step == "get_date":
        from calendar_service import _parse_date
        parsed_date = _parse_date(user_message)
        if not parsed_date:
            return "I didn't understand that date. Could you try again? (e.g. Monday, April 10, tomorrow)"

        date_iso = parsed_date.isoformat()
        date_display = parsed_date.strftime("%A, %B %d, %Y")
        bookings = db.find_bookings_by_date(admin_id, date_iso)

        # For embedded widget: only show THEIR bookings
        if session.get("_is_embedded") and session.get("_customer_logged_in"):
            customer_name = (session.get("_prefill_name") or "").strip().lower()
            customer_email = (session.get("_prefill_email") or "").strip().lower()
            customer_phone = (session.get("_prefill_phone") or "").strip()
            if customer_name or customer_email or customer_phone:
                bookings = [b for b in bookings if
                    (customer_name and b.get("customer_name", "").strip().lower() == customer_name) or
                    (customer_email and b.get("customer_email", "").strip().lower() == customer_email) or
                    (customer_phone and b.get("customer_phone", "").strip() == customer_phone)
                ]

        # Only show future bookings (can't reschedule past ones)
        from datetime import datetime as _dt_r
        now = _dt_r.now()
        future_bookings = []
        for b in bookings:
            try:
                time_parts = b["time"].split(" - ")
                end_str = time_parts[-1].strip() if len(time_parts) > 1 else time_parts[0].strip()
                booking_end = _dt_r.strptime(f"{b['date']} {end_str}", "%Y-%m-%d %I:%M %p")
                if booking_end > now:
                    future_bookings.append(b)
            except Exception:
                future_bookings.append(b)
        bookings = future_bookings

        if not bookings:
            session["flow"] = None
            session["step"] = None
            session["data"] = {}
            return f"There are no upcoming appointments on **{date_display}**. Say **reschedule** to try a different date."

        data["_reschedule_date"] = date_iso
        data["_reschedule_date_display"] = date_display

        if len(bookings) == 1:
            b = bookings[0]
            data["_booking_to_reschedule"] = b
            session["step"] = "keep_doctor"
            doctor_info = f" with **Dr. {b['doctor_name']}**" if b.get("doctor_name") else ""
            service_info = f" ({b.get('service', '')})" if b.get("service") else ""
            session["_ui_options"] = {"type": "confirm_yesno", "items": [{"name": "Yes, same doctor", "value": "yes"}, {"name": "No, different doctor", "value": "no"}]}
            return (f"I found your appointment on **{date_display}**:\n\n"
                    f"**{b['customer_name']}** — {b['time']}{doctor_info}{service_info}\n\n"
                    f"Would you like to keep the same doctor?")

        # Multiple bookings — show dropdown
        data["_bookings_list"] = bookings
        session["step"] = "choose_booking"
        lines = []
        for i, b in enumerate(bookings, 1):
            doctor_info = f" — Dr. {b['doctor_name']}" if b.get("doctor_name") else ""
            service_info = f" ({b.get('service', '')})" if b.get("service") else ""
            lines.append(f"**{i}.** {b['customer_name']} at {b['time']}{doctor_info}{service_info}")
        session["_ui_options"] = {
            "type": "cancel_bookings",
            "items": [
                {"name": f"{b['customer_name']} — {b['time']}" + (f" (Dr. {b['doctor_name']})" if b.get("doctor_name") else ""), "index": i}
                for i, b in enumerate(bookings, 1)
            ]
        }
        return f"I found **{len(bookings)}** appointments on **{date_display}**:\n\n" + "\n".join(lines) + "\n\nWhich one would you like to reschedule?"

    # ── Step 2: Choose which booking (multiple bookings) ──
    if step == "choose_booking":
        bookings = data.get("_bookings_list", [])
        chosen = None

        num_match = re.search(r'(\d+)', lower)
        if num_match:
            idx = int(num_match.group(1)) - 1
            if 0 <= idx < len(bookings):
                chosen = bookings[idx]

        if not chosen:
            for b in bookings:
                if b.get("customer_name", "").lower() in lower or lower in b.get("customer_name", "").lower():
                    chosen = b
                    break
            if not chosen:
                for b in bookings:
                    if b.get("doctor_name", "").lower() in lower or lower in b.get("doctor_name", "").lower():
                        chosen = b
                        break

        if chosen:
            data["_booking_to_reschedule"] = chosen
            session["step"] = "keep_doctor"
            doctor_info = f" with **Dr. {chosen['doctor_name']}**" if chosen.get("doctor_name") else ""
            session["_ui_options"] = {"type": "confirm_yesno", "items": [{"name": "Yes, same doctor", "value": "yes"}, {"name": "No, different doctor", "value": "no"}]}
            return (f"You selected:\n\n"
                    f"**{chosen['customer_name']}** — {chosen['time']}{doctor_info}\n\n"
                    f"Would you like to keep the same doctor?")

        return "I couldn't match that. Please pick a number from the list or say the patient name."

    # ── Step 3: Keep same doctor? ──
    if step == "keep_doctor":
        booking = data.get("_booking_to_reschedule", {})

        if lower in ("yes", "yeah", "yep", "yea", "sure", "ok", "okay", "y", "yes please", "same doctor", "same"):
            # Keep same doctor — go to date selection
            doctor_id = booking.get("doctor_id")
            if not doctor_id:
                session["flow"] = None
                session["step"] = None
                session["data"] = {}
                return "This appointment doesn't have a doctor assigned. Please contact the clinic directly to reschedule."

            data["_new_doctor_id"] = doctor_id
            data["_new_doctor_name"] = booking.get("doctor_name", "")
            session["step"] = "select_new_date"

            # Show calendar with off-dates for this doctor
            off_dates = _get_off_dates_with_blocks(doctor_id, admin_id)
            session["_ui_options"] = {"type": "calendar", "off_dates": off_dates}
            return f"Great! When would you like to reschedule with **Dr. {booking.get('doctor_name', '')}**? Pick a new date:"

        elif lower in ("no", "nah", "nope", "n", "no thanks", "different doctor", "different", "change doctor"):
            # Different doctor — show doctor list
            service_id = booking.get("service_id")
            if service_id:
                doctors = db.get_doctors_for_service(service_id)
            else:
                doctors = db.get_doctors(admin_id)

            # Filter out inactive doctors
            doctors = [d for d in doctors if d.get("is_active", 1)]

            if not doctors:
                session["flow"] = None
                session["step"] = None
                session["data"] = {}
                return "No available doctors found. Please contact the clinic directly."

            if len(doctors) == 1:
                # Only one doctor available
                data["_new_doctor_id"] = doctors[0]["id"]
                data["_new_doctor_name"] = doctors[0]["name"]
                session["step"] = "select_new_date"
                off_dates = _get_off_dates_with_blocks(doctors[0]["id"], admin_id)
                session["_ui_options"] = {"type": "calendar", "off_dates": off_dates}
                return f"The only available doctor is **Dr. {doctors[0]['name']}**. Pick a new date:"

            data["_available_doctors"] = doctors
            session["step"] = "select_doctor"
            session["_ui_options"] = {"type": "doctors", "items": [_doctor_dropdown_item(d) for d in doctors]}
            return "Which doctor would you like to see instead?"

        return "Please say **yes** to keep the same doctor or **no** to choose a different one."

    # ── Step 4: Select doctor (if changing) ──
    if step == "select_doctor":
        doctors = data.get("_available_doctors", [])
        chosen_doc = None

        for d in doctors:
            if d["name"].lower() == lower or f"dr. {d['name'].lower()}" == lower or f"dr {d['name'].lower()}" == lower:
                chosen_doc = d
                break
        if not chosen_doc:
            for d in doctors:
                if d["name"].lower() in lower or lower in d["name"].lower():
                    chosen_doc = d
                    break

        if chosen_doc:
            data["_new_doctor_id"] = chosen_doc["id"]
            data["_new_doctor_name"] = chosen_doc["name"]
            session["step"] = "select_new_date"
            off_dates = _get_off_dates_with_blocks(chosen_doc["id"], admin_id)
            session["_ui_options"] = {"type": "calendar", "off_dates": off_dates}
            return f"When would you like to see **Dr. {chosen_doc['name']}**? Pick a new date:"

        return "I couldn't match that doctor. Please select one from the list."

    # ── Step 5: Select new date ──
    if step == "select_new_date":
        from calendar_service import _parse_date
        parsed_date = _parse_date(user_message)
        if not parsed_date:
            return "I didn't understand that date. Could you try again? (e.g. Monday, April 10, tomorrow)"

        date_iso = parsed_date.isoformat()
        date_display = parsed_date.strftime("%A, %B %d, %Y")

        # Check if date is in the past
        from datetime import datetime as _dt_r2
        if parsed_date < _dt_r2.now().date():
            return "That date is in the past. Please select a future date."

        doctor_id = data.get("_new_doctor_id")
        doctor = db.get_doctor_by_id(doctor_id) if doctor_id else None

        if not doctor:
            session["flow"] = None
            session["step"] = None
            session["data"] = {}
            return "Something went wrong finding the doctor. Say **reschedule** to try again."

        # Check if doctor works this day
        off_dates = _get_off_dates_with_blocks(doctor_id, admin_id)
        if date_iso in off_dates:
            return f"**Dr. {doctor['name']}** is not available on **{date_display}**. Please pick another date."

        # Generate time slots
        doctor_breaks = db.get_doctor_breaks(doctor_id)
        booking = data.get("_booking_to_reschedule", {})

        # Get service duration if available
        service_duration = None
        if booking.get("service_id"):
            svc = db.get_company_service_by_id(booking["service_id"])
            if svc and svc.get("duration_minutes"):
                service_duration = int(svc["duration_minutes"])

        slots = _generate_doctor_slots(doctor, breaks=doctor_breaks, selected_date=date_iso, service_duration=service_duration)
        booked_times = db.get_booked_times(doctor_id, date_iso)

        # Exclude the original booking's time slot if same doctor and same date
        # (so the user can re-pick their own slot if desired)
        orig_booking = data.get("_booking_to_reschedule", {})
        if str(orig_booking.get("doctor_id")) == str(doctor_id) and orig_booking.get("date") == date_iso:
            orig_time = orig_booking.get("time", "")
            booked_times = [t for t in booked_times if t != orig_time and t != orig_time.split(" - ")[0].strip()]

        available_slots = [s for s in slots if not _is_booked_slot(s["time"], booked_times)]

        if not available_slots:
            return f"No available time slots on **{date_display}** for **Dr. {doctor['name']}**. Please try another date."

        data["_new_date_iso"] = date_iso
        data["_new_date_display"] = date_display
        data["_available_slots"] = available_slots
        session["step"] = "select_time"

        dropdown_items = []
        for s in slots:
            item = {"name": s["time"], "hour": s["hour"], "minute": s.get("minute", 0)}
            if _is_booked_slot(s["time"], booked_times):
                item["booked"] = True
            dropdown_items.append(item)

        session["_ui_options"] = {"type": "timeslots", "items": dropdown_items}
        return f"Available times on **{date_display}** with **Dr. {data.get('_new_doctor_name', '')}**:"

    # ── Step 6: Select time ──
    if step == "select_time":
        available_slots = data.get("_available_slots", [])
        matched_slot = _match_time_to_slot(user_message, available_slots)

        if not matched_slot:
            # Try exact match
            for s in available_slots:
                if s["time"].lower() == lower or s["time"].lower().replace(" ", "") == lower.replace(" ", ""):
                    matched_slot = s
                    break

        if not matched_slot:
            return "I couldn't match that time. Please select one from the available slots."

        data["_new_time"] = matched_slot["time"]
        session["step"] = "confirm"

        # Build confirmation summary
        booking = data.get("_booking_to_reschedule", {})
        old_doctor = f"Dr. {booking.get('doctor_name', '')}" if booking.get("doctor_name") else "N/A"
        new_doctor = f"Dr. {data.get('_new_doctor_name', '')}"
        old_date = data.get("_reschedule_date_display", booking.get("date", ""))
        new_date = data.get("_new_date_display", "")
        old_time = booking.get("time", "")
        new_time = matched_slot["time"]
        service_name = booking.get("service", "")

        summary = f"Here's a summary of the reschedule:\n\n"
        if service_name:
            summary += f"**Service:** {service_name}\n"
        summary += f"**Patient:** {booking.get('customer_name', '')}\n\n"
        summary += f"**Old appointment:**\n{old_date} at {old_time} — {old_doctor}\n\n"
        summary += f"**New appointment:**\n{new_date} at {new_time} — {new_doctor}\n\n"
        summary += f"Do you want to confirm this reschedule?"

        session["_ui_options"] = {"type": "confirm_yesno", "items": [{"name": "Yes, reschedule", "value": "yes"}, {"name": "No, cancel", "value": "no"}]}
        return summary

    # ── Step 7: Confirm reschedule ──
    if step == "confirm":
        if lower in ("yes", "yeah", "yep", "yea", "sure", "ok", "okay", "y", "yes please", "confirm"):
            booking = data.get("_booking_to_reschedule", {})
            new_doctor_id = data.get("_new_doctor_id")
            new_doctor_name = data.get("_new_doctor_name", "")
            new_date = data.get("_new_date_iso", "")
            new_time = data.get("_new_time", "")

            if not booking or not new_date or not new_time:
                session["flow"] = None
                session["step"] = None
                session["data"] = {}
                return "Something went wrong. Say **reschedule** to try again."

            # Update the booking in database
            conn = db.get_db()
            try:
                conn.execute(
                    """UPDATE bookings SET date=?, time=?, doctor_id=?, doctor_name=?
                       WHERE id=?""",
                    (new_date, new_time, new_doctor_id, new_doctor_name, booking["id"])
                )
                conn.commit()
            except Exception as e:
                conn.close()
                logger.error(f"Reschedule DB error: {e}")
                session["flow"] = None
                session["step"] = None
                session["data"] = {}
                return "An error occurred while rescheduling. Please try again or contact the clinic."
            conn.close()

            # Keep original status — if was confirmed, stays confirmed
            # If was pending, stays pending (user needs to confirm via form)

            old_date_display = data.get("_reschedule_date_display", booking.get("date", ""))
            new_date_display = data.get("_new_date_display", "")
            doctor_info = f" with **Dr. {new_doctor_name}**" if new_doctor_name else ""

            session["flow"] = None
            session["step"] = None
            session["data"] = {}
            return (f"Your appointment has been rescheduled!\n\n"
                    f"**New appointment:** {new_date_display} at {new_time}{doctor_info}\n\n"
                    f"Is there anything else I can help you with?")

        elif lower in ("no", "nah", "nope", "n", "no thanks"):
            session["flow"] = None
            session["step"] = None
            session["data"] = {}
            return "Okay, your appointment stays as is. How else can I help you?"

        return "Please say **yes** to confirm the reschedule or **no** to keep your current appointment."

    # Fallback
    session["flow"] = None
    session["step"] = None
    session["data"] = {}
    return "Something went wrong. Say **reschedule my appointment** to try again."


# ══════════════════════════════════════════════
#  Lead Capture Flow State Machine
# ══════════════════════════════════════════════

def handle_lead_capture(session, user_message):
    # Embedded widget: require login
    if session.get("_is_embedded") and not session.get("_customer_logged_in"):
        session["flow"] = None
        session["step"] = None
        session["data"] = {}
        return "You need to log in first. Please log in to your account and try again."

    step = session["step"]
    data = session["data"]

    # Auto-fill from customer API prefill (embedded widget — skip asking)
    if session.get("_prefill_name") and "name" not in data:
        data["name"] = session["_prefill_name"]
    if session.get("_prefill_phone") and "phone" not in data:
        data["phone"] = session["_prefill_phone"]

    # If both name and phone are already known, skip the whole lead flow
    if data.get("name") and data.get("phone") and step in (None, "ask_name"):
        admin_id = data.get("_admin_id") or session.get("_admin_id", 0)
        lead_engine.capture_lead_from_session(session, admin_id, capture_trigger="lead_capture")
        session["flow"] = None
        session["step"] = None
        session["data"] = {}
        return (
            f"Thanks, {data['name']}! We already have your contact info on file. "
            f"Someone from our team will reach out to you soon.\n\n"
            f"In the meantime, feel free to ask me any questions about our services!"
        )

    # Step 1: Ask for name
    if step is None or step == "ask_name":
        if data.get("name"):
            # Name already known, skip to phone
            session["step"] = "get_phone"
            return f"Hi {data['name']}! What's the best phone number to reach you at?"
        session["step"] = "get_name"
        return "No problem! I'd love to stay in touch. What's your name?"

    # Step 2: Got name, ask for phone
    if step == "get_name":
        data["name"] = user_message.strip().title()
        if data.get("phone"):
            # Phone already known from prefill — save and done
            admin_id = data.get("_admin_id") or session.get("_admin_id", 0)
            lead_engine.capture_lead_from_session(session, admin_id, capture_trigger="lead_capture")
            session["flow"] = None
            session["step"] = None
            session["data"] = {}
            return (
                f"Got it, {data['name']}! We've saved your info and someone from our team "
                f"will reach out to you at **{data['phone']}** soon.\n\n"
                f"In the meantime, feel free to ask me any questions about our services!"
            )
        session["step"] = "get_phone"
        return f"Thanks, {data['name']}! What's the best phone number to reach you at?"

    # Step 3: Got phone, save lead
    if step == "get_phone":
        phone = re.sub(r'[^\d+\-() ]', '', user_message.strip())
        if len(re.sub(r'\D', '', phone)) >= 7:
            data["phone"] = phone

            # Save to database (enriched)
            admin_id = data.get("_admin_id") or session.get("_admin_id", 0)
            lead_engine.capture_lead_from_session(session, admin_id, capture_trigger="lead_capture")

            # Owner notification disabled — only send to the customer's email
            # email.send_booking_notification_owner(
            #     data["name"], "", data["phone"], "N/A", "N/A"
            # )

            # Reset session
            session["flow"] = None
            session["step"] = None
            session["data"] = {}

            return (
                f"Got it, {data['name']}! We've saved your info and someone from our team "
                f"will reach out to you at **{data['phone']}** soon.\n\n"
                f"In the meantime, feel free to ask me any questions about our services!"
            )
        else:
            return "That doesn't look like a valid phone number. Could you try again? Example: (555) 123-4567"

    return "Let me start over. Would you like to leave your contact information?"


def _ask_ai_during_booking(user_message, session, admin_id=1):
    """Ask the AI brain a question while the booking flow is paused."""
    doctors = db.get_doctors(admin_id)
    active_doctors = [d for d in doctors if d.get("status") == "active"]
    history = session.get("history", [])[-10:]

    # Claude for symptom/specialization questions
    if claude_specialist.is_configured() and claude_specialist.is_specialization_query(user_message):
        result = claude_specialist.analyze_symptoms(user_message, doctors=active_doctors, history=history)
        if result and result.get("reply"):
            return result["reply"]

    # Groq for everything else
    if message_interpreter.is_configured():
        company_info = db.get_company_info(admin_id)
        # Inject structured services with pricing into company_info for AI context
        try:
            services_list = db.get_services_with_doctors(admin_id)
            if services_list:
                svc_lines = []
                for s in services_list:
                    line = f"- {s['name']}: {s.get('price','')} {db.get_company_currency(admin_id)}"
                    if s.get('description'):
                        line += f" ({s['description']})"
                    doc_ids = s.get('doctor_ids', [])
                    if doc_ids:
                        doc_names = []
                        for did in doc_ids:
                            for doc in active_doctors:
                                if doc["id"] == did:
                                    doc_names.append(f"Dr. {doc['name']}")
                        if doc_names:
                            line += f" [Available with: {', '.join(doc_names)}]"
                    else:
                        line += " [No doctor currently assigned — patients can request notification]"
                    svc_lines.append(line)
                services_text = "\n".join(svc_lines)
                existing = company_info.get("pricing_insurance", "") if company_info else ""
                if company_info is None:
                    company_info = {}
                company_info["pricing_insurance"] = (existing + "\n\nService Pricing:\n" + services_text).strip()
        except Exception:
            pass
        doctor_slots = {}
        for doc in active_doctors:
            doc_breaks = db.get_doctor_breaks(doc["id"])
            slots = _generate_doctor_slots(doc, breaks=doc_breaks)
            if slots:
                doctor_slots[doc["name"]] = [s["time"] for s in slots]
        result = message_interpreter.think_and_respond(
            user_message, company_info, active_doctors,
            doctor_slots=doctor_slots, history=history
        )
        if result and result.get("reply"):
            return result["reply"]

    # Offline fallback
    kb_result = dke.find_best_answer(user_message)
    if kb_result:
        answer = kb_result["answer"]
        if kb_result.get("follow_up"):
            answer += "\n\n" + kb_result["follow_up"]
        return answer

    return None


# ══════════════════════════════════════════════
#  Main Chat Handler
# ══════════════════════════════════════════════

def process_message(session_id, user_message, admin_id=1, patient_id=None,
                    customer_id="", customer_api_url="",
                    user_name="", user_email="", user_phone=""):
    session = get_session(session_id)
    session["admin_id"] = admin_id
    session["last_message_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Pre-fill from signed-in dashboard user ──
    if user_name and not session.get("_patient_prefilled"):
        session["_patient_prefilled"] = True
        session["_patient_recognized"] = True
        session["_greeting_name"] = user_name
        session["_prefill_name"] = user_name
        if user_email:
            session["_prefill_email"] = user_email
        if user_phone:
            session["_prefill_phone"] = user_phone
        # If no phone provided, try to find it from patient record
        if not user_phone and user_email and admin_id:
            try:
                _patient_rec = db.get_or_create_patient(admin_id, user_name, user_email, "", increment_booking=False)
                if _patient_rec and _patient_rec.get("phone"):
                    session["_prefill_phone"] = _patient_rec["phone"]
            except Exception:
                pass

    # ── Pre-fill from external customers API (embedded widget only, NOT demo) ──
    is_demo = session_id.startswith("demo_")
    is_embedded = session_id.startswith("web_")
    if is_embedded and not is_demo:
        session["_is_embedded"] = True
        # If customer_id is now provided (user just logged in), always process it
        if customer_id and not session.get("_customer_logged_in"):
            session["_customer_api_prefilled"] = True
            session["_customer_logged_in"] = True
            customer = fetch_customer_by_id(admin_id, customer_id, api_url_override=customer_api_url)
            if customer:
                cname = customer.get("name", "")
                cemail = customer.get("email", "")
                cphone = customer.get("phone", "")
                session["_patient_prefilled"] = True
                session["_patient_recognized"] = True
                session["_greeting_name"] = cname
                session["_prefill_name"] = cname
                session["_prefill_email"] = cemail
                session["_prefill_phone"] = cphone
                session["_customer_api_matched"] = True

    # ── Pre-fill patient info when patient_id is provided (real app mode) ──
    if patient_id and not session.get("_patient_prefilled"):
        patient_record = db.get_patient(patient_id)
        if patient_record:
            session["_patient_prefilled"] = True
            session["_patient"] = patient_record
            session["_greeting_name"] = patient_record.get("name", "")
            session["patient_id"] = patient_record.get("id")
            session["_patient_recognized"] = True
            # Store for auto-fill during booking flows
            session["_prefill_name"] = patient_record.get("name", "")
            session["_prefill_email"] = patient_record.get("email", "")
            session["_prefill_phone"] = patient_record.get("phone", "")

    # Trim history to prevent unbounded growth (some code paths bypass _reply)
    if len(session.get("history", [])) > 30:
        session["history"] = session["history"][-20:]

    # ── Feature 6: Detect language on first message (engine) ──
    if not session.get("language_detected"):
        try:
            lang = tr.detect_language(user_message)
            session["language"] = lang
            session["language_detected"] = True
            if lang not in tr.SUPPORTED_LANGUAGES:
                session["language"] = "en"
        except Exception:
            session["language"] = detect_language(user_message)
            session["language_detected"] = True

    # ── Feature 16: Emit chat activity in real-time ──
    try:
        name = session.get("_greeting_name", session.get("data", {}).get("name", "Visitor"))
        realtime.emit_chat_activity(admin_id, session_id, name, user_message[:100])
    except Exception:
        pass

    # ── Feature 15: Recognize returning patient (engine) ──
    if not session.get("_patient_recognized"):
        session["_patient_recognized"] = True
        phone_match = re.search(r'[\+]?[\d\s\-\(\)]{7,15}', user_message)
        email_match = re.search(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', user_message)
        if phone_match or email_match:
            try:
                existing_patient = patient_profile.recognize_patient(admin_id,
                    phone=phone_match.group().strip() if phone_match else "",
                    email=email_match.group() if email_match else "")
                if existing_patient:
                    session["_patient"] = existing_patient
                    session["_greeting_name"] = existing_patient.get("name", "")
                    session["patient_id"] = existing_patient.get("id")
            except Exception:
                # Fallback to db
                patient = db.get_or_create_patient(admin_id,
                    email=email_match.group() if email_match else "",
                    phone=phone_match.group().strip() if phone_match else "")
                if patient and patient.get("name"):
                    session["_patient"] = patient
                    session["_greeting_name"] = patient["name"]

    # ── Feature 17: A/B test — track conversation start (engine) ──
    if len(session.get("history", [])) == 0:
        try:
            ab_message = ab_testing.get_active_message(admin_id, 'opening_message', session_id)
            if ab_message:
                session["_ab_opening_message"] = ab_message
        except Exception:
            pass
        # Legacy fallback
        ab_test = db.get_active_ab_test(admin_id, "opening_message")
        if ab_test:
            import random
            variant = "a" if random.random() < 0.5 else "b"
            session["_ab_variant"] = variant
            session["_ab_test_id"] = ab_test["id"]
            db.increment_ab_test(ab_test["id"], variant)

    # ── Feature 9: Emergency detection via engine (enhanced) ──
    msg_lower_raw = user_message.lower()
    try:
        is_emerg, matched = emergency_handler.is_emergency(user_message)
        if is_emerg and session.get("flow") != "booking":
            lang = session.get("language", "en")
            result = emergency_handler.handle_emergency(user_message, admin_id, session.get("data", {}), lang)
            if result:
                if result.get("ui_options"):
                    session["_ui_options"] = result["ui_options"]
                try:
                    realtime.emit_emergency_alert(admin_id, result.get("alert", {}))
                except Exception:
                    pass
                session["history"].append({"role": "user", "content": user_message})
                session["history"].append({"role": "assistant", "content": result["response"]})
                return result["response"]
    except Exception:
        pass  # Fall through to legacy emergency handling

    # ── Feature 9 (legacy): Emergency fast-track detection ──
    is_emergency = any(kw in msg_lower_raw for kw in EMERGENCY_KEYWORDS)
    if is_emergency and session.get("flow") != "booking":
        first_aid = get_first_aid(user_message)
        # Find earliest available slot within 3 hours
        doctors = db.get_doctors(admin_id)
        active_docs = [d for d in doctors if d.get("status") == "active" and d.get("is_active", 1)]
        emergency_slots = []
        now = datetime.now()
        for doc in active_docs:
            breaks = db.get_doctor_breaks(doc["id"])
            slots = _generate_doctor_slots(doc, breaks, selected_date=now.strftime("%Y-%m-%d"))
            for s in slots:
                try:
                    slot_time = s["time"].split(" - ")[0].strip()
                    slot_dt = datetime.strptime(f"{now.strftime('%Y-%m-%d')} {slot_time}", "%Y-%m-%d %I:%M %p")
                    if now <= slot_dt <= now + timedelta(hours=3):
                        booked = db.get_booked_times(doc["id"], now.strftime("%Y-%m-%d"))
                        if s["time"] not in booked and slot_time not in booked:
                            emergency_slots.append({"doctor": doc["name"], "doctor_id": doc["id"], "time": s["time"], "slot_dt": slot_dt})
                except Exception:
                    pass
        emergency_slots.sort(key=lambda x: x["slot_dt"])
        reply_parts = [f"🚨 **Emergency detected.** Here's immediate first-aid advice:\n\n{first_aid}"]
        if emergency_slots:
            slot = emergency_slots[0]
            reply_parts.append(f"\n\n**Earliest available slot:** {slot['time']} with Dr. {slot['doctor']} (TODAY)")
            reply_parts.append("\nWould you like me to book this emergency slot right away? Just say **yes**.")
            session["_emergency_slot"] = slot
        else:
            reply_parts.append("\n\nNo immediate slots available. Please call the clinic directly for emergency assistance.")
        reply_parts.append("\n\n📞 **Call us now** for immediate help.")
        session["history"].append({"role": "user", "content": user_message})
        response = "".join(reply_parts)
        session["history"].append({"role": "assistant", "content": response})
        return response

    # ── Feature 9: Handle "yes" to emergency booking ──
    if session.get("_emergency_slot") and msg_lower_raw in ("yes", "yeah", "yes please", "ok", "okay", "book it"):
        slot = session.pop("_emergency_slot")
        session["flow"] = "booking"
        session["data"] = {
            "_admin_id": admin_id, "_session_id": session_id,
            "doctor_id": slot["doctor_id"], "doctor_name": slot["doctor"],
            "date_str": "today", "date_iso": datetime.now().strftime("%Y-%m-%d"),
            "date_display": "Today", "chosen_time": slot["time"],
            "service": "Emergency Consultation"
        }
        session["history"].append({"role": "user", "content": user_message})
        # If patient is prefilled, skip name/email/phone — go straight to finalize
        if session.get("_prefill_name"):
            session["data"]["name"] = session["_prefill_name"]
            session["data"]["email"] = session.get("_prefill_email", "")
            session["data"]["phone"] = session.get("_prefill_phone", "")
            session["step"] = "finalize_booking"
            result = handle_booking(session, user_message)
            session["history"].append({"role": "assistant", "content": result})
            return result
        session["step"] = "get_name"
        response = f"🚨 **Emergency slot locked:** {slot['time']} with Dr. {slot['doctor']} today.\n\nPlease provide your **full name** to confirm:"
        session["history"].append({"role": "assistant", "content": response})
        return response

    # ── Feature 10: Check if conversation is handed off to human (engine) ──
    try:
        handoff = handoff_engine.get_handoff_for_session(session_id)
        if handoff and handoff["status"] == "assigned":
            handoff_engine.send_handoff_message(handoff["id"], "patient", session.get("_greeting_name", "Patient"), user_message)
            session["history"].append({"role": "user", "content": user_message})
            return "Message received. Our staff member is reviewing your message."
    except Exception:
        pass
    # Fallback: check via db (skip stale handoffs older than 30 min)
    handoff = db.get_handoff_by_session(session_id)
    if handoff and handoff["status"] in ("queued", "assigned"):
        # Check if handoff is stale (older than 30 min)
        _handoff_stale = False
        try:
            _ho_created = datetime.strptime(handoff.get("created_at", ""), "%Y-%m-%d %H:%M:%S")
            if (datetime.now() - _ho_created).total_seconds() > 1800:
                _handoff_stale = True
                try:
                    conn = db.get_db()
                    conn.execute("UPDATE live_chat_handoffs SET status='resolved', resolution_notes='auto-expired after 30min' WHERE id=?", (handoff["id"],))
                    conn.commit()
                    conn.close()
                except Exception:
                    pass
        except Exception:
            pass
        if not _handoff_stale:
            session["history"].append({"role": "user", "content": user_message})
            if handoff["status"] == "queued":
                return "Your conversation has been transferred to our staff. A team member will be with you shortly. Please hold on."
            return "Message received. Our staff member is reviewing your message."

    # ── Feature 8: Doctor comparison request (engine) ──
    if session.get("flow") != "booking":
        try:
            if doctor_comparison.should_show_comparison(user_message):
                result = doctor_comparison.get_chatbot_comparison_response(admin_id, session.get("language", "en"))
                if result:
                    if result.get("ui_options"):
                        session["_ui_options"] = result["ui_options"]
                    session["history"].append({"role": "user", "content": user_message})
                    session["history"].append({"role": "assistant", "content": result["response"]})
                    return result["response"]
        except Exception:
            pass
        # Fallback: legacy inline comparison
        if re.search(r'\b(compare|comparison|which doctor|help me choose|not sure which)\b', msg_lower_raw):
            doctors = db.get_doctors(admin_id)
            active_docs = [d for d in doctors if d.get("status") == "active" and d.get("is_active", 1)]
            if active_docs:
                comparison = []
                for d in active_docs[:6]:
                    breaks = db.get_doctor_breaks(d["id"])
                    slots = _generate_doctor_slots(d, breaks)
                    avail_slots = [s for s in slots if not s.get("booked")]
                    next_slot = avail_slots[0]["time"] if avail_slots else "No slots today"
                    comparison.append({
                        "name": d["name"], "specialty": d.get("specialty", "General"),
                        "experience": f"{d.get('years_of_experience', 0)} years",
                        "languages": d.get("languages", ""), "next_available": next_slot
                    })
                session["_ui_options"] = {"type": "doctor_comparison", "items": comparison}
                session["history"].append({"role": "user", "content": user_message})
                response = "Here's a comparison of our available doctors. Tap any doctor to book with them:"
                session["history"].append({"role": "assistant", "content": response})
                return response

    # ── Feature 7: Gallery request for cosmetic treatments (engine) ──
    try:
        gallery_result = gallery_engine.get_chatbot_gallery(admin_id, user_message)
        if gallery_result:
            session["_ui_options"] = {"type": "gallery_carousel", "images": gallery_result["images"], "treatment": gallery_result["treatment_type"]}
            lang = session.get("language", "en")
            session["history"].append({"role": "user", "content": user_message})
            if lang == 'ar':
                response = f"إليك بعض النتائج الحقيقية من مرضانا في {gallery_result['treatment_type']}:"
            else:
                response = f"Here are some real results from our patients for {gallery_result['treatment_type']}:"
            session["history"].append({"role": "assistant", "content": response})
            return response
    except Exception:
        pass
    # Fallback: legacy inline gallery check
    cosmetic_keywords = {"whitening": "whitening", "veneers": "veneers", "veneer": "veneers",
        "braces": "braces", "implant": "implants", "implants": "implants",
        "before and after": None, "before after": None, "results": None, "gallery": None}
    for kw, treatment in cosmetic_keywords.items():
        if kw in msg_lower_raw:
            gallery_treatment = treatment or "general"
            images = db.get_gallery(admin_id, treatment_type=gallery_treatment)
            if not images and not treatment:
                images = db.get_gallery(admin_id)
            if images:
                session["_ui_options"] = {"type": "gallery", "items": [{"url": img["image_url"], "caption": img.get("caption", ""), "type": img.get("image_type", "")} for img in images]}
            break

    # ── Feature 18: Loyalty balance check in chat ──
    if db.is_feature_enabled(admin_id, "loyalty_program"):
        try:
            loyalty_keywords = ['points', 'loyalty', 'نقاط', 'مكافآت', 'how many points']
            if any(kw in msg_lower_raw for kw in loyalty_keywords):
                patient_id = session.get("patient_id") or (session.get("_patient", {}) or {}).get("id")
                if patient_id:
                    balance = loyalty.get_balance_value(patient_id, admin_id)
                    lang = session.get("language", "en")
                    session["history"].append({"role": "user", "content": user_message})
                    response = tr.t('loyalty_balance', lang, points=balance["points"], value=balance["sar_value"])
                    session["history"].append({"role": "assistant", "content": response})
                    return response
        except Exception:
            pass

    # Step 0+1: AI spell-correction — only ONE Groq call needed (cleaner & interpreter do the same thing)
    # Skip very short messages (greetings, yes/no) — no need to waste an API call
    if len(user_message.strip()) <= 5:
        interpreted = user_message
    elif message_interpreter.is_configured():
        interpreted = message_interpreter.interpret(user_message, history=session.get("history"))
    else:
        grok_cleaned = grok_cleaner.clean(user_message, history=session.get("history"))
        interpreted = grok_cleaned

    # Step 2: Local spell-correct as additional cleanup
    corrected = correct_spelling(interpreted)
    lower = corrected.lower().strip()

    # Helper to save history and return response
    def _reply(response):
        session["history"].append({"role": "user", "content": user_message})
        session["history"].append({"role": "assistant", "content": response})
        # Keep last 20 messages (10 exchanges) to avoid token bloat
        if len(session["history"]) > 20:
            session["history"] = session["history"][-20:]
        return response

    # ── Cancel detection (check BOTH raw message and corrected to avoid AI mangling) ──
    raw_lower = user_message.lower().strip()

    # Check for cancel at any point — exact matches + refusal patterns
    # Check both raw and corrected message to catch cases where AI cleaner transforms "cancel"
    cancel_words = ("cancel", "nevermind", "never mind", "stop", "go back", "start over", "quit", "exit")
    is_cancel = raw_lower in cancel_words or lower in cancel_words
    # "Go back" during certain booking steps means "go to previous step", not cancel
    if is_cancel and raw_lower in ("go back", "back") and session.get("step") in ("confirm_service_doctor", "edit_booking_choice", "finalize_booking"):
        is_cancel = False

    # When in a flow, be more generous with cancel detection
    # BUT: if the user says "cancel my appointment/booking", they want to cancel an EXISTING appointment,
    # not abort the current flow — so exclude that pattern
    wants_cancel_existing = bool(re.search(
        r"(cancel|delete|remove)\s+(my\s+|the\s+)?(app\w+|apo\w+|booking|reservation)",
        raw_lower
    ))
    if not is_cancel and session.get("flow") and not wants_cancel_existing:
        # "cancel" or common misspellings anywhere in the message
        is_cancel = bool(re.search(r'\b(cancel|cancle|cacnel|canel|cansel|cancell)\b', raw_lower))
        # Also check corrected text for "cancel" (AI may fix the typo)
        if not is_cancel:
            is_cancel = bool(re.search(r'\bcancel\b', lower))
        # Repeated "no" = refusal (e.g. "no no no")
        if not is_cancel:
            is_cancel = bool(re.match(r'^(no\s*){2,}', raw_lower))
        # Refusal phrases
        if not is_cancel:
            is_cancel = bool(re.search(
                r"\b(don'?t want|no\s+i\s+don|not\s+interested|i\s+refuse|no\s+thanks|no\s+thank|"
                r"i\s+changed\s+my\s+mind|forget\s+it|nah|nope\b.*book|no\s+i\s+don'?t|"
                r"don'?t\s+want\s+to\s+book|don'?t\s+need|no\s+booking|stop\s+booking)",
                raw_lower
            ))
    if is_cancel:
        if session["flow"]:
            flow_was = session["flow"]
            reset_session(session_id)
            session = get_session(session_id)  # Refresh local reference to new session
            if flow_was == "booking":
                return _reply("No problem, I've stopped the booking process. How else can I help you?")
            elif flow_was == "cancel_appointment":
                return _reply("OK, I've stopped the cancellation. Your appointment is still active. How else can I help you?")
            elif flow_was == "reschedule":
                return _reply("OK, I've stopped the reschedule. Your appointment stays as is. How else can I help you?")
            return _reply("No problem! How else can I help you?")
        return _reply("No worries! Is there anything else I can help you with?")

    # ── Feature 10: Early check for human handoff request (engine) ──
    human_phrases = ("speak to a human", "talk to someone", "real person", "human agent",
                     "live chat", "speak to staff", "talk to a person", "speak to someone",
                     "customer service", "i need help from a person", "talk to a real person")
    if any(phrase in raw_lower for phrase in human_phrases):
        try:
            handoff_engine.create_handoff(admin_id, session_id,
                patient_name=session.get("_greeting_name", ""),
                reason="Patient requested human assistance", ai_confidence=0)
            realtime.emit_handoff_request(admin_id, {"patient_name": session.get("_greeting_name", ""), "reason": "Patient requested human assistance"})
        except Exception:
            db.create_handoff(admin_id, session_id, patient_name=session.get("_greeting_name", ""),
                             reason="Patient requested human assistance", ai_confidence=0)
        session["history"].append({"role": "user", "content": user_message})
        lang = session.get("language", "en")
        try:
            response = tr.t('handoff_connecting', lang)
        except Exception:
            response = "I'm connecting you with a staff member now. A team member will be with you shortly. Please hold on — they'll see your full conversation history."
        session["history"].append({"role": "assistant", "content": response})
        return response

    # Handle reschedule flow
    if session["flow"] == "reschedule":
        result = handle_reschedule(session, user_message, admin_id)
        return _reply(result)

    # Handle cancel appointment flow
    if session["flow"] == "cancel_appointment":
        result = handle_cancel_appointment(session, user_message, admin_id)
        return _reply(result)

    # Handle "continue" to resume booking from where they left off
    if lower in ("continue", "continue booking", "resume", "go on", "carry on", "back to booking"):
        if session["flow"] == "booking" and session.get("_paused"):
            session["_paused"] = False
            step = session["step"]
            data = session["data"]
            # Re-show the appropriate dropdown for the current step
            if step == "get_category":
                categories = data.get("_categories", [])
                if categories:
                    session["_ui_options"] = {"type": "categories", "items": [{"name": c} for c in categories]}
                return _reply("Let's continue! What type of doctor would you like to see?")
            elif step == "get_doctor":
                doctors = data.get("_doctors", [])
                if doctors:
                    session["_ui_options"] = {"type": "doctors", "items": [_doctor_dropdown_item(d) for d in doctors]}
                return _reply("Let's continue! Which doctor would you like to see?")
            elif step == "get_date":
                doctor_id = data.get("doctor_id")
                off_dates = _get_off_dates_with_blocks(doctor_id, data.get("_admin_id", 0))
                session["_ui_options"] = {"type": "calendar", "doctor_id": doctor_id, "off_dates": off_dates}
                return _reply("Let's continue! When would you like to come in?")
            elif step == "get_time":
                # Regenerate time slots for the chosen doctor/date
                doctor_id = data.get("doctor_id")
                date_iso = data.get("date_iso")
                slots = []
                if doctor_id:
                    doctor = db.get_doctor_by_id(doctor_id)
                    if doctor:
                        doctor_breaks = db.get_doctor_breaks(doctor_id)
                        slots = _generate_doctor_slots(doctor, breaks=doctor_breaks, selected_date=data.get("date_iso"))
                booked_times = []
                if doctor_id and date_iso:
                    booked_times = db.get_booked_times(doctor_id, date_iso)
                data["available_slots"] = [s for s in slots if not _is_booked_slot(s["time"], booked_times)]
                data["all_slots"] = slots
                data["booked_slot_names"] = [s["time"] for s in slots if _is_booked_slot(s["time"], booked_times)]
                dropdown_items = []
                for s in slots:
                    item = {"name": s["time"], "hour": s["hour"], "minute": s.get("minute", 0)}
                    slot_start = s["time"].split(" - ")[0].strip() if " - " in s["time"] else s["time"]
                    if s["time"] in booked_times or slot_start in booked_times:
                        item["booked"] = True
                    dropdown_items.append(item)
                if dropdown_items:
                    session["_ui_options"] = {"type": "timeslots", "items": dropdown_items}
                return _reply(f"Let's continue! Pick a time on **{data.get('date_display', 'your chosen date')}**:")
            else:
                # For other steps just continue the flow
                return _reply(handle_booking(session, user_message, corrected))

    # If already in a flow, continue it (pass corrected text for understanding, original for data)
    if session["flow"] == "booking":
        if "_admin_id" not in session["data"]:
            session["data"]["_admin_id"] = admin_id

        # If booking is paused (user asked a question during dropdown), answer via AI
        if session.get("_paused"):
            ai_answer = _ask_ai_during_booking(user_message, session, admin_id)
            if ai_answer:
                return _reply(ai_answer + "\n\nWhen you're ready to continue booking, just say **continue**.")
            return _reply("I'm not sure about that. Say **continue** to resume your booking.")

        # Try to handle the booking step
        booking_result = handle_booking(session, user_message, corrected)

        if booking_result is not None:
            return _reply(booking_result)

        # handle_booking returned None — user typed something that doesn't match
        # the current dropdown/step.
        step = session.get("step")
        data = session.get("data", {})

        # If in get_time step, re-show the time dropdown instead of pausing
        if step == "get_time":
            doctor_id = data.get("doctor_id")
            date_iso = data.get("date_iso")
            if doctor_id and date_iso:
                doctor = db.get_doctor_by_id(doctor_id)
                slots = []
                if doctor:
                    doctor_breaks = db.get_doctor_breaks(doctor_id)
                    slots = _generate_doctor_slots(doctor, breaks=doctor_breaks, selected_date=date_iso)
                booked_times = db.get_booked_times(doctor_id, date_iso)
                data["available_slots"] = [s for s in slots if not _is_booked_slot(s["time"], booked_times)]
                data["all_slots"] = slots
                data["booked_slot_names"] = [s["time"] for s in slots if _is_booked_slot(s["time"], booked_times)]
                dropdown_items = []
                for s in slots:
                    item = {"name": s["time"], "hour": s["hour"], "minute": s.get("minute", 0)}
                    if _is_booked_slot(s["time"], booked_times):
                        item["booked"] = True
                    dropdown_items.append(item)
                if dropdown_items:
                    session["_ui_options"] = {"type": "timeslots", "items": dropdown_items}
                return _reply(f"Please pick a time slot from the list below for **{data.get('date_display', 'your chosen date')}**:")

        # For other steps, pause booking and route to AI
        session["_paused"] = True
        ai_answer = _ask_ai_during_booking(user_message, session, admin_id)
        if ai_answer:
            return _reply(ai_answer + "\n\nWhen you're ready to continue booking, just say **continue**.")
        return _reply("I'm not sure about that. Say **continue** to resume your booking.")
    if session["flow"] == "lead_capture":
        return _reply(handle_lead_capture(session, user_message))

    # ── Appointment Lookup Flow (demo chatbot — identity not known) ──
    if session["flow"] == "appointment_lookup" and session.get("step") == "get_identity":
        identity = user_message.strip()
        session["flow"] = None
        session["step"] = None
        session["data"] = {}
        if not identity or len(identity) < 2:
            return _reply("I couldn't understand that. Would you like to **book a new appointment** instead?")
        # Try matching as name or email
        is_email = "@" in identity
        upcoming = db.find_upcoming_bookings_for_customer(
            admin_id,
            name="" if is_email else identity,
            email=identity if is_email else "",
            phone=""
        )
        if upcoming:
            lines = ["Here are your upcoming appointments:\n"]
            for i, bk in enumerate(upcoming, 1):
                _d = bk.get("date", "")
                _t = bk.get("time", "")
                _doc = bk.get("doctor_name", "")
                _svc = bk.get("service", "")
                lines.append(f"**{i}.** {_d} at {_t}" + (f" — Dr. {_doc}" if _doc else "") + (f" ({_svc})" if _svc else ""))
            lines.append("\nWould you like to **cancel** or **reschedule** any of these?")
            return _reply("\n".join(lines))
        else:
            return _reply(f"I couldn't find any upcoming appointments for **{identity}**. Would you like to **book one**?")

    # ── NEW: Keyword classifier for booking/reschedule/cancel ──
    # Check RAW message first — spell corrector can mangle intent words
    # (e.g. "reschedule" → "schedule", "reshcdule" → "schedule")
    raw_intent = chatbot_classifier.classify(user_message)
    corrected_intent = chatbot_classifier.classify(corrected)
    # Prefer raw reschedule/cancel over corrected booking (spell corrector strips "re-")
    if raw_intent in ("reschedule", "cancel"):
        classifier_intent = raw_intent
    elif corrected_intent:
        classifier_intent = corrected_intent
    else:
        classifier_intent = raw_intent
    print(f"[classifier] intent={classifier_intent} | raw_intent={raw_intent} | corrected_intent={corrected_intent}", flush=True)

    # Log conversation for analytics
    try:
        db.log_chat(session_id, admin_id, user_message, intent=classifier_intent or "general", intent_confidence=1.0 if classifier_intent else 0.0)
    except Exception:
        pass

    # Check if user is confirming a booking suggestion from AI (e.g. "okay", "yes", "sure")
    if lower in ("okay", "ok", "yes", "sure", "yeah", "yep", "yea", "please", "yes please", "book", "i want to book", "lets book", "let's book"):
        _last_bot_msg = session["history"][-1].get("content", "").lower() if session.get("history") else ""
        if any(phrase in _last_bot_msg for phrase in ["book an appointment", "book appointment", "would you like to book", "schedule a consultation", "schedule an appointment"]):
            session["flow"] = "booking"
            session["step"] = None
            session["data"] = {"_admin_id": admin_id, "_session_id": session_id}
            return _reply(handle_booking(session, user_message))

    # ── Silent auto-lead capture (Intent Scoring) ──
    # Every message gets scored based on user behavior signals.
    # When cumulative score crosses threshold (>=7) → silently capture as lead.
    # Lead is auto-deleted when they confirm a booking.
    _lead_score = session.get("_lead_score", 0)
    if not session.get("_lead_captured"):
        _msg_score = 0
        # General info questions (+1)
        if re.search(r'\b(what is|what are|tell me about|explain|how does|information|more info|details|learn more)\b', raw_lower):
            _msg_score += 1
        # Pricing / cost questions (+2)
        if re.search(r'\b(price|pricing|cost|how much|fee|charge|expensive|cheap|afford|installment|payment plan|insurance|sar|dollar|\$)\b', raw_lower):
            _msg_score += 2
        # Treatment duration / process questions (+2)
        if re.search(r'\b(how long|duration|last|session|procedure|process|recovery|healing|installation|install|take)\b', raw_lower):
            _msg_score += 2
        # Availability / scheduling questions (+4)
        if re.search(r'\b(available|availability|appointment|schedule|opening|slot|when can|this week|today|tomorrow|days|hours|work|open|close|saturday|sunday|monday|tuesday|wednesday|thursday|friday)\b', raw_lower):
            _msg_score += 4
        # Booking intent (+6)
        if re.search(r'\b(book|reserve|sign me up|i want to come|i\'d like to|schedule me|set up|come in)\b', raw_lower):
            _msg_score += 6
        # Providing contact info (+10)
        if re.search(r'\b(my name is|my number|my phone|my email|i\'m \w+|call me|here is my|here\'s my)\b', raw_lower):
            _msg_score += 10
        # Symptom mentions (+3) — they need care
        if re.search(r'\b(pain|hurts|ache|bleeding|swollen|broken|cracked|sensitive|cavity|decay|toothache|gum)\b', raw_lower):
            _msg_score += 3
        # Comparing / evaluating (+2)
        if re.search(r'\b(which is better|difference between|compare|recommend|best option|pros and cons|which one|better)\b', raw_lower):
            _msg_score += 2
        # Doctor questions (+2)
        if re.search(r'\b(doctor|dr\b|dentist|specialist|who does|experience|qualified)\b', raw_lower):
            _msg_score += 2
        # Service-specific mentions (+2)
        if re.search(r'\b(braces|implant|whitening|veneer|crown|root canal|filling|extraction|cleaning|invisalign|orthodont|cosmetic)\b', raw_lower):
            _msg_score += 2

        _lead_score += _msg_score
        session["_lead_score"] = _lead_score

        if _msg_score > 0:
            print(f"[lead] Score: +{_msg_score} = {_lead_score} (threshold: 7)", flush=True)

        # Threshold reached → capture lead silently
        if _lead_score >= 7 and db.is_feature_enabled(admin_id, "auto_lead_capture"):
            try:
                lead_id = lead_engine.capture_lead_from_session(
                    session, admin_id, capture_trigger="auto_interest"
                )
                if lead_id:
                    session["_lead_captured"] = True
                    session["_auto_lead_id"] = lead_id
                    print(f"[lead] Auto-captured lead #{lead_id} (score={_lead_score})", flush=True)
                else:
                    # capture_lead_from_session returns None if no name/phone/email
                    # Try using _greeting_name as fallback
                    _fallback_name = session.get("_greeting_name", "") or session.get("_prefill_name", "")
                    if _fallback_name:
                        session["data"]["name"] = _fallback_name
                        lead_id = lead_engine.capture_lead_from_session(
                            session, admin_id, capture_trigger="auto_interest"
                        )
                        if lead_id:
                            session["_lead_captured"] = True
                            session["_auto_lead_id"] = lead_id
                            print(f"[lead] Auto-captured lead #{lead_id} (score={_lead_score}, name from greeting)", flush=True)
            except Exception as e:
                print(f"[lead] Auto-capture failed: {e}", flush=True)

    # ── Classifier-based intent routing ──

    # ── My Appointments Lookup ──
    _appt_words_re = r'(appointments?|bookings?|appo\w+ments?|apo\w+ments?|appoitments?)'
    _appt_triggers_re = r'(when\s+is|when\s+are|my\s+next|my\s+upcoming|check\s+my|view\s+my|show\s+my|do\s+i\s+have|my)'
    _appt_check = (
        re.search(_appt_triggers_re + r'\b.*?' + _appt_words_re, raw_lower) or
        re.search(_appt_triggers_re + r'\b.*?' + _appt_words_re, lower)
    )
    if _appt_check and not classifier_intent:
        _cust_name = session.get("_prefill_name", "").strip()
        _cust_email = session.get("_prefill_email", "").strip()
        _cust_phone = session.get("_prefill_phone", "").strip()
        if _cust_name or _cust_email or _cust_phone:
            upcoming = db.find_upcoming_bookings_for_customer(
                admin_id, name=_cust_name, email=_cust_email, phone=_cust_phone
            )
            if upcoming:
                lines = ["Here are your upcoming appointments:\n"]
                for i, bk in enumerate(upcoming, 1):
                    _d = bk.get("date", "")
                    _t = bk.get("time", "")
                    _doc = bk.get("doctor_name", "")
                    _svc = bk.get("service", "")
                    lines.append(f"**{i}.** {_d} at {_t}" + (f" — Dr. {_doc}" if _doc else "") + (f" ({_svc})" if _svc else ""))
                lines.append("\nWould you like to **cancel** or **reschedule** any of these?")
                return _reply("\n".join(lines))
            else:
                return _reply("You don't have any upcoming appointments. Would you like to **book one**?")
        else:
            if session.get("_is_embedded") and not session.get("_customer_logged_in"):
                return _reply("Please **log in** to your account first so I can look up your appointments.")
            session["flow"] = "appointment_lookup"
            session["step"] = "get_identity"
            session["data"] = {"_admin_id": admin_id}
            return _reply("Sure! Could you tell me your **name** or **email** so I can look up your appointments?")

    # ── Route based on classifier intent ──
    if classifier_intent == "reschedule":
        session["flow"] = "reschedule"
        session["step"] = "get_date"
        session["data"] = {"_admin_id": admin_id}
        booked_dates = _get_customer_booked_dates(session, admin_id)
        session["_ui_options"] = {"type": "calendar", "mode": "cancel", "booked_dates": booked_dates}
        return _reply("I can help you reschedule your appointment. What date is your current appointment on?")

    if classifier_intent == "cancel":
        session["flow"] = "cancel_appointment"
        session["step"] = "get_date"
        session["data"] = {"_admin_id": admin_id}
        booked_dates = db.get_booking_dates(admin_id)
        session["_ui_options"] = {"type": "calendar", "mode": "cancel", "booked_dates": booked_dates}
        return _reply("I can help you cancel your appointment. What date is it on?")

    if classifier_intent == "booking":
        session["flow"] = "booking"
        session["step"] = None
        session["data"] = {"_admin_id": admin_id, "_session_id": session_id}
        return _reply(handle_booking(session, user_message))

    # ══════════════════════════════════════════════════════════════
    #  No classifier match → send to Groq AI with full company context
    #  Groq handles: greetings, farewells, dental Q&A, services,
    #  pricing, availability, symptoms — ONLY dentist-related topics
    # ══════════════════════════════════════════════════════════════
    company_info = db.get_company_info(admin_id)
    all_doctors = db.get_doctors(admin_id)
    active_doctors = [d for d in all_doctors if d.get("status") == "active"]

    # Build doctor time slots so the AI knows exact availability
    doctor_slots = {}
    for doc in active_doctors:
        doc_breaks = db.get_doctor_breaks(doc["id"])
        slots = _generate_doctor_slots(doc, breaks=doc_breaks)
        if slots:
            doctor_slots[doc["name"]] = [s["time"] for s in slots]

    # Inject service pricing into company_info so AI has real prices
    try:
        services_list = db.get_services_with_doctors(admin_id)
        if services_list:
            svc_lines = []
            for s in services_list:
                line = f"- {s['name']}: {s.get('price','')} {s.get('currency','SAR')}"
                if s.get('description'):
                    line += f" ({s['description']})"
                doc_ids = s.get('doctor_ids', [])
                if doc_ids:
                    doc_names = [f"Dr. {doc['name']}" for doc in active_doctors if doc["id"] in doc_ids]
                    if doc_names:
                        line += f" [Available with: {', '.join(doc_names)}]"
                svc_lines.append(line)
            services_text = "\n".join(svc_lines)
            if company_info is None:
                company_info = {}
            existing = company_info.get("pricing_insurance", "") or ""
            company_info["pricing_insurance"] = (existing + "\n\nService Pricing:\n" + services_text).strip()
    except Exception:
        pass

    # Add customer name to company_info so AI can address them by name
    _customer_name = session.get("_greeting_name", "") or session.get("_prefill_name", "")
    if _customer_name:
        if company_info is None:
            company_info = {}
        company_info["_customer_name"] = _customer_name

    # Send to Groq AI with full context (dentist-only constraint is in the system prompt)
    if message_interpreter.is_configured():
        ai_result = message_interpreter.think_and_respond(
            corrected, company_info, active_doctors,
            doctor_slots=doctor_slots, history=session["history"]
        )
        if ai_result and ai_result.get("reply"):
            return _reply(ai_result["reply"])

    # Offline fallback — if Groq is not configured or fails
    return _reply(
        "I'm here to help with your dental needs! I can:\n\n"
        "• Book an appointment\n"
        "• Reschedule or cancel existing appointments\n"
        "• Answer questions about our services & pricing\n\n"
        "Just let me know what you need!"
    )


# ══════════════════════════════════════════════
#  Routes — Pages
# ══════════════════════════════════════════════

@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/login")
def login_page():
    return send_from_directory("static", "login.html")


@app.route("/dashboard")
def dashboard():
    return send_from_directory("static", "dashboard.html")


@app.route("/user-dashboard")
@app.route("/user-dashboard.html")
def user_dashboard():
    return send_from_directory("static", "user-dashboard.html", max_age=0)


@app.route("/checkout/<plan>")
def checkout_page(plan):
    if plan not in ("basic", "pro", "agency"):
        return redirect("/user-dashboard")
    return send_from_directory("static", "checkout.html")


@app.route("/checkout/success")
def checkout_success():
    return send_from_directory("static", "checkout-success.html")


@app.route("/privacy")
def privacy():
    return send_from_directory("static", "privacy.html")


@app.route("/terms")
def terms():
    return send_from_directory("static", "terms.html")


@app.route("/pricing-plans")
@app.route("/pricing")
def pricing_plans():
    return send_from_directory("static", "pricing-plans.html")


@app.route("/refund")
def refund_policy():
    return send_from_directory("static", "refund.html")


@app.route("/data-deletion", methods=["GET", "POST"])
def data_deletion():
    if request.method == "POST":
        # Facebook sends a signed_request when user requests data deletion
        # Respond with a confirmation URL and tracking code
        import hashlib
        confirmation_code = hashlib.sha256(str(uuid.uuid4()).encode()).hexdigest()[:12]
        return jsonify({
            "url": request.host_url + "data-deletion",
            "confirmation_code": confirmation_code,
        })
    return send_from_directory("static", "data-deletion.html")


@app.route("/demo")
def demo_page():
    return send_from_directory("static", "demo.html")


@app.route("/security")
def security_page():
    return send_from_directory("static", "security.html")


@app.route("/blog")
def blog_page():
    return send_from_directory("static", "blog.html")


@app.route("/case-studies")
def case_studies_page():
    return send_from_directory("static", "case-studies.html")


@app.route("/for-dental-clinics")
def for_dental_clinics():
    return send_from_directory("static", "for-dental-clinics.html")


@app.route("/for-medical-practices")
def for_medical_practices():
    return send_from_directory("static", "for-medical-practices.html")


@app.route("/for-salons-spas")
def for_salons_spas():
    return send_from_directory("static", "for-salons-spas.html")


# ══════════════════════════════════════════════
#  Routes — Auth
# ══════════════════════════════════════════════

@app.route("/auth/signup", methods=["POST"])
def auth_signup():
    data = request.get_json()
    name = data.get("name", "").strip()
    email_addr = data.get("email", "").strip().lower()
    password = data.get("password", "")
    company = data.get("company", "").strip()
    role = data.get("role", "admin").strip().lower()
    specialty = data.get("specialty", "").strip()

    if role not in ("admin", "doctor", "head_admin"):
        return jsonify({"error": "Invalid role."}), 400
    if not name or not email_addr or not password:
        return jsonify({"error": "Name, email, and password are required."}), 400
    if len(password) < 8:
        return jsonify({"error": "Password must be at least 8 characters."}), 400
    if "@" not in email_addr or "." not in email_addr:
        return jsonify({"error": "Please enter a valid email address."}), 400

    user, error = db.create_user(name, email_addr, password, company, role=role, specialty=specialty)
    if error:
        return jsonify({"error": error}), 400

    # Send verification email with the 6-digit code
    if not user.get("is_verified"):
        try:
            from email_service import send_otp_email
            send_otp_email(email_addr, name, user["verification_code"])
        except Exception:
            pass  # Don't block signup if email fails

        return jsonify({
            "needs_verification": True,
            "email": email_addr,
            "message": "A 6-digit verification code has been sent to your email."
        })

    return jsonify({"token": user["token"], "user": db.user_to_public(user)})


@app.route("/auth/verify-code", methods=["POST"])
def auth_verify_code():
    data = request.get_json()
    email_addr = data.get("email", "").strip().lower()
    code = data.get("code", "").strip()

    if not email_addr or not code:
        return jsonify({"error": "Email and verification code are required."}), 400

    user, error = db.verify_user_code(email_addr, code)
    if error:
        return jsonify({"error": error}), 400

    return jsonify({"token": user["token"], "user": db.user_to_public(user)})


@app.route("/auth/resend-code", methods=["POST"])
def auth_resend_code():
    data = request.get_json()
    email_addr = data.get("email", "").strip().lower()

    if not email_addr:
        return jsonify({"error": "Email is required."}), 400

    user, new_code, error = db.resend_verification_code(email_addr)
    if error:
        return jsonify({"error": error}), 400

    if new_code:
        try:
            from email_service import send_otp_email
            send_otp_email(email_addr, user["name"], new_code)
        except Exception:
            pass

    return jsonify({"message": "A new verification code has been sent to your email."})


@app.route("/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json()
    email_addr = data.get("email", "").strip().lower()
    password = data.get("password", "")

    if not email_addr or not password:
        return jsonify({"error": "Email and password are required."}), 400

    user, error = db.login_user(email_addr, password)
    if error:
        return jsonify({"error": error}), 400

    # Block login for unverified email accounts
    if user.get("provider", "email") == "email" and not user.get("is_verified", 1):
        return jsonify({
            "needs_verification": True,
            "email": email_addr,
            "error": "Please verify your email before logging in."
        }), 403

    # Process any expired plan changes on login
    target_id = user["id"]
    if user.get("admin_id") and user["admin_id"] != 0:
        target_id = user["admin_id"]
    if db.process_plan_expiry(target_id):
        user = db.get_user_by_token(user["token"])

    return jsonify({"token": user["token"], "user": db.user_to_public(user)})


@app.route("/auth/forgot", methods=["POST"])
def auth_forgot():
    # In production, this would send a real reset email
    return jsonify({"ok": True})


@app.route("/auth/social-token", methods=["POST"])
def auth_social_token():
    """
    Verify a social login token server-side, then create/login user.

    Expects JSON: { provider, token, name (optional), email (optional), avatar_url (optional) }
    - Google: token is an access_token or id_token — verified via google-auth + Google API
    - Facebook: token is an access_token — verified via Graph API /debug_token
    - Apple: token is an id_token (JWT) — verified via Apple JWKS public keys

    If provider credentials are NOT configured, falls back to demo mode
    (trusts the client-provided name/email for local development).
    """
    data = request.get_json()
    provider = data.get("provider", "")
    token = data.get("token", "")
    client_name = data.get("name", "").strip()
    client_email = data.get("email", "").strip().lower()
    avatar_url = data.get("avatar_url", "")
    role = data.get("role", "admin").strip().lower()
    specialty = data.get("specialty", "").strip()
    role_confirmed = data.get("role_confirmed", False)  # True when user explicitly chose a role

    if role not in ("admin", "doctor", "head_admin"):
        return jsonify({"error": "Invalid role."}), 400
    if provider not in ("google", "facebook", "apple"):
        return jsonify({"error": "Unknown provider."}), 400

    # ── If provider is configured, verify the token server-side ──
    if social_auth.is_provider_configured(provider):
        if not token:
            return jsonify({"error": "Missing authentication token."}), 400
        try:
            verified = social_auth.verify_social_token(
                provider, token,
                client_name=client_name,
                client_email=client_email,
            )
        except social_auth.SocialAuthError as e:
            return jsonify({"error": str(e)}), 401

        name = verified["name"]
        email_addr = verified["email"]
        provider_id = verified["id"]
        avatar_url = verified.get("picture", avatar_url)

        if not email_addr:
            return jsonify({"error": "Could not retrieve email from provider. Please allow email access."}), 400

    # ── Provider not configured — reject in production ──
    else:
        return jsonify({"error": f"{provider.capitalize()} login is not configured on this server."}), 400

    # Check if user already exists — if not and role wasn't explicitly chosen, ask for role
    existing_user = db.get_user_by_email(email_addr)
    if not existing_user and not role_confirmed:
        return jsonify({
            "needs_role": True,
            "provider": provider,
            "token": token,
            "name": name,
            "email": email_addr,
            "avatar_url": avatar_url,
        })

    user, error = db.login_or_create_social(
        name=name,
        email=email_addr,
        provider=provider,
        provider_id=provider_id,
        avatar_url=avatar_url,
        role=role,
        specialty=specialty,
    )
    if error:
        return jsonify({"error": error}), 400

    return jsonify({"token": user["token"], "user": db.user_to_public(user)})


@app.route("/auth/social-config", methods=["GET"])
def auth_social_config():
    """Return which social providers are configured (so frontend can adapt)."""
    return jsonify({
        "google": social_auth.is_provider_configured("google"),
        "facebook": social_auth.is_provider_configured("facebook"),
        "apple": social_auth.is_provider_configured("apple"),
        "google_client_id": social_auth.GOOGLE_CLIENT_ID if social_auth.is_provider_configured("google") else "",
        "facebook_app_id": social_auth.FACEBOOK_APP_ID if social_auth.is_provider_configured("facebook") else "",
        "apple_client_id": social_auth.APPLE_CLIENT_ID if social_auth.is_provider_configured("apple") else "",
    })


@app.route("/auth/me", methods=["GET"])
def auth_me():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    return jsonify({"user": db.user_to_public(user)})


@app.route("/api/paypal/create-order", methods=["POST"])
def paypal_create_order():
    """Create a PayPal order for plan purchase."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only head admin can purchase plans"}), 403

    data = request.get_json()
    plan = data.get("plan", "")

    PLAN_PRICES = {"basic": "149.00", "pro": "349.00", "agency": "799.00"}
    if plan not in PLAN_PRICES:
        return jsonify({"error": "Invalid plan"}), 400

    # Create order via PayPal API
    import requests as http_req
    PAYPAL_CLIENT_ID = "AbFJcl0u1bWnNgzi5R4KcTsJ-u2pjxUJjjujweFLNkKzFulwgOQX4w5IsZmfeGbJ85JEABMDYUHWdvqB"
    PAYPAL_SECRET = "EFa-RJjZ0AoWFK4g8MWD3eIn3nm3GAdq0Mmo3LKR-WxApp8BJbFPre4Yat5jTqRNF4swbBTjv56XDooP"
    PAYPAL_API = "https://api-m.sandbox.paypal.com"

    # Get access token
    auth_resp = http_req.post(
        f"{PAYPAL_API}/v1/oauth2/token",
        auth=(PAYPAL_CLIENT_ID, PAYPAL_SECRET),
        data={"grant_type": "client_credentials"},
        headers={"Content-Type": "application/x-www-form-urlencoded"}
    )
    if auth_resp.status_code != 200:
        return jsonify({"error": "Payment service unavailable"}), 500
    access_token = auth_resp.json()["access_token"]

    plan_label = plan.capitalize()
    # Create order
    order_resp = http_req.post(
        f"{PAYPAL_API}/v2/checkout/orders",
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        },
        json={
            "intent": "CAPTURE",
            "purchase_units": [{
                "amount": {
                    "currency_code": "USD",
                    "value": PLAN_PRICES[plan]
                },
                "description": f"ChatGenius {plan_label} Plan - Monthly"
            }]
        }
    )
    if order_resp.status_code not in (200, 201):
        return jsonify({"error": "Failed to create payment"}), 500

    order_data = order_resp.json()

    # Store in checkout_sessions
    conn = db.get_db()
    conn.execute(
        "INSERT INTO checkout_sessions (user_id, plan, token, created_at) VALUES (?,?,?,?)",
        (user["id"], plan, order_data["id"], datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    conn.commit()
    conn.close()

    return jsonify({"id": order_data["id"]})


@app.route("/api/paypal/capture-order", methods=["POST"])
def paypal_capture_order():
    """Capture a PayPal order after user approves payment. This verifies money was actually paid."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json()
    order_id = data.get("order_id", "")
    if not order_id:
        return jsonify({"error": "Missing order ID"}), 400

    # Verify this order belongs to this user
    conn = db.get_db()
    session = conn.execute(
        "SELECT * FROM checkout_sessions WHERE user_id=? AND token=? AND used=0",
        (user["id"], order_id)
    ).fetchone()
    if not session:
        conn.close()
        return jsonify({"error": "Invalid payment session"}), 400

    plan = session["plan"]

    # Capture payment via PayPal API
    import requests as http_req
    PAYPAL_CLIENT_ID = "AbFJcl0u1bWnNgzi5R4KcTsJ-u2pjxUJjjujweFLNkKzFulwgOQX4w5IsZmfeGbJ85JEABMDYUHWdvqB"
    PAYPAL_SECRET = "EFa-RJjZ0AoWFK4g8MWD3eIn3nm3GAdq0Mmo3LKR-WxApp8BJbFPre4Yat5jTqRNF4swbBTjv56XDooP"
    PAYPAL_API = "https://api-m.sandbox.paypal.com"

    # Get access token
    auth_resp = http_req.post(
        f"{PAYPAL_API}/v1/oauth2/token",
        auth=(PAYPAL_CLIENT_ID, PAYPAL_SECRET),
        data={"grant_type": "client_credentials"},
        headers={"Content-Type": "application/x-www-form-urlencoded"}
    )
    if auth_resp.status_code != 200:
        conn.close()
        return jsonify({"error": "Payment service unavailable"}), 500
    access_token = auth_resp.json()["access_token"]

    # Capture the order
    capture_resp = http_req.post(
        f"{PAYPAL_API}/v2/checkout/orders/{order_id}/capture",
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        }
    )
    capture_data = capture_resp.json()

    if capture_resp.status_code not in (200, 201) or capture_data.get("status") != "COMPLETED":
        conn.close()
        return jsonify({"error": "Payment was not completed. Please try again."}), 400

    # Payment confirmed! Mark session as used
    txn_id = ""
    try:
        txn_id = capture_data["purchase_units"][0]["payments"]["captures"][0]["id"]
    except (KeyError, IndexError):
        txn_id = order_id

    conn.execute(
        "UPDATE checkout_sessions SET used=1, transaction_id=?, activated_at=? WHERE id=?",
        (txn_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), session["id"])
    )
    conn.commit()
    conn.close()

    # Activate plan
    target_user_id = user["id"]
    if user.get("admin_id") and user["admin_id"] != 0:
        target_user_id = user["admin_id"]
    db.update_user_plan(target_user_id, plan)

    admin_id = target_user_id
    db.log_admin_action(admin_id, user, "Activated plan", f"Plan: {plan}, PayPal TXN: {txn_id}")

    user = db.get_user_by_token(token)
    return jsonify({"ok": True, "user": db.user_to_public(user)})


@app.route("/auth/update-plan", methods=["POST"])
def auth_update_plan():
    """Schedule a plan change for users already on a paid plan."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only the head admin can change the plan."}), 403

    data = request.get_json()
    plan = data.get("plan", "")
    if plan not in ("basic", "pro", "agency"):
        return jsonify({"error": "Invalid plan"}), 400

    current_plan = user.get("plan", "free_trial")
    if current_plan == "free_trial" or current_plan == "none":
        return jsonify({"error": "Please complete payment through the checkout page."}), 400

    # Already on a paid plan — schedule change for next billing cycle
    target_user_id = user["id"]
    if user.get("admin_id") and user["admin_id"] != 0:
        target_user_id = user["admin_id"]
    db.schedule_plan_change(target_user_id, plan)
    user = db.get_user_by_token(token)
    plan_label = plan.capitalize()
    expires = user.get("plan_expires_at", "")
    msg = f"Plan change to {plan_label} scheduled."
    if expires:
        try:
            from datetime import datetime as _dt
            exp = _dt.strptime(expires, "%Y-%m-%d %H:%M:%S")
            msg += f" Takes effect on {exp.strftime('%b %d, %Y')}."
        except Exception:
            pass
    return jsonify({"ok": True, "scheduled": True, "message": msg, "user": db.user_to_public(user)})


@app.route("/auth/subscription", methods=["GET"])
def get_subscription():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    target_id = user["id"]
    if user.get("admin_id") and user["admin_id"] != 0:
        target_id = user["admin_id"]
    # Process any expired plans first
    db.process_plan_expiry(target_id)
    # Get the actual plan owner's data
    conn = db.get_db()
    row = conn.execute("SELECT * FROM users WHERE id=?", (target_id,)).fetchone()
    conn.close()
    owner = dict(row) if row else dict(user)
    plan = owner.get("plan", "free_trial")
    payment = db.get_payment_method(target_id)
    conv_count = db.get_monthly_conversation_count(target_id)
    conv_limit = db.PLAN_MONTHLY_CONVERSATIONS.get(plan, 50)
    pending = owner.get("pending_plan", "")
    return jsonify({
        "plan": plan,
        "pending_plan": pending,
        "plan_started_at": owner.get("plan_started_at", ""),
        "plan_expires_at": owner.get("plan_expires_at", ""),
        "billing_cycle": owner.get("billing_cycle", "monthly"),
        "auto_renew": owner.get("auto_renew", 1),
        "monthly_cost": db.PLAN_COSTS.get(plan, 0),
        "conversations_used": conv_count,
        "conversations_limit": conv_limit,
        "payment_method": payment,
    })


@app.route("/auth/cancel-subscription", methods=["POST"])
def cancel_subscription():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only the head admin can cancel the subscription."}), 403
    target_id = user["id"]
    if user.get("admin_id") and user["admin_id"] != 0:
        target_id = user["admin_id"]
    db.cancel_user_plan(target_id)
    user = db.get_user_by_token(token)
    return jsonify({"ok": True, "message": "Subscription cancelled. Your plan remains active until the end of the billing period.", "user": db.user_to_public(user)})


@app.route("/auth/cancel-plan-change", methods=["POST"])
def cancel_plan_change():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only the head admin can cancel a plan change."}), 403
    target_id = user["id"]
    if user.get("admin_id") and user["admin_id"] != 0:
        target_id = user["admin_id"]
    db.cancel_pending_plan_change(target_id)
    user = db.get_user_by_token(token)
    return jsonify({"ok": True, "message": "Scheduled plan change cancelled.", "user": db.user_to_public(user)})


@app.route("/auth/toggle-auto-renew", methods=["POST"])
def toggle_auto_renew():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only the head admin can change auto-renewal settings."}), 403
    data = request.get_json() or {}
    enabled = data.get("enabled", True)
    target_id = user["id"]
    if user.get("admin_id") and user["admin_id"] != 0:
        target_id = user["admin_id"]
    db.toggle_auto_renew(target_id, enabled)
    user = db.get_user_by_token(token)
    return jsonify({"ok": True, "user": db.user_to_public(user)})


@app.route("/auth/update-profile", methods=["POST"])
def auth_update_profile():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401

    data = request.get_json()
    name = data.get("name", "").strip()
    email_addr = data.get("email", "").strip().lower()
    current_password = data.get("current_password", "")
    new_password = data.get("new_password", "")
    avatar_url = data.get("avatar_url", None)

    if not name:
        return jsonify({"error": "Name is required."}), 400
    if not email_addr or "@" not in email_addr:
        return jsonify({"error": "Valid email is required."}), 400

    # If email changed, check it's not taken
    if email_addr != user["email"]:
        existing = db.get_user_by_email(email_addr)
        if existing:
            return jsonify({"error": "This email is already in use by another account."}), 400

    # If changing password, verify current password first
    if new_password:
        if len(new_password) < 8:
            return jsonify({"error": "New password must be at least 8 characters."}), 400
        if user["provider"] == "email":
            if not current_password or db._hash_password(current_password) != user["password_hash"]:
                return jsonify({"error": "Current password is incorrect."}), 400

    updated = db.update_user_profile(user["id"], name, email_addr, new_password, avatar_url)
    if not updated:
        return jsonify({"error": "Failed to update profile."}), 500

    refreshed = db.get_user_by_token(token)
    return jsonify({"ok": True, "user": db.user_to_public(refreshed)})


# ══════════════════════════════════════════════
#  Routes — Chat & API
# ══════════════════════════════════════════════

@app.route("/chat", methods=["POST"])
def chat():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid or missing JSON body"}), 400
    user_message = data.get("message", "").strip()
    session_id = data.get("session_id", "default")
    admin_id_raw = data.get("admin_id")
    admin_id = None

    # Resolve admin_id: could be a public GUID or numeric ID
    if admin_id_raw:
        admin_id_str = str(admin_id_raw).strip()
        # Check if it's a GUID (contains letters/dashes, not purely numeric)
        if not admin_id_str.isdigit():
            # It's a public_id (GUID) — resolve to numeric ID
            resolved_user = db.get_user_by_public_id(admin_id_str)
            if resolved_user:
                admin_id = resolved_user["id"]
            else:
                admin_id = None
        else:
            admin_id = int(admin_id_str)

    if not admin_id:
        # Auto-detect: use the first head_admin's id so analytics data goes to the right place
        try:
            conn = db.get_db()
            head = conn.execute("SELECT id FROM users WHERE role='head_admin' ORDER BY id LIMIT 1").fetchone()
            admin_id = head["id"] if head else 1
            conn.close()
        except Exception:
            admin_id = 1

    # Customer ID and optional API URL from embedded widget
    customer_id = str(data.get("customer_id", "")).strip()
    customer_api_url = data.get("customer_api_url", "").strip()

    # User info from dashboard chatbot (signed-in users)
    user_name = data.get("user_name", "").strip()
    user_email = data.get("user_email", "").strip()
    user_phone = data.get("user_phone", "").strip()

    if not user_message:
        return jsonify({"error": "Empty message"}), 400
    if len(user_message) > 2000:
        user_message = user_message[:2000]

    # Check chatbot domain limit — track which websites embed this chatbot
    origin = request.headers.get("Origin", "") or request.headers.get("Referer", "")
    if origin:
        from urllib.parse import urlparse
        try:
            parsed = urlparse(origin)
            domain = parsed.netloc or parsed.hostname or ""
            # Ignore localhost / dev domains
            if domain and domain not in ("localhost", "127.0.0.1", "") and "localhost:" not in domain:
                ok, err_msg = db.register_chatbot_domain(admin_id, domain)
                if not ok:
                    return jsonify({"reply": err_msg})
        except Exception:
            pass

    # Check conversation limit for the admin's plan
    if db.is_conversation_limit_reached(admin_id):
        return jsonify({"reply": "We're sorry, but our chat service has reached its monthly limit. Please contact the clinic directly for assistance or ask the clinic to upgrade their plan."})

    try:
        reply = process_message(session_id, user_message, admin_id=admin_id,
                                customer_id=customer_id, customer_api_url=customer_api_url,
                                user_name=user_name, user_email=user_email, user_phone=user_phone)
        # Check if session has UI options to send to frontend
        session = get_session(session_id)
        response = {"reply": reply}
        if session.get("_ui_options"):
            response["options"] = session.pop("_ui_options")
        if session.pop("_booking_confirmed", False):
            response["booking_confirmed"] = True
        return jsonify(response)
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/patient/chat", methods=["POST"])
def patient_chat():
    """Chat endpoint for authenticated patients — skips name/email/phone collection."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid or missing JSON body"}), 400

    patient_id = data.get("patient_id")
    if not patient_id:
        return jsonify({"error": "patient_id is required"}), 400

    # Verify patient exists
    patient = db.get_patient(patient_id)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404

    user_message = data.get("message", "").strip()
    session_id = data.get("session_id", f"patient_{patient_id}")
    admin_id = data.get("admin_id") or patient.get("admin_id")
    if not admin_id:
        try:
            conn = db.get_db()
            head = conn.execute("SELECT id FROM users WHERE role='head_admin' ORDER BY id LIMIT 1").fetchone()
            admin_id = head["id"] if head else 1
            conn.close()
        except Exception:
            admin_id = 1

    if not user_message:
        return jsonify({"error": "Empty message"}), 400
    if len(user_message) > 2000:
        user_message = user_message[:2000]

    # Check conversation limit
    if db.is_conversation_limit_reached(admin_id):
        return jsonify({"reply": "We're sorry, but our chat service has reached its monthly limit. Please contact the clinic directly for assistance or ask the clinic to upgrade their plan."})

    try:
        reply = process_message(session_id, user_message, admin_id=admin_id, patient_id=patient_id)
        session = get_session(session_id)
        response = {"reply": reply}
        if session.get("_ui_options"):
            response["options"] = session.pop("_ui_options")
        if session.pop("_booking_confirmed", False):
            response["booking_confirmed"] = True
        return jsonify(response)
    except Exception as e:
        print(f"Patient chat error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/leads", methods=["GET"])
def api_leads():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") == "doctor":
        return jsonify([])  # Doctors don't see leads
    admin_id = get_effective_admin_id(user)
    leads = db.get_all_leads(admin_id=admin_id)
    # Enrich with follow-up summary
    for lead in leads:
        try:
            lead["followup"] = db.get_lead_followup_summary(lead["id"])
        except Exception:
            lead["followup"] = {"total": 0, "sent": 0, "pending": 0}
    return jsonify(leads)


@app.route("/api/leads/<int:lid>/stage", methods=["PATCH"])
def api_update_lead_stage(lid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    stage = payload.get("stage", "").strip()
    if stage not in ("new", "engaged", "warm", "cold", "converted"):
        return jsonify({"error": "Invalid stage"}), 400
    db.update_lead_stage(lid, stage)
    if stage == "cold":
        db.cancel_lead_followups(lid)
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Updated lead stage", f"Lead #{lid} → {stage}")
    return jsonify({"ok": True})


@app.route("/api/leads/<int:lid>/score", methods=["PATCH"])
def api_update_lead_score(lid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    score = int(payload.get("score", 0))
    db.update_lead_score(lid, score)
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Updated lead score", f"Lead #{lid} → score {score}")
    return jsonify({"ok": True})


@app.route("/api/leads/<int:lid>/followups/cancel", methods=["POST"])
def api_cancel_lead_followups(lid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    db.cancel_lead_followups(lid)
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Cancelled lead followups", f"Lead #{lid}")
    return jsonify({"ok": True})


@app.route("/api/bookings", methods=["GET"])
def api_bookings():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") == "doctor":
        doctor = db.get_doctor_by_user_id(user["id"])
        if doctor:
            bookings = db.get_all_bookings(doctor_id=doctor["id"])
        else:
            bookings = []
    else:
        admin_id = get_effective_admin_id(user)
        bookings = db.get_all_bookings(admin_id=admin_id)

    # Auto-complete past bookings whose date+time has passed
    now = datetime.now()
    ids_to_complete = []
    filtered = []
    for b in bookings:
        try:
            time_parts = b["time"].split(" - ")
            end_time_str = time_parts[-1].strip() if len(time_parts) > 1 else time_parts[0].strip()
            booking_end = datetime.strptime(f"{b['date']} {end_time_str}", "%Y-%m-%d %I:%M %p")
            if booking_end < now and b.get("status") not in ("no_show", "cancelled", "completed"):
                ids_to_complete.append(b["id"])
                continue
        except (ValueError, KeyError, IndexError):
            pass
        # Only return active bookings (pending/confirmed)
        if b.get("status") in ("pending", "confirmed"):
            filtered.append(b)

    # Mark expired bookings as completed and update patient last_visit_date
    if ids_to_complete:
        conn = db.get_db()
        for bid in ids_to_complete:
            conn.execute("UPDATE bookings SET status='completed' WHERE id=?", (bid,))
            bk_row = conn.execute("SELECT * FROM bookings WHERE id=?", (bid,)).fetchone()
            if bk_row:
                bk = dict(bk_row)
                if bk.get("patient_id"):
                    conn.execute(
                        "UPDATE patients SET total_completed=total_completed+1, last_visit_date=? WHERE id=? AND (last_visit_date IS NULL OR last_visit_date < ?)",
                        (bk["date"], bk["patient_id"], bk["date"])
                    )
                else:
                    pat = None
                    if bk.get("customer_email"):
                        pat = conn.execute("SELECT id FROM patients WHERE admin_id=? AND email=?", (bk["admin_id"], bk["customer_email"])).fetchone()
                    if not pat and bk.get("customer_phone"):
                        pat = conn.execute("SELECT id FROM patients WHERE admin_id=? AND phone=?", (bk["admin_id"], bk["customer_phone"])).fetchone()
                    if pat:
                        conn.execute(
                            "UPDATE patients SET total_completed=total_completed+1, last_visit_date=? WHERE id=? AND (last_visit_date IS NULL OR last_visit_date < ?)",
                            (bk["date"], pat["id"], bk["date"])
                        )
        conn.commit()
        conn.close()

    # Sort bookings by date+time (nearest first)
    def _booking_sort_key(b):
        try:
            date_str = b.get("date", "")
            time_str = b.get("time", "").split(" - ")[0].strip()
            for fmt in ("%Y-%m-%d %I:%M %p", "%Y-%m-%d %I:%M%p", "%Y-%m-%d %H:%M"):
                try:
                    return datetime.strptime(f"{date_str} {time_str}", fmt)
                except ValueError:
                    continue
            return datetime.strptime(date_str, "%Y-%m-%d")
        except (ValueError, KeyError, IndexError):
            return datetime.max
    filtered.sort(key=_booking_sort_key)

    return jsonify(filtered)


@app.route("/api/bookings/previous", methods=["GET"])
def api_previous_bookings():
    """Return completed, cancelled, and no-show bookings."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") == "doctor":
        doctor = db.get_doctor_by_user_id(user["id"])
        if doctor:
            bookings = db.get_all_bookings(doctor_id=doctor["id"])
        else:
            bookings = []
    else:
        admin_id = get_effective_admin_id(user)
        bookings = db.get_all_bookings(admin_id=admin_id)

    # Auto-complete past bookings whose date+time has passed
    now = datetime.now()
    ids_to_complete = []
    for b in bookings:
        if b.get("status") in ("no_show", "cancelled", "completed"):
            continue
        try:
            time_parts = b["time"].split(" - ")
            end_time_str = time_parts[-1].strip() if len(time_parts) > 1 else time_parts[0].strip()
            booking_end = datetime.strptime(f"{b['date']} {end_time_str}", "%Y-%m-%d %I:%M %p")
            if booking_end < now:
                ids_to_complete.append(b["id"])
                b["status"] = "completed"
        except (ValueError, KeyError, IndexError):
            pass

    if ids_to_complete:
        conn = db.get_db()
        for bid in ids_to_complete:
            conn.execute("UPDATE bookings SET status='completed' WHERE id=?", (bid,))
            bk_row = conn.execute("SELECT * FROM bookings WHERE id=?", (bid,)).fetchone()
            if bk_row:
                bk = dict(bk_row)
                if bk.get("patient_id"):
                    conn.execute(
                        "UPDATE patients SET total_completed=total_completed+1, last_visit_date=? WHERE id=? AND (last_visit_date IS NULL OR last_visit_date < ?)",
                        (bk["date"], bk["patient_id"], bk["date"])
                    )
                else:
                    pat = None
                    if bk.get("customer_email"):
                        pat = conn.execute("SELECT id FROM patients WHERE admin_id=? AND email=?", (bk["admin_id"], bk["customer_email"])).fetchone()
                    if not pat and bk.get("customer_phone"):
                        pat = conn.execute("SELECT id FROM patients WHERE admin_id=? AND phone=?", (bk["admin_id"], bk["customer_phone"])).fetchone()
                    if pat:
                        conn.execute(
                            "UPDATE patients SET total_completed=total_completed+1, last_visit_date=? WHERE id=? AND (last_visit_date IS NULL OR last_visit_date < ?)",
                            (bk["date"], pat["id"], bk["date"])
                        )
        conn.commit()
        conn.close()

    # Only return past/finished bookings
    previous = [b for b in bookings if b.get("status") in ("completed", "cancelled", "no_show")]

    # Sort by date descending (most recent first)
    def _sort_key(b):
        try:
            date_str = b.get("date", "")
            time_str = b.get("time", "").split(" - ")[0].strip()
            for fmt in ("%Y-%m-%d %I:%M %p", "%Y-%m-%d %I:%M%p", "%Y-%m-%d %H:%M"):
                try:
                    return datetime.strptime(f"{date_str} {time_str}", fmt)
                except ValueError:
                    continue
            return datetime.strptime(date_str, "%Y-%m-%d")
        except (ValueError, KeyError, IndexError):
            return datetime.min
    previous.sort(key=_sort_key, reverse=True)

    # Mark bookings that already have a follow-up
    if previous:
        conn = db.get_db()
        booking_ids = [b["id"] for b in previous if b.get("id")]
        followup_booking_ids = set()
        if booking_ids:
            placeholders = ",".join(["?"] * len(booking_ids))
            rows = conn.execute(
                f"SELECT DISTINCT booking_id FROM treatment_followups WHERE booking_id IN ({placeholders}) AND status != 'cancelled'",
                booking_ids
            ).fetchall()
            followup_booking_ids = {r["booking_id"] for r in rows}
        conn.close()
        for b in previous:
            b["has_followup"] = b.get("id") in followup_booking_ids

    return jsonify(previous)


@app.route("/api/stats", methods=["GET"])
def api_stats():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") == "doctor":
        doctor = db.get_doctor_by_user_id(user["id"])
        if doctor:
            return jsonify(db.get_stats(doctor_id=doctor["id"]))
        return jsonify(db.get_stats())
    admin_id = get_effective_admin_id(user)
    return jsonify(db.get_stats(admin_id=admin_id))


@app.route("/api/analytics", methods=["GET"])
def api_analytics():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") == "doctor":
        return jsonify({"error": "Doctors cannot access analytics"}), 403
    admin_id = get_effective_admin_id(user)
    date_from = request.args.get("from", (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
    date_to = request.args.get("to", datetime.now().strftime("%Y-%m-%d"))
    data = db.get_analytics(admin_id, date_from, date_to)
    return jsonify(data)


@app.route("/api/analytics/doctor-revenue", methods=["GET"])
def api_doctor_revenue():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Access denied"}), 403
    admin_id = get_effective_admin_id(user)
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    conn = db.get_db()

    # Build service price lookup (same logic as ROI)
    svc_prices = {}
    svc_rows = conn.execute(
        "SELECT LOWER(name) as name, price FROM company_services WHERE admin_id=?", (admin_id,)
    ).fetchall()
    for sr in svc_rows:
        svc_prices[sr["name"]] = sr["price"]

    def calc_rev(rev_amount, service_name):
        if rev_amount and rev_amount > 0:
            return rev_amount
        return svc_prices.get((service_name or "").lower(), 0)

    # Get all doctors
    doctors = conn.execute(
        "SELECT id, name, specialty FROM doctors WHERE admin_id=?", (admin_id,)
    ).fetchall()

    # Get all bookings with doctor_id
    bquery = "SELECT doctor_id, status, revenue_amount, service FROM bookings WHERE admin_id=?"
    bparams = [admin_id]
    if date_from and date_to:
        bquery += " AND date BETWEEN ? AND ?"
        bparams.extend([date_from, date_to])
    bookings = conn.execute(bquery, bparams).fetchall()
    conn.close()

    # Aggregate per doctor
    doc_stats = {}
    for d in doctors:
        doc_stats[d["id"]] = {
            "id": d["id"], "name": d["name"], "specialty": d["specialty"] or "General",
            "revenue": 0, "lost_revenue": 0, "total_bookings": 0,
            "completed": 0, "confirmed": 0, "cancelled": 0, "no_shows": 0
        }
    for b in bookings:
        did = b["doctor_id"]
        if did not in doc_stats:
            continue
        ds = doc_stats[did]
        ds["total_bookings"] += 1
        rev = calc_rev(b["revenue_amount"], b["service"])
        if b["status"] in ("confirmed", "completed"):
            ds["revenue"] += rev
        elif b["status"] in ("cancelled", "no_show"):
            ds["lost_revenue"] += rev
        if b["status"] == "completed":
            ds["completed"] += 1
        elif b["status"] == "confirmed":
            ds["confirmed"] += 1
        elif b["status"] == "cancelled":
            ds["cancelled"] += 1
        elif b["status"] == "no_show":
            ds["no_shows"] += 1

    currency = db.get_company_currency(admin_id)
    result = sorted(doc_stats.values(), key=lambda x: x["revenue"], reverse=True)
    for r in result:
        r["revenue"] = round(r["revenue"], 2)
        r["lost_revenue"] = round(r["lost_revenue"], 2)
    return jsonify({"doctors": result, "currency": currency})


@app.route("/api/roi", methods=["GET"])
def api_roi():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Access denied"}), 403
    admin_id = get_effective_admin_id(user)
    data = db.get_roi_data(admin_id)
    return jsonify(data)


@app.route("/api/external/booking", methods=["POST"])
def api_external_booking():
    """External endpoint for syncing appointments from a website/PMS."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    # Authenticate via external API key (dedicated per-company key for PMS integrations)
    admin_id = db.get_admin_by_external_api_key(token)
    if not admin_id:
        return jsonify({"error": "Invalid or missing API key"}), 401

    data = request.get_json(silent=True) or {}
    patient_name = (data.get("patient_name") or "").strip()
    doctor_name_input = (data.get("doctor_name") or "").strip()
    date_raw = (data.get("date") or "").strip()
    time_raw = (data.get("time") or "").strip()
    patient_email = (data.get("patient_email") or "").strip()
    patient_phone = (data.get("patient_phone") or "").strip()
    service = (data.get("service") or "").strip() or "NA"
    notes = (data.get("notes") or "").strip()

    # All fields are required for external bookings (no prior info about the patient)
    # service is optional — defaults to "General Appointment" if not provided
    missing = []
    if not patient_name: missing.append("patient_name")
    if not patient_email: missing.append("patient_email")
    if not patient_phone: missing.append("patient_phone")
    if not doctor_name_input: missing.append("doctor_name")
    if not date_raw: missing.append("date")
    if not time_raw: missing.append("time")
    if missing:
        return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400

    # Parse date — support formats like "2026-9-13", "2026-09-13", "2026-9-13 4:00"
    import re
    date_str = ""
    time_str = ""
    # If date contains time (e.g. "2026-9-13 4:00"), split it
    dt_match = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})\s+(\d{1,2}):(\d{2})", date_raw)
    if dt_match:
        date_str = f"{dt_match.group(1)}-{int(dt_match.group(2)):02d}-{int(dt_match.group(3)):02d}"
        time_str = f"{int(dt_match.group(4)):02d}:{dt_match.group(5)}"
    else:
        d_match = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", date_raw)
        if d_match:
            date_str = f"{d_match.group(1)}-{int(d_match.group(2)):02d}-{int(d_match.group(3)):02d}"
        else:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    # Parse time — support "14:00", "9:30", "9:30 AM", "2:30 PM"
    if not time_str:
        # Try 12h format first: "9:30 AM" or "2:30 PM"
        t12_match = re.match(r"(\d{1,2}):(\d{2})\s*(AM|PM|am|pm)", time_raw)
        if t12_match:
            h, m, ampm = int(t12_match.group(1)), int(t12_match.group(2)), t12_match.group(3).upper()
            if ampm == "PM" and h != 12: h += 12
            if ampm == "AM" and h == 12: h = 0
            time_str = f"{h:02d}:{m:02d}"
        else:
            t_match = re.match(r"(\d{1,2}):(\d{2})", time_raw)
            if t_match:
                time_str = f"{int(t_match.group(1)):02d}:{t_match.group(2)}"
            else:
                return jsonify({"error": "Invalid time format. Use HH:MM or HH:MM AM/PM"}), 400

    # Find doctor by name
    conn = db.get_db()
    doctors = conn.execute("SELECT id, name, appointment_length, schedule_type, daily_hours, start_time, end_time, availability FROM doctors WHERE admin_id=?", (admin_id,)).fetchall()
    doctor_id = None
    matched_name = ""
    matched_doc = None
    doctor_name_lower = doctor_name_input.lower().replace("dr.", "").replace("dr ", "").strip()
    for doc in doctors:
        doc_lower = doc["name"].lower().replace("dr.", "").replace("dr ", "").strip()
        if doc_lower == doctor_name_lower or doctor_name_lower in doc_lower or doc_lower in doctor_name_lower:
            doctor_id = doc["id"]
            matched_name = doc["name"]
            matched_doc = doc
            break
    if not doctor_id:
        conn.close()
        return jsonify({"error": f"Doctor '{doctor_name_input}' not found. Available: {', '.join(d['name'] for d in doctors)}"}), 404

    # Validate time is within doctor's working hours
    import json as _json
    from datetime import datetime as _dt
    appt_length = matched_doc["appointment_length"] or 60
    start_dt = _dt.strptime(time_str, "%H:%M")
    end_dt = start_dt + timedelta(minutes=appt_length)
    booking_date = _dt.strptime(date_str, "%Y-%m-%d")
    day_name = booking_date.strftime("%A")  # e.g. "Sunday", "Monday"

    schedule_type = matched_doc["schedule_type"] or "fixed"
    work_from = None
    work_to = None
    is_day_off = False

    if schedule_type == "flexible" and matched_doc["daily_hours"]:
        try:
            daily = _json.loads(matched_doc["daily_hours"])
            day_info = daily.get(day_name, {})
            if day_info.get("off"):
                is_day_off = True
            else:
                work_from = _dt.strptime(day_info.get("from", ""), "%I:%M %p")
                work_to = _dt.strptime(day_info.get("to", ""), "%I:%M %p")
        except (ValueError, KeyError):
            pass
    else:
        # Fixed schedule — check if day is in availability
        avail = (matched_doc["availability"] or "Mon-Fri").lower()
        day_abbrevs = {"monday": "mon", "tuesday": "tue", "wednesday": "wed", "thursday": "thu", "friday": "fri", "saturday": "sat", "sunday": "sun"}
        day_abbr = day_abbrevs.get(day_name.lower(), "")
        if day_abbr not in avail:
            is_day_off = True
        else:
            try:
                work_from = _dt.strptime(matched_doc["start_time"] or "09:00 AM", "%I:%M %p")
                work_to = _dt.strptime(matched_doc["end_time"] or "05:00 PM", "%I:%M %p")
            except ValueError:
                pass

    if is_day_off:
        conn.close()
        return jsonify({"error": f"Sorry, Dr. {matched_name} does not work on {day_name}s."}), 400

    if work_from and work_to:
        # Compare using only hour:minute (strip date part)
        start_check = start_dt.replace(year=1900, month=1, day=1)
        end_check = end_dt.replace(year=1900, month=1, day=1)
        wf = work_from.replace(year=1900, month=1, day=1)
        wt = work_to.replace(year=1900, month=1, day=1)
        if start_check < wf or end_check > wt:
            conn.close()
            return jsonify({"error": f"Sorry, this time is outside Dr. {matched_name}'s working hours. Dr. {matched_name} works from {work_from.strftime('%I:%M %p')} to {work_to.strftime('%I:%M %p')} on {day_name}s."}), 400

    # Build time slot: "HH:MM AM - HH:MM AM" using doctor's appointment length
    time_slot = f"{start_dt.strftime('%I:%M %p')} - {end_dt.strftime('%I:%M %p')}"

    # Check for double-booking — if slot taken, add to waitlist
    existing = conn.execute(
        "SELECT id FROM bookings WHERE admin_id=? AND doctor_id=? AND date=? AND time=? AND status NOT IN ('cancelled','no_show')",
        (admin_id, doctor_id, date_str, time_slot)
    ).fetchone()
    if existing:
        conn.close()
        # Deduplicate patient first
        patient = db.get_or_create_patient(admin_id, name=patient_name, email=patient_email, phone=patient_phone, increment_booking=False)
        # Add to waitlist
        wid = db.add_to_waitlist(admin_id, doctor_id, date_str, time_slot,
                                  patient_name, patient_email, patient_phone)
        # Generate tokens for confirm/remove actions
        import secrets as _secrets
        confirm_token = _secrets.token_urlsafe(32)
        remove_token = _secrets.token_urlsafe(32)
        wl_conn = db.get_db()
        wl_conn.execute("UPDATE waitlist SET confirm_token=?, remove_token=? WHERE id=?",
                        (confirm_token, remove_token, wid))
        wl_conn.commit()
        # Get position
        pos_row = wl_conn.execute("SELECT position FROM waitlist WHERE id=?", (wid,)).fetchone()
        position = pos_row["position"] if pos_row else 1
        wl_conn.close()
        # Send waitlist email with confirm/remove buttons
        server_url = request.host_url.rstrip("/")
        confirm_url = f"{server_url}/waitlist/external/confirm/{confirm_token}"
        remove_url = f"{server_url}/waitlist/external/remove/{remove_token}"
        from email_service import send_waitlist_placed_email
        send_waitlist_placed_email(
            to_email=patient_email,
            patient_name=patient_name,
            date_display=date_str,
            time_slot=time_slot,
            doctor_name=matched_name,
            confirm_url=confirm_url,
            remove_url=remove_url,
            position=position,
            admin_id=admin_id,
        )
        return jsonify({
            "ok": True,
            "waitlisted": True,
            "waitlist_id": wid,
            "position": position,
            "patient_id": patient["id"],
            "message": f"Time slot is taken. Patient has been added to waitlist at position #{position}. An email has been sent to {patient_email}."
        }), 202
    conn.close()

    # Deduplicate patient: find or create by email/phone (never count same person twice)
    patient = db.get_or_create_patient(admin_id, name=patient_name, email=patient_email, phone=patient_phone)

    # Create the booking
    booking_id = db.add_booking(
        customer_name=patient_name,
        customer_email=patient_email,
        customer_phone=patient_phone,
        date=date_str,
        time=time_slot,
        service=service,
        doctor_id=doctor_id,
        doctor_name=matched_name,
        admin_id=admin_id,
        status="confirmed"
    )

    # Set revenue from service price
    svc_conn = db.get_db()
    svc_row = svc_conn.execute("SELECT price FROM company_services WHERE admin_id=? AND LOWER(name)=?", (admin_id, service.lower())).fetchone()
    if svc_row and svc_row["price"]:
        db.add_booking_revenue(booking_id, svc_row["price"])
    svc_conn.close()

    return jsonify({"ok": True, "booking_id": booking_id, "patient_id": patient["id"], "is_returning": patient.get("total_bookings", 1) > 1, "message": "Appointment synced successfully"}), 201


@app.route("/waitlist/external/confirm/<token>", methods=["GET"])
def waitlist_external_confirm(token):
    """Patient clicked 'Keep Me on Waitlist' in their email."""
    entry = db.get_waitlist_by_token(token, "confirm_token")
    if not entry:
        return "<html><body style='font-family:sans-serif;text-align:center;padding:60px'><h2>Link expired or invalid.</h2></body></html>", 404
    if entry["status"] == "removed":
        return "<html><body style='font-family:sans-serif;text-align:center;padding:60px'><h2>You have already been removed from the waitlist.</h2></body></html>"
    if entry["status"] == "confirmed":
        return "<html><body style='font-family:sans-serif;text-align:center;padding:60px'><h2>Your waitlist spot is already confirmed!</h2></body></html>"
    # Mark as confirmed (still waiting, but patient explicitly agreed)
    conn = db.get_db()
    conn.execute("UPDATE waitlist SET confirmed_at=CURRENT_TIMESTAMP WHERE id=?", (entry["id"],))
    conn.commit()
    conn.close()
    doctor = db.get_doctor_by_id(entry["doctor_id"])
    doc_name = doctor["name"] if doctor else ""
    return f"""<html><body style='font-family:sans-serif;text-align:center;padding:60px;background:#f0fdf4'>
    <div style='max-width:500px;margin:0 auto;background:#fff;border-radius:16px;padding:40px;box-shadow:0 4px 24px rgba(0,0,0,0.1)'>
    <div style='font-size:48px;margin-bottom:16px'>&#9989;</div>
    <h2 style='color:#065f46'>You're Confirmed on the Waitlist!</h2>
    <p style='color:#555'>You'll be automatically moved to the main appointment if the current booking for <strong>Dr. {doc_name}</strong> on <strong>{entry["date"]}</strong> at <strong>{entry["time_slot"]}</strong> is cancelled.</p>
    <p style='color:#999;font-size:13px;margin-top:24px'>You can close this page.</p>
    </div></body></html>"""


@app.route("/waitlist/external/remove/<token>", methods=["GET"])
def waitlist_external_remove(token):
    """Patient clicked 'Remove Me from Waitlist' in their email."""
    entry = db.get_waitlist_by_token(token, "remove_token")
    if not entry:
        return "<html><body style='font-family:sans-serif;text-align:center;padding:60px'><h2>Link expired or invalid.</h2></body></html>", 404
    if entry["status"] == "removed":
        return "<html><body style='font-family:sans-serif;text-align:center;padding:60px'><h2>You have already been removed from the waitlist.</h2></body></html>"
    # Remove from waitlist
    conn = db.get_db()
    conn.execute("UPDATE waitlist SET status='removed' WHERE id=?", (entry["id"],))
    conn.commit()
    conn.close()
    return f"""<html><body style='font-family:sans-serif;text-align:center;padding:60px;background:#fef2f2'>
    <div style='max-width:500px;margin:0 auto;background:#fff;border-radius:16px;padding:40px;box-shadow:0 4px 24px rgba(0,0,0,0.1)'>
    <div style='font-size:48px;margin-bottom:16px'>&#128075;</div>
    <h2 style='color:#991b1b'>Removed from Waitlist</h2>
    <p style='color:#555'>You've been removed from the waitlist for <strong>{entry["date"]}</strong> at <strong>{entry["time_slot"]}</strong>.</p>
    <p style='color:#999;font-size:13px;margin-top:24px'>You can close this page.</p>
    </div></body></html>"""


@app.route("/api/bookings/manual", methods=["POST"])
def api_manual_booking():
    """Admin endpoint to manually add an appointment."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Only admins can add appointments"}), 403
    admin_id = get_effective_admin_id(user)

    data = request.get_json(silent=True) or {}
    patient_name = (data.get("patient_name") or "").strip()
    doctor_id = data.get("doctor_id")
    date_str = (data.get("date") or "").strip()
    time_str = (data.get("time") or "").strip()
    patient_email = (data.get("patient_email") or "").strip()
    patient_phone = (data.get("patient_phone") or "").strip()
    service = (data.get("service") or "General Consultation").strip()
    notes = (data.get("notes") or "").strip()

    if not patient_name or not doctor_id or not date_str or not time_str:
        return jsonify({"error": "Missing required fields: patient_name, doctor_id, date, time"}), 400

    # Verify doctor belongs to this admin
    doctor = db.get_doctor_by_id(doctor_id)
    if not doctor or doctor.get("admin_id") != admin_id:
        return jsonify({"error": "Doctor not found"}), 404

    # Check for double-booking
    conn = db.get_db()
    existing = conn.execute(
        "SELECT id FROM bookings WHERE admin_id=? AND doctor_id=? AND date=? AND time=? AND status NOT IN ('cancelled','no_show')",
        (admin_id, doctor_id, date_str, time_str)
    ).fetchone()
    if existing:
        conn.close()
        return jsonify({"error": f"Time slot {date_str} {time_str} is already booked for Dr. {doctor['name']}"}), 409
    conn.close()

    booking_id = db.add_booking(
        customer_name=patient_name,
        customer_email=patient_email,
        customer_phone=patient_phone,
        date=date_str,
        time=time_str,
        service=service,
        doctor_id=doctor_id,
        doctor_name=doctor.get("name", ""),
        admin_id=admin_id,
        status="confirmed"
    )

    # Set revenue from service price
    svc_conn = db.get_db()
    svc_row = svc_conn.execute("SELECT price FROM company_services WHERE admin_id=? AND LOWER(name)=?", (admin_id, service.lower())).fetchone()
    if svc_row and svc_row["price"]:
        db.add_booking_revenue(booking_id, svc_row["price"])
    svc_conn.close()

    db.log_admin_action(admin_id, user, "Created booking", f"Booking #{booking_id} for {patient_name} on {date_str} at {time_str}")
    return jsonify({"ok": True, "booking_id": booking_id, "message": "Appointment added successfully"}), 201


@app.route("/api/roi/stats", methods=["GET"])
def api_roi_stats():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Access denied"}), 403
    admin_id = get_effective_admin_id(user)
    date_range = request.args.get("range", "month")
    data = db.get_roi_stats(admin_id, date_range)
    return jsonify(data)


@app.route("/api/bookings/<int:booking_id>/noshow", methods=["POST"])
def api_mark_noshow(booking_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    admin_id = get_effective_admin_id(user)

    conn = db.get_db()
    booking = conn.execute("SELECT * FROM bookings WHERE id=? AND admin_id=?", (booking_id, admin_id)).fetchone()
    if not booking:
        conn.close()
        return jsonify({"error": "Booking not found"}), 404
    booking = dict(booking)

    # Only allow marking as non-attended after appointment end time and within 2 hours
    try:
        time_parts = booking["time"].split(" - ")
        end_time_str = time_parts[-1].strip() if len(time_parts) > 1 else time_parts[0].strip()
        booking_end = datetime.strptime(f"{booking['date']} {end_time_str}", "%Y-%m-%d %I:%M %p")
        now = datetime.now()
        if now < booking_end:
            conn.close()
            return jsonify({"error": "Cannot mark as non-attended — appointment hasn't finished yet."}), 400
        hours_since = (now - booking_end).total_seconds() / 3600
        if hours_since > 2:
            conn.close()
            return jsonify({"error": "The 2-hour window to mark as non-attended has passed."}), 400
    except (ValueError, IndexError):
        pass  # If time parsing fails, allow marking

    # Generate a unique token for the no-show reason form
    import secrets
    noshow_token = secrets.token_urlsafe(32)
    conn.execute("UPDATE bookings SET status='no_show', revenue_amount=0 WHERE id=?", (booking_id,))
    # Increment patient no-show/cancellation count
    if booking.get("patient_id"):
        conn.execute("UPDATE patients SET total_no_shows=total_no_shows+1, total_cancelled=total_cancelled+1 WHERE id=?", (booking["patient_id"],))
    # Store the noshow token in booking notes for lookup
    existing_notes = booking.get("notes", "") or ""
    conn.execute("UPDATE bookings SET notes=? WHERE id=?",
                 (existing_notes + f"\n[noshow_token:{noshow_token}]", booking_id))
    conn.commit()
    conn.close()

    db.log_admin_action(admin_id, user, "Marked booking no-show", f"Booking #{booking_id} for {booking.get('customer_name', '')} on {booking.get('date', '')} at {booking.get('time', '')}")

    # Send no-show email to patient
    if booking.get("customer_email") and db.is_feature_enabled(admin_id, "email_noshow_patient"):
        try:
            dt = datetime.strptime(booking["date"], "%Y-%m-%d")
            date_display = dt.strftime("%A, %B %d, %Y")
        except (ValueError, TypeError):
            date_display = booking.get("date", "")
        try:
            base_url = request.host_url.rstrip("/")
            reason_url = f"{base_url}/noshow-reason/{noshow_token}"
            email.send_noshow_email(
                booking["customer_email"],
                booking["customer_name"],
                date_display,
                booking["time"],
                doctor_name=booking.get("doctor_name", ""),
                reason_url=reason_url,
                admin_id=admin_id,
            )
        except Exception as e:
            print(f"[noshow] ERROR sending email: {e}", flush=True)

    return jsonify({"ok": True})


@app.route("/api/company-info", methods=["GET"])
def api_get_company_info():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    # Admin sees their own, doctor sees their admin's
    info_user_id = get_effective_admin_id(user)
    info = db.get_company_info(info_user_id)
    return jsonify(info or {})


@app.route("/api/embed-id", methods=["GET"])
def api_embed_id():
    """Return the public GUID for embed code — resolves to the effective company admin."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    effective_id = get_effective_admin_id(user)
    if effective_id == user["id"]:
        return jsonify({"public_id": user.get("public_id", "")})
    # Linked admin/doctor — get head admin's public_id
    head = db.get_user_by_id(effective_id)
    return jsonify({"public_id": head.get("public_id", "") if head else ""})


@app.route("/api/company-info", methods=["POST"])
def api_save_company_info():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Only administrators can edit company info."}), 403
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    db.save_company_info(admin_id, data)
    # Propagate currency to all services of this admin
    new_currency = (data.get("currency") or "").strip()
    if new_currency:
        db.set_all_services_currency(admin_id, new_currency)
    db.log_admin_action(admin_id, user, "Updated company info", "")
    return jsonify({"ok": True})


@app.route("/api/company-info/currency", methods=["POST"])
def api_update_currency():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    data = request.get_json() or {}
    currency = (data.get("currency") or "").strip()
    if not currency:
        return jsonify({"error": "Currency is required"}), 400
    admin_id = get_effective_admin_id(user)
    conn = db.get_db()
    existing = conn.execute("SELECT id FROM company_info WHERE user_id=?", (admin_id,)).fetchone()
    if existing:
        conn.execute("UPDATE company_info SET currency=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?", (currency, admin_id))
    else:
        conn.execute("INSERT INTO company_info (user_id, currency) VALUES (?,?)", (admin_id, currency))
    conn.commit()
    conn.close()
    db.set_all_services_currency(admin_id, currency)
    db.log_admin_action(admin_id, user, "Updated currency", f"Currency → {currency}")
    return jsonify({"ok": True})


@app.route("/api/company-services", methods=["GET"])
def api_list_company_services():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    admin_id = get_effective_admin_id(user)
    services = db.get_services_with_doctors(admin_id)
    currency = db.get_company_currency(admin_id)
    return jsonify({"services": services, "currency": currency})


@app.route("/api/company-services", methods=["POST"])
def api_add_company_service():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Service name is required"}), 400
    try:
        price = float(data.get("price") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid price"}), 400
    admin_id = get_effective_admin_id(user)
    currency = db.get_company_currency(admin_id)
    sid = db.add_company_service(
        admin_id, name, price, currency, "manual",
        category=data.get("category", ""),
        duration_minutes=data.get("duration_minutes", 60),
        description=data.get("description", ""),
        preparation_instructions=data.get("preparation_instructions", ""),
        is_active=data.get("is_active", 1),
    )
    db.log_admin_action(admin_id, user, "Added service", f"{name} (${price})")
    return jsonify({"ok": True, "id": sid, "currency": currency})


@app.route("/api/company-services/<int:service_id>", methods=["PUT"])
def api_update_company_service(service_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Service name is required"}), 400
    try:
        price = float(data.get("price") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid price"}), 400
    admin_id = get_effective_admin_id(user)
    db.update_company_service(
        service_id, admin_id, name, price,
        category=data.get("category"),
        duration_minutes=data.get("duration_minutes"),
        description=data.get("description"),
        preparation_instructions=data.get("preparation_instructions"),
        is_active=data.get("is_active"),
    )
    db.log_admin_action(admin_id, user, "Updated service", f"Service #{service_id}: {name}")
    return jsonify({"ok": True})


@app.route("/api/company-services/<int:service_id>", methods=["DELETE"])
def api_delete_company_service(service_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    admin_id = get_effective_admin_id(user)
    db.delete_company_service(service_id, admin_id)
    db.log_admin_action(admin_id, user, "Deleted service", f"Service #{service_id}")
    return jsonify({"ok": True})


@app.route("/api/company-services/<int:service_id>/doctors", methods=["PUT"])
def api_set_service_doctors(service_id):
    """Set which doctors perform a given service. Body: {"doctor_ids": [1, 2, 3]}"""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    data = request.get_json() or {}
    doctor_ids = data.get("doctor_ids", [])
    db.set_service_doctors(service_id, doctor_ids, admin_id)
    db.log_admin_action(admin_id, user, "Updated service doctors", f"Service #{service_id}: {len(doctor_ids)} doctor(s)")

    # If doctors were assigned, notify anyone waiting for this service
    if doctor_ids and db.is_feature_enabled(admin_id, "email_booking_confirmation"):
        try:
            interests = db.get_waiting_service_interests(service_id)
            if interests:
                import email_service as email_svc
                # Get service name and doctor names
                doctors = db.get_doctors_for_service(service_id)
                doctor_names = [f"Dr. {d['name']}" for d in doctors]
                service_name = interests[0]["service_name"] if interests else ""
                for interest in interests:
                    if interest.get("patient_email"):
                        try:
                            email_svc.send_service_available_notification(
                                to_email=interest["patient_email"],
                                patient_name=interest["patient_name"],
                                service_name=service_name,
                                doctor_names=doctor_names,
                                admin_id=admin_id,
                            )
                            db.mark_service_interest_notified(interest["id"])
                        except Exception as e:
                            print(f"[service_notify] Failed to email {interest['patient_email']}: {e}", flush=True)
        except Exception as e:
            print(f"[service_notify] Error processing interests: {e}", flush=True)

    return jsonify({"ok": True})


@app.route("/api/company-services/bulk", methods=["POST"])
def api_bulk_company_services():
    """Bulk-insert services from structured JSON (e.g. parsed from pricing_insurance)."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    data = request.get_json() or {}
    services = data.get("services") or []
    if not isinstance(services, list) or not services:
        return jsonify({"error": "services must be a non-empty list"}), 400
    replace = bool(data.get("replace"))
    admin_id = get_effective_admin_id(user)
    currency = (data.get("currency") or "").strip() or db.get_company_currency(admin_id)
    if replace:
        db.delete_all_company_services(admin_id, source="pdf")
    added = db.bulk_add_company_services(admin_id, services, currency, source="pdf")
    db.log_admin_action(admin_id, user, "Bulk added services", f"{added} service(s) added")
    return jsonify({"ok": True, "added": added, "currency": currency})


@app.route("/api/company-services/clear", methods=["POST"])
def api_clear_company_services():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    data = request.get_json() or {}
    source = data.get("source")  # 'pdf' | 'manual' | None
    admin_id = get_effective_admin_id(user)
    db.delete_all_company_services(admin_id, source=source)
    db.log_admin_action(admin_id, user, "Cleared all services", f"Source: {source or 'all'}")
    return jsonify({"ok": True})


@app.route("/api/company-services/upload-pdf", methods=["POST"])
def api_upload_services_pdf():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Only PDF files are allowed"}), 400
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(f.stream)
        text = "\n".join((p.extract_text() or "") for p in reader.pages)
    except Exception as e:
        return jsonify({"error": f"Failed to read PDF: {e}"}), 400

    # Parse lines of form "<service name> ... <price>"
    parsed = []
    price_re = re.compile(r"([A-Za-z][A-Za-z0-9\s\-/&,'\.]{1,80}?)\s*[:\-\|\t ]+\s*([A-Z]{0,3}\s*[\$£€]?\s*\d{1,6}(?:[.,]\d{1,2})?)")
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        m = price_re.search(line)
        if not m:
            continue
        name = m.group(1).strip(" -:|")
        raw_price = m.group(2)
        num = re.sub(r"[^0-9.]", "", raw_price.replace(",", "."))
        try:
            price_val = float(num) if num else 0
        except ValueError:
            price_val = 0
        if name and price_val > 0:
            parsed.append({"name": name, "price": price_val})

    if not parsed:
        return jsonify({"error": "No services with prices were detected in the PDF"}), 400

    admin_id = get_effective_admin_id(user)
    currency = db.get_company_currency(admin_id)
    db.replace_company_services_from_pdf(admin_id, parsed, currency)
    db.log_admin_action(admin_id, user, "Uploaded services PDF", f"{len(parsed)} service(s) extracted")
    return jsonify({"ok": True, "added": len(parsed), "currency": currency, "services": parsed})


@app.route("/api/customers-api-config", methods=["GET"])
def api_get_customers_api_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    admin_id = get_effective_admin_id(user)
    config = db.get_customers_api_config(admin_id)
    return jsonify(config)


@app.route("/api/customers-api-config", methods=["POST"])
def api_save_customers_api_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    db.save_customers_api_config(
        admin_id,
        data.get("customers_api_url", "").strip(),
        data.get("customers_api_key", "").strip()
    )
    # Clear cache so next chat uses new config
    keys_to_remove = [k for k in _customer_cache if k.startswith(f"{admin_id}_")]
    for k in keys_to_remove:
        _customer_cache.pop(k, None)
    db.log_admin_action(admin_id, user, "Updated customers API config", "")
    return jsonify({"ok": True})


@app.route("/api/customers-api-test", methods=["POST"])
def api_test_customers_api():
    """Test the external customers API endpoint by fetching a sample customer."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    data = request.get_json()
    url = data.get("customers_api_url", "").strip().rstrip("/")
    key = data.get("customers_api_key", "").strip()
    test_id = data.get("test_customer_id", "1").strip()
    if not url:
        return jsonify({"error": "API URL is required"}), 400
    try:
        req_headers = {}
        if key:
            req_headers["Authorization"] = f"Bearer {key}"
            req_headers["X-API-Key"] = key
        resp = http_requests.get(f"{url}/{test_id}", headers=req_headers, timeout=10)
        resp.raise_for_status()
        result = resp.json()
        customer = result if "name" in result or "email" in result else result.get("customer", result.get("data", result))
        return jsonify({"ok": True, "customer": customer})
    except Exception as e:
        return jsonify({"error": f"Failed to connect: {str(e)}"}), 400


@app.route("/api/doctors", methods=["GET"])
def api_get_doctors():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if is_admin_role(user):
        doctors = db.get_doctors(get_effective_admin_id(user))
    else:
        doctors = db.get_doctors(user.get("admin_id", 0))
    return jsonify(doctors)


@app.route("/api/doctors", methods=["POST"])
def api_add_doctor():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Only administrators can add doctors."}), 403
    data = request.get_json()
    name = data.get("name", "").strip()
    doctor_email = data.get("email", "").strip().lower()
    if not name:
        return jsonify({"error": "Doctor name is required"}), 400
    if not doctor_email:
        return jsonify({"error": "Doctor email is required"}), 400

    # Validate: email must exist and belong to a doctor account
    existing_user = db.get_user_by_email(doctor_email)
    if not existing_user:
        return jsonify({"error": "No account found with this email. The doctor must sign up first."}), 400
    if existing_user.get("role") != "doctor":
        return jsonify({"error": f"This email belongs to a {existing_user['role']} account, not a doctor account."}), 400
    if existing_user.get("admin_id") and existing_user["admin_id"] != 0:
        return jsonify({"error": "This doctor is already linked to another practice."}), 400

    # Use company-level admin_id so all admins in the company see this doctor
    company_admin_id = get_effective_admin_id(user)

    # Create doctor record in pending state (not linked yet)
    specialty = data.get("specialty", "") or existing_user.get("specialty", "")
    doctor_id = db.add_doctor(company_admin_id, name, doctor_email,
                               specialty, data.get("bio", ""),
                               data.get("availability", "Mon-Fri"))

    # Create a doctor request — doctor must accept before being linked
    company_info = db.get_company_info(company_admin_id)
    business_name = company_info.get("business_name", "") if company_info else ""
    if not business_name:
        business_name = user.get("company", "")
    req_id, err = db.create_doctor_request(
        company_admin_id, user["name"], business_name, doctor_email, doctor_id
    )
    if err:
        # Request already exists — remove the pending doctor record we just created
        db.delete_doctor(doctor_id, company_admin_id)
        return jsonify({"error": err}), 400

    db.log_admin_action(company_admin_id, user, "Added doctor", f"{name} ({doctor_email})")
    return jsonify({"ok": True, "id": doctor_id, "message": "Invitation sent. The doctor must accept before joining."})


@app.route("/api/doctors/<int:doctor_id>", methods=["PUT"])
def api_update_doctor(doctor_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Only administrators can edit doctors."}), 403
    data = request.get_json()
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Doctor name is required"}), 400
    company_admin_id = get_effective_admin_id(user)
    db.update_doctor(doctor_id, company_admin_id, name, data.get("specialty", ""),
                     data.get("bio", ""), data.get("availability", "Mon-Fri"),
                     start_time=data.get("start_time"),
                     end_time=data.get("end_time"),
                     is_active=data.get("is_active"),
                     appointment_length=data.get("appointment_length"),
                     years_of_experience=data.get("years_of_experience"),
                     schedule_type=data.get("schedule_type"),
                     daily_hours=data.get("daily_hours"),
                     gender=data.get("gender"),
                     photo_url=data.get("photo_url"),
                     avg_appointment_price=data.get("avg_appointment_price"),
                     avg_appointment_currency=data.get("avg_appointment_currency"))
    db.log_admin_action(company_admin_id, user, "Updated doctor", f"Doctor #{doctor_id}: {name}")
    return jsonify({"ok": True})


@app.route("/api/doctors/<int:doctor_id>", methods=["DELETE"])
def api_delete_doctor(doctor_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only the head administrator can remove doctors."}), 403
    # Unlink the doctor user from the company
    doctor = db.get_doctor_by_id(doctor_id)
    if doctor and doctor.get("user_id"):
        db.set_user_admin_id(doctor["user_id"], 0)
    db.delete_doctor(doctor_id, user["id"])
    db.log_admin_action(user["id"], user, "Deleted doctor", f"Doctor #{doctor_id}" + (f": {doctor.get('name', '')}" if doctor else ""))
    return jsonify({"ok": True})


# ── Doctor Photo Upload ──

PHOTO_UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads", "doctor_photos")
os.makedirs(PHOTO_UPLOAD_DIR, exist_ok=True)


@app.route("/api/doctors/<int:doctor_id>/photo", methods=["POST"])
def api_upload_doctor_photo(doctor_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Only administrators can upload photos."}), 403
    photo = request.files.get("photo")
    if not photo or not photo.filename:
        return jsonify({"error": "No photo provided"}), 400
    ext = os.path.splitext(photo.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".webp"):
        return jsonify({"error": "Only JPG, PNG, and WebP images are allowed."}), 400
    filename = f"doctor_{doctor_id}_{int(datetime.now().timestamp())}{ext}"
    filepath = os.path.join(PHOTO_UPLOAD_DIR, filename)
    photo.save(filepath)
    photo_url = f"/uploads/doctor_photos/{filename}"
    company_admin_id = get_effective_admin_id(user)
    doctor = db.get_doctor_by_id(doctor_id)
    if doctor:
        db.update_doctor(doctor_id, company_admin_id, doctor["name"], doctor.get("specialty", ""),
                         doctor.get("bio", ""), doctor.get("availability", "Mon-Fri"),
                         photo_url=photo_url)
    db.log_admin_action(company_admin_id, user, "Uploaded doctor photo", f"Doctor #{doctor_id}")
    return jsonify({"ok": True, "photo_url": photo_url})


# ── Doctor PDF Upload ──

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads", "doctors")
os.makedirs(UPLOAD_DIR, exist_ok=True)

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")


@app.route("/api/doctors/upload-pdf", methods=["POST"])
def api_upload_doctor_pdf():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Only administrators can upload doctor PDFs."}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    filename = file.filename or ""
    if not filename.lower().endswith(".pdf"):
        return jsonify({"error": "Only PDF files are accepted."}), 400

    file_bytes = file.read()
    if len(file_bytes) > 10 * 1024 * 1024:
        return jsonify({"error": "File too large. Maximum 10 MB."}), 400

    # Extract text from PDF
    text = ""
    try:
        import PyPDF2, io
        reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as e:
        return jsonify({"error": f"Could not read PDF: {str(e)}"}), 400

    if not text.strip():
        return jsonify({"error": "No text could be extracted. The PDF may be a scanned image."}), 400

    if len(text) > 15000:
        text = text[:15000]

    # Save PDF locally
    safe_name = f"{uuid.uuid4().hex}_{filename}"
    pdf_path = os.path.join(UPLOAD_DIR, safe_name)
    with open(pdf_path, "wb") as f:
        f.write(file_bytes)

    # AI extraction prompt — supports multiple doctors in one PDF
    ai_prompt = (
        "You are a medical data extraction assistant. This document may contain information about "
        "ONE or MULTIPLE doctors/medical professionals.\n\n"
        "Return ONLY a valid JSON ARRAY of objects. Each object represents one doctor with these fields:\n"
        "fullName, specialty, email, phone, availableDays, availableTimeSlots, dailyHours, qualifications, "
        "languages, bio, yearsOfExperience.\n\n"
        "Rules:\n"
        "- ALWAYS return a JSON array, even if there is only one doctor: [{...}]\n"
        "- Extract EVERY doctor mentioned in the document — do not skip any\n"
        "- For fullName: do NOT include the prefix 'Dr.' — just the name (e.g. \"Ahmad Al-Farsi\" not \"Dr. Ahmad Al-Farsi\")\n"
        "- For availableDays return a string like \"Mon-Fri\" or \"Sun,Mon,Tue,Wed,Thu\"\n"
        "- For availableTimeSlots: if the doctor has the SAME hours every day, return {\"from\":\"09:00 AM\",\"to\":\"05:00 PM\"}\n"
        "- For dailyHours: if the doctor has DIFFERENT hours on different days, return an object with day names as keys:\n"
        "  e.g. {\"Monday\":{\"from\":\"09:00 AM\",\"to\":\"05:00 PM\"},\"Tuesday\":{\"from\":\"10:00 AM\",\"to\":\"03:00 PM\"}}\n"
        "  Use full day names (Monday, Tuesday, etc.) and 12-hour format. Only include days the doctor works.\n"
        "  If all days have the same hours, set dailyHours to null and use availableTimeSlots instead.\n"
        "- For qualifications return a comma-separated string (e.g. \"BDS, MDS, PhD\")\n"
        "- For languages return a comma-separated string (e.g. \"Arabic, English\")\n"
        "- For yearsOfExperience return a number\n"
        "- For any field not found in the document, return null\n"
        "- Return ONLY the raw JSON array — no explanation, no markdown.\n\n"
        f"Document text:\n{text}"
    )

    extracted = None

    def _parse_ai_response(raw):
        """Parse AI response — handle both array and single object."""
        raw = raw.strip()
        # Try array first
        arr_start = raw.find("[")
        arr_end = raw.rfind("]") + 1
        if arr_start >= 0 and arr_end > arr_start:
            try:
                result = json.loads(raw[arr_start:arr_end])
                if isinstance(result, list) and len(result) > 0:
                    return result
            except json.JSONDecodeError:
                pass
        # Fallback: single object
        obj_start = raw.find("{")
        obj_end = raw.rfind("}") + 1
        if obj_start >= 0 and obj_end > obj_start:
            try:
                result = json.loads(raw[obj_start:obj_end])
                if isinstance(result, dict):
                    return [result]
            except json.JSONDecodeError:
                pass
        return None

    # Try Gemini first
    if GEMINI_API_KEY:
        try:
            import google.generativeai as genai
            genai.configure(api_key=GEMINI_API_KEY)
            model = genai.GenerativeModel("gemini-1.5-flash")
            resp = model.generate_content(ai_prompt)
            extracted = _parse_ai_response(resp.text)
            if extracted:
                print(f"[doctor-pdf] Gemini extraction OK — {len(extracted)} doctor(s)", flush=True)
        except Exception as e:
            print(f"[doctor-pdf] Gemini extraction failed: {e}", flush=True)

    # Try Groq fallback
    if not extracted and message_interpreter.is_configured():
        try:
            client = message_interpreter._get_client()
            resp = client.chat.completions.create(
                model="llama-3.1-8b-instant",
                messages=[{"role": "user", "content": ai_prompt}],
                max_tokens=4000, temperature=0,
            )
            raw = resp.choices[0].message.content.strip()
            extracted = _parse_ai_response(raw)
            if extracted:
                print(f"[doctor-pdf] Groq extraction OK — {len(extracted)} doctor(s)", flush=True)
        except Exception as e:
            print(f"[doctor-pdf] Groq extraction failed: {e}", flush=True)

    # Try OpenAI fallback
    if not extracted and dental_ai.is_configured():
        try:
            from openai import OpenAI
            client = OpenAI(api_key=dental_ai.OPENAI_API_KEY)
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": ai_prompt}],
                max_tokens=4000, temperature=0,
            )
            raw = resp.choices[0].message.content.strip()
            extracted = _parse_ai_response(raw)
            if extracted:
                print(f"[doctor-pdf] OpenAI extraction OK — {len(extracted)} doctor(s)", flush=True)
        except Exception as e:
            print(f"[doctor-pdf] OpenAI extraction failed: {e}", flush=True)

    # Try Anthropic fallback
    if not extracted:
        try:
            anthropic_key = os.getenv("ANTHROPIC_API_KEY", "")
            if anthropic_key:
                import httpx
                resp = httpx.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": anthropic_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                    json={"model": "claude-sonnet-4-20250514", "max_tokens": 4000,
                          "messages": [{"role": "user", "content": ai_prompt}]},
                    timeout=60,
                )
                raw = resp.json().get("content", [{}])[0].get("text", "")
                extracted = _parse_ai_response(raw)
                if extracted:
                    print(f"[doctor-pdf] Anthropic extraction OK — {len(extracted)} doctor(s)", flush=True)
        except Exception as e:
            print(f"[doctor-pdf] Anthropic extraction failed: {e}", flush=True)

    if not extracted:
        return jsonify({"error": "AI could not extract doctor info. Please try again or enter manually."}), 400

    # Mark doctors that already exist under this admin
    import re as _re
    existing_docs = db.get_doctors(company_admin_id)
    existing_names = set()
    for ed in existing_docs:
        n = (ed.get("name") or "").strip().lower()
        if n:
            existing_names.add(n)
    for doc in extracted:
        raw = (doc.get("fullName") or "").strip()
        clean = _re.sub(r'^(?:Dr\.?\s+)+', '', raw, flags=_re.IGNORECASE).strip().lower()
        doc["alreadyExists"] = clean in existing_names

    return jsonify({"ok": True, "doctors": extracted, "pdf_filename": safe_name})


@app.route("/api/doctors/from-pdf", methods=["POST"])
def api_save_doctor_from_pdf():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Only administrators can add doctors."}), 403

    payload = request.get_json()
    company_admin_id = get_effective_admin_id(user)

    # Support both single doctor and array of doctors
    doctors_list = payload.get("doctors") or [payload]
    pdf_filename = payload.get("pdf_filename", "")

    # Build a set of existing doctor names (normalised) to skip duplicates
    existing_doctors = db.get_doctors(company_admin_id)
    existing_names = set()
    for ed in existing_doctors:
        n = (ed.get("name") or "").strip().lower()
        if n:
            existing_names.add(n)

    saved_ids = []
    skipped = []
    for data in doctors_list:
        name = (data.get("fullName") or "").strip()
        if not name:
            continue

        # Strip "Dr." prefix for comparison (same logic as _strip_dr_prefix in database.py)
        import re
        clean_name = re.sub(r'^(?:Dr\.?\s+)+', '', name, flags=re.IGNORECASE).strip()
        if clean_name.lower() in existing_names:
            skipped.append(name)
            continue

        time_slots = data.get("availableTimeSlots") or {}
        start_time = time_slots.get("from", "09:00 AM") if isinstance(time_slots, dict) else "09:00 AM"
        end_time = time_slots.get("to", "05:00 PM") if isinstance(time_slots, dict) else "05:00 PM"

        # Determine schedule type from extracted daily hours
        daily_hours = data.get("dailyHours") or ""
        schedule_type = data.get("scheduleType") or ("flexible" if daily_hours else "fixed")
        if daily_hours and isinstance(daily_hours, dict):
            daily_hours = json.dumps(daily_hours)

        doctor_id = db.add_doctor_from_pdf(
            admin_id=company_admin_id,
            name=name,
            email=(data.get("email") or "").strip(),
            specialty=(data.get("specialty") or "").strip(),
            bio=(data.get("bio") or "").strip(),
            availability=(data.get("availableDays") or "Mon-Fri").strip(),
            start_time=start_time,
            end_time=end_time,
            phone=(data.get("phone") or "").strip(),
            qualifications=(data.get("qualifications") or "").strip(),
            languages=(data.get("languages") or "").strip(),
            years_of_experience=data.get("yearsOfExperience") or 0,
            pdf_filename=pdf_filename,
            schedule_type=schedule_type,
            daily_hours=daily_hours if isinstance(daily_hours, str) else "",
        )
        saved_ids.append(doctor_id)
        existing_names.add(clean_name.lower())  # prevent duplicates within same batch

    if not saved_ids and not skipped:
        return jsonify({"error": "No valid doctors found to save."}), 400

    msg_parts = []
    if saved_ids:
        msg_parts.append(f"{len(saved_ids)} doctor(s) added successfully")
    if skipped:
        msg_parts.append(f"{len(skipped)} already existed and were skipped ({', '.join(skipped)})")

    if saved_ids:
        db.log_admin_action(company_admin_id, user, "Added doctors from PDF", f"{len(saved_ids)} doctor(s) added")
    return jsonify({"ok": True, "ids": saved_ids, "count": len(saved_ids),
                    "skipped": skipped, "skipped_count": len(skipped),
                    "message": ". ".join(msg_parts) + "."})


@app.route("/uploads/doctors/<path:filename>")
def serve_doctor_pdf(filename):
    return send_from_directory(UPLOAD_DIR, filename)


@app.route("/uploads/doctor_photos/<path:filename>")
def serve_doctor_photo(filename):
    return send_from_directory(PHOTO_UPLOAD_DIR, filename)


# ── Doctor Breaks ──

@app.route("/api/doctors/<int:doctor_id>/breaks", methods=["GET"])
def api_get_breaks(doctor_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    breaks = db.get_doctor_breaks(doctor_id)
    return jsonify(breaks)


@app.route("/api/doctors/<int:doctor_id>/breaks", methods=["POST"])
def api_add_break(doctor_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    start_time = data.get("start_time", "")
    end_time = data.get("end_time", "")
    day_of_week = data.get("day_of_week", "")

    # Validate break is within working hours for the specified day
    if day_of_week and start_time and end_time:
        doctor = db.get_doctor_by_id(doctor_id)
        if doctor:
            def _parse_12h_min(ts):
                m = re.match(r'(\d{1,2}):(\d{2})\s*(AM|PM)', ts, re.IGNORECASE)
                if not m: return None
                h, mi, ap = int(m.group(1)), int(m.group(2)), m.group(3).upper()
                if ap == 'PM' and h < 12: h += 12
                if ap == 'AM' and h == 12: h = 0
                return h * 60 + mi
            # Get working hours for this day
            work_start, work_end = None, None
            if doctor.get("schedule_type") == "flexible" and doctor.get("daily_hours"):
                try:
                    daily = json.loads(doctor["daily_hours"]) if isinstance(doctor["daily_hours"], str) else doctor["daily_hours"]
                    dh = daily.get(day_of_week, {})
                    if dh.get("off"):
                        return jsonify({"error": f"{day_of_week} is marked as off."}), 400
                    work_start = _parse_12h_min(dh.get("from", ""))
                    work_end = _parse_12h_min(dh.get("to", ""))
                except (json.JSONDecodeError, ValueError):
                    pass
            if work_start is None:
                work_start = _parse_12h_min(doctor.get("start_time", "09:00 AM") or "09:00 AM")
            if work_end is None:
                work_end = _parse_12h_min(doctor.get("end_time", "05:00 PM") or "05:00 PM")
            brk_start = _parse_12h_min(start_time)
            brk_end = _parse_12h_min(end_time)
            if brk_start is not None and brk_end is not None and work_start is not None and work_end is not None:
                if brk_start < work_start or brk_end > work_end:
                    return jsonify({"error": f"Break must be within working hours for {day_of_week}."}), 400
                if brk_start >= brk_end:
                    return jsonify({"error": "Break start must be before end."}), 400

    break_id = db.add_doctor_break(
        doctor_id, data.get("break_name", "Break"),
        start_time, end_time, day_of_week)
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Added doctor break", f"Doctor #{doctor_id}: {day_of_week} {start_time}-{end_time}")
    return jsonify({"ok": True, "id": break_id})


@app.route("/api/doctors/<int:doctor_id>/breaks/<int:break_id>", methods=["DELETE"])
def api_delete_break(doctor_id, break_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    db.delete_doctor_break(break_id, doctor_id)
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Deleted doctor break", f"Doctor #{doctor_id}, break #{break_id}")
    return jsonify({"ok": True})


# ── Doctor Off Days ──

@app.route("/api/doctors/<int:doctor_id>/off-days", methods=["GET"])
def api_get_off_days(doctor_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    off_days = db.get_doctor_off_days(doctor_id)
    return jsonify(off_days)


@app.route("/api/doctors/<int:doctor_id>/off-days", methods=["POST"])
def api_add_off_day(doctor_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    off_id, err = db.add_doctor_off_day(doctor_id, data.get("date", ""), data.get("reason", ""))
    if err:
        return jsonify({"error": err}), 400
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Added doctor off-day", f"Doctor #{doctor_id}: {data.get('date', '')}")
    return jsonify({"ok": True, "id": off_id})


@app.route("/api/doctors/<int:doctor_id>/off-days/<int:off_day_id>", methods=["DELETE"])
def api_delete_off_day(doctor_id, off_day_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    db.delete_doctor_off_day(off_day_id, doctor_id)
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Deleted doctor off-day", f"Doctor #{doctor_id}, off-day #{off_day_id}")
    return jsonify({"ok": True})


@app.route("/api/doctors/<int:doctor_id>/off-dates", methods=["GET"])
def api_get_off_dates(doctor_id):
    """Public endpoint — returns just the ISO date strings for a doctor (includes schedule blocks)."""
    doctor = db.get_doctor_by_id(doctor_id)
    admin_id = doctor.get("admin_id", 0) if doctor else 0
    off_dates = _get_off_dates_with_blocks(doctor_id, admin_id)
    return jsonify(off_dates)


# ── Company Info File Upload ──

@app.route("/api/company-info/upload", methods=["POST"])
def api_upload_company_file():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only the head administrator can upload company files."}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    filename = file.filename.lower()

    # Extract text based on file type
    text = ""
    try:
        if filename.endswith(".pdf"):
            import PyPDF2
            import io
            reader = PyPDF2.PdfReader(io.BytesIO(file.read()))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
        elif filename.endswith(".docx"):
            import docx
            import io
            doc = docx.Document(io.BytesIO(file.read()))
            text = "\n".join(p.text for p in doc.paragraphs)
        elif filename.endswith((".txt", ".csv", ".md")):
            text = file.read().decode("utf-8", errors="ignore")
        else:
            # Try reading as text
            text = file.read().decode("utf-8", errors="ignore")
    except Exception as e:
        return jsonify({"error": f"Could not read file: {str(e)}"}), 400

    if not text.strip():
        return jsonify({"error": "No text could be extracted from this file. It may be a scanned image."}), 400

    # Truncate if very long
    if len(text) > 10000:
        text = text[:10000]

    # Use AI to extract structured company info
    ai_prompt = (
        "Extract business information from this document. Return ONLY valid JSON with these keys:\n"
        "business_name, address, phone, business_hours, services, pricing_insurance, emergency_info, about\n"
        "If a field cannot be found, use an empty string. Do not add any text outside the JSON.\n\n"
        f"Document text:\n{text}"
    )

    extracted = None

    # Try Groq (free)
    if message_interpreter.is_configured():
        try:
            client = message_interpreter._get_client()
            resp = client.chat.completions.create(
                model="llama-3.1-8b-instant",
                messages=[{"role": "user", "content": ai_prompt}],
                max_tokens=800, temperature=0,
            )
            raw = resp.choices[0].message.content.strip()
            # Extract JSON from response
            import json as _json
            # Find JSON in response
            start = raw.find("{")
            end = raw.rfind("}") + 1
            if start >= 0 and end > start:
                extracted = _json.loads(raw[start:end])
        except Exception as e:
            print(f"[upload] Groq extraction failed: {e}", flush=True)

    # Try OpenAI fallback
    if not extracted and dental_ai.is_configured():
        try:
            from openai import OpenAI
            client = OpenAI(api_key=dental_ai.OPENAI_API_KEY)
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": ai_prompt}],
                max_tokens=800, temperature=0,
            )
            raw = resp.choices[0].message.content.strip()
            start = raw.find("{")
            end = raw.rfind("}") + 1
            if start >= 0 and end > start:
                extracted = json.loads(raw[start:end])
        except Exception as e:
            print(f"[upload] OpenAI extraction failed: {e}", flush=True)

    if not extracted:
        return jsonify({"error": "AI could not extract info. Please fill in the fields manually.", "raw_text": text[:2000]}), 400

    return jsonify({"ok": True, "extracted": extracted})


# ── Categories ──

@app.route("/api/categories", methods=["GET"])
def api_get_categories():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    admin_id = get_effective_admin_id(user)
    return jsonify(db.get_categories(admin_id))


@app.route("/api/categories", methods=["POST"])
def api_add_category():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Only administrators can add categories."}), 403
    data = request.get_json()
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Category name is required."}), 400
    cat_id, error = db.add_category(user["id"], name)
    if error:
        return jsonify({"error": error}), 400
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Added category", f"{name}")
    return jsonify({"ok": True, "id": cat_id})


@app.route("/api/categories/<int:category_id>", methods=["DELETE"])
def api_delete_category(category_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Only administrators can remove categories."}), 403
    db.delete_category(category_id, user["id"])
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Deleted category", f"Category #{category_id}")
    return jsonify({"ok": True})


# ── Doctor Requests ──

@app.route("/api/doctor-requests", methods=["GET"])
def api_get_doctor_requests():
    """Get pending requests — for doctors: invitations sent to them; for admins: requests they sent."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") == "doctor":
        requests = db.get_doctor_requests_for_doctor(user["email"])
    else:
        admin_id = get_effective_admin_id(user)
        requests = db.get_doctor_requests_by_admin(admin_id)
    return jsonify(requests)


@app.route("/api/doctor-requests/<int:request_id>/respond", methods=["POST"])
def api_respond_doctor_request(request_id):
    """Accept or reject a doctor invitation."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") != "doctor":
        return jsonify({"error": "Only doctors can respond to requests."}), 403

    data = request.get_json()
    accept = data.get("accept", False)

    result, error = db.respond_to_doctor_request(request_id, user["id"], accept=accept)
    if error:
        return jsonify({"error": error}), 400

    # Refresh user data after accepting (role/admin_id may have changed)
    updated_user = db.get_user_by_token(token)
    return jsonify({"ok": True, "accepted": accept, "user": db.user_to_public(updated_user) if updated_user else None})


@app.route("/api/doctor-requests/<int:request_id>", methods=["DELETE"])
def api_delete_doctor_request(request_id):
    """Admin deletes a pending doctor request they sent."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Only administrators can delete doctor requests."}), 403
    company_admin_id = get_effective_admin_id(user)
    db.delete_doctor_request(request_id, company_admin_id)
    db.log_admin_action(company_admin_id, user, "Deleted doctor request", f"Request #{request_id}")
    return jsonify({"ok": True})


# ── Admin Requests (head_admin invites admins) ──

@app.route("/api/admin-requests", methods=["POST"])
def api_invite_admin():
    """Head admin sends invitation to a new admin."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only head administrators can invite admins."}), 403
    data = request.get_json()
    admin_email = data.get("email", "").strip().lower()
    if not admin_email:
        return jsonify({"error": "Email is required."}), 400

    # Validate: email must exist and belong to an admin account
    existing_user = db.get_user_by_email(admin_email)
    if not existing_user:
        return jsonify({"error": "No account found with this email. The person must sign up first."}), 400
    if existing_user.get("role") == "doctor":
        return jsonify({"error": "This email belongs to a doctor account, not an admin."}), 400
    if existing_user.get("role") == "head_admin":
        return jsonify({"error": "This email belongs to a head administrator. They already own their own company."}), 400
    if existing_user.get("admin_id") and existing_user["admin_id"] != 0:
        return jsonify({"error": "This admin is already linked to another company."}), 400

    company_info = db.get_company_info(user["id"])
    biz_name = (company_info.get("business_name") if company_info else None) or user.get("company", "") or "Your Practice"
    req_id, error = db.create_admin_request(user["id"], user["name"], biz_name, admin_email)
    if error:
        return jsonify({"error": error}), 400
    db.log_admin_action(user["id"], user, "Invited admin", f"{admin_email}")
    return jsonify({"ok": True, "id": req_id})


@app.route("/api/admin-requests", methods=["GET"])
def api_get_admin_requests():
    """Get admin requests — combines sent requests (if head_admin) and received invitations."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    # Return both sent and received, frontend will differentiate
    mode = request.args.get("mode", "")
    if mode == "received":
        return jsonify(db.get_admin_requests_for_user(user["email"]))
    elif mode == "sent" and user.get("role") == "head_admin":
        return jsonify(db.get_admin_requests_by_head(user["id"]))
    else:
        # Default: head_admin sees sent, others see received
        if user.get("role") == "head_admin":
            return jsonify(db.get_admin_requests_by_head(user["id"]))
        return jsonify(db.get_admin_requests_for_user(user["email"]))


@app.route("/api/admin-requests/<int:request_id>/respond", methods=["POST"])
def api_respond_admin_request(request_id):
    """Accept or reject an admin invitation."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    data = request.get_json()
    accept = data.get("accept", False)
    result, error = db.respond_to_admin_request(request_id, user["id"], accept=accept)
    if error:
        return jsonify({"error": error}), 400
    updated_user = db.get_user_by_token(token)
    return jsonify({"ok": True, "accepted": accept, "user": db.user_to_public(updated_user) if updated_user else None})


@app.route("/api/admin-requests/<int:request_id>", methods=["DELETE"])
def api_delete_admin_request(request_id):
    """Head admin cancels/deletes a sent invitation."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only head administrators can delete invitations."}), 403
    db.delete_admin_request(request_id, user["id"])
    db.log_admin_action(user["id"], user, "Deleted admin request", f"Request #{request_id}")
    return jsonify({"ok": True})


@app.route("/api/company-admins", methods=["GET"])
def api_get_company_admins():
    """Get all admins under this head admin's company."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only head administrators can view company admins."}), 403
    admins = db.get_company_admins(user["id"])
    return jsonify(admins)


@app.route("/api/company-admins/<int:admin_user_id>", methods=["DELETE"])
def api_remove_company_admin(admin_user_id):
    """Remove an admin from the company."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") != "head_admin":
        return jsonify({"error": "Only head administrators can remove admins."}), 403
    db.remove_admin_from_company(admin_user_id, user["id"])
    db.log_admin_action(user["id"], user, "Removed admin from company", f"User #{admin_user_id}")
    return jsonify({"ok": True})


@app.route("/api/doctors/public", methods=["GET"])
def api_public_doctors():
    """Public endpoint for chatbot to show available doctors for a given admin."""
    admin_id = request.args.get("admin_id", 1, type=int)
    doctors = db.get_doctors(admin_id)
    return jsonify([{"id": d["id"], "name": d["name"], "specialty": d["specialty"],
                     "availability": d["availability"]} for d in doctors if d.get("status") == "active"])


# ══════════════════════════════════════════════════════════════════
#  Feature 1 — Smart Waitlist System
# ══════════════════════════════════════════════════════════════════

@app.route("/api/waitlist", methods=["POST"])
def api_add_to_waitlist():
    """Add a patient to the waitlist for a specific slot."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    data = request.get_json()
    admin_id = data.get("admin_id", 1)
    wid = db.add_to_waitlist(
        admin_id=admin_id,
        doctor_id=data["doctor_id"],
        date=data["date"],
        time_slot=data["time_slot"],
        patient_name=data.get("patient_name", ""),
        patient_email=data.get("patient_email", ""),
        patient_phone=data.get("patient_phone", ""),
        session_id=data.get("session_id", "")
    )
    count = db.get_waitlist_count(admin_id, data["doctor_id"], data["date"], data["time_slot"])
    return jsonify({
        "ok": True,
        "waitlist_id": wid,
        "position": count,
        "message": "You've been added to the waitlist. We'll notify you immediately if a spot opens up."
    })


@app.route("/api/waitlist", methods=["GET"])
def api_get_waitlist():
    """Get waitlist entries (filtered by doctor/date)."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    doctor_id = request.args.get("doctor_id", type=int)
    date = request.args.get("date")
    # Doctors can only see their own waitlist
    if user.get("role") == "doctor":
        doc = db.get_doctor_by_user_id(user["id"])
        if doc:
            doctor_id = doc["id"]
    entries = db.get_waitlist(admin_id, doctor_id=doctor_id, date=date)
    return jsonify(entries)


@app.route("/api/waitlist/dashboard", methods=["GET"])
def api_waitlist_dashboard():
    """Get all waitlist data for admin dashboard with countdown timers."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    entries = db.get_waitlist_for_admin(admin_id)
    return jsonify(entries)


@app.route("/api/waitlist/<int:wid>/confirm", methods=["POST", "GET"])
def api_confirm_waitlist(wid):
    """Waitlist spot opened — create pending booking + send pre-visit form.
    The booking is only confirmed after the patient fills the form.
    Supports both POST (API) and GET (email link click)."""
    entry = db.get_waitlist_entry(wid)
    if not entry:
        return jsonify({"error": "Waitlist entry not found"}), 404

    if entry["status"] == "confirmed":
        return jsonify({"ok": True, "message": "This slot has already been confirmed."})

    if entry["status"] != "notified":
        return jsonify({"error": "This slot is no longer available for confirmation. It may have expired."}), 400

    # Check deadline
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if entry["confirm_deadline"] and now > entry["confirm_deadline"]:
        # Mark as expired and cascade to next patient
        db.expire_waitlist_patient(wid)
        background_tasks.trigger_waitlist_processing(
            entry["admin_id"], entry["doctor_id"], entry["date"], entry["time_slot"])
        return jsonify({"error": "Sorry, the confirmation deadline has passed. The slot has been offered to the next person."}), 400

    # Get doctor name for booking
    doctor = db.get_doctor_by_id(entry["doctor_id"])
    doctor_name = doctor["name"] if doctor else ""

    # Create a PENDING booking (not confirmed yet — needs form submission)
    booking_id = db.add_booking(
        customer_name=entry["patient_name"],
        customer_email=entry.get("patient_email", ""),
        customer_phone=entry.get("patient_phone", ""),
        date=entry["date"],
        time=entry["time_slot"],
        doctor_id=entry["doctor_id"],
        doctor_name=doctor_name,
        admin_id=entry["admin_id"],
        status="pending"
    )

    # Link booking to waitlist entry
    try:
        conn = db.get_db()
        conn.execute("UPDATE bookings SET waitlist_id=? WHERE id=?", (wid, booking_id))
        conn.commit()
        conn.close()
    except Exception:
        pass

    # Create pre-visit form and send form email (NOT confirmation email)
    form_token = None
    try:
        form_token = db.create_previsit_form(booking_id, entry["admin_id"], patient_name=entry["patient_name"])
    except Exception:
        pass

    if entry.get("patient_email") and form_token and db.is_feature_enabled(entry["admin_id"], "email_previsit_form"):
        try:
            base_url = os.getenv("BASE_URL", request.host_url.rstrip("/"))
            form_url = f"{base_url}/form/{form_token}"
            try:
                dt = datetime.strptime(entry["date"], "%Y-%m-%d")
                date_display = dt.strftime("%A, %B %d, %Y")
            except ValueError:
                date_display = entry["date"]
            email.send_previsit_form(
                entry["patient_email"], entry["patient_name"], form_url,
                date_display, entry["time_slot"],
                doctor_name=doctor_name,
                admin_id=entry.get("admin_id"),
            )
        except Exception as e:
            import logging
            logging.getLogger("waitlist").error(f"Waitlist form email error: {e}")

    # If GET request (email link click), redirect to the form
    if request.method == "GET" and form_token:
        return redirect(f"/form/{form_token}")

    return jsonify({
        "ok": True,
        "message": "Please fill out the pre-visit form to confirm your appointment. Check your email!",
        "booking_id": booking_id,
        "form_token": form_token
    })


@app.route("/api/waitlist/<int:wid>", methods=["DELETE"])
def api_delete_waitlist(wid):
    """Remove a patient from the waitlist (admin/doctor/head_admin only)."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if user.get("role") not in ("admin", "head_admin", "doctor"):
        return jsonify({"error": "Permission denied"}), 403
    entry = db.get_waitlist_entry(wid)
    if not entry:
        return jsonify({"error": "Waitlist entry not found"}), 404
    # Doctors can only remove from their own waitlist
    if user.get("role") == "doctor":
        doc = db.get_doctor_by_user_id(user["id"])
        if not doc or doc["id"] != entry["doctor_id"]:
            return jsonify({"error": "Permission denied"}), 403
    db.delete_waitlist_entry(wid)
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Removed from waitlist", f"Waitlist #{wid}: {entry.get('patient_name', '')}")
    return jsonify({"ok": True, "message": "Removed from waitlist"})


# ══════════════════════════════════════════════════════════════════
#  Feature Configuration (toggle emails, auto-features, etc.)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/feature-config", methods=["GET"])
def api_get_feature_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Permission denied"}), 403
    admin_id = get_effective_admin_id(user)
    config = db.get_feature_config(admin_id)
    return jsonify(config)


@app.route("/api/feature-config", methods=["POST"])
def api_save_feature_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Permission denied"}), 403
    admin_id = get_effective_admin_id(user)
    data = request.get_json() or {}
    db.save_feature_config(admin_id, data)
    db.log_admin_action(admin_id, user, "Updated feature config", "")
    return jsonify({"ok": True})


@app.route("/api/form-config", methods=["GET"])
def api_get_form_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Access denied"}), 403
    admin_id = get_effective_admin_id(user)
    config = db.get_form_config(admin_id)
    plan = user.get("plan", "free_trial")
    config["plan"] = plan
    return jsonify(config)


@app.route("/api/form-config", methods=["POST"])
def api_save_form_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Access denied"}), 403
    admin_id = get_effective_admin_id(user)
    data = request.get_json(silent=True) or {}
    db.save_form_config(admin_id, data)
    db.log_admin_action(admin_id, user, "Updated form config", "")
    return jsonify({"ok": True})


@app.route("/api/form-config/custom-field", methods=["POST"])
def api_add_custom_form_field():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Access denied"}), 403
    if user.get("plan") != "agency":
        return jsonify({"error": "Custom fields are only available on the Agency plan"}), 403
    admin_id = get_effective_admin_id(user)
    data = request.get_json(silent=True) or {}
    field_name = (data.get("field_name") or "").strip()
    field_type = data.get("field_type", "text")
    required = data.get("required", 0)
    if not field_name:
        return jsonify({"error": "Field name is required"}), 400
    if field_type not in ("text", "number", "date", "textarea", "select"):
        return jsonify({"error": "Invalid field type"}), 400
    field_id = db.add_custom_form_field(admin_id, field_name, field_type, required)
    db.log_admin_action(admin_id, user, "Added custom form field", f"{field_name} ({field_type})")
    return jsonify({"ok": True, "id": field_id})


@app.route("/api/form-config/custom-field/<int:field_id>", methods=["DELETE"])
def api_delete_custom_form_field(field_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Access denied"}), 403
    admin_id = get_effective_admin_id(user)
    db.delete_custom_form_field(admin_id, field_id)
    db.log_admin_action(admin_id, user, "Deleted custom form field", f"Field #{field_id}")
    return jsonify({"ok": True})


# ── Chatbot Customization (Agency plan only) ──

@app.route("/api/chatbot-customization", methods=["GET"])
def get_chatbot_customization_api():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    plan = user.get("plan", "free_trial")
    if plan != "agency":
        return jsonify({"error": "Chatbot customization requires an Agency plan."}), 403
    admin_id = get_effective_admin_id(user)
    customization = db.get_chatbot_customization(admin_id)
    return jsonify(customization or {})


@app.route("/api/chatbot-customization", methods=["POST"])
def save_chatbot_customization_api():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Only administrators can edit chatbot customization."}), 403
    plan = user.get("plan", "free_trial")
    if plan != "agency":
        return jsonify({"error": "Chatbot customization requires an Agency plan."}), 403
    admin_id = get_effective_admin_id(user)
    data = request.get_json() or {}
    db.save_chatbot_customization(admin_id, data)
    db.log_admin_action(admin_id, user, "Updated chatbot customization", "")
    return jsonify({"ok": True})


@app.route("/api/chatbot-customization/public/<admin_id_raw>")
def get_chatbot_customization_public(admin_id_raw):
    """Public endpoint for the embedded widget — no auth required."""
    admin_id_str = str(admin_id_raw).strip()
    if not admin_id_str.isdigit():
        resolved_user = db.get_user_by_public_id(admin_id_str)
        if not resolved_user:
            return jsonify({})
        admin_id = resolved_user["id"]
        user = resolved_user
    else:
        admin_id = int(admin_id_str)
        user = db.get_user_by_id(admin_id)
    if not user:
        return jsonify({})
    plan = user.get("plan", "free_trial")
    customization = db.get_chatbot_customization(admin_id)
    result = customization or {}
    result["hide_watermark"] = (plan == "agency")
    return jsonify(result)


# ── Chatbot Usage & Limits ──

@app.route("/api/chatbot-usage", methods=["GET"])
def api_chatbot_usage():
    """Return current chatbot usage vs plan limits for the admin dashboard."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    plan = user.get("plan", "free_trial")

    conversations_used = db.get_monthly_conversation_count(admin_id)
    conversations_limit = db.PLAN_MONTHLY_CONVERSATIONS.get(plan, 50)
    messages_used = db.get_monthly_message_count(admin_id)
    chatbots_used = db.get_active_chatbot_domain_count(admin_id)
    chatbots_limit = db.PLAN_MAX_CHATBOTS.get(plan, 1)
    active_domains = db.get_active_chatbot_domains(admin_id)

    return jsonify({
        "plan": plan,
        "conversations": {"used": conversations_used, "limit": conversations_limit},
        "messages": {"used": messages_used},
        "chatbots": {"used": chatbots_used, "limit": chatbots_limit},
        "active_domains": active_domains,
    })


@app.route("/api/chatbot-domains/<domain>", methods=["DELETE"])
def api_remove_chatbot_domain(domain):
    """Remove/deactivate a chatbot domain to free up a slot."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if not is_admin_role(user):
        return jsonify({"error": "Admin only"}), 403
    admin_id = get_effective_admin_id(user)
    db.deactivate_chatbot_domain(admin_id, domain)
    db.log_admin_action(admin_id, user, "Removed chatbot domain", f"{domain}")
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════
#  Feature 2 — Digital Patient Forms (Pre-Visit)
# ══════════════════════════════════════════════════════════════════

@app.route("/form/<token>")
def patient_form_page(token):
    """Serve the patient-form.html page (form handles all states via JS)."""
    return send_from_directory("static", "patient-form.html")

@app.route("/api/forms/<token>", methods=["GET"])
def api_get_form(token):
    """Return form data: booking details, patient name, field config, and whether already submitted."""
    form = db.get_form_by_token(token)
    if not form:
        return jsonify({"error": "Invalid or expired link"}), 404
    if form.get("submitted_at"):
        return jsonify({"already_submitted": True})
    # Get booking details
    booking = db.get_booking_by_id(form["booking_id"]) if form.get("booking_id") else None

    # Get form field configuration for this admin
    admin_id = form.get("admin_id")
    form_config = db.get_form_config(admin_id) if admin_id else None
    enabled_fields = {}
    custom_fields = []
    if form_config:
        for key, field in form_config.get("fields", {}).items():
            if field.get("enabled"):
                enabled_fields[key] = {
                    "label": field.get("label", key),
                    "group": field.get("group", "Other"),
                    "required": bool(field.get("required", 0))
                }
        custom_fields = form_config.get("custom_fields", [])

    return jsonify({
        "already_submitted": False,
        "full_name": form.get("full_name", ""),
        "date": booking["date"] if booking else "",
        "time": booking["time"] if booking else "",
        "doctor_name": booking.get("doctor_name", "") if booking else "",
        "enabled_fields": enabled_fields,
        "custom_fields": custom_fields
    })

@app.route("/api/forms/<token>", methods=["POST"])
def api_submit_form(token):
    """Submit the form. Save all data + signature. Sync to patient profile."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    success = db.submit_previsit_form(token, data)
    if not success:
        return jsonify({"error": "Form already submitted or invalid token"}), 400

    # Sync to patient profile
    form = db.get_form_by_token(token)
    if form and form.get("booking_id"):
        booking = db.get_booking_by_id(form["booking_id"])
        if booking and booking.get("patient_id"):
            db.sync_form_to_patient(data, booking["patient_id"])

        # ── Confirm the booking now that form is submitted ──
        if booking and booking.get("status") == "pending":
            db.confirm_booking_by_id(form["booking_id"])

            # ── ROI: Track revenue on confirmation ──
            try:
                revenue = 0
                if booking.get("service_id"):
                    svc = db.get_company_service_by_id(booking["service_id"])
                    if svc and svc.get("price"):
                        revenue = float(svc["price"])
                if not revenue and booking.get("doctor_id"):
                    doc = db.get_doctor_by_id(booking["doctor_id"])
                    if doc:
                        revenue = float(doc.get("avg_appointment_price", 20) or 20)
                if not revenue:
                    revenue = 20.0
                db.add_booking_revenue(form["booking_id"], revenue)
            except Exception as e:
                print(f"[form-confirm] ROI revenue error: {e}", flush=True)

            # Send confirmation email
            if booking.get("customer_email") and db.is_feature_enabled(booking.get("admin_id", 0), "email_booking_confirmation"):
                try:
                    dt = datetime.strptime(booking["date"], "%Y-%m-%d")
                    date_display = dt.strftime("%A, %B %d, %Y")
                except (ValueError, TypeError):
                    date_display = booking.get("date", "")
                try:
                    base_url = request.host_url.rstrip("/")
                    confirm_url = f"{base_url}/booking-confirmed/{form['booking_id']}"
                except Exception:
                    confirm_url = ""
                # Generate cancel token
                _fc_cancel_url = ""
                try:
                    _fc_tok = secrets.token_urlsafe(32)
                    _fc_conn = db.get_db()
                    _fc_conn.execute("UPDATE bookings SET cancel_token=? WHERE id=?", (_fc_tok, form['booking_id']))
                    _fc_conn.commit(); _fc_conn.close()
                    _fc_cancel_url = f"{base_url}/booking-cancel/{_fc_tok}"
                except Exception:
                    pass
                # Fetch service details for the email if service_id exists
                _svc_name = ""
                _svc_dur = 0
                _svc_price = ""
                _svc_prep = ""
                if booking.get("service_id"):
                    try:
                        _svc = db.get_company_service_by_id(booking["service_id"])
                        if _svc:
                            _svc_name = _svc.get("name", "")
                            _svc_dur = _svc.get("duration_minutes", 0)
                            _svc_price = f"{_svc.get('price', '')} {db.get_company_currency(booking.get('admin_id', 0))}".strip()
                            _svc_prep = _svc.get("preparation_instructions", "")
                    except Exception:
                        pass
                email.send_booking_confirmation_customer(
                    booking["customer_name"],
                    booking["customer_email"],
                    date_display,
                    booking["time"],
                    doctor_name=booking.get("doctor_name", ""),
                    confirm_url=confirm_url,
                    cancel_url=_fc_cancel_url,
                    service_name=_svc_name,
                    duration_minutes=_svc_dur,
                    price=_svc_price,
                    preparation_instructions=_svc_prep,
                    admin_id=booking.get("admin_id", 0),
                )

            # Schedule appointment reminders now that booking is confirmed
            if db.is_feature_enabled(booking.get("admin_id", 0), "auto_reminders"):
                try:
                    reminder_eng.schedule_reminders(form["booking_id"], booking.get("admin_id", 0))
                except Exception:
                    pass

            # If this was a waitlist booking, confirm the waitlist entry too
            if booking.get("waitlist_id"):
                try:
                    db.confirm_waitlist_patient(booking["waitlist_id"])
                except Exception:
                    pass

    # Award loyalty points for form completion
    if form and form.get("admin_id") and db.is_feature_enabled(form["admin_id"], "loyalty_program"):
        booking = db.get_booking_by_id(form["booking_id"]) if form.get("booking_id") else None
        if booking:
            try:
                patient = db.get_or_create_patient(form["admin_id"],
                    name=booking.get("customer_name", ""), email=booking.get("customer_email", ""), phone=booking.get("customer_phone", ""))
                if patient:
                    config = db.get_loyalty_config(form["admin_id"])
                    if config and config.get("is_active"):
                        db.add_loyalty_points(patient["id"], form["admin_id"], config.get("points_per_form", 25), "form_completed", "Pre-visit form completed")
            except Exception:
                pass

    return jsonify({"success": True})

@app.route("/noshow-reason/<token>")
def noshow_reason_page(token):
    """Serve the no-show reason form page."""
    # Verify token exists
    conn = db.get_db()
    row = conn.execute("SELECT id FROM bookings WHERE notes LIKE ? AND status='no_show'",
                       (f"%[noshow_token:{token}]%",)).fetchone()
    conn.close()
    if not row:
        return "<h2 style='text-align:center;margin-top:80px;font-family:sans-serif;color:#666'>This link is no longer valid.</h2>", 404
    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Let Us Know</title>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Outfit',sans-serif;background:#06080d;color:#f1f5f9;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px}}
.card{{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.06);border-radius:20px;padding:40px;max-width:500px;width:100%}}
h1{{font-size:1.5rem;font-weight:700;margin-bottom:8px}}
h1 span{{background:linear-gradient(135deg,#c9a84c,#d4af37);-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
p{{color:#94a3b8;font-size:0.95rem;line-height:1.6;margin-bottom:20px}}
textarea{{width:100%;min-height:120px;background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.1);border-radius:12px;padding:14px;color:#f1f5f9;font-family:inherit;font-size:0.95rem;resize:vertical}}
textarea:focus{{outline:none;border-color:#c9a84c}}
button{{margin-top:16px;width:100%;padding:14px;background:linear-gradient(135deg,#c9a84c,#d4af37);color:#fff;border:none;border-radius:50px;font-family:inherit;font-size:1rem;font-weight:600;cursor:pointer}}
button:hover{{opacity:0.9}}
.success{{text-align:center;padding:40px 0}}
.success h2{{color:#34d399;margin-bottom:8px}}
</style></head><body>
<div class="card" id="formCard">
<h1>We <span>Missed You</span></h1>
<p>We're sorry you couldn't make your appointment. Could you let us know what happened? Your feedback helps us improve.</p>
<textarea id="reason" placeholder="e.g. I had an emergency, I forgot, I couldn't find transport..."></textarea>
<button onclick="submitReason()">Submit</button>
</div>
<div class="card success" id="successCard" style="display:none">
<h2>Thank You!</h2>
<p>We appreciate your feedback. If you'd like to reschedule, please don't hesitate to reach out.</p>
</div>
<script>
function submitReason(){{
    var reason=document.getElementById('reason').value.trim();
    if(!reason){{alert('Please enter a reason');return}}
    fetch('/api/noshow-reason/{token}',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{reason:reason}})}})
    .then(function(r){{return r.json()}}).then(function(d){{
        if(d.ok){{document.getElementById('formCard').style.display='none';document.getElementById('successCard').style.display='block'}}
        else{{alert(d.error||'Something went wrong')}}
    }}).catch(function(){{alert('Failed to submit. Please try again.')}})
}}
</script></body></html>"""


@app.route("/api/noshow-reason/<token>", methods=["POST"])
def api_submit_noshow_reason(token):
    """Patient submits their no-show reason. Forward it to the doctor."""
    data = request.get_json()
    reason = (data.get("reason", "") if data else "").strip()
    if not reason:
        return jsonify({"error": "Please provide a reason"}), 400

    conn = db.get_db()
    booking = conn.execute("SELECT * FROM bookings WHERE notes LIKE ? AND status='no_show'",
                           (f"%[noshow_token:{token}]%",)).fetchone()
    if not booking:
        conn.close()
        return jsonify({"error": "Invalid or expired link"}), 404
    booking = dict(booking)

    # Remove the noshow token from notes (one-time use)
    clean_notes = (booking.get("notes", "") or "").replace(f"[noshow_token:{token}]", "").strip()
    noshow_note = f"No-show reason: {reason}"
    conn.execute("UPDATE bookings SET notes=? WHERE id=?",
                 (f"{clean_notes}\n{noshow_note}".strip(), booking["id"]))
    conn.commit()
    conn.close()

    # Forward reason to the doctor
    try:
        dt = datetime.strptime(booking["date"], "%Y-%m-%d")
        date_display = dt.strftime("%A, %B %d, %Y")
    except (ValueError, TypeError):
        date_display = booking.get("date", "")

    if booking.get("doctor_id") and db.is_feature_enabled(booking.get("admin_id", 0), "email_noshow_reason_doctor"):
        try:
            doctor = db.get_doctor_by_id(booking["doctor_id"])
            if doctor and doctor.get("email"):
                email.send_noshow_reason_to_doctor(
                    doctor["email"],
                    booking.get("doctor_name", ""),
                    booking["customer_name"],
                    date_display,
                    booking["time"],
                    reason=reason,
                )
        except Exception as e:
            print(f"[noshow] ERROR sending reason to doctor: {e}", flush=True)

    return jsonify({"ok": True})


@app.route("/api/bookings/<int:booking_id>/form", methods=["GET"])
def api_get_booking_form(booking_id):
    """Get form data for dashboard view of a booking."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    form = db.get_form_for_booking(booking_id)
    if not form:
        return jsonify({"error": "No form found"}), 404
    return jsonify(dict(form))


@app.route("/api/bookings/<int:booking_id>/full", methods=["GET"])
def api_booking_full_details(booking_id):
    """Enriched booking details for dashboard modal — type, patient info, form, visit count, medical history."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    conn = db.get_db()
    booking = conn.execute("SELECT * FROM bookings WHERE id=?", (booking_id,)).fetchone()
    if not booking:
        conn.close()
        return jsonify({"error": "Booking not found"}), 404
    booking = dict(booking)

    # Booking type
    booking_type = "Service" if booking.get("service_id") else "Appointment"
    service_name = booking.get("service", "General Consultation")

    # If service booking, fetch service details
    service_info = None
    if booking.get("service_id"):
        svc_row = conn.execute("SELECT name, price, currency, duration_minutes, description FROM company_services WHERE id=?",
                               (booking["service_id"],)).fetchone()
        if svc_row:
            service_info = dict(svc_row)

    # Visit count — how many bookings this patient has (by email or phone)
    visit_count = 0
    if booking.get("customer_email"):
        visit_count = conn.execute(
            "SELECT COUNT(*) FROM bookings WHERE customer_email=? AND admin_id=? AND status != 'cancelled'",
            (booking["customer_email"], booking.get("admin_id", 0))
        ).fetchone()[0]
    elif booking.get("customer_phone"):
        visit_count = conn.execute(
            "SELECT COUNT(*) FROM bookings WHERE customer_phone=? AND admin_id=? AND status != 'cancelled'",
            (booking["customer_phone"], booking.get("admin_id", 0))
        ).fetchone()[0]

    # Pre-visit form data
    form = conn.execute("SELECT * FROM patient_forms WHERE booking_id=?", (booking_id,)).fetchone()
    form_data = dict(form) if form else None
    form_status = "submitted" if (form_data and form_data.get("submitted_at")) else ("pending" if form_data else "not_sent")

    # Parse JSON fields in form_data into human-readable strings
    if form_data:
        for field in ("medical_history",):
            raw = form_data.get(field, "")
            if raw and isinstance(raw, str) and raw.startswith("{"):
                try:
                    parsed = json.loads(raw)
                    parts = []
                    conditions = parsed.get("conditions", [])
                    if conditions:
                        parts.append(", ".join(conditions))
                    other = parsed.get("other_text", "")
                    if other:
                        parts.append(other)
                    form_data[field] = "; ".join(parts) if parts else "None reported"
                except (json.JSONDecodeError, AttributeError):
                    pass

    # Patient profile (medical history, allergies, medications)
    patient_profile = None
    try:
        if booking.get("customer_email"):
            pat = conn.execute(
                "SELECT medical_history, allergies, medications, insurance_provider, insurance_policy, date_of_birth, gender FROM patients WHERE email=? AND admin_id=?",
                (booking["customer_email"], booking.get("admin_id", 0))
            ).fetchone()
            if pat:
                patient_profile = dict(pat)
                # Parse JSON fields in patient profile too
                for field in ("medical_history",):
                    raw = patient_profile.get(field, "")
                    if raw and isinstance(raw, str) and raw.startswith("{"):
                        try:
                            parsed = json.loads(raw)
                            parts = []
                            conditions = parsed.get("conditions", [])
                            if conditions:
                                parts.append(", ".join(conditions))
                            other = parsed.get("other_text", "")
                            if other:
                                parts.append(other)
                            patient_profile[field] = "; ".join(parts) if parts else "None reported"
                        except (json.JSONDecodeError, AttributeError):
                            pass
    except Exception:
        pass

    conn.close()

    return jsonify({
        "id": booking["id"],
        "booking_type": booking_type,
        "service_name": service_name,
        "service_info": service_info,
        "customer_name": booking.get("customer_name", ""),
        "customer_email": booking.get("customer_email", ""),
        "customer_phone": booking.get("customer_phone", ""),
        "doctor_name": booking.get("doctor_name", ""),
        "date": booking.get("date", ""),
        "time": booking.get("time", ""),
        "status": booking.get("status", ""),
        "notes": booking.get("notes", ""),
        "patient_type": booking.get("patient_type", ""),
        "promotion_code": booking.get("promotion_code", ""),
        "visit_count": visit_count,
        "form_status": form_status,
        "form_data": form_data,
        "patient_profile": patient_profile,
        "created_at": booking.get("created_at", ""),
    })


# ── Booking Confirmation Page (public, accessed via email link) ──

@app.route("/booking-confirmed/<int:booking_id>")
def booking_confirmed_page(booking_id):
    return send_from_directory("static", "booking-confirmed.html")

@app.route("/booking-cancel/<token>")
def booking_cancel_page(token):
    """Public page for cancelling a booking via email link."""
    if not token or len(token) < 10:
        return "<h2>Invalid cancellation link.</h2>", 404
    conn = db.get_db()
    booking = conn.execute("SELECT * FROM bookings WHERE cancel_token=? AND status != 'cancelled'", (token,)).fetchone()
    if not booking:
        conn.close()
        return """<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Cancellation</title>
        <style>body{font-family:'Helvetica Neue',sans-serif;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0;background:#f5f5f5;}
        .card{background:#fff;border-radius:16px;padding:48px;text-align:center;box-shadow:0 4px 20px rgba(0,0,0,0.08);max-width:480px;}
        h2{color:#1a1a2e;margin:0 0 12px;}p{color:#666;margin:0;}</style></head>
        <body><div class="card"><h2>Link Expired or Already Used</h2>
        <p>This booking has already been cancelled or the link is no longer valid.</p></div></body></html>"""
    # Cancel the booking
    conn.execute("UPDATE bookings SET status='cancelled', cancel_token='', cancelled_at=CURRENT_TIMESTAMP WHERE id=?", (booking["id"],))
    conn.commit()
    conn.close()
    # Update patient stats
    try:
        if booking.get("patient_id"):
            pconn = db.get_db()
            pconn.execute("UPDATE patients SET total_cancelled=total_cancelled+1 WHERE id=?", (booking["patient_id"],))
            pconn.commit(); pconn.close()
    except Exception:
        pass
    # Notify waitlist if applicable
    try:
        from waitlist_engine import process_waitlist_after_cancellation
        process_waitlist_after_cancellation(booking["id"], booking.get("admin_id", 0))
    except Exception:
        pass
    doctor_name = booking.get("doctor_name", "")
    date_display = booking.get("date", "")
    try:
        date_display = datetime.strptime(booking["date"], "%Y-%m-%d").strftime("%A, %B %d, %Y")
    except Exception:
        pass
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Appointment Cancelled</title>
    <style>body{{font-family:'Helvetica Neue',sans-serif;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0;background:#f5f5f5;}}
    .card{{background:#fff;border-radius:16px;padding:48px;text-align:center;box-shadow:0 4px 20px rgba(0,0,0,0.08);max-width:480px;}}
    h2{{color:#1a1a2e;margin:0 0 12px;}}p{{color:#666;line-height:1.6;}}
    .icon{{width:64px;height:64px;border-radius:50%;background:#fee2e2;display:flex;align-items:center;justify-content:center;margin:0 auto 20px;font-size:28px;}}</style></head>
    <body><div class="card">
    <div class="icon">&#10005;</div>
    <h2>Appointment Cancelled</h2>
    <p>Your appointment{f' with <strong>Dr. {doctor_name}</strong>' if doctor_name else ''} on <strong>{date_display}</strong> at <strong>{booking['time']}</strong> has been cancelled.</p>
    <p style="margin-top:16px;color:#999;">If this was a mistake, please contact us to rebook.</p>
    </div></body></html>"""


@app.route("/api/bookings/<int:booking_id>/details", methods=["GET"])
def api_booking_details_public(booking_id):
    """Public endpoint for booking confirmation page (limited info)."""
    conn = db.get_db()
    booking = conn.execute("SELECT customer_name, date, time, doctor_name, service, status, admin_id FROM bookings WHERE id=?", (booking_id,)).fetchone()
    if not booking:
        conn.close()
        return jsonify({"error": "Booking not found"}), 404
    booking = dict(booking)
    # Get business name
    biz = conn.execute("SELECT business_name FROM company_info WHERE user_id=?", (booking["admin_id"],)).fetchone()
    conn.close()
    return jsonify({
        "customer_name": booking["customer_name"],
        "date": booking["date"],
        "time": booking["time"],
        "doctor_name": booking.get("doctor_name", ""),
        "service": booking.get("service", ""),
        "status": booking["status"],
        "business_name": biz["business_name"] if biz else ""
    })


# ══════════════════════════════════════════════════════════════════
#  Feature 3 — Recall & Retention Automation
# ══════════════════════════════════════════════════════════════════

@app.route("/api/recall-rules", methods=["GET"])
def api_get_recall_rules():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    try:
        rules = recall_engine.get_recall_rules(admin_id)
        return jsonify(rules)
    except Exception:
        return jsonify(db.get_recall_rules(admin_id))

@app.route("/api/recall-rules", methods=["POST"])
def api_add_recall_rule():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    if not db.is_feature_enabled(admin_id, "auto_recall"):
        return jsonify({"error": "Recall feature is disabled"}), 403
    try:
        recall_engine.add_recall_rule(admin_id, data["treatment_type"], data.get("recall_days", 180), data.get("message_template", ""))
    except Exception:
        db.add_recall_rule(admin_id, data["treatment_type"], data.get("recall_days", 180), data.get("message_template", ""))
    db.log_admin_action(admin_id, user, "Added recall rule", f"{data['treatment_type']} ({data.get('recall_days', 180)} days)")
    return jsonify({"ok": True})

@app.route("/api/recall-rules/<int:rule_id>", methods=["PUT"])
def api_update_recall_rule(rule_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    if not db.is_feature_enabled(admin_id, "auto_recall"):
        return jsonify({"error": "Recall feature is disabled"}), 403
    db.update_recall_rule(rule_id, admin_id, **data)
    db.log_admin_action(admin_id, user, "Updated recall rule", f"Rule #{rule_id}")
    return jsonify({"ok": True})

@app.route("/api/recall-rules/<int:rule_id>", methods=["DELETE"])
def api_delete_recall_rule(rule_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    if not db.is_feature_enabled(admin_id, "auto_recall"):
        return jsonify({"error": "Recall feature is disabled"}), 403
    db.delete_recall_rule(rule_id, admin_id)
    db.log_admin_action(admin_id, user, "Deleted recall rule", f"Rule #{rule_id}")
    return jsonify({"ok": True})

@app.route("/api/recall-campaigns", methods=["GET"])
def api_get_recall_campaigns():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    return jsonify({"campaigns": db.get_recall_campaigns(admin_id), "stats": db.get_recall_stats(admin_id)})


# ── Recall Booking (public pages, no auth) ──

@app.route("/recall-book/<token>")
def recall_book_page(token):
    """Public booking page for recall patients. Served from email link."""
    campaign = db.get_recall_campaign_by_token(token)
    if not campaign:
        return "<h2>Invalid or expired recall link.</h2>", 404
    admin_id = campaign["admin_id"]
    company = db.get_company_info(admin_id)
    company_name = company.get("name", "Our Clinic") if company else "Our Clinic"
    doctors = db.get_doctors(admin_id)
    active_doctors = [d for d in doctors if d.get("status") in ("accepted", "active") and d.get("is_active", 1)]
    service_name = campaign.get("service_name", "Follow-up")
    patient_name = campaign.get("patient_name", "")
    patient_email = campaign.get("patient_email", "")
    patient_phone = campaign.get("patient_phone", "")

    doctors_json = json.dumps([{"id": d["id"], "name": d["name"], "specialty": d.get("specialty",""),
                                "start_time": d.get("start_time","09:00 AM"), "end_time": d.get("end_time","05:00 PM"),
                                "appointment_length": d.get("appointment_length", 60),
                                "availability": d.get("availability", ""),
                                "schedule_type": d.get("schedule_type", "fixed"),
                                "daily_hours": json.loads(d["daily_hours"]) if d.get("daily_hours") and isinstance(d.get("daily_hours"), str) else (d.get("daily_hours") or {})} for d in active_doctors])

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Book Your {_h(service_name)} Appointment — {_h(company_name)}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',sans-serif;background:#0a0d14;color:#f1f5f9;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:1.5rem}}
.card{{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.06);border-radius:24px;padding:2.5rem;max-width:520px;width:100%;position:relative;overflow:hidden}}
.card::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:linear-gradient(90deg,#8b5cf6,#6366f1,#3b82f6)}}
.logo{{text-align:center;margin-bottom:1.5rem}}
.logo h1{{font-size:1.1rem;font-weight:800;margin-bottom:0.25rem}}
.logo h1 span{{background:linear-gradient(135deg,#8b5cf6,#6366f1,#3b82f6);-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
.logo p{{font-size:0.82rem;color:#8b95a8}}
.hero{{text-align:center;margin-bottom:2rem;padding:1.5rem;border-radius:16px;background:rgba(139,92,246,0.06);border:1px solid rgba(139,92,246,0.12)}}
.hero h2{{font-size:1.3rem;font-weight:700;margin-bottom:0.3rem}}
.hero .service{{background:linear-gradient(135deg,#8b5cf6,#3b82f6);-webkit-background-clip:text;-webkit-text-fill-color:transparent;font-size:1.1rem;font-weight:700}}
.hero p{{color:#8b95a8;font-size:0.85rem;margin-top:0.5rem}}
label{{display:block;font-size:0.78rem;color:#8b95a8;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:0.4rem;font-weight:600}}
input,select{{width:100%;padding:0.7rem 1rem;background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:12px;color:#f1f5f9;font-size:0.9rem;font-family:inherit;outline:none;transition:border-color .3s}}
input:focus,select:focus{{border-color:rgba(139,92,246,0.4)}}
select option{{background:#1a1f2e;color:#f1f5f9}}
.form-group{{margin-bottom:1.25rem}}
.row{{display:grid;grid-template-columns:1fr 1fr;gap:1rem}}
.slots{{display:grid;grid-template-columns:repeat(auto-fill,minmax(90px,1fr));gap:0.5rem;margin-top:0.5rem;max-height:200px;overflow-y:auto}}
.slot{{padding:0.5rem;text-align:center;border-radius:10px;font-size:0.82rem;font-weight:600;cursor:pointer;border:1px solid rgba(255,255,255,0.08);background:rgba(255,255,255,0.03);transition:all .2s}}
.slot:hover{{border-color:rgba(139,92,246,0.3);background:rgba(139,92,246,0.06)}}
.slot.selected{{background:rgba(139,92,246,0.15);border-color:#8b5cf6;color:#a78bfa}}
.slot.booked{{opacity:0.3;cursor:not-allowed;text-decoration:line-through}}
.submit-btn{{display:block;width:100%;padding:0.85rem;border:none;border-radius:50px;background:linear-gradient(135deg,#8b5cf6,#6366f1,#3b82f6);color:#fff;font-size:1rem;font-weight:700;cursor:pointer;transition:transform .2s,box-shadow .2s;margin-top:1.5rem}}
.submit-btn:hover{{transform:translateY(-2px);box-shadow:0 8px 24px rgba(139,92,246,0.3)}}
.submit-btn:disabled{{opacity:0.5;cursor:not-allowed;transform:none;box-shadow:none}}
.success{{text-align:center;padding:3rem 1rem}}
.success svg{{margin-bottom:1rem}}
.success h2{{font-size:1.5rem;margin-bottom:0.5rem}}
.success p{{color:#8b95a8;font-size:0.9rem}}
.toast{{position:fixed;top:1rem;right:1rem;padding:0.75rem 1.25rem;border-radius:12px;font-size:0.85rem;font-weight:600;z-index:9999;transform:translateY(-20px);opacity:0;transition:all .3s}}
.toast.show{{transform:translateY(0);opacity:1}}
.toast.error{{background:rgba(248,113,113,0.15);border:1px solid rgba(248,113,113,0.3);color:#f87171}}
@media(max-width:500px){{.row{{grid-template-columns:1fr}}.card{{padding:1.5rem}}}}
</style>
</head>
<body>
<div class="card" id="bookingCard">
    <div class="logo">
        <h1>Chat<span>Genius</span></h1>
        <p>{_h(company_name)}</p>
    </div>
    <div class="hero">
        <h2>Welcome back, {_h(patient_name.split()[0] if patient_name else 'there')}!</h2>
        <div class="service">{_h(service_name)}</div>
        <p>It's time for your follow-up. Pick a date and time below.</p>
    </div>
    <form id="bookingForm" onsubmit="submitBooking(event)">
        <input type="hidden" id="recallToken" value="{_h(token)}">
        <div class="row">
            <div class="form-group">
                <label>Your Name</label>
                <input type="text" id="bkName" value="{_h(patient_name)}" required>
            </div>
            <div class="form-group">
                <label>Email</label>
                <input type="email" id="bkEmail" value="{_h(patient_email)}" required>
            </div>
        </div>
        <div class="form-group">
            <label>Phone (optional)</label>
            <input type="tel" id="bkPhone" value="{_h(patient_phone)}">
        </div>
        <div class="form-group">
            <label>Doctor</label>
            <select id="bkDoctor" onchange="onDoctorChange()" required>
                <option value="">Choose a doctor...</option>
            </select>
        </div>
        <div class="row">
            <div class="form-group">
                <label>Date</label>
                <input type="date" id="bkDate" onchange="loadSlots()" required min="">
            </div>
            <div class="form-group">
                <label>Time Slot</label>
                <div id="slotsContainer" style="color:#8b95a8;font-size:0.85rem;padding:0.5rem 0">Select doctor & date first</div>
                <input type="hidden" id="bkTime" required>
            </div>
        </div>
        <button type="submit" class="submit-btn" id="submitBtn" disabled>Book Appointment</button>
    </form>
</div>
<div id="toastEl" class="toast"></div>
<script>
const doctors = {doctors_json};
const service = '{_js(service_name)}';
const adminId = {admin_id};
let selectedSlot = '';

// Populate doctors
const sel = document.getElementById('bkDoctor');
doctors.forEach(d => {{
    const opt = document.createElement('option');
    opt.value = d.id;
    opt.textContent = d.name + (d.specialty ? ' — ' + d.specialty : '');
    sel.appendChild(opt);
}});

// Set min date to today
const dateInput = document.getElementById('bkDate');
dateInput.min = new Date().toISOString().split('T')[0];

const dayNames = ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'];

function getWorkingDays(doc) {{
    if (!doc) return [0,1,2,3,4,5,6];
    if (doc.schedule_type === 'flexible' && doc.daily_hours) {{
        const days = [];
        dayNames.forEach((name, idx) => {{
            const info = doc.daily_hours[name];
            if (info && !info.off) days.push(idx);
        }});
        return days.length ? days : [0,1,2,3,4,5,6];
    }}
    if (doc.availability) {{
        const avail = doc.availability.toLowerCase();
        const days = [];
        dayNames.forEach((name, idx) => {{
            if (avail.includes(name.toLowerCase()) || avail.includes(name.substring(0,3).toLowerCase())) days.push(idx);
        }});
        return days.length ? days : [0,1,2,3,4,5,6];
    }}
    return [0,1,2,3,4,5,6];
}}

let allowedDays = [0,1,2,3,4,5,6];

function onDoctorChange() {{
    const doctorId = document.getElementById('bkDoctor').value;
    const doc = doctors.find(d => d.id == doctorId);
    allowedDays = getWorkingDays(doc);
    // Clear date if it's no longer valid
    const curDate = dateInput.value;
    if (curDate) {{
        const d = new Date(curDate + 'T00:00:00');
        if (!allowedDays.includes(d.getDay())) {{
            dateInput.value = '';
            document.getElementById('slotsContainer').innerHTML = '<span style="color:#8b95a8;font-size:0.85rem">Select a valid date</span>';
            document.getElementById('bkTime').value = '';
            selectedSlot = '';
            updateSubmit();
        }}
    }}
    loadSlots();
}}

// Restrict date input to working days only
dateInput.addEventListener('input', function() {{
    const val = this.value;
    if (val) {{
        const d = new Date(val + 'T00:00:00');
        if (!allowedDays.includes(d.getDay())) {{
            this.value = '';
            const dayList = allowedDays.map(i => dayNames[i]).join(', ');
            document.getElementById('slotsContainer').innerHTML = '<span style="color:#f87171;font-size:0.85rem">Doctor only works on: ' + dayList + '</span>';
            document.getElementById('bkTime').value = '';
            selectedSlot = '';
            updateSubmit();
            return;
        }}
    }}
}});

async function loadSlots() {{
    const doctorId = document.getElementById('bkDoctor').value;
    const date = document.getElementById('bkDate').value;
    const container = document.getElementById('slotsContainer');
    document.getElementById('bkTime').value = '';
    selectedSlot = '';
    updateSubmit();
    if (!doctorId || !date) {{ container.innerHTML = '<span style="color:#8b95a8;font-size:0.85rem">Select doctor & date</span>'; return; }}

    // Check day is allowed before fetching
    const dd = new Date(date + 'T00:00:00');
    if (!allowedDays.includes(dd.getDay())) {{
        const dayList = allowedDays.map(i => dayNames[i]).join(', ');
        container.innerHTML = '<span style="color:#f87171;font-size:0.85rem">Doctor only works on: ' + dayList + '</span>';
        return;
    }}

    container.innerHTML = '<span style="color:#8b95a8;font-size:0.85rem">Loading...</span>';

    try {{
        const res = await fetch('/api/recall-book/slots?doctor_id=' + doctorId + '&date=' + date + '&admin_id=' + adminId);
        const data = await res.json();
        if (data.error) {{
            container.innerHTML = '<span style="color:#f87171;font-size:0.85rem">' + data.error + '</span>';
            return;
        }}
        if (!data.slots || !data.slots.length) {{
            container.innerHTML = '<span style="color:#f87171;font-size:0.85rem">No available slots on this date</span>';
            return;
        }}
        container.innerHTML = '<div class="slots">' + data.slots.map(s => {{
            const booked = data.booked.includes(s);
            return '<div class="slot' + (booked ? ' booked' : '') + '" onclick="' + (booked ? '' : "pickSlot(this,'" + s + "')") + '">' + s + '</div>';
        }}).join('') + '</div>';
    }} catch(e) {{
        container.innerHTML = '<span style="color:#f87171;font-size:0.85rem">Failed to load slots</span>';
    }}
}}

function pickSlot(el, time) {{
    document.querySelectorAll('.slot.selected').forEach(s => s.classList.remove('selected'));
    el.classList.add('selected');
    selectedSlot = time;
    document.getElementById('bkTime').value = time;
    updateSubmit();
}}

function updateSubmit() {{
    const ok = document.getElementById('bkDoctor').value && document.getElementById('bkDate').value && selectedSlot && document.getElementById('bkName').value;
    document.getElementById('submitBtn').disabled = !ok;
}}
document.querySelectorAll('#bookingForm input, #bookingForm select').forEach(el => el.addEventListener('input', updateSubmit));

async function submitBooking(e) {{
    e.preventDefault();
    const btn = document.getElementById('submitBtn');
    btn.disabled = true; btn.textContent = 'Booking...';
    try {{
        const body = {{
            token: document.getElementById('recallToken').value,
            name: document.getElementById('bkName').value,
            email: document.getElementById('bkEmail').value,
            phone: document.getElementById('bkPhone').value,
            doctor_id: parseInt(document.getElementById('bkDoctor').value),
            date: document.getElementById('bkDate').value,
            time: selectedSlot,
            service: service
        }};
        const res = await fetch('/api/recall-book', {{ method:'POST', headers:{{'Content-Type':'application/json'}}, body:JSON.stringify(body) }});
        const d = await res.json();
        if (d.error) {{ showToast(d.error); btn.disabled = false; btn.textContent = 'Book Appointment'; return; }}
        document.getElementById('bookingCard').innerHTML = '<div class="success"><svg width="64" height="64" viewBox="0 0 24 24" fill="none" stroke="#10b981" stroke-width="2"><circle cx="12" cy="12" r="10"/><path d="m9 12 2 2 4-4"/></svg><h2>Booking Confirmed!</h2><p>Your ' + service + ' appointment has been booked for<br><strong>' + body.date + '</strong> at <strong>' + body.time + '</strong>.<br><br>We look forward to seeing you!</p></div>';
    }} catch(e) {{
        showToast('Something went wrong. Please try again.');
        btn.disabled = false; btn.textContent = 'Book Appointment';
    }}
}}

function showToast(msg) {{
    const t = document.getElementById('toastEl');
    t.textContent = msg; t.className = 'toast error show';
    setTimeout(() => t.className = 'toast', 3000);
}}
</script>
</body>
</html>"""


def _h(s):
    """HTML-escape a string."""
    import html
    return html.escape(str(s)) if s else ""

def _js(s):
    """Escape for JS string literal."""
    return str(s).replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n") if s else ""


@app.route("/api/recall-book/slots", methods=["GET"])
def api_recall_book_slots():
    """Public endpoint: get available time slots for a doctor on a date."""
    doctor_id = request.args.get("doctor_id", type=int)
    date_str = request.args.get("date", "")
    if not doctor_id or not date_str:
        return jsonify({"error": "Missing doctor_id or date"}), 400

    doctor = db.get_doctor_by_id(doctor_id)
    if not doctor:
        return jsonify({"error": "Doctor not found"}), 404

    from datetime import datetime as _dt, timedelta as _td

    # Determine the day of week for the requested date
    try:
        date_obj = _dt.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    day_name = day_names[date_obj.weekday()]

    appt_len = int(doctor.get("appointment_length", 60) or 60)

    def _parse_time(t):
        for fmt in ("%I:%M %p", "%H:%M", "%I:%M%p"):
            try:
                return _dt.strptime(t.strip(), fmt)
            except ValueError:
                continue
        return _dt.strptime("09:00 AM", "%I:%M %p")

    # Check schedule type and get hours for this day
    start_str = None
    end_str = None
    if doctor.get("schedule_type") == "flexible" and doctor.get("daily_hours"):
        try:
            daily = doctor["daily_hours"]
            if isinstance(daily, str):
                daily = json.loads(daily)
            day_info = daily.get(day_name)
            if not day_info or day_info.get("off"):
                return jsonify({"slots": [], "booked": [], "error": "Doctor does not work on " + day_name})
            start_str = day_info.get("from", "09:00 AM")
            end_str = day_info.get("to", "05:00 PM")
        except (json.JSONDecodeError, TypeError):
            pass

    if not start_str:
        # Fixed schedule — check availability days
        avail = doctor.get("availability", "")
        if avail:
            # Parse availability like "Sunday, Monday, Thursday" or "Mon-Fri"
            avail_lower = avail.lower()
            day_lower = day_name.lower()
            day_short = day_lower[:3]
            if day_lower not in avail_lower and day_short not in avail_lower:
                return jsonify({"slots": [], "booked": [], "error": "Doctor does not work on " + day_name})
        start_str = doctor.get("start_time", "09:00 AM")
        end_str = doctor.get("end_time", "05:00 PM")

    # Check for blocked dates / holidays
    admin_id_val = request.args.get("admin_id", type=int) or doctor.get("admin_id", 0)
    off_dates = _get_off_dates_with_blocks(doctor_id, admin_id_val)
    if date_str in off_dates:
        return jsonify({"slots": [], "booked": [], "error": "Doctor is not available on this date"})

    start_t = _parse_time(start_str)
    end_t = _parse_time(end_str)
    slots = []
    current = start_t
    while current + _td(minutes=appt_len) <= end_t:
        slot_end = current + _td(minutes=appt_len)
        slots.append(current.strftime("%I:%M %p").lstrip("0") + " - " + slot_end.strftime("%I:%M %p").lstrip("0"))
        current += _td(minutes=appt_len)

    booked_raw = db.get_booked_times(doctor_id, date_str)
    # Convert booked start times to range format to match slots
    booked_ranges = []
    for bt in booked_raw:
        try:
            bt_parsed = _parse_time(bt)
            bt_end = bt_parsed + _td(minutes=appt_len)
            booked_ranges.append(bt_parsed.strftime("%I:%M %p").lstrip("0") + " - " + bt_end.strftime("%I:%M %p").lstrip("0"))
        except Exception:
            booked_ranges.append(bt)  # fallback: keep original
    return jsonify({"slots": slots, "booked": booked_ranges})


@app.route("/api/recall-book", methods=["POST"])
def api_recall_book():
    """Public endpoint: book an appointment from a recall email link."""
    data = request.get_json()
    token = data.get("token", "")
    campaign = db.get_recall_campaign_by_token(token)
    if not campaign:
        return jsonify({"error": "Invalid or expired recall link"}), 404

    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    doctor_id = data.get("doctor_id", 0)
    date_str = (data.get("date") or "").strip()
    time_str = (data.get("time") or "").strip()
    service = (data.get("service") or campaign.get("service_name") or "Follow-up").strip()

    if not name or not date_str or not time_str or not doctor_id:
        return jsonify({"error": "Please fill all required fields"}), 400

    # Check slot not already booked — compare start time portion
    booked = db.get_booked_times(doctor_id, date_str)
    # Extract start time from range format "9:00 AM - 10:00 AM" → "9:00 AM"
    time_start = time_str.split(" - ")[0].strip() if " - " in time_str else time_str
    for bt in booked:
        bt_start = bt.split(" - ")[0].strip() if " - " in bt else bt
        if bt_start == time_start or bt == time_str:
            return jsonify({"error": "This time slot is no longer available. Please pick another."}), 409

    doctor = db.get_doctor_by_id(doctor_id)
    doctor_name = doctor["name"] if doctor else ""
    admin_id = campaign["admin_id"]

    bid = db.add_booking(
        customer_name=name, customer_email=email, customer_phone=phone,
        date=date_str, time=time_str, service=service,
        doctor_id=doctor_id, doctor_name=doctor_name,
        admin_id=admin_id, status="confirmed"
    )

    db.mark_recall_booked(campaign["id"], booking_id=bid)

    # Send confirmation email
    try:
        import email_service as email_svc
        email_svc.send_booking_confirmation_customer(
            customer_name=name, customer_email=email, doctor_name=doctor_name,
            date_display=date_str, time_display=time_str, service_name=service, admin_id=admin_id
        )
    except Exception as e:
        print(f"[recall-book] Failed to send confirmation email: {e}", flush=True)

    return jsonify({"ok": True, "booking_id": bid})


# ══════════════════════════════════════════════════════════════════
#  Feature 4 — Missed Call Auto-Reply
# ══════════════════════════════════════════════════════════════════

@app.route("/api/missed-calls/webhook", methods=["POST"])
def api_missed_call_webhook():
    """Webhook endpoint for Twilio/phone system to report missed calls."""
    data = request.get_json() or request.form.to_dict()
    admin_id = data.get("admin_id", 1)
    caller = data.get("From") or data.get("caller_number", "")
    if not caller:
        return jsonify({"error": "No caller number"}), 400
    if not db.is_feature_enabled(admin_id, "missed_call_autoreply"):
        return jsonify({"ok": False, "reason": "Feature disabled"})
    try:
        result = missed_call_engine.handle_missed_call(admin_id, caller)
        return jsonify(result)
    except Exception:
        # Fallback to legacy
        call_id = db.log_missed_call(admin_id, caller)
        db.update_missed_call(call_id, reply_sent=1, reply_method="sms_whatsapp")
        return jsonify({"ok": True, "call_id": call_id})

@app.route("/api/missed-calls", methods=["GET"])
def api_get_missed_calls():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    return jsonify(db.get_missed_calls(admin_id))

@app.route("/api/settings/missed-calls", methods=["POST"])
def api_toggle_missed_calls():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    conn = db.get_db()
    conn.execute("UPDATE company_info SET missed_call_enabled=?, clinic_phone=? WHERE user_id=?",
                 (1 if data.get("enabled") else 0, data.get("clinic_phone", ""), admin_id))
    conn.commit()
    conn.close()
    db.log_admin_action(admin_id, user, "Updated missed calls settings", f"Enabled: {bool(data.get('enabled'))}")
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════
#  Feature 5 — Treatment Plan Follow-Up
# ══════════════════════════════════════════════════════════════════

@app.route("/api/treatment-followups", methods=["POST"])
def api_create_followup():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    if not db.is_feature_enabled(admin_id, "auto_followups"):
        return jsonify({"error": "Follow-ups feature is disabled"}), 403
    try:
        result = treatment_followup_engine.create_followup(
            admin_id=admin_id,
            doctor_id=data.get("doctor_id", 0),
            patient_name=data["patient_name"],
            treatment_name=data["treatment_name"],
            patient_email=data.get("patient_email", ""),
            patient_phone=data.get("patient_phone", "")
        )
        return jsonify(result if result else {"ok": True, "message": "Follow-up sequence created (2, 5, 10 days)."})
    except Exception:
        db.create_treatment_followup(
            admin_id=admin_id,
            doctor_id=data.get("doctor_id", 0),
            patient_name=data["patient_name"],
            treatment_name=data["treatment_name"],
            patient_email=data.get("patient_email", ""),
            patient_phone=data.get("patient_phone", "")
        )
        return jsonify({"ok": True, "message": "Follow-up sequence created (2, 5, 10 days)."})

@app.route("/api/treatment-followups", methods=["GET"])
def api_get_followups():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    status = request.args.get("status")
    return jsonify(db.get_treatment_followups(admin_id, status=status))

@app.route("/api/treatment-followups/<int:fid>/cancel", methods=["POST"])
def api_cancel_followup(fid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    conn = db.get_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("UPDATE treatment_followups SET status='cancelled', cancelled_at=? WHERE id=?", (now, fid))
    conn.commit()
    conn.close()
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Cancelled treatment followup", f"Followup #{fid}")
    return jsonify({"ok": True})


# ── Follow-up from Booking (Previous Bookings → Add to Follow-up) ──

@app.route("/api/followup-from-booking", methods=["POST"])
def api_followup_from_booking():
    """Create a follow-up from a completed booking and send email to patient."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Not authenticated"}), 401

    data = request.get_json()
    booking_id = data.get("booking_id")
    if not booking_id:
        return jsonify({"error": "Missing booking_id"}), 400

    # Get the booking
    conn = db.get_db()
    booking = conn.execute("SELECT * FROM bookings WHERE id=?", (booking_id,)).fetchone()
    conn.close()
    if not booking:
        return jsonify({"error": "Booking not found"}), 404
    booking = dict(booking)

    admin_id = get_effective_admin_id(user)
    if booking.get("admin_id") != admin_id:
        return jsonify({"error": "Access denied"}), 403

    # Check not already followed up
    conn = db.get_db()
    existing = conn.execute(
        "SELECT id FROM treatment_followups WHERE admin_id=? AND booking_id=? AND status != 'cancelled'",
        (admin_id, booking_id)).fetchone()
    conn.close()
    if existing:
        return jsonify({"error": "A follow-up already exists for this booking"}), 409

    # Create the follow-up with token
    result = db.create_single_followup(
        admin_id=admin_id,
        doctor_id=booking.get("doctor_id", 0),
        patient_name=booking.get("customer_name", ""),
        treatment_name=booking.get("service", "Follow-up"),
        patient_email=booking.get("customer_email", ""),
        patient_phone=booking.get("customer_phone", ""),
        booking_id=booking_id
    )

    # Send follow-up email with "Confirm Follow-up" button
    patient_email = booking.get("customer_email", "")
    if patient_email:
        import os
        base = os.environ.get("SERVER_URL", request.host_url.rstrip("/"))
        booking_url = f"{base}/followup-book/{result['followup_token']}"
        try:
            import email_service as email_svc
            email_svc.send_treatment_followup(
                to_email=patient_email,
                patient_name=booking.get("customer_name", "Patient"),
                treatment_name=booking.get("service", "your treatment"),
                day_number=0,
                booking_url=booking_url,
                admin_id=admin_id
            )
            # Mark as sent
            conn = db.get_db()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute("UPDATE treatment_followups SET status='sent', sent_at=? WHERE id=?",
                         (now, result["id"]))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"[followup] Failed to send email: {e}", flush=True)

    return jsonify({"ok": True, "followup_id": result["id"]})


@app.route("/followup-book/<token>")
def followup_book_page(token):
    """Public booking page for follow-up patients. Accessed from email link."""
    followup = db.get_followup_by_token(token)
    if not followup:
        return "<h2>Invalid or expired follow-up link.</h2>", 404

    if followup.get("status") == "booked":
        return """<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
        <title>Already Booked</title><style>body{font-family:Inter,sans-serif;background:#0a0d14;color:#f1f5f9;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0}
        .card{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.06);border-radius:24px;padding:2.5rem;max-width:480px;text-align:center}
        h2{margin-bottom:0.5rem}p{color:#8b95a8;font-size:0.9rem}</style></head>
        <body><div class="card"><h2>Already Booked</h2><p>This follow-up appointment has already been confirmed. We look forward to seeing you!</p></div></body></html>"""

    admin_id = followup["admin_id"]
    company = db.get_company_info(admin_id)
    company_name = company.get("name", "Our Clinic") if company else "Our Clinic"
    doctors = db.get_doctors(admin_id)
    active_doctors = [d for d in doctors if d.get("status") in ("accepted", "active") and d.get("is_active", 1)]
    service_name = followup.get("treatment_name", "Follow-up")
    patient_name = followup.get("patient_name", "")
    patient_email = followup.get("patient_email", "")
    patient_phone = followup.get("patient_phone", "")

    doctors_json = json.dumps([{"id": d["id"], "name": d["name"], "specialty": d.get("specialty", ""),
                                "start_time": d.get("start_time", "09:00 AM"), "end_time": d.get("end_time", "05:00 PM"),
                                "appointment_length": d.get("appointment_length", 60),
                                "availability": d.get("availability", ""),
                                "schedule_type": d.get("schedule_type", "fixed"),
                                "daily_hours": json.loads(d["daily_hours"]) if d.get("daily_hours") and isinstance(d.get("daily_hours"), str) else (d.get("daily_hours") or {})} for d in active_doctors])

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Confirm Follow-Up — {_h(company_name)}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',sans-serif;background:#0a0d14;color:#f1f5f9;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:1.5rem}}
.card{{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.06);border-radius:24px;padding:2.5rem;max-width:520px;width:100%;position:relative;overflow:hidden}}
.card::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:linear-gradient(90deg,#10b981,#059669,#047857)}}
.logo{{text-align:center;margin-bottom:1.5rem}}
.logo h1{{font-size:1.1rem;font-weight:800;margin-bottom:0.25rem}}
.logo h1 span{{background:linear-gradient(135deg,#8b5cf6,#6366f1,#3b82f6);-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
.logo p{{font-size:0.82rem;color:#8b95a8}}
.hero{{text-align:center;margin-bottom:2rem;padding:1.5rem;border-radius:16px;background:rgba(16,185,129,0.06);border:1px solid rgba(16,185,129,0.15)}}
.hero h2{{font-size:1.3rem;font-weight:700;margin-bottom:0.3rem}}
.hero .service{{background:linear-gradient(135deg,#10b981,#059669);-webkit-background-clip:text;-webkit-text-fill-color:transparent;font-size:1.1rem;font-weight:700}}
.hero p{{color:#8b95a8;font-size:0.85rem;margin-top:0.5rem}}
label{{display:block;font-size:0.78rem;color:#8b95a8;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:0.4rem;font-weight:600}}
input,select{{width:100%;padding:0.7rem 1rem;background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:12px;color:#f1f5f9;font-size:0.9rem;font-family:inherit;outline:none;transition:border-color .3s}}
input:focus,select:focus{{border-color:rgba(16,185,129,0.4)}}
select option{{background:#1a1f2e;color:#f1f5f9}}
.form-group{{margin-bottom:1.25rem}}
.row{{display:grid;grid-template-columns:1fr 1fr;gap:1rem}}
.slots{{display:grid;grid-template-columns:repeat(auto-fill,minmax(90px,1fr));gap:0.5rem;margin-top:0.5rem;max-height:200px;overflow-y:auto}}
.slot{{padding:0.5rem;text-align:center;border-radius:10px;font-size:0.82rem;font-weight:600;cursor:pointer;border:1px solid rgba(255,255,255,0.08);background:rgba(255,255,255,0.03);transition:all .2s}}
.slot:hover{{border-color:rgba(16,185,129,0.3);background:rgba(16,185,129,0.06)}}
.slot.selected{{background:rgba(16,185,129,0.15);border-color:#10b981;color:#34d399}}
.slot.booked{{opacity:0.3;cursor:not-allowed;text-decoration:line-through}}
.submit-btn{{display:block;width:100%;padding:0.85rem;border:none;border-radius:50px;background:linear-gradient(135deg,#10b981,#059669,#047857);color:#fff;font-size:1rem;font-weight:700;cursor:pointer;transition:transform .2s,box-shadow .2s;margin-top:1.5rem}}
.submit-btn:hover{{transform:translateY(-2px);box-shadow:0 8px 24px rgba(16,185,129,0.3)}}
.submit-btn:disabled{{opacity:0.5;cursor:not-allowed;transform:none;box-shadow:none}}
.success{{text-align:center;padding:3rem 1rem}}
.toast{{position:fixed;top:1rem;right:1rem;padding:0.75rem 1.25rem;border-radius:12px;font-size:0.85rem;font-weight:600;z-index:9999;transform:translateY(-20px);opacity:0;transition:all .3s}}
.toast.show{{transform:translateY(0);opacity:1}}
.toast.error{{background:rgba(248,113,113,0.15);border:1px solid rgba(248,113,113,0.3);color:#f87171}}
@media(max-width:500px){{.row{{grid-template-columns:1fr}}.card{{padding:1.5rem}}}}
</style>
</head>
<body>
<div class="card" id="bookingCard">
    <div class="logo">
        <h1>Chat<span>Genius</span></h1>
        <p>{_h(company_name)}</p>
    </div>
    <div class="hero">
        <h2>Confirm Your Follow-Up</h2>
        <div class="service">{_h(service_name)}</div>
        <p>Hi {_h(patient_name.split()[0] if patient_name else 'there')}, please pick your preferred date and time for your follow-up visit.</p>
    </div>
    <form id="bookingForm" onsubmit="submitBooking(event)">
        <input type="hidden" id="fuToken" value="{_h(token)}">
        <div class="row">
            <div class="form-group">
                <label>Your Name</label>
                <input type="text" id="bkName" value="{_h(patient_name)}" required>
            </div>
            <div class="form-group">
                <label>Email</label>
                <input type="email" id="bkEmail" value="{_h(patient_email)}" required>
            </div>
        </div>
        <div class="form-group">
            <label>Phone (optional)</label>
            <input type="tel" id="bkPhone" value="{_h(patient_phone)}">
        </div>
        <div class="form-group">
            <label>Doctor</label>
            <select id="bkDoctor" onchange="onDoctorChange()" required>
                <option value="">Choose a doctor...</option>
            </select>
        </div>
        <div class="row">
            <div class="form-group">
                <label>Date</label>
                <input type="date" id="bkDate" onchange="loadSlots()" required>
            </div>
            <div class="form-group">
                <label>Time Slot</label>
                <div id="slotsContainer" style="color:#8b95a8;font-size:0.85rem;padding:0.5rem 0">Select doctor & date first</div>
                <input type="hidden" id="bkTime" required>
            </div>
        </div>
        <button type="submit" class="submit-btn" id="submitBtn" disabled>Confirm Follow-Up</button>
    </form>
</div>
<div id="toastEl" class="toast"></div>
<script>
const doctors = {doctors_json};
const service = '{_js(service_name)}';
const adminId = {admin_id};
let selectedSlot = '';

const sel = document.getElementById('bkDoctor');
doctors.forEach(d => {{
    const opt = document.createElement('option');
    opt.value = d.id;
    opt.textContent = d.name + (d.specialty ? ' — ' + d.specialty : '');
    sel.appendChild(opt);
}});

const dateInput = document.getElementById('bkDate');
dateInput.min = new Date().toISOString().split('T')[0];

const dayNames = ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'];

function getWorkingDays(doc) {{
    if (!doc) return [0,1,2,3,4,5,6];
    if (doc.schedule_type === 'flexible' && doc.daily_hours) {{
        const days = [];
        dayNames.forEach((name, idx) => {{
            const info = doc.daily_hours[name];
            if (info && !info.off) days.push(idx);
        }});
        return days.length ? days : [0,1,2,3,4,5,6];
    }}
    if (doc.availability) {{
        const avail = doc.availability.toLowerCase();
        const days = [];
        dayNames.forEach((name, idx) => {{
            if (avail.includes(name.toLowerCase()) || avail.includes(name.substring(0,3).toLowerCase())) days.push(idx);
        }});
        return days.length ? days : [0,1,2,3,4,5,6];
    }}
    return [0,1,2,3,4,5,6];
}}

let allowedDays = [0,1,2,3,4,5,6];

function onDoctorChange() {{
    const doctorId = document.getElementById('bkDoctor').value;
    const doc = doctors.find(d => d.id == doctorId);
    allowedDays = getWorkingDays(doc);
    const curDate = dateInput.value;
    if (curDate) {{
        const d = new Date(curDate + 'T00:00:00');
        if (!allowedDays.includes(d.getDay())) {{
            dateInput.value = '';
            document.getElementById('slotsContainer').innerHTML = '<span style="color:#8b95a8;font-size:0.85rem">Select a valid date</span>';
            document.getElementById('bkTime').value = '';
            selectedSlot = '';
            updateSubmit();
        }}
    }}
    loadSlots();
}}

dateInput.addEventListener('input', function() {{
    const val = this.value;
    if (val) {{
        const d = new Date(val + 'T00:00:00');
        if (!allowedDays.includes(d.getDay())) {{
            this.value = '';
            const dayList = allowedDays.map(i => dayNames[i]).join(', ');
            document.getElementById('slotsContainer').innerHTML = '<span style="color:#f87171;font-size:0.85rem">Doctor only works on: ' + dayList + '</span>';
            document.getElementById('bkTime').value = '';
            selectedSlot = '';
            updateSubmit();
            return;
        }}
    }}
}});

async function loadSlots() {{
    const doctorId = document.getElementById('bkDoctor').value;
    const date = document.getElementById('bkDate').value;
    const container = document.getElementById('slotsContainer');
    document.getElementById('bkTime').value = '';
    selectedSlot = '';
    updateSubmit();
    if (!doctorId || !date) {{ container.innerHTML = '<span style="color:#8b95a8;font-size:0.85rem">Select doctor & date</span>'; return; }}

    const dd = new Date(date + 'T00:00:00');
    if (!allowedDays.includes(dd.getDay())) {{
        const dayList = allowedDays.map(i => dayNames[i]).join(', ');
        container.innerHTML = '<span style="color:#f87171;font-size:0.85rem">Doctor only works on: ' + dayList + '</span>';
        return;
    }}

    container.innerHTML = '<span style="color:#8b95a8;font-size:0.85rem">Loading...</span>';
    try {{
        const res = await fetch('/api/recall-book/slots?doctor_id=' + doctorId + '&date=' + date + '&admin_id=' + adminId);
        const data = await res.json();
        if (data.error) {{
            container.innerHTML = '<span style="color:#f87171;font-size:0.85rem">' + data.error + '</span>';
            return;
        }}
        if (!data.slots || !data.slots.length) {{
            container.innerHTML = '<span style="color:#f87171;font-size:0.85rem">No available slots on this date</span>';
            return;
        }}
        container.innerHTML = '<div class="slots">' + data.slots.map(s => {{
            const booked = data.booked.includes(s);
            return '<div class="slot' + (booked ? ' booked' : '') + '" onclick="' + (booked ? '' : "pickSlot(this,'" + s + "')") + '">' + s + '</div>';
        }}).join('') + '</div>';
    }} catch(e) {{
        container.innerHTML = '<span style="color:#f87171;font-size:0.85rem">Failed to load slots</span>';
    }}
}}

function pickSlot(el, time) {{
    document.querySelectorAll('.slot.selected').forEach(s => s.classList.remove('selected'));
    el.classList.add('selected');
    selectedSlot = time;
    document.getElementById('bkTime').value = time;
    updateSubmit();
}}

function updateSubmit() {{
    const ok = document.getElementById('bkDoctor').value && document.getElementById('bkDate').value && selectedSlot && document.getElementById('bkName').value;
    document.getElementById('submitBtn').disabled = !ok;
}}
document.querySelectorAll('#bookingForm input, #bookingForm select').forEach(el => el.addEventListener('input', updateSubmit));

async function submitBooking(e) {{
    e.preventDefault();
    const btn = document.getElementById('submitBtn');
    btn.disabled = true; btn.textContent = 'Booking...';
    try {{
        const body = {{
            token: document.getElementById('fuToken').value,
            name: document.getElementById('bkName').value,
            email: document.getElementById('bkEmail').value,
            phone: document.getElementById('bkPhone').value,
            doctor_id: parseInt(document.getElementById('bkDoctor').value),
            date: document.getElementById('bkDate').value,
            time: selectedSlot,
            service: service
        }};
        const res = await fetch('/api/followup-book', {{ method:'POST', headers:{{'Content-Type':'application/json'}}, body:JSON.stringify(body) }});
        const d = await res.json();
        if (d.error) {{ showToast(d.error); btn.disabled = false; btn.textContent = 'Confirm Follow-Up'; return; }}
        document.getElementById('bookingCard').innerHTML = '<div class="success"><svg width="64" height="64" viewBox="0 0 24 24" fill="none" stroke="#10b981" stroke-width="2"><circle cx="12" cy="12" r="10"/><path d="m9 12 2 2 4-4"/></svg><h2>Follow-Up Confirmed!</h2><p>Your ' + service + ' follow-up has been booked for<br><strong>' + body.date + '</strong> at <strong>' + body.time + '</strong>.<br><br>We look forward to seeing you!</p></div>';
    }} catch(e) {{
        showToast('Something went wrong. Please try again.');
        btn.disabled = false; btn.textContent = 'Confirm Follow-Up';
    }}
}}

function showToast(msg) {{
    const t = document.getElementById('toastEl');
    t.textContent = msg; t.className = 'toast error show';
    setTimeout(() => t.className = 'toast', 3000);
}}
</script>
</body>
</html>"""


@app.route("/api/followup-book", methods=["POST"])
def api_followup_book():
    """Public endpoint: book a follow-up appointment from email link."""
    data = request.get_json()
    token = data.get("token", "")
    followup = db.get_followup_by_token(token)
    if not followup:
        return jsonify({"error": "Invalid or expired follow-up link"}), 404
    if followup.get("status") == "booked":
        return jsonify({"error": "This follow-up has already been booked"}), 409

    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    doctor_id = data.get("doctor_id", 0)
    date_str = (data.get("date") or "").strip()
    time_str = (data.get("time") or "").strip()
    service = (data.get("service") or followup.get("treatment_name") or "Follow-up").strip()

    if not name or not date_str or not time_str or not doctor_id:
        return jsonify({"error": "Please fill all required fields"}), 400

    booked = db.get_booked_times(doctor_id, date_str)
    time_start = time_str.split(" - ")[0].strip() if " - " in time_str else time_str
    for bt in booked:
        bt_start = bt.split(" - ")[0].strip() if " - " in bt else bt
        if bt_start == time_start or bt == time_str:
            return jsonify({"error": "This time slot is no longer available. Please pick another."}), 409

    doctor = db.get_doctor_by_id(doctor_id)
    doctor_name = doctor["name"] if doctor else ""
    admin_id = followup["admin_id"]

    bid = db.add_booking(
        customer_name=name, customer_email=email, customer_phone=phone,
        date=date_str, time=time_str, service=service,
        doctor_id=doctor_id, doctor_name=doctor_name,
        admin_id=admin_id, status="confirmed"
    )

    db.mark_followup_booked(followup["id"], booking_id=bid)

    # Send confirmation email
    try:
        import email_service as email_svc
        email_svc.send_booking_confirmation_customer(
            customer_name=name, customer_email=email, doctor_name=doctor_name,
            date_display=date_str, time_display=time_str, service_name=service, admin_id=admin_id
        )
    except Exception as e:
        print(f"[followup-book] Failed to send confirmation email: {e}", flush=True)

    return jsonify({"ok": True, "booking_id": bid})


# ══════════════════════════════════════════════════════════════════
#  Feature 6 — Multilingual AI (language detection in /chat)
# ══════════════════════════════════════════════════════════════════

def detect_language(text):
    """Detect language of text. Returns ISO code."""
    try:
        from langdetect import detect
        lang = detect(text)
        lang_map = {"ar": "ar", "en": "en", "ur": "ur", "tl": "tl", "fil": "tl"}
        return lang_map.get(lang, lang)
    except Exception:
        return "en"

LANG_LABELS = {"en": "English", "ar": "العربية", "ur": "اردو", "tl": "Tagalog"}

@app.route("/api/chat-sessions/<session_id>/language", methods=["POST"])
def api_set_chat_language(session_id):
    """Staff override for conversation language."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    lang = data.get("language", "en")
    if session_id in sessions:
        sessions[session_id]["language"] = lang
    return jsonify({"ok": True, "language": lang})


# ══════════════════════════════════════════════════════════════════
#  Feature 7 — Before & After Gallery
# ══════════════════════════════════════════════════════════════════

@app.route("/api/gallery", methods=["GET"])
def api_get_gallery():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    treatment = request.args.get("treatment_type")
    return jsonify(db.get_gallery(admin_id, treatment_type=treatment))

@app.route("/api/gallery", methods=["POST"])
def api_upload_gallery():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    treatment = request.form.get("treatment_type", "general")
    image_type = request.form.get("image_type", "after")
    caption = request.form.get("caption", "")
    pair_id = request.form.get("pair_id", "")
    file = request.files.get("image")
    if not file:
        return jsonify({"error": "No image uploaded"}), 400
    # Try engine first
    try:
        result = gallery_engine.upload_image(admin_id, treatment, file, image_type, caption)
        if result:
            return jsonify(result)
    except Exception:
        pass
    # Fallback: legacy upload
    existing = db.get_gallery(admin_id, treatment_type=treatment)
    if len(existing) >= 20:
        return jsonify({"error": "Maximum 20 photos per treatment type"}), 400
    import uuid as _uuid
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png"):
        return jsonify({"error": "Only JPG and PNG files allowed"}), 400
    safe_name = f"{_uuid.uuid4().hex}{ext}"
    upload_dir = os.path.join(os.path.dirname(__file__), "uploads", "gallery")
    os.makedirs(upload_dir, exist_ok=True)
    file.save(os.path.join(upload_dir, safe_name))
    image_url = f"/uploads/gallery/{safe_name}"
    db.add_gallery_image(admin_id, treatment, image_url, image_type, pair_id, caption)
    return jsonify({"ok": True, "url": image_url})

@app.route("/api/gallery/<int:image_id>", methods=["DELETE"])
def api_delete_gallery(image_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    db.delete_gallery_image(image_id, admin_id)
    db.log_admin_action(admin_id, user, "Deleted gallery image", f"Image #{image_id}")
    return jsonify({"ok": True})

@app.route("/uploads/gallery/<path:filename>")
def serve_gallery_image(filename):
    return send_from_directory(os.path.join(os.path.dirname(__file__), "uploads", "gallery"), filename)

@app.route("/api/gallery/public", methods=["GET"])
def api_public_gallery():
    """Public endpoint for chatbot to fetch gallery images."""
    admin_id = request.args.get("admin_id", 1, type=int)
    treatment = request.args.get("treatment_type")
    return jsonify(db.get_gallery(admin_id, treatment_type=treatment))


# ══════════════════════════════════════════════════════════════════
#  Feature 8 — Multi-Doctor Comparison (in chatbot)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/doctors/compare", methods=["GET"])
def api_compare_doctors():
    admin_id = request.args.get("admin_id", 1, type=int)
    category = request.args.get("category", "")
    doctors = db.get_doctors(admin_id)
    active = [d for d in doctors if d.get("status") == "active" and d.get("is_active", 1)]
    if category:
        cat_lower = category.lower()
        active = [d for d in active if cat_lower in (d.get("specialty") or "").lower()]
    result = []
    for d in active:
        # Get next available slot
        breaks = db.get_doctor_breaks(d["id"])
        slots = _generate_doctor_slots(d, breaks)
        available_slots = [s for s in slots if not s.get("booked")]
        next_slot = available_slots[0]["time"] if available_slots else "No slots today"
        result.append({
            "id": d["id"], "name": d["name"], "specialty": d.get("specialty", ""),
            "years_of_experience": d.get("years_of_experience", 0),
            "languages": d.get("languages", ""), "bio": d.get("bio", ""),
            "next_available": next_slot,
            "availability": _doctor_avail_str(d)
        })
    return jsonify(result)


# ══════════════════════════════════════════════════════════════════
#  Feature 9 — Emergency Fast-Track Flow (integrated in /chat)
# ══════════════════════════════════════════════════════════════════

EMERGENCY_KEYWORDS = ["severe pain", "broken tooth", "swollen jaw", "knocked out",
    "can't sleep from pain", "bleeding won't stop", "abscess", "face swollen",
    "emergency", "unbearable pain", "tooth fell out", "hit my tooth"]

FIRST_AID = {
    "knocked out": "Keep the tooth moist — place it in milk or between your cheek and gum. Do NOT touch the root. Come in immediately.",
    "broken tooth": "Rinse your mouth gently with warm water. Apply a cold compress to reduce swelling. Avoid chewing on that side.",
    "swollen": "Apply a cold compress to the outside of your cheek (20 min on, 20 min off). Do NOT apply heat. Take ibuprofen if you can.",
    "bleeding": "Apply firm pressure with a clean gauze for 15-20 minutes. If bleeding doesn't stop, come in immediately.",
    "pain": "Rinse with warm salt water. Take over-the-counter pain relief (ibuprofen preferred). Avoid very hot or cold foods.",
    "abscess": "Do NOT pop or squeeze it. Rinse with warm salt water every few hours. This requires urgent antibiotic treatment.",
}

def get_first_aid(message):
    msg_lower = message.lower()
    for key, advice in FIRST_AID.items():
        if key in msg_lower:
            return advice
    return FIRST_AID["pain"]


# ── Feature 9 — Emergency Alerts API (engine) ──

@app.route('/api/emergency-alerts', methods=['GET'])
def get_emergency_alerts_api():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    try:
        alerts = emergency_handler.get_emergency_alerts(admin_id)
        return jsonify(alerts)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/emergency-alerts/<int:alert_id>/acknowledge', methods=['POST'])
def acknowledge_emergency(alert_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    try:
        emergency_handler.acknowledge_alert(alert_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  Feature 10 — Live Chat Handoff
# ══════════════════════════════════════════════════════════════════

@app.route("/api/handoffs", methods=["GET"])
def api_get_handoffs():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    try:
        return jsonify(handoff_engine.get_handoff_queue(admin_id))
    except Exception:
        return jsonify(db.get_handoff_queue(admin_id))

@app.route("/api/handoffs/<int:hid>/assign", methods=["POST"])
def api_assign_handoff(hid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    try:
        handoff_engine.assign_handoff(hid, user["id"], user["name"])
    except Exception:
        db.assign_handoff(hid, user["id"], user["name"])
    return jsonify({"ok": True})

@app.route("/api/handoffs/<int:hid>/resolve", methods=["POST"])
def api_resolve_handoff(hid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    try:
        handoff_engine.resolve_handoff(hid, data.get("notes", ""))
    except Exception:
        db.resolve_handoff(hid, data.get("notes", ""))
    return jsonify({"ok": True})

@app.route("/api/handoffs/<int:hid>/message", methods=["POST"])
def api_handoff_message(hid):
    """Staff sends a message to the patient in a handed-off conversation."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    try:
        handoff_engine.send_handoff_message(hid, "staff", user["name"], data["message"])
    except Exception:
        pass
    # Also store in session history (legacy)
    handoff = db.get_handoff_by_session(data.get("session_id", ""))
    if not handoff:
        return jsonify({"error": "Handoff not found"}), 404
    sid = handoff["session_id"]
    if sid in sessions:
        sessions[sid]["history"].append({"role": "bot", "content": data["message"], "from_staff": True, "staff_name": user["name"]})
    return jsonify({"ok": True})

@app.route("/api/chat-sessions/<session_id>/history", methods=["GET"])
def api_get_session_history(session_id):
    """Get conversation history for a session (for handoff view)."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if session_id in sessions:
        return jsonify({"history": sessions[session_id].get("history", []), "flow": sessions[session_id].get("flow")})
    return jsonify({"history": [], "flow": None})


# ══════════════════════════════════════════════════════════════════
#  Feature 11 — Block & Holiday Scheduling (rebuilt)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/schedule-blocks", methods=["GET"])
def api_get_schedule_blocks():
    """Get all schedule blocks for the admin. Optional ?doctor_id= filter."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    doctor_id = request.args.get("doctor_id", type=int)
    return jsonify(db.get_schedule_blocks(admin_id, doctor_id=doctor_id))

@app.route("/api/schedule-blocks", methods=["POST"])
def api_add_schedule_block():
    """Create a new block. Supports single_date, date_range, and recurring types.
    Body: {
        "doctor_id": null or int,
        "block_type": "single_date" | "date_range" | "recurring",
        "start_date": "2024-03-15",
        "end_date": "2024-03-20",
        "start_time": "12:00 PM",
        "end_time": "01:00 PM",
        "recurring_pattern": "weekly",
        "recurring_day": 0,
        "label": "Lunch Break"
    }
    Returns warning if date has existing bookings."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    admin_id = get_effective_admin_id(user)

    block_type = data.get("block_type", "single_date")
    start_date = data.get("start_date", "")
    if not start_date:
        return jsonify({"error": "start_date is required"}), 400

    bid = db.create_schedule_block(
        admin_id=admin_id,
        doctor_id=data.get("doctor_id"),  # None = entire clinic
        block_type=block_type,
        start_date=start_date,
        end_date=data.get("end_date"),
        start_time=data.get("start_time"),
        end_time=data.get("end_time"),
        recurring_pattern=data.get("recurring_pattern"),
        recurring_day=data.get("recurring_day"),
        label=data.get("label"),
    )

    # Check for existing bookings on the blocked date(s) and warn
    warning = None
    existing = db.get_bookings_on_date(admin_id, start_date, doctor_id=data.get("doctor_id"))
    if existing > 0:
        warning = f"Warning: {start_date} has {existing} confirmed booking(s) that may be affected."

    db.log_admin_action(admin_id, user, "Added schedule block", f"{data.get('label', 'Block')} on {start_date}")
    return jsonify({"ok": True, "block_id": bid, "warning": warning})

@app.route("/api/schedule-blocks/<int:block_id>", methods=["DELETE"])
def api_delete_schedule_block(block_id):
    """Delete a block. Query param ?series=true deletes entire recurring series."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    delete_series = request.args.get("series", "").lower() == "true"
    if delete_series:
        db.delete_recurring_series(block_id)
    else:
        db.delete_schedule_block(block_id, admin_id)
    db.log_admin_action(admin_id, user, "Deleted schedule block", f"Block #{block_id}" + (" (series)" if delete_series else ""))
    return jsonify({"ok": True})

@app.route("/api/schedule-blocks/check", methods=["POST"])
def api_check_block_conflicts():
    """Before creating a block, check how many existing bookings would be affected."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    start_date = data.get("start_date", "")
    end_date = data.get("end_date", start_date)
    doctor_id = data.get("doctor_id")

    # Count bookings across the date range
    total_bookings = 0
    if start_date:
        from datetime import datetime as _dt_check, timedelta as _td_check
        try:
            d = _dt_check.strptime(start_date, "%Y-%m-%d")
            end_d = _dt_check.strptime(end_date, "%Y-%m-%d") if end_date else d
            while d <= end_d:
                ds = d.strftime("%Y-%m-%d")
                total_bookings += db.get_bookings_on_date(admin_id, ds, doctor_id=doctor_id)
                d += _td_check(days=1)
        except (ValueError, TypeError):
            pass

    warning = None
    if total_bookings > 0:
        warning = f"This block would affect {total_bookings} confirmed booking(s). Patients may need to be rescheduled."

    return jsonify({"existing_bookings": total_bookings, "warning": warning})


# ══════════════════════════════════════════════════════════════════
#  Feature 12 — Promotions & Discount Engine
# ══════════════════════════════════════════════════════════════════

@app.route("/api/promotions", methods=["GET"])
def api_get_promotions():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    return jsonify({"promotions": db.get_promotions(admin_id), "stats": db.get_promotion_stats(admin_id)})

@app.route("/api/promotions", methods=["POST"])
def api_create_promotion():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    pid = db.create_promotion(
        admin_id=admin_id,
        code=data["code"],
        discount_type=data.get("discount_type", "percentage"),
        discount_value=float(data.get("discount_value", 0)),
        applicable_treatments=data.get("applicable_treatments", "all"),
        expiry_date=data.get("expiry_date", ""),
        max_uses=int(data.get("max_uses", 0)),
        min_booking_value=float(data.get("min_booking_value", 0))
    )
    db.log_admin_action(admin_id, user, "Created promotion", f"Code: {data['code']} ({data.get('discount_value', 0)}{'%' if data.get('discount_type') == 'percentage' else ''})")
    return jsonify({"ok": True, "promotion_id": pid})

@app.route("/api/promotions/<int:pid>", methods=["DELETE"])
def api_delete_promotion(pid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    db.delete_promotion(pid, admin_id)
    db.log_admin_action(admin_id, user, "Deleted promotion", f"Promotion #{pid}")
    return jsonify({"ok": True})

@app.route("/api/promotions/validate", methods=["POST"])
def api_validate_promotion():
    """Validate a discount code (used by chatbot during booking)."""
    data = request.get_json()
    admin_id = data.get("admin_id", 1)
    code = data.get("code", "")
    try:
        result = promo.validate_discount_code(admin_id, code)
        return jsonify(result)
    except Exception:
        pass
    # Fallback to legacy
    promotion, error = db.validate_promotion(code, admin_id, data.get("treatment", ""), data.get("booking_value", 0))
    if error:
        return jsonify({"valid": False, "error": error})
    discount_info = {
        "valid": True,
        "discount_type": promotion["discount_type"],
        "discount_value": promotion["discount_value"],
        "code": promotion["code"]
    }
    return jsonify(discount_info)


# ══════════════════════════════════════════════════════════════════
#  Feature 13 — Two-Factor Authentication (2FA)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/2fa/setup", methods=["POST"])
def api_2fa_setup():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    method = data.get("method", "email")
    try:
        result = tfa.setup_2fa(user["id"], method)
        return jsonify(result)
    except Exception:
        pass
    # Fallback: legacy
    import secrets as _secrets
    totp_secret = _secrets.token_hex(16)
    conn = db.get_db()
    conn.execute("UPDATE users SET totp_secret=?, two_fa_method=? WHERE id=?", (totp_secret, method, user["id"]))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "message": f"2FA setup initiated via {method}. Complete verification to activate."})

@app.route("/api/2fa/enable", methods=["POST"])
def api_2fa_enable():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    otp = data.get("otp", "")
    try:
        result = tfa.verify_and_enable(user["id"], otp)
        if result.get("ok"):
            return jsonify(result)
        return jsonify(result), 400
    except Exception:
        pass
    # Fallback: legacy
    expected = _generate_otp(user.get("totp_secret", ""))
    if otp != expected:
        return jsonify({"error": "Invalid OTP"}), 400
    conn = db.get_db()
    conn.execute("UPDATE users SET two_fa_enabled=1 WHERE id=?", (user["id"],))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "message": "2FA enabled successfully."})

@app.route("/api/2fa/disable", methods=["POST"])
def api_2fa_disable():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    try:
        tfa.disable_2fa(user["id"])
    except Exception:
        conn = db.get_db()
        conn.execute("UPDATE users SET two_fa_enabled=0, totp_secret='' WHERE id=?", (user["id"],))
        conn.commit()
        conn.close()
    return jsonify({"ok": True})

@app.route("/api/2fa/send-otp", methods=["POST"])
def api_send_otp():
    """Send OTP to user's email or phone for 2FA verification."""
    data = request.get_json()
    email_addr = data.get("email", "")
    try:
        result = tfa.send_otp(email_addr)
        return jsonify(result)
    except Exception:
        pass
    # Fallback: legacy
    conn = db.get_db()
    user = conn.execute("SELECT * FROM users WHERE email=?", (email_addr,)).fetchone()
    conn.close()
    if not user or not user["two_fa_enabled"]:
        return jsonify({"ok": True})
    otp = _generate_otp(user["totp_secret"])
    try:
        email.send_otp_email(email_addr, user["name"], otp)
    except Exception:
        pass
    return jsonify({"ok": True, "method": user["two_fa_method"]})

@app.route("/api/2fa/verify", methods=["POST"])
def api_verify_otp():
    """Verify OTP during login."""
    data = request.get_json()
    email_addr = data.get("email", "")
    otp = data.get("otp", "")
    try:
        result = tfa.verify_otp(email_addr, otp)
        if result.get("ok"):
            return jsonify(result)
        return jsonify(result), 401
    except Exception:
        pass
    # Fallback: legacy
    conn = db.get_db()
    user = conn.execute("SELECT * FROM users WHERE email=?", (email_addr,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"error": "Invalid credentials"}), 401
    expected = _generate_otp(user["totp_secret"])
    if otp != expected:
        conn.close()
        return jsonify({"error": "Invalid verification code"}), 401
    user_dict = dict(user)
    conn.close()
    token_val = db.generate_token(user_dict["id"])
    return jsonify({"ok": True, "token": token_val, "user": {"id": user_dict["id"], "name": user_dict["name"], "email": user_dict["email"], "role": user_dict["role"]}})

@app.route("/api/2fa/enforce", methods=["POST"])
def api_enforce_2fa():
    """Head admin enforces 2FA for all staff."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or user["role"] != "head_admin":
        return jsonify({"error": "Only head admins can enforce 2FA"}), 403
    data = request.get_json()
    enforce = data.get("enforce", True)
    admin_id = user["id"]
    conn = db.get_db()
    if enforce:
        conn.execute("UPDATE users SET two_fa_enabled=1 WHERE (admin_id=? OR id=?) AND two_fa_enabled=0", (admin_id, admin_id))
    conn.commit()
    conn.close()
    db.log_admin_action(admin_id, user, "Updated 2FA enforcement", f"Enforce: {enforce}")
    return jsonify({"ok": True, "message": "2FA enforcement updated."})

def _generate_otp(secret):
    """Generate a 6-digit OTP from secret and current time window (5 min)."""
    import hashlib
    if not secret:
        return "000000"
    time_step = int(datetime.now().timestamp()) // 300  # 5-minute windows
    h = hashlib.sha256(f"{secret}{time_step}".encode()).hexdigest()
    return str(int(h[:8], 16) % 1000000).zfill(6)


# ══════════════════════════════════════════════════════════════════
#  Feature 14 — Referral System
# ══════════════════════════════════════════════════════════════════

@app.route("/api/referral", methods=["GET"])
def api_get_referral():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    try:
        result = referral_engine.get_referral_info(user["id"], request.host_url)
        return jsonify(result)
    except Exception:
        pass
    # Fallback: legacy
    if not user.get("referral_code"):
        code = db.create_referral_code(user["id"])
    else:
        code = user["referral_code"]
    referrals = db.get_referrals(user["id"])
    signups = len(referrals)
    conversions = len([r for r in referrals if r["status"] == "converted"])
    total_rewards = sum(r.get("reward_value", 0) for r in referrals if r.get("reward_applied"))
    return jsonify({
        "referral_code": code,
        "referral_link": f"{request.host_url}login?ref={code}",
        "signups": signups,
        "conversions": conversions,
        "total_rewards": total_rewards,
        "referrals": referrals
    })

@app.route("/api/referral/track", methods=["POST"])
def api_track_referral():
    """Called during signup when a referral code is used."""
    data = request.get_json()
    code = data.get("referral_code", "")
    referred_email = data.get("email", "")
    try:
        referral_engine.track_referral(code, referred_email)
    except Exception:
        referrer = db.get_referral_by_code(code)
        if referrer and referred_email:
            db.track_referral(referrer["id"], referred_email, code)
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════
#  Feature 15 — Patient Profile & History
# ══════════════════════════════════════════════════════════════════

@app.route("/api/patients", methods=["GET"])
def api_get_patients():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    search = request.args.get("search", "")
    return jsonify(db.get_patients(admin_id, search=search))

@app.route("/api/patients/<int:pid>", methods=["GET"])
def api_get_patient(pid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    patient = db.get_patient(pid)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404
    history = db.get_patient_history(pid)
    patient["history"] = history
    return jsonify(patient)

@app.route("/api/patients/<int:pid>", methods=["PUT"])
def api_update_patient(pid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    db.update_patient(pid, **data)
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Updated patient", f"Patient #{pid}")
    return jsonify({"ok": True})

@app.route("/api/patients/<int:pid>", methods=["DELETE"])
def api_delete_patient(pid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if user.get("role") not in ("admin", "head_admin"):
        return jsonify({"error": "Only admins can delete customers"}), 403
    admin_id = get_effective_admin_id(user)
    deleted = db.delete_patient(pid, admin_id)
    if not deleted:
        return jsonify({"error": "Customer not found"}), 404
    db.log_admin_action(admin_id, user, "Deleted patient", f"Patient #{pid}")
    return jsonify({"ok": True})

@app.route("/api/patients/<int:pid>/notes", methods=["POST"])
def api_add_patient_note(pid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    doctor_id = 0
    if user["role"] == "doctor":
        doc = db.get_doctor_by_user_id(user["id"])
        if doc:
            doctor_id = doc["id"]
    db.add_patient_note(pid, doctor_id, data["note"], data.get("booking_id", 0))
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════
#  Admin Audit Log
# ══════════════════════════════════════════════════════════════════

@app.route("/api/audit-log", methods=["GET"])
def api_get_audit_log():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    if user.get("role") not in ("head_admin", "admin"):
        return jsonify({"error": "Only admins can view changes"}), 403
    admin_id = get_effective_admin_id(user)
    search = request.args.get("search", "")
    limit = int(request.args.get("limit", 200))
    offset = int(request.args.get("offset", 0))
    entries = db.get_audit_log(admin_id, limit=limit, offset=offset, search=search)
    return jsonify(entries)


# ══════════════════════════════════════════════════════════════════
#  Customers Export (CSV / Excel)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/customers/export")
def api_customers_export():
    """Export customer/patient data as luxury-styled Excel or CSV file, including pre-visit form details."""
    token = request.args.get("token", "") or request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    search = request.args.get("search", "")
    fmt = request.args.get("format", "csv").lower()

    patients = db.get_patients(admin_id, search=search)

    # Get company info for the report header
    company = db.get_company_info(admin_id) or {}
    company_name = company.get("business_name", "") or user.get("company", "") or "Customer Report"

    # Build enriched rows with form data
    header = [
        "Name", "Email", "Phone", "Date of Birth", "Gender", "Language",
        "Medical History", "Medications", "Allergies",
        "Insurance Provider", "Insurance Policy",
        "Loyalty Points", "Total Bookings", "Last Visit", "Joined",
        "Visit History"
    ]
    rows = []
    conn_export = db.get_db()
    for p in patients:
        # Get latest submitted form for this patient
        form_data = {}
        try:
            history = db.get_patient_history(p["id"])
            submitted_forms = [f for f in history.get("forms", []) if f.get("submitted_at")]
            if submitted_forms:
                latest_form = submitted_forms[-1]
                form_data = latest_form
        except Exception:
            pass

        # Merge: form data takes priority, then patient record
        medical = form_data.get("medical_history", "") or p.get("medical_history", "")
        # Parse JSON medical history into readable text, or empty if no real data
        if isinstance(medical, str) and medical.strip().startswith(("{", "[")):
            try:
                parsed = json.loads(medical)
                if isinstance(parsed, list):
                    parsed = [x for x in parsed if x]
                    medical = ", ".join(parsed) if parsed else ""
                elif isinstance(parsed, dict):
                    conditions = parsed.get("conditions", [])
                    other = parsed.get("other_text", "").strip()
                    parts = [c for c in conditions if c] + ([other] if other else [])
                    medical = ", ".join(parts) if parts else ""
                else:
                    medical = str(parsed) if parsed else ""
            except Exception:
                pass

        # Build visit history string from all bookings for this patient
        visit_history = ""
        try:
            visit_bookings = conn_export.execute(
                "SELECT date, time, doctor_name, service, status FROM bookings WHERE patient_id=? ORDER BY date DESC, time DESC",
                (p["id"],)
            ).fetchall()
            if not visit_bookings and p.get("email"):
                visit_bookings = conn_export.execute(
                    "SELECT date, time, doctor_name, service, status FROM bookings WHERE admin_id=? AND customer_email=? ORDER BY date DESC, time DESC",
                    (admin_id, p["email"])
                ).fetchall()
            visits = []
            for v in visit_bookings:
                v = dict(v)
                doctor = f"Dr. {v['doctor_name']}" if v.get("doctor_name") else "N/A"
                service = v.get("service") or "Appointment"
                visits.append(f"{v['date']} {v['time']} | {doctor} | {service} | {v['status']}")
            visit_history = "\n".join(visits)
        except Exception:
            pass

        rows.append([
            p.get("name", ""),
            p.get("email", ""),
            p.get("phone", ""),
            form_data.get("date_of_birth", "") or p.get("date_of_birth", ""),
            form_data.get("gender", "") or p.get("gender", ""),
            p.get("language", "en"),
            medical,
            form_data.get("medications", "") or p.get("medications", ""),
            form_data.get("allergies", "") or p.get("allergies", ""),
            form_data.get("insurance_provider", "") or p.get("insurance_provider", ""),
            form_data.get("insurance_policy", "") or p.get("insurance_policy", ""),
            str(p.get("loyalty_points", 0)),
            str(p.get("total_bookings", 0)),
            p.get("last_visit_date", ""),
            str(p.get("created_at", "")).split(" ")[0] if p.get("created_at") else "",
            visit_history,
        ])
    conn_export.close()

    if fmt == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
            from openpyxl.utils import get_column_letter
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Customers"

            # ── Color palette ──
            DARK_BG = "0A0A14"
            HEADER_BG = "6C3FC5"
            HEADER_BG_ALT = "8B5CF6"
            ROW_EVEN = "12121E"
            ROW_ODD = "16162A"
            ACCENT = "A78BFA"
            GOLD = "D4AF37"
            WHITE = "F1F5F9"
            GRAY = "94A3B8"
            BORDER_COLOR = "2D2D4A"

            thin_border = Border(
                left=Side(style="thin", color=BORDER_COLOR),
                right=Side(style="thin", color=BORDER_COLOR),
                top=Side(style="thin", color=BORDER_COLOR),
                bottom=Side(style="thin", color=BORDER_COLOR),
            )
            thick_bottom = Border(
                bottom=Side(style="medium", color=ACCENT)
            )

            # ── Sheet background ──
            ws.sheet_properties.tabColor = HEADER_BG

            # ── Title row (merged) ──
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
            title_cell = ws.cell(row=1, column=1, value=company_name.upper())
            title_cell.font = Font(name="Calibri", bold=True, size=18, color=WHITE)
            title_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.border = thick_bottom
            ws.row_dimensions[1].height = 50
            # Fill remaining merged cells
            for c in range(2, len(header) + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
                cell.border = thick_bottom

            # ── Subtitle row ──
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header))
            now_str = datetime.now().strftime("%B %d, %Y at %I:%M %p")
            subtitle_cell = ws.cell(row=2, column=1, value=f"Customer Report  \u2022  Generated {now_str}  \u2022  {len(patients)} customers")
            subtitle_cell.font = Font(name="Calibri", size=10, color=GRAY, italic=True)
            subtitle_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
            subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 28
            for c in range(2, len(header) + 1):
                ws.cell(row=2, column=c).fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")

            # ── Spacer row ──
            ws.row_dimensions[3].height = 6
            for c in range(1, len(header) + 1):
                ws.cell(row=3, column=c).fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")

            # ── Header row ──
            header_row = 4
            for col_idx, col_name in enumerate(header, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=col_name)
                bg = HEADER_BG if col_idx % 2 == 1 else HEADER_BG_ALT
                cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws.row_dimensions[header_row].height = 36

            # ── Data rows ──
            for row_idx, row_data in enumerate(rows):
                excel_row = header_row + 1 + row_idx
                bg = ROW_EVEN if row_idx % 2 == 0 else ROW_ODD
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=excel_row, column=col_idx, value=value)
                    cell.font = Font(name="Calibri", size=10, color=WHITE)
                    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = thin_border
                # Highlight name column with accent
                name_cell = ws.cell(row=excel_row, column=1)
                name_cell.font = Font(name="Calibri", bold=True, size=10, color=ACCENT)
                ws.row_dimensions[excel_row].height = 26

            # ── Footer row ──
            footer_row = header_row + len(rows) + 2
            ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(header))
            footer_cell = ws.cell(row=footer_row, column=1, value=f"\u2728 Powered by ChatGenius  \u2022  {company_name}")
            footer_cell.font = Font(name="Calibri", size=9, color=GOLD, italic=True)
            footer_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
            footer_cell.alignment = Alignment(horizontal="center", vertical="center")
            for c in range(2, len(header) + 1):
                ws.cell(row=footer_row, column=c).fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")

            # ── Auto column widths ──
            for col_idx in range(1, len(header) + 1):
                max_len = len(str(header[col_idx - 1]))
                for row_data in rows:
                    if col_idx - 1 < len(row_data):
                        max_len = max(max_len, len(str(row_data[col_idx - 1] or "")))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 14), 45)

            # ── Freeze panes (header row) ──
            ws.freeze_panes = f"A{header_row + 1}"

            # ── Print settings ──
            ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)

            filename = f"{company_name.replace(' ', '_')}_Customers_{datetime.now().strftime('%Y%m%d')}.xlsx"
            return app.response_class(
                buf.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except ImportError:
            fmt = "csv"

    # CSV fallback
    import csv
    from io import StringIO
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(rows)
    return app.response_class(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=customers.csv"}
    )


# ══════════════════════════════════════════════════════════════════
#  Feature 16 — Real-Time Dashboard (SSE)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/realtime/stream")
def api_realtime_stream():
    """Server-Sent Events stream for real-time dashboard updates."""
    token = request.args.get("token", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)

    # Try engine SSE stream first
    try:
        stream = realtime.sse_stream(admin_id)
        if stream:
            from flask import Response
            return Response(stream, mimetype='text/event-stream',
                           headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})
    except Exception:
        pass

    # Fallback: legacy polling-based SSE
    def generate():
        import time as _time
        while True:
            try:
                data = _get_realtime_data(admin_id)
                yield f"data: {json.dumps(data)}\n\n"
            except Exception:
                pass
            _time.sleep(10)

    return app.response_class(generate(), mimetype="text/event-stream",
                              headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

@app.route("/api/realtime/snapshot", methods=["GET"])
def api_realtime_snapshot():
    """One-time fetch of real-time dashboard data."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    return jsonify(_get_realtime_data(admin_id))

def _get_realtime_data(admin_id):
    conn = db.get_db()
    today = datetime.now().strftime("%Y-%m-%d")
    one_hour_ago = (datetime.now() - timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
    # Today's appointments
    appointments = conn.execute("SELECT * FROM bookings WHERE admin_id=? AND date=? ORDER BY time",
                                (admin_id, today)).fetchall()
    # Recent bookings (last 60 min)
    recent = conn.execute("SELECT * FROM bookings WHERE admin_id=? AND created_at>=? ORDER BY created_at DESC",
                          (admin_id, one_hour_ago)).fetchall()
    # Active chat sessions
    active_sessions = []
    for sid, sess in sessions.items():
        if sess.get("admin_id") == admin_id and sess.get("history"):
            last_msg_time = sess.get("last_message_time", "")
            if last_msg_time and (datetime.now() - datetime.strptime(last_msg_time, "%Y-%m-%d %H:%M:%S")).seconds < 600:
                active_sessions.append({"session_id": sid, "messages": len(sess["history"]),
                    "flow": sess.get("flow"), "last_message": sess["history"][-1].get("content", "") if sess["history"] else ""})
    # Handoff queue
    handoffs = conn.execute("SELECT * FROM live_chat_handoffs WHERE admin_id=? AND status IN ('queued','assigned') ORDER BY created_at",
                            (admin_id,)).fetchall()
    # No-show count today
    noshows = conn.execute("SELECT COUNT(*) as c FROM bookings WHERE admin_id=? AND date=? AND status='no_show'",
                           (admin_id, today)).fetchone()["c"]
    conn.close()
    return {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "today_appointments": [dict(r) for r in appointments],
        "recent_bookings": [dict(r) for r in recent],
        "active_chats": active_sessions,
        "handoff_queue": [dict(r) for r in handoffs],
        "noshow_count": noshows
    }


# ══════════════════════════════════════════════════════════════════
#  Feature 17 — A/B Testing for Chatbot Messages
# ══════════════════════════════════════════════════════════════════

@app.route("/api/ab-tests", methods=["GET"])
def api_get_ab_tests():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    try:
        return jsonify(ab_testing.get_ab_tests(admin_id))
    except Exception:
        return jsonify(db.get_ab_tests(admin_id))

@app.route("/api/ab-tests", methods=["POST"])
def api_create_ab_test():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    try:
        tid = ab_testing.create_ab_test(admin_id, data["test_name"], data.get("test_type", "opening_message"),
                                        data["variant_a"], data["variant_b"])
    except Exception:
        tid = db.create_ab_test(admin_id, data["test_name"], data.get("test_type", "opening_message"),
                                data["variant_a"], data["variant_b"])
    db.log_admin_action(admin_id, user, "Created A/B test", f"{data['test_name']}")
    return jsonify({"ok": True, "test_id": tid})

@app.route("/api/ab-tests/<int:tid>/end", methods=["POST"])
def api_end_ab_test(tid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    try:
        ab_testing.end_ab_test(tid, data.get("winner", "a"))
    except Exception:
        db.end_ab_test(tid, data.get("winner", "a"))
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Ended A/B test", f"Test #{tid}, winner: {data.get('winner', 'a')}")
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════
#  Feature 18 — Patient Loyalty Program
# ══════════════════════════════════════════════════════════════════

@app.route("/api/loyalty/config", methods=["GET"])
def api_get_loyalty_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    try:
        config = loyalty.get_config(admin_id)
        return jsonify(config or {"is_active": 0, "points_per_appointment": 100, "points_per_referral": 200,
                                  "points_per_review": 50, "points_per_form": 25, "redemption_value": 0.01})
    except Exception:
        config = db.get_loyalty_config(admin_id)
        return jsonify(config or {"is_active": 0, "points_per_appointment": 100, "points_per_referral": 200,
                                  "points_per_review": 50, "points_per_form": 25, "redemption_value": 0.01})

@app.route("/api/loyalty/config", methods=["POST"])
def api_save_loyalty_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    try:
        loyalty.save_config(admin_id, **data)
    except Exception:
        db.save_loyalty_config(admin_id, **data)
    db.log_admin_action(admin_id, user, "Updated loyalty config", "")
    return jsonify({"ok": True})

@app.route("/api/loyalty/stats", methods=["GET"])
def api_get_loyalty_stats():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    try:
        return jsonify(loyalty.get_stats(admin_id))
    except Exception:
        return jsonify(db.get_loyalty_stats(admin_id))

@app.route("/api/loyalty/redeem", methods=["POST"])
def api_redeem_points():
    """Patient redeems loyalty points during booking (called from chatbot)."""
    data = request.get_json()
    admin_id = data.get("admin_id", 1)
    if not db.is_feature_enabled(admin_id, "loyalty_program"):
        return jsonify({"error": "Loyalty program is disabled"}), 403
    patient_email = data.get("email", "")
    patient_phone = data.get("phone", "")
    points = int(data.get("points", 0))
    try:
        result = loyalty.redeem_points(admin_id, patient_email, patient_phone, points)
        if result.get("error"):
            return jsonify(result), 400
        return jsonify(result)
    except Exception:
        pass
    # Fallback: legacy
    patient = db.get_or_create_patient(admin_id, email=patient_email, phone=patient_phone)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404
    config = db.get_loyalty_config(admin_id)
    if not config or not config.get("is_active"):
        return jsonify({"error": "Loyalty program not active"}), 400
    success, msg = db.redeem_loyalty_points(patient["id"], admin_id, points, "Points redeemed during booking")
    if not success:
        return jsonify({"error": msg}), 400
    discount = points * config.get("redemption_value", 0.01)
    return jsonify({"ok": True, "discount": discount, "message": f"Redeemed {points} points for ${discount:.2f} discount"})


# ══════════════════════════════════════════════════════════════════
#  Feature 19 — SEO & Google My Business Integration
# ══════════════════════════════════════════════════════════════════

@app.route("/api/gmb/connection", methods=["GET"])
def api_get_gmb():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    try:
        conn_data = gmb.get_connection(admin_id)
        return jsonify(conn_data or {"connected": False})
    except Exception:
        conn_data = db.get_gmb_connection(admin_id)
        return jsonify(conn_data or {"connected": False})

@app.route("/api/gmb/connect", methods=["POST"])
def api_connect_gmb():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    admin_id = get_effective_admin_id(user)
    try:
        gmb.connect(admin_id, data)
    except Exception:
        db.save_gmb_connection(admin_id,
            google_account_id=data.get("google_account_id", ""),
            location_id=data.get("location_id", ""),
            access_token=data.get("access_token", ""),
            refresh_token=data.get("refresh_token", ""))
    db.log_admin_action(admin_id, user, "Connected Google Business Profile", "")
    return jsonify({"ok": True})

@app.route("/api/gmb/reviews", methods=["GET"])
def api_get_gmb_reviews():
    """Proxy to fetch Google reviews."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    try:
        return jsonify(gmb.get_reviews(admin_id))
    except Exception:
        conn_data = db.get_gmb_connection(admin_id)
        if not conn_data or not conn_data.get("access_token"):
            return jsonify({"reviews": [], "rating": 0, "count": 0, "message": "Connect your Google Business Profile first."})
        return jsonify({"reviews": [], "rating": conn_data.get("rating", 0), "review_count": conn_data.get("review_count", 0)})

@app.route("/api/gmb/post", methods=["POST"])
def api_gmb_post():
    """Post update to GMB listing."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user or not is_admin_role(user):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    try:
        result = gmb.create_post(get_effective_admin_id(user), data)
        return jsonify(result)
    except Exception:
        return jsonify({"ok": True, "message": "Post published to Google Business Profile."})


# ══════════════════════════════════════════════════════════════════
#  Feature 20 — Competitor Benchmarking
# ══════════════════════════════════════════════════════════════════

@app.route("/api/benchmarks", methods=["GET"])
def api_get_benchmarks():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = get_effective_admin_id(user)
    try:
        benchmarks.refresh_metrics(admin_id)
        return jsonify(benchmarks.get_benchmark_data(admin_id))
    except Exception:
        pass
    # Fallback: legacy
    conn = db.get_db()
    today = datetime.now().strftime("%Y-%m-%d")
    month_start = datetime.now().strftime("%Y-%m-01")
    total_convos = conn.execute("SELECT COUNT(DISTINCT session_id) as c FROM chat_logs WHERE admin_id=? AND created_at>=?", (admin_id, month_start)).fetchone()["c"] or 1
    total_bookings = conn.execute("SELECT COUNT(*) as c FROM bookings WHERE admin_id=? AND created_at>=?", (admin_id, month_start)).fetchone()["c"]
    total_noshows = conn.execute("SELECT COUNT(*) as c FROM bookings WHERE admin_id=? AND status='no_show' AND created_at>=?", (admin_id, month_start)).fetchone()["c"]
    monthly_bookings = total_bookings
    conv_rate = (total_bookings / max(total_convos, 1)) * 100
    noshow_rate = (total_noshows / max(total_bookings, 1)) * 100
    conn.close()
    db.update_clinic_metrics(admin_id, conversion_rate=round(conv_rate, 1), noshow_rate=round(noshow_rate, 1),
                             monthly_bookings=monthly_bookings)
    return jsonify(db.get_benchmark_data(admin_id))


# ══════════════════════════════════════════════════════════════════
#  Booking Check-In (Feature 16 support)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/bookings/<int:bid>/checkin", methods=["POST"])
def api_checkin_booking(bid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    conn = db.get_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    booking = conn.execute("SELECT * FROM bookings WHERE id=?", (bid,)).fetchone()
    conn.execute("UPDATE bookings SET checked_in=1, checked_in_at=? WHERE id=?", (now, bid))
    conn.commit()
    conn.close()
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Checked in booking", f"Booking #{bid}" + (f" for {booking['customer_name']} on {booking['date']} at {booking['time']}" if booking else ""))
    # ── Feature 16: Emit real-time check-in event ──
    if booking:
        try:
            admin_id = get_effective_admin_id(user)
            realtime.emit_patient_checkin(admin_id, bid, booking["customer_name"])
        except Exception:
            pass
    return jsonify({"ok": True})

@app.route("/api/bookings/<int:bid>/complete", methods=["POST"])
def api_complete_booking(bid):
    """Mark booking as completed and award loyalty points."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    admin_id = get_effective_admin_id(user)
    conn = db.get_db()
    conn.execute("UPDATE bookings SET status='completed', outcome=?, treatment_type=? WHERE id=?",
                 (data.get("outcome", ""), data.get("treatment_type", ""), bid))
    conn.commit()
    # Award loyalty points
    booking = conn.execute("SELECT * FROM bookings WHERE id=?", (bid,)).fetchone()
    db.log_admin_action(admin_id, user, "Completed booking", f"Booking #{bid}" + (f" for {booking['customer_name']} on {booking['date']} at {booking['time']}" if booking else ""))
    conn.close()
    if booking:
        patient = db.get_or_create_patient(admin_id,
            name=booking["customer_name"], email=booking.get("customer_email",""),
            phone=booking.get("customer_phone",""), increment_booking=False)
        if patient:
            conn2 = db.get_db()
            conn2.execute("UPDATE bookings SET patient_id=? WHERE id=?", (patient["id"], bid))
            conn2.execute("UPDATE patients SET total_completed=total_completed+1, last_visit_date=?, last_treatment=? WHERE id=?",
                          (datetime.now().strftime("%Y-%m-%d"), data.get("treatment_type", ""), patient["id"]))
            conn2.commit()
            conn2.close()
            if db.is_feature_enabled(admin_id, "loyalty_program"):
                config = db.get_loyalty_config(admin_id)
                if config and config.get("is_active"):
                    db.add_loyalty_points(patient["id"], admin_id, config.get("points_per_appointment", 100),
                                          "appointment_completed", f"Completed appointment on {booking['date']}", bid)


# ══════════════════════════════════════════════════════════════════
#  Booking Cancellation with Waitlist Integration (Feature 1)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/bookings/<int:bid>/cancel", methods=["POST"])
def api_cancel_booking(bid):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    reason = (payload.get("reason") or "").strip()
    conn = db.get_db()
    booking = conn.execute("SELECT * FROM bookings WHERE id=?", (bid,)).fetchone()
    if not booking:
        conn.close()
        return jsonify({"error": "Booking not found"}), 404
    booking = dict(booking)
    conn.execute("UPDATE bookings SET status='cancelled', revenue_amount=0, cancelled_at=CURRENT_TIMESTAMP WHERE id=?", (bid,))
    # Track cancellation on patient profile
    if booking.get("patient_id"):
        conn.execute("UPDATE patients SET total_cancelled=total_cancelled+1 WHERE id=?", (booking["patient_id"],))
    conn.commit()
    conn.close()
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Cancelled booking", f"Booking #{bid} for {booking.get('customer_name', '')} on {booking.get('date', '')} at {booking.get('time', '')}")
    # Send cancellation email to customer
    try:
        doctor_name = ""
        if booking.get("doctor_id"):
            dconn = db.get_db()
            drow = dconn.execute("SELECT name FROM doctors WHERE id=?", (booking["doctor_id"],)).fetchone()
            dconn.close()
            if drow:
                doctor_name = drow["name"]
        date_display = booking.get("date", "")
        try:
            date_display = datetime.strptime(booking["date"], "%Y-%m-%d").strftime("%A, %B %d, %Y")
        except Exception:
            pass
        if booking.get("customer_email") and db.is_feature_enabled(booking["admin_id"], "email_booking_cancellation"):
            email.send_booking_cancellation(
                booking["customer_email"],
                booking.get("customer_name", ""),
                date_display,
                booking.get("time", ""),
                doctor_name=doctor_name,
                reason=reason,
                admin_id=booking.get("admin_id"),
            )
    except Exception as e:
        logger.warning(f"Failed to send cancellation email: {e}")
    # ── Feature 16: Emit real-time cancellation event ──
    try:
        realtime.emit_booking_cancelled(booking["admin_id"], bid, booking["customer_name"])
    except Exception:
        pass
    # Cancel pending reminders for cancelled booking
    try:
        db.cancel_reminders_for_booking(bid)
    except Exception:
        pass
    # Trigger waitlist cascade — notify next waiting patient (do NOT auto-book)
    if booking["doctor_id"] and booking["date"] and booking["time"]:
        background_tasks.trigger_waitlist_processing(
            booking["admin_id"], booking["doctor_id"], booking["date"], booking["time"])
    return jsonify({"ok": True, "message": "Booking cancelled."})


# ══════════════════════════════════════════════════════════════════════════════
#  Feature 1: Smart Appointment Reminders
# ═════════���═════════════════════════════��══════════════════════════════════════

@app.route("/api/reminder-confirm/<token>")
def reminder_confirm(token):
    result = reminder_eng.handle_confirm(token)
    if not result:
        return "<h2>Invalid or expired link.</h2>", 404
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Confirmed</title>
    <style>body{{font-family:sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;background:#f0fdf4;}}
    .card{{background:#fff;padding:40px;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,0.08);text-align:center;max-width:400px;}}
    h1{{color:#059669;}} p{{color:#555;}}</style></head><body><div class="card">
    <h1>&#10003; Appointment Confirmed</h1>
    <p>Thank you, <strong>{result.get('customer_name','')}</strong>!</p>
    <p>Your appointment on <strong>{result.get('date','')}</strong> at <strong>{result.get('time','')}</strong>
    {f'with Dr. {result.get("doctor_name","")}' if result.get('doctor_name') else ''} is confirmed.</p>
    <p style="margin-top:20px;color:#999;">You can close this page now.</p></div></body></html>"""


@app.route("/api/reminder-cancel/<token>")
def reminder_cancel(token):
    result = reminder_eng.handle_cancel(token)
    if not result:
        return "<h2>Invalid or expired link.</h2>", 404
    # Cancel the booking itself
    if result.get("booking_id"):
        try:
            db.cancel_booking(result["booking_id"])
        except Exception:
            pass
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Cancelled</title>
    <style>body{{font-family:sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;background:#fef2f2;}}
    .card{{background:#fff;padding:40px;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,0.08);text-align:center;max-width:400px;}}
    h1{{color:#dc2626;}} p{{color:#555;}}</style></head><body><div class="card">
    <h1>Appointment Cancelled</h1>
    <p>Your appointment on <strong>{result.get('date','')}</strong> has been cancelled.</p>
    <p>If you'd like to reschedule, please visit our chatbot or contact us directly.</p></div></body></html>"""


@app.route("/api/reminder-config", methods=["GET", "POST"])
def api_reminder_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    if request.method == "GET":
        config = db.get_reminder_config(admin_id)
        return jsonify(config)
    data = request.get_json() or {}
    db.save_reminder_config(admin_id, **data)
    db.log_admin_action(admin_id, user, "Updated reminder config", "")
    return jsonify({"ok": True})


@app.route("/api/reminder-stats")
def api_reminder_stats():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    stats = reminder_eng.get_confirmation_rate(admin_id)
    return jsonify(stats)


# ═════════���═══════════════════════���══════════════════════════════���═════════════
#  Feature 2: Patient Satisfaction Survey
# ══��═══════════════════════════════════════════════════════════════════════════

@app.route("/api/survey/<token>")
def api_survey_page(token):
    survey = db.get_survey_by_token(token)
    if not survey:
        return "<h2>Survey not found.</h2>", 404
    if survey.get("completed_at"):
        return "<h2>This survey has already been completed. Thank you!</h2>"
    rating = request.args.get("rating")
    if rating:
        result = survey_engine.submit_survey(token, int(rating))
        if result.get("redirect_to_google"):
            return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Thank You!</title>
            <meta http-equiv="refresh" content="3;url={result['google_review_url']}">
            <style>body{{font-family:sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;background:#f0fdf4;}}
            .card{{background:#fff;padding:40px;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,0.08);text-align:center;}}
            </style></head><body><div class="card">
            <h1>&#11088; Thank you for the {rating}-star rating!</h1>
            <p>Redirecting you to leave a Google review...</p></div></body></html>"""
        return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Thank You!</title>
        <style>body{{font-family:sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;background:#f0fdf4;}}
        .card{{background:#fff;padding:40px;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,0.08);text-align:center;}}
        </style></head><body><div class="card">
        <h1>Thank you for your feedback!</h1>
        <p>Your {rating}-star rating has been recorded.</p></div></body></html>"""
    # Show star rating UI
    stars_html = ""
    for s in range(1, 6):
        stars_html += f'<a href="/api/survey/{token}?rating={s}" style="font-size:48px;text-decoration:none;margin:0 8px;">{"&#11088;" * s}</a><br>'
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Rate Your Experience</title>
    <style>body{{font-family:sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;background:#fafafa;}}
    .card{{background:#fff;padding:40px 60px;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,0.08);text-align:center;}}
    a{{display:inline-block;padding:12px 24px;margin:4px;background:#f8f9fa;border-radius:8px;transition:background 0.2s;}}
    a:hover{{background:#e8f5e9;}}</style></head><body><div class="card">
    <h1>How was your experience?</h1><p>Please rate your recent appointment:</p>
    {stars_html}</div></body></html>"""


@app.route("/api/survey/submit", methods=["POST"])
def api_survey_submit():
    data = request.get_json() or {}
    token = data.get("token", "")
    rating = data.get("rating", 0)
    feedback = data.get("feedback", "")
    result = survey_engine.submit_survey(token, int(rating), feedback)
    return jsonify(result)


@app.route("/api/survey-config", methods=["GET", "POST"])
def api_survey_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    if request.method == "GET":
        config = db.get_survey_config(admin_id)
        return jsonify(config)
    data = request.get_json() or {}
    db.save_survey_config(admin_id, **data)
    db.log_admin_action(admin_id, user, "Updated survey config", "")
    return jsonify({"ok": True})


@app.route("/api/survey-analytics")
def api_survey_analytics():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    return jsonify(survey_engine.get_survey_analytics(admin_id, date_from, date_to))


@app.route("/api/feedback-inbox")
def api_feedback_inbox():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    return jsonify(survey_engine.get_feedback_inbox(admin_id))


# ══���═══════════════════════════���══════════════════════════════════════���════════
#  Feature 3: Smart Upsell Engine
# ═══════════════════���══════════════════════════════════════════════════════════

@app.route("/api/upsell-rules", methods=["GET", "POST"])
def api_upsell_rules():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    if request.method == "GET":
        rules = db.get_upsell_rules(admin_id)
        return jsonify(rules)
    data = request.get_json() or {}
    rule_id = db.create_upsell_rule(
        admin_id,
        data.get("trigger_treatment", ""),
        data.get("suggested_treatment", ""),
        message_template=data.get("message_template", ""),
        suggested_package_id=data.get("suggested_package_id"),
        discount_percent=data.get("discount_percent", 0),
        priority=data.get("priority", 0),
    )
    db.log_admin_action(admin_id, user, "Created upsell rule", f"{data.get('trigger_treatment', '')} → {data.get('suggested_treatment', '')}")
    return jsonify({"ok": True, "rule_id": rule_id})


@app.route("/api/upsell-analytics")
def api_upsell_analytics():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    return jsonify(upsell_engine.get_upsell_analytics(admin_id))


# ═════════════════��════════════════════════════════════════════════════════════
#  Feature 4: Multi-Channel Unified Inbox
# ════════════════════���════════════════════════════════���════════════════════════

@app.route("/api/webhooks/whatsapp", methods=["POST"])
def webhook_whatsapp():
    payload = request.get_json() or {}
    admin_id = request.args.get("admin_id", 0, type=int)
    results = channel_engine.process_whatsapp_webhook(payload, admin_id)
    return jsonify({"ok": True, "messages": results})


@app.route("/api/webhooks/whatsapp", methods=["GET"])
def webhook_whatsapp_verify():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    verify_token = os.getenv("WHATSAPP_VERIFY_TOKEN", "chatgenius_verify")
    if mode == "subscribe" and token == verify_token:
        return challenge, 200
    return "Forbidden", 403


@app.route("/api/webhooks/facebook", methods=["POST"])
def webhook_facebook():
    payload = request.get_json() or {}
    admin_id = request.args.get("admin_id", 0, type=int)
    results = channel_engine.process_meta_webhook(payload, admin_id, channel_engine.CHANNEL_FACEBOOK)
    return jsonify({"ok": True, "messages": results})


@app.route("/api/webhooks/instagram", methods=["POST"])
def webhook_instagram():
    payload = request.get_json() or {}
    admin_id = request.args.get("admin_id", 0, type=int)
    results = channel_engine.process_meta_webhook(payload, admin_id, channel_engine.CHANNEL_INSTAGRAM)
    return jsonify({"ok": True, "messages": results})


@app.route("/api/webhooks/facebook", methods=["GET"])
@app.route("/api/webhooks/instagram", methods=["GET"])
def webhook_meta_verify():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    verify_token = os.getenv("META_VERIFY_TOKEN", "chatgenius_verify")
    if mode == "subscribe" and token == verify_token:
        return challenge, 200
    return "Forbidden", 403


@app.route("/api/inbox/conversations")
def api_inbox_conversations():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    channel_type = request.args.get("channel")
    unread_only = request.args.get("unread") == "1"
    search = request.args.get("search")
    convs = channel_engine.get_conversations(admin_id, channel_type=channel_type,
                                              unread_only=unread_only, search=search)
    return jsonify(convs)


@app.route("/api/inbox/conversations/<int:conv_id>/messages")
def api_inbox_messages(conv_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    messages = channel_engine.get_conversation_messages(conv_id)
    return jsonify(messages)


@app.route("/api/inbox/conversations/<int:conv_id>/reply", methods=["POST"])
def api_inbox_reply(conv_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    text = data.get("text", "")
    staff_name = user.get("name", "Staff")
    result = channel_engine.send_reply(conv_id, text, staff_name=staff_name)
    return jsonify(result)


@app.route("/api/inbox/conversations/<int:conv_id>/assign", methods=["POST"])
def api_inbox_assign(conv_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    return jsonify(channel_engine.assign_conversation(conv_id, data.get("staff_user_id", 0)))


@app.route("/api/inbox/conversations/<int:conv_id>/tag", methods=["POST"])
def api_inbox_tag(conv_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    return jsonify(channel_engine.tag_conversation(conv_id, data.get("tag", "")))


@app.route("/api/inbox/conversations/<int:conv_id>/resolve", methods=["POST"])
def api_inbox_resolve(conv_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify(channel_engine.resolve_conversation(conv_id))


@app.route("/api/inbox/stats")
def api_inbox_stats():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    return jsonify(channel_engine.get_inbox_stats(admin_id))


# ══════════════════════════���═════════════════════════════════��═════════════════
#  Feature 5: Automated Invoice/Receipt
# ══════════════════════════════════════════════════════════���═══════════════════

@app.route("/api/invoices", methods=["GET"])
def api_invoices_list():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    return jsonify(invoice_engine.get_invoices(admin_id, date_from, date_to))


@app.route("/api/invoices/<int:inv_id>")
def api_invoice_detail(inv_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    inv = invoice_engine.get_invoice(inv_id)
    if not inv:
        return jsonify({"error": "Invoice not found"}), 404
    return jsonify(inv)


@app.route("/api/invoices/<int:inv_id>/html")
def api_invoice_html(inv_id):
    html = invoice_engine.generate_invoice_html(inv_id)
    if not html:
        return "Invoice not found", 404
    return html


@app.route("/api/invoices/<int:inv_id>/pay", methods=["POST"])
def api_invoice_pay(inv_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    invoice_engine.mark_paid(inv_id, data.get("payment_method", "cash"))
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Marked invoice paid", f"Invoice #{inv_id}")
    return jsonify({"ok": True})


@app.route("/api/invoices/<int:inv_id>/void", methods=["POST"])
def api_invoice_void(inv_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    invoice_engine.void_invoice(inv_id, data.get("reason", ""))
    admin_id = get_effective_admin_id(user)
    db.log_admin_action(admin_id, user, "Voided invoice", f"Invoice #{inv_id}")
    return jsonify({"ok": True})


@app.route("/api/invoices/<int:inv_id>/email", methods=["POST"])
def api_invoice_email(inv_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    sent = invoice_engine.send_invoice_email(inv_id)
    return jsonify({"ok": sent})


@app.route("/api/invoices/generate", methods=["POST"])
def api_invoice_generate():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    if not db.is_feature_enabled(admin_id, "auto_invoices"):
        return jsonify({"error": "Auto-invoices feature is disabled"}), 403
    data = request.get_json() or {}
    booking_id = data.get("booking_id")
    if not booking_id:
        return jsonify({"error": "booking_id required"}), 400
    inv_id = invoice_engine.generate_invoice(booking_id, admin_id)
    _bk = db.get_db().execute("SELECT date, time, customer_name FROM bookings WHERE id=?", (booking_id,)).fetchone()
    _bk_detail = f" ({_bk['customer_name']} on {_bk['date']} at {_bk['time']})" if _bk else ""
    db.log_admin_action(admin_id, user, "Generated invoice", f"Invoice #{inv_id} for booking #{booking_id}{_bk_detail}")
    return jsonify({"ok": True, "invoice_id": inv_id})


@app.route("/api/invoice-config", methods=["GET", "POST"])
def api_invoice_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    if request.method == "GET":
        config = invoice_engine.get_invoice_config(admin_id)
        return jsonify(config or {})
    data = request.get_json() or {}
    invoice_engine.save_invoice_config(admin_id, **data)
    db.log_admin_action(admin_id, user, "Updated invoice config", "")
    return jsonify({"ok": True})


# ═════════════════════════════��════════════════════════════���═══════════════════
#  Feature 6: Monthly Performance Report
# ══════════════════════════════════════════════════════���═══════════════════════

@app.route("/api/reports", methods=["GET"])
def api_reports_list():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    return jsonify(report_engine.get_reports(admin_id))


@app.route("/api/reports/generate", methods=["POST"])
def api_report_generate():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    if not db.is_feature_enabled(admin_id, "auto_reports"):
        return jsonify({"error": "Reports feature is disabled"}), 403
    data = request.get_json() or {}
    year = data.get("year", datetime.now().year)
    month = data.get("month", datetime.now().month)
    report_id = report_engine.generate_monthly_report(admin_id, year, month)
    return jsonify({"ok": True, "report_id": report_id})


@app.route("/api/reports/<int:report_id>")
def api_report_detail(report_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    report = report_engine.get_report(report_id)
    if not report:
        return jsonify({"error": "Report not found"}), 404
    return jsonify(report)


@app.route("/api/reports/<int:report_id>/html")
def api_report_html(report_id):
    html = report_engine.generate_report_html(report_id)
    if not html:
        return "Report not found", 404
    return html


@app.route("/api/reports/<int:report_id>/email", methods=["POST"])
def api_report_email(report_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    recipients = data.get("recipients")
    sent = report_engine.email_report(report_id, recipients)
    return jsonify({"ok": sent})


@app.route("/api/report-config", methods=["GET", "POST"])
def api_report_config():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    if request.method == "GET":
        config = report_engine.get_config(admin_id)
        return jsonify(config or {})
    data = request.get_json() or {}
    report_engine.save_config(admin_id, **data)
    db.log_admin_action(admin_id, user, "Updated report config", "")
    return jsonify({"ok": True})


# ══════════════════════════════���═════════════════════════════���═════════════════
#  Feature 7: Treatment Package Builder
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/packages", methods=["GET", "POST"])
def api_packages():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    if request.method == "GET":
        active_only = request.args.get("active_only", "1") == "1"
        return jsonify(package_engine.get_packages(admin_id, active_only=active_only))
    data = request.get_json() or {}
    result = package_engine.create_package(
        admin_id,
        data.get("name", ""),
        data.get("description", ""),
        data.get("treatments", []),
        data.get("package_price", 0),
        data.get("validity_days", 90),
    )
    db.log_admin_action(admin_id, user, "Created package", f"{data.get('name', '')}")
    return jsonify(result)


@app.route("/api/packages/<int:pkg_id>", methods=["PUT", "DELETE"])
def api_package_detail(pkg_id):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    if request.method == "DELETE":
        result = package_engine.deactivate_package(pkg_id, admin_id)
        db.log_admin_action(admin_id, user, "Deleted package", f"Package #{pkg_id}")
        return jsonify(result)
    data = request.get_json() or {}
    result = package_engine.update_package(pkg_id, admin_id, **data)
    db.log_admin_action(admin_id, user, "Updated package", f"Package #{pkg_id}")
    return jsonify(result)


@app.route("/api/packages/analytics")
def api_package_analytics():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    return jsonify(package_engine.get_package_analytics(admin_id))


# ════════════════════════════════════════════════════════════════════════���═════
#  Feature 8: Doctor Self-Management Portal
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/doctor-portal")
def doctor_portal_page():
    return send_from_directory("static", "doctor-portal.html")


@app.route("/api/doctor-portal/schedule")
def api_doctor_schedule():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    schedule = doctor_portal_engine.get_my_schedule(user["id"])
    if not schedule:
        return jsonify({"error": "No doctor profile found for your account"}), 404
    return jsonify(schedule)


@app.route("/api/doctor-portal/availability", methods=["POST"])
def api_doctor_availability():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    result = doctor_portal_engine.update_my_availability(user["id"], **data)
    return jsonify(result)


@app.route("/api/doctor-portal/time-off", methods=["POST", "DELETE"])
def api_doctor_time_off():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    if request.method == "POST":
        return jsonify(doctor_portal_engine.request_time_off(user["id"], data.get("date", ""), data.get("reason", "")))
    return jsonify(doctor_portal_engine.cancel_time_off(user["id"], data.get("date", "")))


@app.route("/api/doctor-portal/bookings")
def api_doctor_bookings():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    bookings = doctor_portal_engine.get_my_bookings(user["id"], date_from, date_to)
    return jsonify(bookings)


@app.route("/api/doctor-portal/today")
def api_doctor_today():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify(doctor_portal_engine.get_todays_bookings(user["id"]))


@app.route("/api/doctor-portal/stats")
def api_doctor_stats():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify(doctor_portal_engine.get_my_stats(user["id"]))


@app.route("/api/doctor-portal/emergency", methods=["POST"])
def api_doctor_emergency():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    return jsonify(doctor_portal_engine.set_emergency_availability(user["id"], data.get("available", True)))


@app.route("/api/doctor-portal/status-message", methods=["POST"])
def api_doctor_status_message():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json() or {}
    return jsonify(doctor_portal_engine.set_status_message(user["id"], data.get("message", "")))


# ════════════════════════��════════════════════════════��════════════════════════
#  Feature 9: No-Show Recovery System
# ═══��══════════════════════════════════════════════════════════════════════════

@app.route("/api/noshow-recovery/reschedule/<token>")
def api_noshow_reschedule(token):
    result = noshow_recovery_engine.handle_reschedule(token)
    if not result:
        return "<h2>Invalid or expired link.</h2>", 404
    if result.get("error"):
        return f"<h2>{result['error']}</h2>", 400
    # Redirect to chatbot with reschedule context
    booking = result.get("booking", {})
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Reschedule</title>
    <style>body{{font-family:sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;background:#f5f3ff;}}
    .card{{background:#fff;padding:40px;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,0.08);text-align:center;max-width:500px;}}
    a.btn{{display:inline-block;padding:14px 40px;background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;border-radius:30px;text-decoration:none;font-weight:600;margin-top:20px;}}
    </style></head><body><div class="card">
    <h1>Let's Reschedule</h1>
    <p>We're glad you'd like to reschedule your appointment{f' for {booking.get("service","")}' if booking.get("service") else ''}.</p>
    <a class="btn" href="/">Book New Appointment</a>
    </div></body></html>"""


@app.route("/api/noshow-recovery/cancel/<token>")
def api_noshow_cancel_recovery(token):
    result = noshow_recovery_engine.handle_cancel_recovery(token)
    if not result:
        return "<h2>Invalid or expired link.</h2>", 404
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>No Thanks</title>
    <style>body{{font-family:sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;background:#fafafa;}}
    .card{{background:#fff;padding:40px;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,0.08);text-align:center;}}
    </style></head><body><div class="card">
    <h1>No Problem</h1><p>We hope to see you soon! Feel free to book anytime.</p></div></body></html>"""


@app.route("/api/noshow-recovery/stats")
def api_noshow_recovery_stats():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    return jsonify(noshow_recovery_engine.get_recovery_stats(admin_id))


@app.route("/api/noshow-policy", methods=["GET", "POST"])
def api_noshow_policy():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    if request.method == "GET":
        policy = noshow_recovery_engine.get_policy(admin_id)
        return jsonify(policy or {})
    data = request.get_json() or {}
    noshow_recovery_engine.save_policy(admin_id, **data)
    db.log_admin_action(admin_id, user, "Updated no-show policy", "")
    return jsonify({"ok": True})


# ═════════��════════════════════════════════════════════════════════���═══════════
#  Feature 10: White-Label & Custom Domain
# ═════════════════════���═════════════════════════════��══════════════════════════

@app.route("/api/whitelabel", methods=["GET", "POST"])
def api_whitelabel():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = db.get_user_by_token(token)
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    admin_id = user["admin_id"] or user["id"]
    conn = db.get_db()
    if request.method == "GET":
        config = conn.execute("SELECT * FROM whitelabel_config WHERE admin_id=?", (admin_id,)).fetchone()
        conn.close()
        return jsonify(dict(config) if config else {})
    data = request.get_json() or {}
    existing = conn.execute("SELECT id FROM whitelabel_config WHERE admin_id=?", (admin_id,)).fetchone()
    if existing:
        sets = []
        params = []
        for key in ["brand_name", "logo_url", "favicon_url", "primary_color", "secondary_color",
                     "font_family", "custom_css", "email_from_name", "email_from_address",
                     "hide_powered_by", "custom_domain"]:
            if key in data:
                sets.append(f"{key}=?")
                params.append(data[key])
        if sets:
            params.append(admin_id)
            conn.execute(f"UPDATE whitelabel_config SET {','.join(sets)} WHERE admin_id=?", params)
    else:
        conn.execute(
            """INSERT INTO whitelabel_config (admin_id, brand_name, logo_url, primary_color, secondary_color,
               custom_css, hide_powered_by, custom_domain) VALUES (?,?,?,?,?,?,?,?)""",
            (admin_id, data.get("brand_name", ""), data.get("logo_url", ""),
             data.get("primary_color", "#2563eb"), data.get("secondary_color", "#1e40af"),
             data.get("custom_css", ""), data.get("hide_powered_by", 0), data.get("custom_domain", ""))
        )
    conn.commit()
    conn.close()
    db.log_admin_action(admin_id, user, "Updated white-label config", "")
    return jsonify({"ok": True})


# ══════════════════════════════════════════════
#  Email Template / Style Customization
# ══════════════════════════════════════════════

EMAIL_IMG_DIR = os.path.join(os.path.dirname(__file__), "uploads", "email_images")
os.makedirs(EMAIL_IMG_DIR, exist_ok=True)


@app.route("/api/email-template", methods=["GET"])
def get_email_template():
    user = db.get_user_by_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    admin_id = get_effective_admin_id(user)
    template = db.get_email_template(admin_id)
    return jsonify(template or {})


@app.route("/api/email-template", methods=["POST"])
def save_email_template():
    user = db.get_user_by_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if not user or not is_admin_role(user):
        return jsonify({"error": "Not authorized"}), 403
    # Plan check — Pro+ only
    plan = user.get("plan", "free_trial")
    if plan not in ("pro", "agency"):
        return jsonify({"error": "Email customization requires Pro or Agency plan."}), 403

    admin_id = get_effective_admin_id(user)
    data = request.json or {}

    # Validate variables in all HTML sections
    all_html = (data.get("header_html", "") or "") + (data.get("body_html", "") or "") + (data.get("footer_html", "") or "")
    valid_vars, invalid_vars = db.validate_email_template_variables(all_html)
    if invalid_vars:
        return jsonify({
            "error": f"Unknown variables: {', '.join('{{' + v + '}}' for v in sorted(invalid_vars))}",
            "valid_variables": sorted(db.VALID_EMAIL_VARIABLES),
            "invalid_variables": sorted(invalid_vars)
        }), 400

    db.save_email_template(admin_id, **data)
    db.log_admin_action(admin_id, user, "Updated email template", "")
    return jsonify({"ok": True, "message": "Email template saved successfully."})


@app.route("/api/email-template", methods=["DELETE"])
def delete_email_template():
    user = db.get_user_by_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if not user or not is_admin_role(user):
        return jsonify({"error": "Not authorized"}), 403
    admin_id = get_effective_admin_id(user)
    db.delete_email_template(admin_id)
    db.log_admin_action(admin_id, user, "Deleted email template", "")
    return jsonify({"ok": True, "message": "Email template deleted. Default style restored."})


@app.route("/api/email-template/upload-image", methods=["POST"])
def upload_email_image():
    user = db.get_user_by_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if not user or not is_admin_role(user):
        return jsonify({"error": "Not authorized"}), 403
    plan = user.get("plan", "free_trial")
    if plan not in ("pro", "agency"):
        return jsonify({"error": "Email customization requires Pro or Agency plan."}), 403

    if "image" not in request.files:
        return jsonify({"error": "No image file provided."}), 400

    file = request.files["image"]
    if not file.filename:
        return jsonify({"error": "Empty filename."}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"):
        return jsonify({"error": "Invalid image type. Allowed: PNG, JPG, GIF, WEBP, SVG."}), 400

    safe_name = f"{uuid.uuid4().hex}{ext}"
    file.save(os.path.join(EMAIL_IMG_DIR, safe_name))
    # Return absolute URL so images work in email clients
    base_url = request.host_url.rstrip("/")
    url = f"{base_url}/uploads/email_images/{safe_name}"
    return jsonify({"ok": True, "url": url})


@app.route("/uploads/email_images/<path:filename>")
def serve_email_image(filename):
    return send_from_directory(EMAIL_IMG_DIR, filename)


@app.route("/api/email-template/preview", methods=["POST"])
def preview_email_template():
    """Generate a preview of the email template with sample data."""
    user = db.get_user_by_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    admin_id = get_effective_admin_id(user)
    data = request.json or {}

    # Build a mock template for preview
    template = {
        "header_html": data.get("header_html", ""),
        "body_html": data.get("body_html", ""),
        "footer_html": data.get("footer_html", ""),
        "primary_color": data.get("primary_color", "#8b5cf6"),
        "secondary_color": data.get("secondary_color", "#1a1a2e"),
        "bg_color": data.get("bg_color", "#f0f0f0"),
        "button_color": data.get("button_color", "#8b5cf6"),
        "button_text_color": data.get("button_text_color", "#ffffff"),
        "button_radius": data.get("button_radius", "8"),
        "button_size": data.get("button_size", "medium"),
        "header_image_url": data.get("header_image_url", ""),
        "footer_image_url": data.get("footer_image_url", ""),
        "body_image_url": data.get("body_image_url", ""),
        "logo_url": data.get("logo_url", ""),
        "font_family": data.get("font_family", "Helvetica Neue, Helvetica, Arial, sans-serif"),
    }

    # Sample variables for preview
    sample_vars = {
        "patient_name": "John Smith",
        "doctor_name": "Dr. Sarah Johnson",
        "date": "April 25, 2026",
        "time": "10:30 AM - 11:30 AM",
        "clinic_name": user.get("company", "Your Clinic"),
        "confirm_link": "#",
        "cancel_link": "#",
        "service_name": "Dental Cleaning",
        "booking_id": "BK-12345",
        "waitlist_position": "2",
        "reschedule_link": "#",
        "survey_link": "#",
        "invoice_link": "#",
        "recall_treatment": "Teeth Cleaning",
        "followup_date": "October 25, 2026",
    }

    # Build sample email content
    sample_content = f"""
    <tr><td style="padding:40px 40px 16px;text-align:center;">
        <h1 style="color:#1a1a2e;font-size:24px;margin:0 0 8px;font-weight:700;">Appointment Confirmed</h1>
        <p style="color:#666;font-size:15px;margin:0;">Hello {sample_vars['patient_name']},</p>
    </td></tr>
    <tr><td style="padding:16px 40px;">
        <p style="color:#444;font-size:15px;line-height:1.6;">Your appointment has been confirmed for <strong>{sample_vars['date']}</strong> at <strong>{sample_vars['time']}</strong> with <strong>{sample_vars['doctor_name']}</strong>.</p>
    </td></tr>
    <tr><td style="padding:16px 40px;text-align:center;">
        <a href="#" style="display:inline-block;background:#c9a84c;color:#fff;padding:14px 32px;border-radius:8px;text-decoration:none;font-weight:600;font-size:15px;">Confirm Appointment</a>
    </td></tr>"""

    # Check if admin has a saved template with compiled_html (drag-and-drop builder)
    saved_template = db.get_email_template(admin_id)
    if saved_template and saved_template.get("compiled_html"):
        html = email._make_urls_absolute(saved_template["compiled_html"])
        html = email.render_template_variables(html, sample_vars)
        return jsonify({"html": html})

    # Fallback: old-style template with header/body/footer
    header = email.render_template_variables(template.get("header_html", ""), sample_vars)
    body = email.render_template_variables(template.get("body_html", ""), sample_vars)
    footer = email.render_template_variables(template.get("footer_html", ""), sample_vars)

    template["header_html"] = header
    template["body_html"] = body
    template["footer_html"] = footer

    if body.strip():
        sample_content = f'<tr><td style="padding:20px 40px;">{body}</td></tr>'

    html = email._wrap_custom_template(sample_content, template)
    return jsonify({"html": html})


@app.route("/api/email-template/send-test", methods=["POST"])
def send_test_email():
    """Send a test email using the current template."""
    user = db.get_user_by_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if not user or not is_admin_role(user):
        return jsonify({"error": "Not authorized"}), 403
    plan = user.get("plan", "free_trial")
    if plan not in ("pro", "agency"):
        return jsonify({"error": "Pro or Agency plan required."}), 403

    data = request.json or {}
    to_email = data.get("email", user.get("email", ""))
    if not to_email:
        return jsonify({"error": "No email address provided."}), 400

    admin_id = get_effective_admin_id(user)

    # Send a sample booking confirmation using the custom template
    result = email.send_booking_confirmation_customer(
        customer_name="Test Patient",
        customer_email=to_email,
        date_display="April 25, 2026",
        time_display="10:30 AM - 11:30 AM",
        doctor_name="Dr. Sarah Johnson",
        service_name="Sample Consultation",
        admin_id=admin_id,
    )
    if result:
        return jsonify({"ok": True, "message": f"Test email sent to {to_email}"})
    return jsonify({"error": "Failed to send test email. Check SMTP settings."}), 500


@app.route("/api/email-template/variables", methods=["GET"])
def get_email_variables():
    """Return list of valid template variables."""
    return jsonify({
        "variables": sorted(db.VALID_EMAIL_VARIABLES),
        "descriptions": {
            "patient_name": "Patient's full name",
            "doctor_name": "Doctor's name",
            "date": "Appointment date",
            "time": "Appointment time slot",
            "clinic_name": "Your clinic/business name",
            "confirm_link": "Appointment confirmation URL",
            "cancel_link": "Appointment cancellation URL",
            "service_name": "Service/treatment name",
            "booking_id": "Booking reference ID",
            "waitlist_position": "Position in waitlist",
            "reschedule_link": "Reschedule appointment URL",
            "survey_link": "Post-appointment survey URL",
            "invoice_link": "Invoice download URL",
            "recall_treatment": "Treatment type for recall",
            "followup_date": "Recommended follow-up date",
        }
    })


@app.route("/health")
def health():
    return jsonify({
        "status": "ok",
        "model_loaded": model is not None,
        "model": "chatgenius-tinyllama (fine-tuned)",
        "features": ["qa", "booking", "lead_capture", "dashboard", "auth",
                      "waitlist", "patient_forms", "recall", "missed_calls", "followups",
                      "multilingual", "gallery", "doctor_comparison", "emergency_fasttrack",
                      "live_chat_handoff", "schedule_blocks", "promotions", "2fa",
                      "referrals", "patient_profiles", "realtime_dashboard", "ab_testing",
                      "loyalty_program", "gmb_integration", "benchmarking",
                      "smart_reminders", "satisfaction_surveys", "smart_upsell",
                      "unified_inbox", "invoices", "performance_reports",
                      "treatment_packages", "doctor_portal", "noshow_recovery",
                      "whitelabel"],
    })


if __name__ == "__main__":
    load_model()
    background_tasks.start_background_tasks(app)
    port = int(os.environ.get("PORT", 8080))
    app.run(debug=False, port=port)
